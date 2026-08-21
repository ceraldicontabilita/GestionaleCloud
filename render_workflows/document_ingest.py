"""Anteprima incrementale e non distruttiva dell'ingresso documentale Drive."""

from __future__ import annotations

import hashlib
import io
import json
import os
import re
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import PurePosixPath
from typing import Any, Iterable
from urllib.parse import quote


SUPPORTED_EXTENSIONS = {
    ".pdf", ".xlsx", ".xls", ".csv", ".xml", ".p7m", ".eml",
}
LIFECYCLE_FOLDER_NAMES = {
    "REVIEW": "90 - DA ELABORARE",
    "ERROR": "98 - ERRORI",
    "DONE": "99 - ELABORATE",
}
LIFECYCLE_ENV_VARS = {
    "REVIEW": "GOOGLE_DRIVE_REVIEW_FOLDER_ID",
    "ERROR": "GOOGLE_DRIVE_ERROR_FOLDER_ID",
    "DONE": "GOOGLE_DRIVE_PROCESSED_FOLDER_ID",
}
CLASSIFICATION_RULES = (
    ("corrispettivo", ("dati rt", "dataorarilevazione", "codicefiscaleesercente", "pivaesercente")),
    ("cedolino", ("cedolino", "busta paga", "libro unico del lavoro", "netto del mese")),
    ("f24", ("modello f24", "f24 semplificato", "f24 ordinario", "delega irrevocabile")),
    ("quietanza_f24", ("quietanza f24", "ricevuta f24", "esito delega")),
    ("dichiarazione_fiscale", ("dichiarazione iva", "modello 770", "redditi sc", "modello irap", "lipe")),
    ("estratto_conto", ("estratto conto", "lista movimenti", "saldo contabile", "saldo disponibile")),
    ("bonifico", ("bonifico", "cro", "trn", "ordinante", "beneficiario")),
    ("cartella_esattoriale", ("cartella di pagamento", "agenzia entrate riscossione", "ader")),
    ("avviso", ("avviso bonario", "avviso di accertamento", "avviso pagopa")),
    ("fattura", ("fattura elettronica", "fatturaelettronica", "fattura", "nota di credito")),
    ("verbale", ("verbale", "violazione", "codice della strada")),
)

# Questa tabella non crea un secondo importatore: indica quale ingresso gia'
# esistente del Gestionale deve ricevere il documento dopo la conferma umana.
# ``CANONICAL_IMPORT_READY`` significa che upload-auto possiede gia' un parser
# specialistico e la deduplica di dominio. ``REVIEW_REQUIRED`` conserva invece
# il documento senza trasformarlo in un fatto contabile ambiguo.
ROUTING_REGISTRY = {
    "corrispettivo": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> corrispettivi",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "fattura": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> fatture XML",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "f24": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> F24 canonico",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "quietanza_f24": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> quietanze F24",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "cedolino": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> Libro Unico",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "estratto_conto": {
        "consumer": "documents_inbox -> /api/documenti/upload-auto -> movimenti bancari",
        "readiness": "CANONICAL_IMPORT_READY",
    },
    "dichiarazione_fiscale": {
        "consumer": "documents_inbox -> registro dichiarazioni fiscali",
        "readiness": "REVIEW_REQUIRED",
    },
    "bonifico": {
        "consumer": "documents_inbox -> archivio bonifici/distinte",
        "readiness": "REVIEW_REQUIRED",
    },
    "cartella_esattoriale": {
        "consumer": "documents_inbox -> atti amministrativi/AdeR",
        "readiness": "REVIEW_REQUIRED",
    },
    "avviso": {
        "consumer": "documents_inbox -> atti amministrativi/PagoPA",
        "readiness": "REVIEW_REQUIRED",
    },
    "verbale": {
        "consumer": "documents_inbox -> verbali/PagoPA",
        "readiness": "REVIEW_REQUIRED",
    },
}


def route_for(document_type: str) -> dict[str, str]:
    return dict(ROUTING_REGISTRY.get(document_type, {
        "consumer": "documents_inbox -> classificazione manuale",
        "readiness": "REVIEW_REQUIRED",
    }))


def _safe_member(name: str) -> str:
    normalized = name.replace("\\", "/")
    parts = PurePosixPath(normalized).parts
    if normalized.startswith("/") or ".." in parts:
        raise ValueError("membro ZIP non sicuro")
    return normalized


def iter_supported_documents(name: str, content: bytes) -> Iterable[tuple[str, bytes]]:
    """Espande uno ZIP in memoria e restituisce solo formati documentali ammessi."""
    suffix = PurePosixPath(name).suffix.lower()
    if suffix != ".zip":
        if suffix in SUPPORTED_EXTENSIONS:
            yield name, content
        return
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        infos = archive.infolist()
        total_uncompressed = sum(info.file_size for info in infos if not info.is_dir())
        if len(infos) > 20_000 or total_uncompressed > 2 * 1024 * 1024 * 1024:
            raise ValueError("archivio ZIP oltre i limiti di sicurezza")
        for info in infos:
            member = _safe_member(info.filename)
            if info.is_dir() or info.flag_bits & 1:
                continue
            if PurePosixPath(member).suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            yield member, archive.read(info)


def _pdf_text(content: bytes) -> str:
    if not content.startswith(b"%PDF"):
        return ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            return "\n".join((page.extract_text() or "") for page in pdf.pages[:2])
    except Exception:
        return ""


def classify_document(name: str, content: bytes) -> dict[str, Any]:
    """Classifica deterministicamente; il dubbio resta in revisione."""
    extension = PurePosixPath(name).suffix.lower()
    searchable = re.sub(r"[_\-]+", " ", name).casefold()
    if extension == ".pdf":
        searchable += " " + _pdf_text(content).casefold()
    elif extension in {".xml", ".p7m", ".eml", ".csv"}:
        searchable += " " + content[:256 * 1024].decode("utf-8", errors="ignore").casefold()
    matches = []
    for document_type, markers in CLASSIFICATION_RULES:
        score = sum(marker in searchable for marker in markers)
        if score:
            matches.append((score, document_type))
    matches.sort(reverse=True)
    if not matches:
        fallback = {
            ".xlsx": "foglio_elettronico", ".xls": "foglio_elettronico",
            ".csv": "dati_tabellari", ".xml": "documento_xml",
            ".p7m": "documento_firmato", ".eml": "messaggio_email",
        }.get(extension)
        result = {
            "document_type": fallback or "documento_non_classificato",
            "status": "DA_VERIFICARE", "confidence": 0.5 if fallback else 0.0,
        }
        result.update(route_for(result["document_type"]))
        return result
    best_score, best_type = matches[0]
    tied = len(matches) > 1 and matches[1][0] == best_score
    result = {
        "document_type": best_type,
        "status": "DA_VERIFICARE" if tied else "CLASSIFICATO",
        "confidence": 0.6 if tied else min(0.7 + best_score * 0.1, 0.99),
    }
    result.update(route_for(best_type))
    return result


def index_hashes_from_xlsx(content: bytes) -> set[str]:
    """Legge gli SHA-256 dell'indice canonico senza importare dati di dominio."""
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if "DOCUMENTI" not in workbook.sheetnames:
        raise ValueError("foglio DOCUMENTI assente dall'indice canonico")
    rows = workbook["DOCUMENTI"].iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("indice documentale vuoto") from exc
    if "SHA-256" not in headers:
        raise ValueError("colonna SHA-256 assente dall'indice canonico")
    index = headers.index("SHA-256")
    hashes = set()
    for row in rows:
        value = str(row[index] or "").strip().lower() if index < len(row) else ""
        if re.fullmatch(r"[0-9a-f]{64}", value):
            hashes.add(value)
    return hashes


def _drive_service(*, write: bool = False):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    credentials = service_account.Credentials.from_service_account_info(
        json.loads(os.environ["GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON"]),
        scopes=[
            "https://www.googleapis.com/auth/drive"
            if write else "https://www.googleapis.com/auth/drive.readonly"
        ],
    )
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _folder_metadata(service, folder_id: str) -> dict[str, Any]:
    return service.files().get(
        fileId=folder_id,
        fields=("id,name,mimeType,parents,appProperties,"
                "capabilities(canEdit,canMoveItemWithinDrive)"),
        supportsAllDrives=True,
    ).execute()


def _resolve_lifecycle_folders(service, inbox_id: str) -> dict[str, str]:
    """Trova le tre destinazioni nello stesso Calderone dell'inbox."""
    inbox = _folder_metadata(service, inbox_id)
    parents = inbox.get("parents") or []
    if len(parents) != 1:
        raise RuntimeError("01 - IN ARRIVO deve avere un solo genitore Calderone")
    root_id = parents[0]
    resolved: dict[str, str] = {}
    for outcome, folder_name in LIFECYCLE_FOLDER_NAMES.items():
        configured = os.environ.get(LIFECYCLE_ENV_VARS[outcome], "").strip()
        if configured:
            metadata = _folder_metadata(service, configured)
            if metadata.get("mimeType") != "application/vnd.google-apps.folder":
                raise RuntimeError(f"destinazione non e una cartella: {folder_name}")
            if metadata.get("name") != folder_name:
                raise RuntimeError(f"nome destinazione non conforme: {folder_name}")
            if root_id not in (metadata.get("parents") or []):
                raise RuntimeError(f"{folder_name} non appartiene allo stesso Calderone")
            resolved[outcome] = configured
            continue
        escaped = folder_name.replace("'", "\\'")
        response = service.files().list(
            q=(
                f"'{root_id}' in parents and trashed=false and "
                "mimeType='application/vnd.google-apps.folder' and "
                f"name='{escaped}'"
            ),
            fields="files(id,name,parents)", pageSize=10,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        matches = response.get("files", [])
        if len(matches) != 1:
            raise RuntimeError(f"cartella univoca non trovata: {folder_name}")
        resolved[outcome] = matches[0]["id"]
    if len(set(resolved.values()) | {inbox_id}) != 4:
        raise RuntimeError("cartelle lifecycle non distinte")
    return resolved


def lifecycle_destination(outcomes: Iterable[str], *, complete: bool) -> str | None:
    """Determina la destinazione del contenitore senza dividerne il contenuto."""
    values = list(outcomes)
    if not complete:
        return None
    if any(value == "ERROR" for value in values):
        return "ERROR"
    if not values or any(value == "REVIEW" for value in values):
        return "REVIEW"
    return "DONE"


def _move_source_atomically(
    service, *, source: dict[str, Any], inbox_id: str,
    destination_id: str, state: str, source_sha256: str,
) -> None:
    """Aggiorna stato e parent in una singola richiesta Drive e verifica l'esito."""
    metadata = source if source.get("capabilities") else _folder_metadata(service, source["id"])
    capabilities = metadata.get("capabilities") or {}
    if not capabilities.get("canEdit") or capabilities.get("canMoveItemWithinDrive") is False:
        raise PermissionError("service account senza permesso di modifica sul file")
    checked_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    properties = dict(metadata.get("appProperties") or {})
    properties.update({
        "gestionale_state": state,
        "gestionale_checked_at": checked_at,
    })
    if re.fullmatch(r"[0-9a-f]{64}", source_sha256):
        properties["gestionale_source_sha256"] = source_sha256
    else:
        properties["gestionale_source_hash_status"] = "unavailable"
    updated = service.files().update(
        fileId=source["id"],
        addParents=destination_id,
        removeParents=inbox_id,
        body={"appProperties": properties},
        fields="id,parents,appProperties",
        supportsAllDrives=True,
    ).execute()
    parents = set(updated.get("parents") or [])
    if destination_id not in parents or inbox_id in parents:
        raise RuntimeError("spostamento Drive non verificato")


def _list_inbox_sources(service, inbox_id: str) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    page_token = None
    while True:
        response = service.files().list(
            q=f"'{inbox_id}' in parents and trashed=false",
            fields=(
                "nextPageToken,files(id,name,mimeType,size,modifiedTime,md5Checksum,"
                "parents,createdTime,appProperties,"
                "capabilities(canEdit,canMoveItemWithinDrive))"
            ),
            orderBy="createdTime desc,name",
            pageSize=1000, pageToken=page_token,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        sources.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            return sources


def _canonical_index_hashes(service) -> set[str]:
    index_file_id = os.environ.get("GOOGLE_DRIVE_DOCUMENT_INDEX_FILE_ID", "").strip()
    if not index_file_id:
        raise RuntimeError("GOOGLE_DRIVE_DOCUMENT_INDEX_FILE_ID non configurato")
    index_content = service.files().get_media(fileId=index_file_id).execute()
    existing_hashes = index_hashes_from_xlsx(index_content)
    if not existing_hashes:
        raise RuntimeError("indice canonico privo di SHA-256: confronto non affidabile")
    return existing_hashes


def _ingest_configuration(confirm: bool) -> tuple[str, str]:
    if confirm is not True:
        raise RuntimeError("ingest bloccato: avviare il task con confirm=true")
    enabled = os.environ.get("ENABLE_RENDER_CANONICAL_INGEST", "").strip().casefold()
    if enabled not in {"1", "true", "yes", "on"}:
        raise RuntimeError("ingest bloccato: ENABLE_RENDER_CANONICAL_INGEST non attivo")
    base_url = os.environ.get(
        "GESTIONALE_CANONICAL_BASE_URL", "https://impresasemplice.online"
    ).strip().rstrip("/")
    secret = os.environ.get("RENDER_INGEST_SHARED_SECRET", "").strip()
    if not base_url.startswith("https://"):
        raise RuntimeError("GESTIONALE_CANONICAL_BASE_URL deve usare HTTPS")
    if len(secret) < 32:
        raise RuntimeError("RENDER_INGEST_SHARED_SECRET assente o troppo corto")
    _drive_move_configuration(confirm)
    return base_url, secret


def _drive_move_configuration(confirm_move: bool) -> None:
    if confirm_move is not True:
        raise RuntimeError("spostamento bloccato: avviare il task con confirm_move=true")
    enabled = os.environ.get("ENABLE_RENDER_DRIVE_MOVES", "").strip().casefold()
    if enabled not in {"1", "true", "yes", "on"}:
        raise RuntimeError("spostamento bloccato: ENABLE_RENDER_DRIVE_MOVES non attivo")


def scan_document_inbox_preview(max_documents: int = 20_000) -> dict[str, Any]:
    """Confronta prima con l'indice; classifica soltanto i documenti nuovi."""
    if not 1 <= int(max_documents) <= 20_000:
        raise ValueError("max_documents deve essere compreso fra 1 e 20000")
    service = _drive_service()
    inbox_id = os.environ["GOOGLE_DRIVE_INBOX_FOLDER_ID"]
    existing_hashes = _canonical_index_hashes(service)
    sources = _list_inbox_sources(service, inbox_id)
    seen = set(existing_hashes)
    classifications: Counter[str] = Counter()
    statuses: Counter[str] = Counter()
    routes: Counter[str] = Counter()
    readiness: Counter[str] = Counter()
    source_count = document_count = duplicate_count = error_count = 0
    for source in sources:
        if document_count >= max_documents:
            break
        suffix = PurePosixPath(source.get("name") or "").suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS | {".zip"}:
            continue
        source_count += 1
        try:
            payload = service.files().get_media(fileId=source["id"]).execute()
            for member, content in iter_supported_documents(source["name"], payload):
                if document_count >= max_documents:
                    break
                document_count += 1
                digest = hashlib.sha256(content).hexdigest()
                if digest in seen:
                    duplicate_count += 1
                    statuses["DUPLICATO_ESATTO"] += 1
                    continue
                seen.add(digest)
                classification = classify_document(member, content)
                classifications[classification["document_type"]] += 1
                statuses[classification["status"]] += 1
                routes[classification["consumer"]] += 1
                readiness[classification["readiness"]] += 1
        except Exception:
            error_count += 1
            statuses["ERRORE"] += 1
    return {
        "mode": "universal_document_preview", "read_only": True,
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "index_sha256_count": len(existing_hashes),
        "source_files": source_count, "documents_seen": document_count,
        "exact_duplicates_existing_or_batch": duplicate_count,
        "new_documents": document_count - duplicate_count,
        "statuses": dict(statuses), "classifications": dict(classifications),
        "routes": dict(routes), "routing_readiness": dict(readiness),
        "source_errors": error_count,
        "writes": 0, "moves": 0, "deletes": 0,
        "next_action": "anteprima e conferma prima dell'ingest canonico",
    }


def ingest_document_inbox(
    *, confirm: bool = False, max_documents: int = 100,
) -> dict[str, Any]:
    """Importa i nuovi documenti e sposta ogni sorgente solo a esito completo."""
    import mimetypes
    import httpx

    base_url, secret = _ingest_configuration(confirm)
    if not 1 <= int(max_documents) <= 1000:
        raise ValueError("max_documents deve essere compreso fra 1 e 1000")

    service = _drive_service(write=True)
    inbox_id = os.environ["GOOGLE_DRIVE_INBOX_FOLDER_ID"]
    folders = _resolve_lifecycle_folders(service, inbox_id)
    canonical_hashes = set(_canonical_index_hashes(service))
    batch_outcomes: dict[str, str] = {}
    sources = _list_inbox_sources(service, inbox_id)
    stats: Counter[str] = Counter()
    examined = sent_to_preview = moves = 0
    headers = {"X-Render-Ingest-Token": secret}

    with httpx.Client(timeout=300.0, follow_redirects=False) as client:
        for source in sources:
            if examined >= max_documents:
                stats["LIMIT_REACHED"] += 1
                break
            suffix = PurePosixPath(source.get("name") or "").suffix.lower()
            outcomes: list[str] = []
            complete = True
            source_sha256 = source.get("md5Checksum") or "unavailable"
            try:
                payload = service.files().get_media(fileId=source["id"]).execute()
                source_sha256 = hashlib.sha256(payload).hexdigest()
                members = (
                    list(iter_supported_documents(source["name"], payload))
                    if suffix in SUPPORTED_EXTENSIONS | {".zip"} else []
                )
                if not members:
                    outcomes.append("REVIEW")
                    stats["DA_VERIFICARE"] += 1
                for member, content in members:
                    if examined >= max_documents:
                        stats["LIMIT_REACHED"] += 1
                        complete = False
                        break
                    examined += 1
                    digest = hashlib.sha256(content).hexdigest()
                    if digest in canonical_hashes:
                        outcomes.append("DONE")
                        stats["DUPLICATO_INDICE"] += 1
                        continue
                    if digest in batch_outcomes:
                        outcomes.append(batch_outcomes[digest])
                        stats["DUPLICATO_BATCH"] += 1
                        continue
                    classification = classify_document(member, content)
                    if (
                        classification["status"] != "CLASSIFICATO"
                        or classification["readiness"] != "CANONICAL_IMPORT_READY"
                    ):
                        outcomes.append("REVIEW")
                        batch_outcomes[digest] = "REVIEW"
                        stats["DA_VERIFICARE"] += 1
                        continue

                    sent_to_preview += 1
                    safe_name = PurePosixPath(member).name or "documento"
                    media_type = mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
                    files = {"file": (safe_name, content, media_type)}
                    try:
                        preview = client.post(
                            f"{base_url}/api/documenti/upload-auto/render/preview",
                            headers=headers, files=files,
                        )
                        preview.raise_for_status()
                        preview_data = preview.json()
                        if preview_data.get("duplicate"):
                            outcome = "DONE"
                            stats["DUPLICATO_GESTIONALE"] += 1
                        else:
                            token = preview_data.get("confirmation_token")
                            if not preview_data.get("success") or not token:
                                outcome = "REVIEW"
                                stats["ANTEPRIMA_BLOCCATA"] += 1
                            else:
                                upload_headers = {
                                    **headers,
                                    "X-Document-Preview-Token": token,
                                    "X-Source-Drive-File-ID": source["id"],
                                    "X-Source-Drive-Parent-ID": inbox_id,
                                    "X-Source-SHA256": source_sha256,
                                    "X-Source-Archive-Member": quote(member, safe=""),
                                }
                                uploaded = client.post(
                                    f"{base_url}/api/documenti/upload-auto/render",
                                    headers=upload_headers, files=files,
                                )
                                uploaded.raise_for_status()
                                upload_data = uploaded.json()
                                if upload_data.get("duplicate"):
                                    outcome = "DONE"
                                    stats["DUPLICATO_GESTIONALE"] += 1
                                elif upload_data.get("success"):
                                    outcome = "DONE"
                                    canonical_hashes.add(digest)
                                    stats["IMPORTATO"] += 1
                                else:
                                    outcome = "ERROR"
                                    stats["IMPORT_NON_RIUSCITO"] += 1
                    except Exception:
                        outcome = "ERROR"
                        stats["ERRORE_DOCUMENTO"] += 1
                    outcomes.append(outcome)
                    batch_outcomes[digest] = outcome
            except Exception:
                outcomes.append("ERROR")
                stats["ERRORE_SORGENTE"] += 1

            destination = lifecycle_destination(outcomes, complete=complete)
            if destination:
                try:
                    _move_source_atomically(
                        service, source=source, inbox_id=inbox_id,
                        destination_id=folders[destination], state=destination,
                        source_sha256=source_sha256,
                    )
                    moves += 1
                    stats[f"SPOSTATO_{destination}"] += 1
                except Exception:
                    stats["SPOSTAMENTO_FALLITO"] += 1
            if not complete:
                break

    return {
        "mode": "canonical_document_ingest",
        "confirmed": True,
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "documents_examined": examined,
        "documents_sent_to_preview": sent_to_preview,
        "results": dict(stats),
        "moves": moves,
        "deletes": 0,
    }


def lifecycle_preflight() -> dict[str, Any]:
    """Verifica in sola lettura cartelle e capacità senza spostare file."""
    service = _drive_service(write=True)
    inbox_id = os.environ["GOOGLE_DRIVE_INBOX_FOLDER_ID"]
    folders = _resolve_lifecycle_folders(service, inbox_id)
    sources = _list_inbox_sources(service, inbox_id)
    editable = sum(bool((source.get("capabilities") or {}).get("canEdit")) for source in sources)
    return {
        "mode": "lifecycle_preflight", "checked_at": datetime.now(timezone.utc).isoformat(),
        "inbox_source_files": len(sources), "editable_source_files": editable,
        "all_sources_editable": editable == len(sources),
        "folders": {state: {"id": folder_id, "name": LIFECYCLE_FOLDER_NAMES[state]}
                    for state, folder_id in folders.items()},
        "writes": 0, "moves": 0, "deletes": 0,
    }


def reconcile_document_lifecycle(
    *, confirm_move: bool = False, max_sources: int = 100,
) -> dict[str, Any]:
    """Sposta il pregresso già indicizzato senza ritrasmettere documenti."""
    _drive_move_configuration(confirm_move)
    if not 1 <= int(max_sources) <= 1000:
        raise ValueError("max_sources deve essere compreso fra 1 e 1000")
    service = _drive_service(write=True)
    inbox_id = os.environ["GOOGLE_DRIVE_INBOX_FOLDER_ID"]
    folders = _resolve_lifecycle_folders(service, inbox_id)
    canonical_hashes = _canonical_index_hashes(service)
    sources = _list_inbox_sources(service, inbox_id)
    stats: Counter[str] = Counter()
    moved = 0
    for source in sources[:max_sources]:
        outcomes: list[str] = []
        source_sha256 = source.get("md5Checksum") or "unavailable"
        try:
            payload = service.files().get_media(fileId=source["id"]).execute()
            source_sha256 = hashlib.sha256(payload).hexdigest()
            suffix = PurePosixPath(source.get("name") or "").suffix.lower()
            members = (
                list(iter_supported_documents(source["name"], payload))
                if suffix in SUPPORTED_EXTENSIONS | {".zip"} else []
            )
            if not members:
                outcomes.append("REVIEW")
            for member, content in members:
                digest = hashlib.sha256(content).hexdigest()
                if digest in canonical_hashes:
                    outcomes.append("DONE")
                    continue
                classification = classify_document(member, content)
                outcomes.append("REVIEW" if (
                    classification["status"] != "CLASSIFICATO"
                    or classification["readiness"] == "REVIEW_REQUIRED"
                ) else "PENDING")
        except Exception:
            outcomes.append("ERROR")
        if "PENDING" in outcomes:
            stats["LASCIATO_IN_ARRIVO"] += 1
            continue
        destination = lifecycle_destination(outcomes, complete=True)
        try:
            _move_source_atomically(
                service, source=source, inbox_id=inbox_id,
                destination_id=folders[destination], state=destination,
                source_sha256=source_sha256,
            )
            moved += 1
            stats[f"SPOSTATO_{destination}"] += 1
        except Exception:
            stats["SPOSTAMENTO_FALLITO"] += 1
    return {
        "mode": "lifecycle_reconciliation", "confirmed": True,
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "source_files_examined": min(len(sources), max_sources),
        "results": dict(stats), "moves": moved, "deletes": 0,
    }
