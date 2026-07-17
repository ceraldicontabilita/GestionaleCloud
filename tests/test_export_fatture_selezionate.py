"""Richiesta utente 16/07/2026: selezionare le fatture in Archivio e
scaricarle in PDF ed Excel. POST /api/fatture-ricevute/export-selezione
con {ids, formato}: Excel con riga totali, PDF prospetto; le fatture
eliminate (soft-delete) non entrano mai nell'export."""
import asyncio
import zipfile
from io import BytesIO

import pytest
from fastapi import HTTPException

from app.routers.fatture_module import export_selezione as mod


def _run(c):
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(c)
    finally:
        loop.close()


class _Cursor:
    def __init__(self, docs):
        self._docs = docs

    async def to_list(self, n=None):
        return list(self._docs[:n] if n else self._docs)


class _Coll:
    def __init__(self, docs=None):
        self.docs = docs or []

    def find(self, query=None, projection=None, *a, **k):
        ids = set(query["id"]["$in"])
        out = [
            d for d in self.docs
            if d.get("id") in ids
            and d.get("entity_status") != "deleted"
            and d.get("status") != "deleted"
        ]
        return _Cursor(out)


class _Db:
    def __init__(self):
        self.collections = {}

    def __getitem__(self, name):
        return self.collections.setdefault(name, _Coll())


def _setup(monkeypatch):
    db = _Db()
    monkeypatch.setattr(mod.Database, "get_db", staticmethod(lambda: db))
    db["invoices"].docs = [
        {"id": "f1", "invoice_number": "10/A", "invoice_date": "2026-03-01",
         "supplier_name": "Dolciaria Acquaviva S.p.A.", "supplier_vat": "01234567890",
         "total_amount": 122.0, "taxable_amount": 100.0, "vat_amount": 22.0,
         "pagato": True, "payment_method": "banca"},
        {"id": "f2", "invoice_number": "11/A", "invoice_date": "2026-03-05",
         "supplier_name": "Enoteca Dante S.r.l.", "supplier_vat": "09876543210",
         "total_amount": 244.0, "taxable_amount": 200.0, "vat_amount": 44.0},
        {"id": "f-del", "invoice_number": "12/A", "invoice_date": "2026-03-06",
         "supplier_name": "Eliminata", "total_amount": 999.0, "status": "deleted"},
    ]
    return db


def test_excel_contiene_le_fatture_e_i_totali(monkeypatch):
    _setup(monkeypatch)

    resp = _run(mod.export_fatture_selezionate({"ids": ["f1", "f2", "f-del"], "formato": "excel"}))

    assert "spreadsheetml" in resp.media_type
    contenuto = BytesIO(b"".join(_run(_leggi_stream(resp))))
    from openpyxl import load_workbook
    wb = load_workbook(contenuto)
    ws = wb.active
    righe = list(ws.iter_rows(values_only=True))
    assert righe[0][:3] == ("Data", "Numero", "Fornitore")
    assert len(righe) == 4  # header + 2 fatture (la eliminata NO) + totali
    assert righe[1][1] == "10/A"
    assert righe[3][0] == "TOTALE"
    assert righe[3][6] == 366.0  # 122 + 244


def test_pdf_generato_con_media_type_corretto(monkeypatch):
    _setup(monkeypatch)

    resp = _run(mod.export_fatture_selezionate({"ids": ["f1", "f2"], "formato": "pdf"}))

    assert resp.media_type == "application/pdf"
    contenuto = b"".join(_run(_leggi_stream(resp)))
    assert contenuto.startswith(b"%PDF")


def test_selezione_vuota_o_formato_invalido(monkeypatch):
    _setup(monkeypatch)

    with pytest.raises(HTTPException) as e1:
        _run(mod.export_fatture_selezionate({"ids": [], "formato": "excel"}))
    assert e1.value.status_code == 400

    with pytest.raises(HTTPException) as e2:
        _run(mod.export_fatture_selezionate({"ids": ["f1"], "formato": "csv"}))
    assert e2.value.status_code == 400

    with pytest.raises(HTTPException) as e3:
        _run(mod.export_fatture_selezionate({"ids": ["inesistente"], "formato": "excel"}))
    assert e3.value.status_code == 404


async def _leggi_stream(resp):
    out = []
    async for chunk in resp.body_iterator:
        out.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return out
