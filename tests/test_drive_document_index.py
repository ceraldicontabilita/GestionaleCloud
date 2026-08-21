import io

import pytest
from openpyxl import Workbook

from app.services.drive_document_index import (
    DECLARATION_HEADERS,
    DUPLICATE_HEADERS,
    F24_HEADERS,
    FOLDER_MIME,
    REQUIRED_HEADERS,
    _discover_index_file_sync,
    _parse_index_xlsx,
    _resolve_path_sync,
    list_documented_tax_payments,
    list_declarations,
    list_f24_rows,
    search_records,
    validate_relations,
)


class _Call:
    def __init__(self, value):
        self.value = value

    def execute(self):
        return self.value


class _Files:
    def __init__(self, children, metadata):
        self.children = children
        self.metadata = metadata

    def list(self, q, **kwargs):
        parent = q.split("'")[1]
        return _Call({"files": self.children.get(parent, [])})

    def get(self, fileId, **kwargs):
        return _Call(self.metadata[fileId])


class _Service:
    def __init__(self, children, metadata):
        self._files = _Files(children, metadata)

    def files(self):
        return self._files


def _xlsx(records, f24_rows=None, declarations=None, duplicates=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DOCUMENTI"
    sheet.append(REQUIRED_HEADERS)
    for record in records:
        sheet.append([record.get(header) for header in REQUIRED_HEADERS])
    for name, headers, values in (
        ("F24_RIGHE", F24_HEADERS, f24_rows or []),
        ("DICHIARAZIONI", DECLARATION_HEADERS, declarations or []),
        ("DUPLICATI_SCARTI", DUPLICATE_HEADERS, duplicates or []),
    ):
        child = workbook.create_sheet(name)
        child.append(headers)
        for record in values:
            child.append([record.get(header) for header in headers])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_discovers_one_exact_index_and_fails_closed_on_duplicates():
    root = {"id": "root", "name": "GESTIONALE", "mimeType": FOLDER_MIME, "trashed": False}
    folder = {"id": "indexes", "name": "INDICI GESTIONALE", "mimeType": FOLDER_MIME}
    index = {"id": "xlsx", "name": "INDICE_DOCUMENTALE_DRIVE.xlsx", "mimeType": "application/xlsx"}
    metadata = {"root": root, "xlsx": {**index, "trashed": False, "webViewLink": "https://drive/xlsx"}}
    service = _Service({"root": [folder], "indexes": [index]}, metadata)

    found_folder, found_index = _discover_index_file_sync(service, "root")
    assert found_folder["id"] == "indexes"
    assert found_index["id"] == "xlsx"

    duplicate_service = _Service({"root": [folder, {**folder, "id": "indexes-2"}]}, metadata)
    with pytest.raises(ValueError, match="ambiguo"):
        _discover_index_file_sync(duplicate_service, "root")


def test_parses_and_filters_document_index():
    records = [{
        "ID documento": "DOC-1", "Dominio": "F24", "Categoria": "QUIETANZA",
        "Anno": "2026", "Nome file": "quietanza.pdf", "Estensione": ".pdf",
        "Dimensione byte": 123, "SHA-256": "a" * 64,
        "Percorso Drive": r"F24\2026\quietanza.pdf", "Stato": "CARICATO_UNICO",
    }, {
        "ID documento": "DOC-2", "Dominio": "PARTENOPAY", "Categoria": "AVVISO",
        "Anno": "2025", "Nome file": "avviso.xml", "Estensione": ".xml",
        "Dimensione byte": 456, "SHA-256": "b" * 64,
        "Percorso Drive": r"PARTENOPAY\AVVISI\avviso.xml", "Stato": "CARICATO_UNICO",
    }]
    parsed = _parse_index_xlsx(_xlsx(records))
    result = search_records(parsed, q="quietanza", domain="f24", year="2026", extension="pdf")
    assert len(parsed) == 2
    assert result == [{
        "document_id": "DOC-1", "domain": "F24", "category": "QUIETANZA",
        "year": "2026", "filename": "quietanza.pdf", "extension": ".pdf",
        "size_bytes": 123, "sha256": "a" * 64,
        "drive_path": r"F24\2026\quietanza.pdf", "source_zip": None,
        "source_path": None, "status": "CARICATO_UNICO", "document_number": None,
    }]


def test_rejects_workbook_without_required_headers():
    workbook = Workbook()
    workbook.active.title = "DOCUMENTI"
    workbook.active.append(["ID documento"])
    output = io.BytesIO()
    workbook.save(output)
    with pytest.raises(ValueError, match="Colonne obbligatorie assenti"):
        _parse_index_xlsx(output.getvalue())


def test_resolves_exact_path_and_rejects_ambiguous_file():
    children = {
        "root": [{"id": "f24", "name": "F24", "mimeType": FOLDER_MIME}],
        "f24": [{"id": "year", "name": "2026", "mimeType": FOLDER_MIME}],
        "year": [{"id": "pdf", "name": "quietanza.pdf", "mimeType": "application/pdf"}],
    }
    metadata = {"pdf": {"id": "pdf", "name": "quietanza.pdf", "trashed": False, "webViewLink": "https://drive/pdf"}}
    service = _Service(children, metadata)
    assert _resolve_path_sync(service, "root", r"F24\2026\quietanza.pdf")["id"] == "pdf"

    children["year"].append({"id": "pdf-2", "name": "QUIETANZA.PDF", "mimeType": "application/pdf"})
    with pytest.raises(ValueError, match="ambiguo"):
        _resolve_path_sync(service, "root", r"F24\2026\quietanza.pdf")


def test_boolean_validation_connects_f24_and_declaration_to_document():
    document = {
        "ID documento": "DOC-1", "Nome file": "modello.pdf", "SHA-256": "a" * 64,
        "Percorso Drive": r"F24\2026\modello.pdf",
    }
    catalog = {
        "documents": [document],
        "f24_rows": [{
            "ID documento": "DOC-1", "SHA-256": "a" * 64,
            "Percorso Drive": r"F24\2026\modello.pdf", "Debito": 10, "Credito": 0,
        }],
        "declarations": [{"Percorso archivio": "02_ANNI/2026/DICHIARAZIONI/modello.pdf"}],
        "duplicates": [],
    }
    validation = validate_relations(catalog)
    assert validation["all_true"] is True
    assert all(validation["checks"].values())
    assert validation["counts"]["f24_documents"] == 1


def test_boolean_validation_rejects_broken_relations():
    catalog = {
        "documents": [{
            "ID documento": "DOC-1", "Nome file": "modello.pdf",
            "SHA-256": "a" * 64, "Percorso Drive": r"F24\modello.pdf",
        }],
        "f24_rows": [{
            "ID documento": "DOC-MISSING", "SHA-256": "b" * 64,
            "Percorso Drive": r"F24\altro.pdf", "Debito": -1,
        }],
        "declarations": [{"Percorso archivio": "missing.pdf"}],
        "duplicates": [],
    }
    validation = validate_relations(catalog)
    assert validation["all_true"] is False
    assert validation["checks"]["all_f24_documents_exist"] is False
    assert validation["checks"]["all_declarations_link_exactly_one_document"] is False


def test_f24_rows_are_drive_first_filterable_and_keep_stable_ids(monkeypatch):
    from app.services import drive_document_index as index

    document = {
        "ID documento": "DOC-F24", "Nome file": "F24_2026.pdf",
        "SHA-256": "a" * 64, "Percorso Drive": r"F24\2026\F24_2026.pdf",
    }
    rows = [{
        "ID documento": "DOC-F24", "Anno pagamento": "2026",
        "Data pagamento": "2026-08-20", "Sezione": "ERARIO",
        "Tipo riga": "DEBITO", "Codice tributo": "1704",
        "Periodo tributo": "2026", "Debito": 51.64, "Credito": 0,
        "SHA-256": "a" * 64, "Pagina": 1,
    }, {
        "ID documento": "DOC-F24", "Anno pagamento": "2026",
        "Data pagamento": "2026-08-20", "Sezione": "ERARIO",
        "Tipo riga": "CREDITO", "Codice tributo": "6099",
        "Periodo tributo": "2025", "Debito": 0, "Credito": 10,
        "SHA-256": "a" * 64, "Pagina": 1,
    }, {
        "ID documento": "DOC-F24", "Anno pagamento": "2019",
        "Data pagamento": "31/12/2019", "Sezione": "ERARIO",
        "Tipo riga": "DEBITO", "Codice tributo": "9001",
        "Periodo tributo": "2015", "Debito": 100, "Credito": 0,
        "SHA-256": "a" * 64, "Pagina": 2,
    }]
    catalog = {"documents": [document], "f24_rows": rows, "declarations": [], "duplicates": []}
    monkeypatch.setattr(index, "load_full_catalog", lambda service=None: ({}, catalog))

    all_rows = list_f24_rows()
    credit_rows = list_f24_rows(year="2026", credits_only=True)
    assert all_rows["total"] == 3
    assert all_rows["items"][0]["payment_date"] == "2026-08-20"
    assert credit_rows["total"] == 1
    credit_from_all = next(item for item in all_rows["items"] if item["tax_code"] == "6099")
    assert credit_rows["items"][0]["id"] == credit_from_all["id"]
    assert credit_rows["items"][0]["evidence_state"] == "MODELLO_F24_NON_PROVA_BANCARIA"


def test_documented_tax_payments_only_include_quietanze_and_keep_bank_unverified(monkeypatch):
    from app.services import drive_document_index as index

    documents = [{
        "ID documento": "DOC-Q", "Nome file": "quietanza.pdf",
        "SHA-256": "a" * 64, "Percorso Drive": r"F24\quietanza.pdf",
    }, {
        "ID documento": "DOC-M", "Nome file": "modello.pdf",
        "SHA-256": "b" * 64, "Percorso Drive": r"F24\modello.pdf",
    }]
    rows = [{
        "ID documento": "DOC-Q", "Tipo documento": "Quietanza AE",
        "Codice tributo": "1001", "Debito": 100, "Credito": 0,
    }, {
        "ID documento": "DOC-M", "Tipo documento": "Modello F24",
        "Codice tributo": "6001", "Debito": 50, "Credito": 0,
    }]
    catalog = {"documents": documents, "f24_rows": rows, "declarations": [], "duplicates": []}
    monkeypatch.setattr(index, "load_full_catalog", lambda service=None: ({}, catalog))

    payload = list_documented_tax_payments()

    assert payload["total"] == 1
    assert payload["items"][0]["tax_code"] == "1001"
    assert payload["items"][0]["payment_status"] == "DOCUMENTATO_DA_QUIETANZA"
    assert payload["items"][0]["bank_status"] == "DA_VERIFICARE"


def test_declarations_use_canonical_types_and_verified_document_identity(monkeypatch):
    from app.services import drive_document_index as index

    document = {
        "ID documento": "DOC-770", "Nome file": "770_2026.pdf",
        "SHA-256": "b" * 64, "Percorso Drive": r"DICHIARAZIONI\770_2026.pdf",
    }
    lipe_document = {
        "ID documento": "DOC-LIPE", "Nome file": "LIPE_2026_407141844.pdf",
        "SHA-256": "c" * 64, "Percorso Drive": r"DICHIARAZIONI\LIPE_2026_407141844.pdf",
    }
    catalog = {
        "documents": [document, lipe_document], "f24_rows": [], "duplicates": [],
        "declarations": [{
            "Anno": "2026", "Tipo": "DICHIARAZIONE", "Protocollo": "T26001",
            "Percorso archivio": "02_ANNI/2026/DICHIARAZIONI/770_2026.pdf",
        }, {
            "Anno": "2026", "Tipo": "DICHIARAZIONE", "Protocollo": "",
            "Percorso archivio": "02_ANNI/2026/DICHIARAZIONI/LIPE_2026_407141844.pdf",
        }],
    }
    monkeypatch.setattr(index, "load_full_catalog", lambda service=None: ({}, catalog))

    payload = list_declarations(year="2026", declaration_type="MODELLO_770")
    assert payload["total_matching"] == 1
    item = payload["results"][0]
    assert item["id"] == "DOC-770"
    assert item["document_type"] == "MODELLO_770"
    assert item["filing_year"] == 2026
    assert item["tax_year"] == 2025
    assert item["relation_state"] == "CONFERMATA_NOME_UNIVOCO_E_INDICE_VERIFICATO"

    all_items = list_declarations(year="2026")["results"]
    lipe = next(entry for entry in all_items if entry["document_type"] == "LIPE")
    assert lipe["tax_year"] == 2026
