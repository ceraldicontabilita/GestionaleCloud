"""Genera il bundle idempotente dei ricettari Excel Ceraldi.

Il server di produzione non legge file dal computer dell'utente: questo script
estrae le quattro fonti locali, conserva hash/foglio/riga e produce il JSON
versionato che viene poi importato nel database tramite l'API amministrativa.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "backend" / "data" / "ricettario_excel_ceraldi.json"


def _text(value: Any) -> str:
    return " ".join(str(value or "").replace("\r", "\n").split()).strip()


def _key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _text(value).casefold().replace("’", "'"))
    normalized = "".join(c for c in normalized if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", normalized)).strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = _text(value).replace(" ", "")
    if not raw:
        return None
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if not raw or raw in {"-", ".", ","}:
        return None
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+", raw):
        raw = raw.replace(".", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def _unit(value: Any) -> str:
    raw = _text(value).casefold().strip(". ")
    mapping = {
        "gr": "g", "grammi": "g", "grammo": "g", "lt": "l", "litri": "l",
        "n": "pz", "n°": "pz", "pezzi": "pz", "pezzo": "pz", "uova": "pz",
    }
    return mapping.get(raw, raw)


def _quantity(value: Any, unit: Any = "") -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = _text(value).replace(" ", "")
    normalized_unit = _unit(unit)
    # Nei ricettari storici "Kg. 1.200" significa 1,2 kg; nelle dosi in
    # grammi "1.300 g" significa invece 1300 g.
    if normalized_unit in {"kg", "l", "lt"} and re.fullmatch(r"\d+[.,]\d{3}", raw):
        try:
            return float(raw.replace(",", "."))
        except ValueError:
            return None
    return _number(value)


def _ingredient(name: Any, quantity: Any = None, unit: Any = "") -> dict | None:
    nome = _text(name).strip(" -–—:;")
    if not nome or _key(nome) in {"ingrediente", "ingredienti", "totale", "totale peso"}:
        return None
    result = {"nome": nome, "quantita": _quantity(quantity, unit), "unita_misura": _unit(unit)}
    if result["quantita"] is not None and result["quantita"].is_integer():
        result["quantita"] = int(result["quantita"])
    return result


_ACTION = re.compile(
    r"\b(procedimento|preparazione|mescol|monta|impasta|aggiung|incorpora|cuoc|forno|"
    r"inforna|friggi|boll|stend|versa|lascia|mettete|prendete|sciogli|taglia|servi)\b",
    re.IGNORECASE,
)


def _ingredient_from_line(line: str, previous_unit: str = "") -> dict | None:
    raw = _text(line).strip("•*- ")
    if not raw or raw.endswith(":") or re.fullmatch(r"\d+", raw):
        return None
    # Nome seguito da dose: "Farina 480 g".
    suffix = re.match(
        r"^(.*?)[\s:]+(\d+(?:[.,]\d+)?)\s*(kg|g|gr|mg|ml|cl|l|lt|pz|n°?|pezzi|uova)\.?$",
        raw,
        re.IGNORECASE,
    )
    if suffix:
        return _ingredient(suffix.group(1), suffix.group(2), suffix.group(3))
    # Notazione dei ricettari storici: "Mandorle amare Kg. 1".
    unit_before = re.match(
        r"^(.*?)\s+(kg|g|gr|mg|ml|cl|l|lt|pz|n°?)\.?\s*(\d+(?:[.,]\d+)?)$",
        raw,
        re.IGNORECASE,
    )
    if unit_before:
        return _ingredient(unit_before.group(1), unit_before.group(3), unit_before.group(2))
    # Il simbolo » ripete l'unità della riga precedente.
    ditto = re.match(r"^(.*?)\s+[»\"]\s*(\d+(?:[.,]\d+)?)$", raw)
    if ditto:
        return _ingredient(ditto.group(1), ditto.group(2), previous_unit)
    # Dose seguita dal nome: "500 g farina" / "4 uova intere".
    prefix = re.match(
        r"^(\d+(?:[.,]\d+)?)\s*(kg|g|gr|mg|ml|cl|l|lt|pz|n°?|pezzi|uova|"
        r"bustin[ae]|fial[ae]|cucchiai(?:o|ni)?|cucchiaini?|pizzic[oa])?\s+(.*)$",
        raw,
        re.IGNORECASE,
    )
    if prefix:
        quantita, unita, nome = prefix.groups()
        if not unita:
            unita = "pz"
        elif _key(unita) in {"uova", "bustina", "bustine", "fiala", "fiale"}:
            nome = f"{unita} {nome}"
            unita = "pz"
        return _ingredient(nome, quantita, unita)
    if re.search(r"\bq\.?\s*b\.?$", raw, re.IGNORECASE) and not _ACTION.search(raw):
        return _ingredient(re.sub(r"\bq\.?\s*b\.?$", "", raw, flags=re.IGNORECASE), None, "q.b.")
    return None


def _ingredients_from_text(value: Any) -> list[dict]:
    rows: list[dict] = []
    previous_unit = ""
    for line in str(value or "").replace("\r", "\n").split("\n"):
        item = _ingredient_from_line(line, previous_unit)
        if item:
            rows.append(item)
            if item.get("unita_misura"):
                previous_unit = item["unita_misura"]
    return rows


def _yield_from_text(value: Any) -> int:
    raw = _text(value)
    matches = re.findall(r"(\d+)\s*(?:pz|pezzi|porzioni|zeppole|palline|brioche|cornetti)", raw, re.I)
    return int(matches[0]) if matches else 0


def _source(path: Path, sha: str, sheet: str, row: int | None, kind: str) -> dict:
    return {"file": path.name, "sha256": sha, "sheet": sheet, "row": row, "tipo": kind}


def _record(
    *, name: Any, ingredients: list[dict], procedure: Any, notes: Any, source: dict,
    reparto: str = "", porzioni: int = 0, confidence: int = 1,
) -> dict | None:
    nome = re.sub(r"^\s*\d+[.)]\s*", "", _text(name)).strip(" -–—")
    if not nome or len(nome) < 2:
        return None
    return {
        "nome": nome,
        "chiave": _key(nome),
        "reparto_hint": reparto,
        "porzioni": int(porzioni or 0),
        "ingredienti_dettaglio": ingredients,
        "procedimento_testo": str(procedure or "").strip(),
        "note": str(notes or "").strip(),
        "fonte": source,
        "confidenza": confidence,
    }


def parse_ricettario_completo_207(path: Path, sha: str) -> list[dict]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb["Ricettario"]
    headers = [row for row in range(4, ws.max_row + 1) if isinstance(ws.cell(row, 1).value, (int, float))]
    result = []
    for pos, start in enumerate(headers):
        end = (headers[pos + 1] - 1) if pos + 1 < len(headers) else ws.max_row
        name = ws.cell(start, 2).value
        ingredients, procedure = [], []
        for row in range(start + 1, end + 1):
            label = _text(ws.cell(row, 2).value)
            if not label or _key(label) in {"ingrediente", "ingredienti"}:
                continue
            if label.startswith("📝"):
                procedure.append(label.lstrip("📝 "))
                continue
            item = _ingredient(label, ws.cell(row, 3).value, ws.cell(row, 4).value)
            if item:
                ingredients.append(item)
        rec = _record(
            name=name, ingredients=ingredients, procedure="\n".join(procedure), notes="",
            source=_source(path, sha, ws.title, start, "strutturato_207"),
            porzioni=_yield_from_text(ws.cell(start, 3).value), confidence=5,
        )
        if rec:
            result.append(rec)
    wb.close()
    return result


def _block_rows(ws, marker: str) -> list[int]:
    return [
        row for row in range(1, ws.max_row)
        if _text(ws.cell(row + 1, 1).value).casefold().startswith(marker.casefold())
        and _text(ws.cell(row, 1).value)
    ]


def parse_ricettario_ceraldi(path: Path, sha: str) -> list[dict]:
    wb = load_workbook(path, read_only=False, data_only=True)
    result: list[dict] = []
    for sheet_name in ("Gelati - Basi", "Gelati - Varianti Galatea"):
        ws = wb[sheet_name]
        starts = _block_rows(ws, "Riferimento:")
        for pos, start in enumerate(starts):
            end = (starts[pos + 1] - 1) if pos + 1 < len(starts) else ws.max_row
            ingredients = []
            for row in range(start + 3, end + 1):
                name = _text(ws.cell(row, 1).value)
                if not name or "totale" in _key(name):
                    continue
                item = _ingredient(name, ws.cell(row, 3).value, "g")
                if item:
                    ingredients.append(item)
            rec = _record(
                name=ws.cell(start, 1).value, ingredients=ingredients, procedure="", notes=ws.cell(start + 1, 1).value,
                source=_source(path, sha, sheet_name, start, "ceraldi_gelati"),
                reparto="pasticceria", confidence=6,
            )
            if rec:
                result.append(rec)

    for sheet_name, reparto in (("Dolci (Pasticceria)", "pasticceria"), ("Salate (Rosticceria)", "rosticceria")):
        ws = wb[sheet_name]
        starts = _block_rows(ws, "Ingrediente")
        for pos, start in enumerate(starts):
            end = (starts[pos + 1] - 1) if pos + 1 < len(starts) else ws.max_row
            ingredients, porzioni = [], 0
            for row in range(start + 2, end + 1):
                label = _text(ws.cell(row, 1).value)
                key = _key(label)
                if not label:
                    continue
                if "pezzi ottenuti" in key or "porzioni" in key:
                    porzioni = int(_number(ws.cell(row, 2).value) or 0)
                    continue
                if any(word in key for word in ("totale peso", "peso per pezzo", "costo", "prezzo")):
                    continue
                item = _ingredient(label, ws.cell(row, 2).value, "g")
                if item:
                    ingredients.append(item)
            rec = _record(
                name=ws.cell(start, 1).value, ingredients=ingredients, procedure="", notes="",
                source=_source(path, sha, sheet_name, start, "ceraldi_reparto"),
                reparto=reparto, porzioni=porzioni, confidence=6,
            )
            if rec:
                result.append(rec)
    wb.close()
    return result


def parse_v6(path: Path, sha: str) -> list[dict]:
    # Il JSON esistente è una trascrizione verificata dello stesso file e
    # conserva gli a-capo che rendono affidabile il parsing delle dosi.
    archive_path = ROOT / "backend" / "data" / "archivio_dolce.json"
    archive = json.loads(archive_path.read_text(encoding="utf-8"))
    if archive.get("meta", {}).get("sourceSha256") != sha:
        raise RuntimeError("Ricettario_Completo_v6.xlsx non coincide con archivio_dolce.json")
    result = []
    for item in [*(archive.get("recipes") or []), *(archive.get("components") or [])]:
        provenance = item.get("provenance") or {}
        rec = _record(
            name=item.get("name"),
            ingredients=_ingredients_from_text(item.get("ingredients")),
            procedure=item.get("procedure"), notes=item.get("notes"),
            source=_source(path, sha, provenance.get("sheet") or "🍰 Ricette", provenance.get("row"), "professionale_v6"),
            reparto="pasticceria", confidence=4,
        )
        if rec:
            result.append(rec)
    return result


def parse_schede_libere(path: Path, sha: str) -> list[dict]:
    wb = load_workbook(path, read_only=False, data_only=True)
    result = []
    for ws in wb.worksheets:
        lines = []
        for row in ws.iter_rows():
            for cell in row:
                value = _text(cell.value)
                if value:
                    lines.append((cell.row, value))
        if not lines:
            continue
        first_row, first = lines[0]
        name = ws.title.strip()
        if not _ingredient_from_line(first) and not _ACTION.search(first):
            name = first
            body = lines[1:]
        else:
            body = lines
        ingredients, procedure, notes = [], [], []
        in_procedure = False
        for row_number, line in body:
            key = _key(line)
            if key in {"procedimento", "preparazione"}:
                in_procedure = True
                continue
            item = _ingredient_from_line(line) if not in_procedure else None
            if item and not _ACTION.search(item["nome"]):
                ingredients.append(item)
                continue
            if _ACTION.search(line) or in_procedure:
                in_procedure = True
                procedure.append(line)
            elif len(line) <= 80:
                notes.append(line)
        rec = _record(
            name=name, ingredients=ingredients, procedure="\n".join(procedure), notes="\n".join(notes),
            source=_source(path, sha, ws.title, first_row, "scheda_operativa"),
            porzioni=max((_yield_from_text(line) for _, line in lines), default=0), confidence=5,
        )
        if rec:
            result.append(rec)
    wb.close()
    return result


def _score_ingredients(record: dict) -> tuple[int, int, int]:
    rows = record.get("ingredienti_dettaglio") or []
    with_quantity = sum(item.get("quantita") is not None for item in rows)
    return int(record.get("confidenza") or 0), with_quantity, len(rows)


def merge(records: list[dict]) -> list[dict]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        if record.get("chiave"):
            groups[record["chiave"]].append(record)

    result = []
    for key, versions in groups.items():
        ingredient_version = max(versions, key=_score_ingredients)
        procedure_version = max(
            versions,
            key=lambda item: (int(item.get("confidenza") or 0), len(item.get("procedimento_testo") or "")),
        )
        name_version = max(versions, key=lambda item: (int(item.get("confidenza") or 0), len(item.get("nome") or "")))
        notes_version = max(versions, key=lambda item: len(item.get("note") or ""))
        reparto = next((item.get("reparto_hint") for item in sorted(versions, key=lambda x: x.get("confidenza", 0), reverse=True) if item.get("reparto_hint")), "")
        porzioni = next((item.get("porzioni") for item in sorted(versions, key=lambda x: x.get("confidenza", 0), reverse=True) if item.get("porzioni")), 0)
        sources = []
        seen_sources = set()
        for item in versions:
            source = item["fonte"]
            source_key = (source["sha256"], source["sheet"], source.get("row"))
            if source_key not in seen_sources:
                seen_sources.add(source_key)
                sources.append(source)
        result.append({
            "chiave": key,
            "nome": name_version["nome"],
            "reparto_hint": reparto,
            "porzioni": porzioni,
            "ingredienti": [item["nome"] for item in ingredient_version.get("ingredienti_dettaglio") or []],
            "ingredienti_dettaglio": ingredient_version.get("ingredienti_dettaglio") or [],
            "procedimento_testo": procedure_version.get("procedimento_testo") or "",
            "note": notes_version.get("note") or "",
            "fonti_excel": sources,
            "versioni_fonte": versions,
        })
    result.sort(key=lambda item: item["chiave"])
    return result


def main(paths: list[Path]) -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    if len(paths) != 4:
        raise SystemExit("Servono i quattro file Excel nell'ordine indicato nella documentazione")
    for path in paths:
        if not path.exists():
            raise SystemExit(f"File non trovato: {path}")
    by_name = {path.name: path for path in paths}
    required = {
        "Ricettario_Ceraldi completo.xlsx",
        "Ricettario_Ceraldi.xlsx",
        "Ricettario_Completo_v6.xlsx",
        "ricette da importare.xlsx",
    }
    if set(by_name) != required:
        raise SystemExit(f"File attesi: {', '.join(sorted(required))}")

    hashes = {name: _sha256(path) for name, path in by_name.items()}
    records = [
        *parse_ricettario_completo_207(by_name["Ricettario_Ceraldi completo.xlsx"], hashes["Ricettario_Ceraldi completo.xlsx"]),
        *parse_ricettario_ceraldi(by_name["Ricettario_Ceraldi.xlsx"], hashes["Ricettario_Ceraldi.xlsx"]),
        *parse_v6(by_name["Ricettario_Completo_v6.xlsx"], hashes["Ricettario_Completo_v6.xlsx"]),
        *parse_schede_libere(by_name["ricette da importare.xlsx"], hashes["ricette da importare.xlsx"]),
    ]
    recipes = merge(records)
    sources = []
    for name, path in by_name.items():
        wb = load_workbook(path, read_only=False, data_only=True)
        sources.append({"file": name, "sha256": hashes[name], "sheets": wb.sheetnames})
        wb.close()
    bundle_without_hash = {
        "meta": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "sources": sources,
            "record_estratti": len(records),
            "ricette_uniche": len(recipes),
            "con_ingredienti": sum(bool(item["ingredienti_dettaglio"]) for item in recipes),
            "con_preparazione": sum(bool(item["procedimento_testo"]) for item in recipes),
        },
        "recipes": recipes,
    }
    canonical = json.dumps(bundle_without_hash, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    bundle_without_hash["meta"]["bundle_sha256"] = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    OUT.write_text(
        json.dumps(bundle_without_hash, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(bundle_without_hash["meta"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("files", nargs=4, type=Path)
    args = parser.parse_args()
    main(args.files)
