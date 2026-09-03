#!/usr/bin/env python3
"""
IMPORTAZIONE DEFINITIVA PRODOTTI ACQUAVIVA
==========================================
1. Legge il listino Excel (333 prodotti con immagini ufficiali)
2. Integra con acquaviva_prodotti (339 prodotti già in DB con prezzi e allergeni)
3. Cancella TUTTI i prodotti acquaviva in prodotti_vendita
4. Ricrea da zero con dati puliti e completi
5. Ricrea i prodotti interni (89 ricette) con le immagini corrette

Lavora LENTAMENTE e verifica ogni passo.
"""
import asyncio
import os
import uuid
from datetime import datetime, timezone
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
import unicodedata
import re


def normalize(s: str) -> str:
    """Normalizza nome per match fuzzy."""
    if not s:
        return ""
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b\d+[\.,]?\d*\s*(g|kg|gr|ml|cl|lt|cm)\b", "", s)
    return s.strip()


def word_overlap(a: str, b: str) -> float:
    """Score di similarità tra nomi."""
    wa = set(w for w in normalize(a).split() if len(w) > 2)
    wb = set(w for w in normalize(b).split() if len(w) > 2)
    if not wa or not wb:
        return 0.0
    common = wa & wb
    score = len(common) / max(len(wa | wb), 1)
    if wa.issubset(wb) or wb.issubset(wa):
        score = max(score, 0.8)
    return score


def parse_codice(codice_str) -> str:
    """Normalizza il codice prodotto."""
    if not codice_str:
        return ""
    s = str(codice_str).strip()
    # Rimuovi ".0" se è un float
    if s.endswith(".0"):
        s = s[:-2]
    return s


def leggi_listino_excel(path: str) -> list[dict]:
    """Legge il listino Excel e restituisce lista di dizionari."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Intestazioni dalla prima riga
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    prodotti = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col_map.get("Nome", 1)]:
            continue

        def get(key, default=""):
            idx = col_map.get(key)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        codice = parse_codice(get("Codice"))
        nome = get("Nome").strip()
        if not nome:
            continue

        # Categoria (prendi prima categoria significativa, ignora duplicati)
        cat_raw = get("Categoria", "")
        cats = [c.strip() for c in cat_raw.split(";") if c.strip()]
        # Preferisci categoria con ">" (sottocategoria)
        categoria = ""
        for c in cats:
            if " > " in c:
                categoria = c
                break
        if not categoria and cats:
            categoria = cats[0]

        # Peso
        grammi_raw = get("Grammi")
        kg_raw = get("KG")
        peso_g = 0.0
        try:
            if grammi_raw and grammi_raw not in ("", "-"):
                peso_g = float(grammi_raw.replace(",", "."))
        except Exception:
            pass

        # Pezzi per confezione
        pz_conf = 0
        try:
            pz_raw = get("Pz_Confezione")
            if pz_raw and pz_raw not in ("", "-"):
                pz_conf = int(float(pz_raw.replace(",", ".")))
        except Exception:
            pass

        # Info cottura
        gradi = get("Gradi_Forno")
        min_forno = get("Min_Forno")
        min_scongelamento = get("Min_Scongelamento")
        istruzioni = ""
        parts = []
        if gradi:
            parts.append(f"Forno: {gradi}°C")
        if min_forno:
            parts.append(f"Cottura: {min_forno} min")
        if min_scongelamento:
            parts.append(f"Scongelamento: {min_scongelamento} min")
        ore_scongelamento = get("Ore_Scongelamento")
        if ore_scongelamento:
            parts.append(f"Scongelamento: {ore_scongelamento} ore")
        lievitazione = get("Lievitazione")
        if lievitazione:
            parts.append(f"Lievitazione: {lievitazione}")
        if parts:
            istruzioni = " | ".join(parts)

        immagine_url = get("URL_Immagine")
        descrizione = get("Descrizione")
        ingredienti_str = get("Ingredienti")
        allergeni_raw = get("Allergeni")

        prodotti.append({
            "codice": codice,
            "nome": nome,
            "categoria": categoria,
            "grammi": peso_g,
            "pz_confezione": pz_conf,
            "istruzioni_cottura": istruzioni,
            "immagine_url": immagine_url,
            "descrizione": descrizione,
            "ingredienti_str": ingredienti_str,
            "allergeni_raw": allergeni_raw,
        })

    print(f"Letti {len(prodotti)} prodotti dal listino Excel")
    return prodotti


async def main():
    client = AsyncIOMotorClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
    db = client[os.environ.get("DB_NAME", "Gestionale")]
    now = datetime.now(timezone.utc).isoformat()

    # ============================================================
    # STEP 1: Leggi il listino Excel
    # ============================================================
    print("\n=== STEP 1: Lettura listino Excel ===")
    excel_prodotti = leggi_listino_excel("/tmp/listino_acquaviva.xlsx")

    # ============================================================
    # STEP 2: Carica acquaviva_prodotti (con prezzi e allergeni)
    # ============================================================
    print("\n=== STEP 2: Carica acquaviva_prodotti dal DB ===")
    db_acquaviva = await db.acquaviva_prodotti.find({}, {"_id": 0}).to_list(1000)
    print(f"Trovati {len(db_acquaviva)} prodotti in acquaviva_prodotti")

    # Crea mappe per match
    # Map codice -> db_product
    db_by_codice = {}
    for p in db_acquaviva:
        codice = parse_codice(str(p.get("codice", "")))
        if codice:
            db_by_codice[codice] = p

    # ============================================================
    # STEP 3: Cancella prodotti acquaviva in prodotti_vendita
    # ============================================================
    print("\n=== STEP 3: Cancellazione prodotti Acquaviva in prodotti_vendita ===")
    pre_count = await db.prodotti_vendita.count_documents({"fonte": "acquaviva"})
    print(f"  Prima della cancellazione: {pre_count} prodotti acquaviva")

    # CANCELLA
    result = await db.prodotti_vendita.delete_many({"fonte": "acquaviva"})
    print(f"  Cancellati: {result.deleted_count} prodotti")

    post_count = await db.prodotti_vendita.count_documents({"fonte": "acquaviva"})
    print(f"  Dopo cancellazione: {post_count} prodotti acquaviva (deve essere 0)")
    assert post_count == 0, "Errore: prodotti non cancellati!"

    # ============================================================
    # STEP 4: Crea mappa Excel per match con DB
    # ============================================================
    print("\n=== STEP 4: Costruzione catalog unificato ===")

    nuovi_prodotti = []
    match_by_codice = 0
    match_by_nome = 0
    solo_excel = 0

    for excel_p in excel_prodotti:
        codice_excel = excel_p["codice"]

        # Tenta match per codice esatto
        db_p = db_by_codice.get(codice_excel)

        if db_p:
            match_by_codice += 1
        else:
            # Tenta match per nome
            best_score = 0.0
            best_db = None
            for dp in db_acquaviva:
                score = word_overlap(excel_p["nome"], dp.get("nome", ""))
                if score > best_score:
                    best_score = score
                    best_db = dp
            if best_score >= 0.7:
                db_p = best_db
                match_by_nome += 1
            else:
                solo_excel += 1

        # Calcolo costo per pezzo
        prezzo_cartone = 0.0
        pz_cart = excel_p["pz_confezione"]
        if db_p:
            prezzo_cartone = float(db_p.get("prezzo_acquisto_confezione", 0) or 0)
            if pz_cart <= 0:
                pz_cart = int(db_p.get("pz_confezione", 0) or 0)

        costo_pezzo = 0.0
        if prezzo_cartone > 0 and pz_cart > 0:
            costo_pezzo = round(prezzo_cartone / pz_cart, 4)

        # Prezzo vendita (mantenuto se esistente)
        prezzo_vendita = float(db_p.get("prezzo_vendita", 0) or 0) if db_p else 0.0

        # Margine
        margine_euro = 0.0
        margine_perc = 0.0
        if prezzo_vendita > 0 and costo_pezzo > 0:
            margine_euro = round(prezzo_vendita - costo_pezzo, 2)
            margine_perc = round((margine_euro / prezzo_vendita) * 100, 1)

        # Allergeni dal DB
        allergeni = []
        if db_p and db_p.get("allergeni"):
            allergeni = db_p["allergeni"]

        # Categoria: preferisci quella dell'Excel (più precisa)
        categoria = excel_p["categoria"]
        if not categoria and db_p:
            categoria = db_p.get("categoria", "")

        # Descrizione: Excel + DB
        descrizione = excel_p["descrizione"]
        if not descrizione and db_p:
            descrizione = db_p.get("descrizione", "")

        # Immagine: sempre dall'Excel (ufficiale)
        immagine_url = excel_p["immagine_url"]
        if not immagine_url and db_p:
            immagine_url = db_p.get("foto_url", "") or db_p.get("immagine_url", "")

        prodotto = {
            "id": str(uuid.uuid4()),
            "nome": excel_p["nome"],
            "codice_prodotto": codice_excel,
            "categoria": categoria,
            "fonte": "acquaviva",
            "fornitore": "Dolciaria Acquaviva",
            "attivo": True,
            "peso_pezzo_g": excel_p["grammi"],
            "pezzi_cartone": pz_cart,
            "costo_produzione_cartone": prezzo_cartone,
            "costo_produzione": costo_pezzo,
            "iva": 10,
            "prezzo_vendita": prezzo_vendita,
            "prezzo_ivato": round(prezzo_vendita * 1.1, 2) if prezzo_vendita > 0 else 0.0,
            "margine_euro": margine_euro,
            "margine_percentuale": margine_perc,
            "immagine_url": immagine_url,
            "descrizione": descrizione,
            "ingredienti": excel_p["ingredienti_str"],
            "istruzioni_cottura": excel_p["istruzioni_cottura"],
            "allergeni": allergeni,
            "created_at": now,
            "updated_at": now,
        }
        nuovi_prodotti.append(prodotto)

    print(f"  Prodotti da inserire: {len(nuovi_prodotti)}")
    print(f"  Match per codice: {match_by_codice}")
    print(f"  Match per nome: {match_by_nome}")
    print(f"  Solo Excel (no prezzo): {solo_excel}")

    # ============================================================
    # STEP 5: Inserisci i nuovi prodotti Acquaviva
    # ============================================================
    print("\n=== STEP 5: Inserimento prodotti Acquaviva ===")
    if nuovi_prodotti:
        result = await db.prodotti_vendita.insert_many(nuovi_prodotti)
        print(f"  Inseriti: {len(result.inserted_ids)} prodotti")

    # ============================================================
    # STEP 6: Ricostruisci prodotti interni da 'ricette'
    # ============================================================
    print("\n=== STEP 6: Ricostruzione prodotti interni da ricette ===")

    ricette = await db.ricette.find({}, {"_id": 0}).to_list(500)
    print(f"  Ricette trovate: {len(ricette)}")

    # Carica prodotti interni esistenti in prodotti_vendita (per preservare prezzi)
    esistenti_interni = await db.prodotti_vendita.find(
        {"fonte": {"$ne": "acquaviva"}},
        {"_id": 0}
    ).to_list(500)
    esistenti_map = {}
    for ep in esistenti_interni:
        key = normalize(ep.get("nome", ""))
        esistenti_map[key] = ep

    print(f"  Prodotti interni esistenti in prodotti_vendita: {len(esistenti_interni)}")

    aggiornati_interni = 0
    for ricetta in ricette:
        nome = ricetta.get("nome", "")
        if not nome:
            continue

        key = normalize(nome)

        # Immagine dalla ricetta
        immagine = ricetta.get("immagine", "") or ""
        if not immagine.startswith("/uploads/") and not immagine.startswith("http"):
            immagine = ""

        # Categoria dalla ricetta
        categoria = ricetta.get("categoria", "")

        # Costo dalla ricetta (se calcolato)
        costo = float(ricetta.get("costo_totale", 0) or 0)

        # Trova prodotto interno già in prodotti_vendita
        ep = esistenti_map.get(key)

        if ep:
            # Aggiorna: immagine corretta dalla ricetta, preserva prezzo
            update = {
                "updated_at": now,
            }
            if immagine:
                update["immagine_url"] = immagine
                update["immagine"] = immagine

            # Rimuovi eventuali campi Acquaviva
            unset = {}
            if ep.get("acquaviva_id"):
                unset["acquaviva_id"] = ""
            if ep.get("codice_prodotto") and ep.get("fonte") != "interno":
                pass  # mantieni codice se è già corretto

            # Fix fonte se sbagliata
            if ep.get("fonte") != "interno":
                update["fonte"] = "interno"

            update_op = {"$set": update}
            if unset:
                update_op["$unset"] = unset

            await db.prodotti_vendita.update_one({"id": ep["id"]}, update_op)
            aggiornati_interni += 1
        else:
            # Prodotto della ricetta non è ancora in prodotti_vendita, crealo
            pv = float(ricetta.get("prezzo_vendita", 0) or 0)
            margine_e = 0.0
            margine_p = 0.0
            if pv > 0 and costo > 0:
                margine_e = round(pv - costo, 2)
                margine_p = round((margine_e / pv) * 100, 1)

            nuovo = {
                "id": str(uuid.uuid4()),
                "nome": nome,
                "categoria": categoria or "Pasticceria",
                "fonte": "interno",
                "attivo": True,
                "costo_produzione": costo,
                "iva": 10,
                "prezzo_vendita": pv,
                "prezzo_ivato": round(pv * 1.1, 2) if pv > 0 else 0.0,
                "margine_euro": margine_e,
                "margine_percentuale": margine_p,
                "immagine_url": immagine,
                "immagine": immagine,
                "allergeni": ricetta.get("allergeni", []),
                "ingredienti": ", ".join([
                    (i.get("nome", "") if isinstance(i, dict) else str(i))
                    for i in ricetta.get("ingredienti", [])
                ]),
                "created_at": now,
                "updated_at": now,
            }
            await db.prodotti_vendita.insert_one(nuovo)
            aggiornati_interni += 1

    print(f"  Prodotti interni aggiornati/creati: {aggiornati_interni}")

    # ============================================================
    # STEP 7: Verifica finale
    # ============================================================
    print("\n=== STEP 7: Verifica finale ===")
    total = await db.prodotti_vendita.count_documents({})
    acquaviva_count = await db.prodotti_vendita.count_documents({"fonte": "acquaviva"})
    interni_count = await db.prodotti_vendita.count_documents({"fonte": "interno"})
    con_img = await db.prodotti_vendita.count_documents({"fonte": "acquaviva", "immagine_url": {"$nin": ["", None]}})
    con_costo = await db.prodotti_vendita.count_documents({"fonte": "acquaviva", "costo_produzione": {"$gt": 0}})
    con_prezzo = await db.prodotti_vendita.count_documents({"fonte": "acquaviva", "prezzo_vendita": {"$gt": 0}})

    print(f"  TOTALE prodotti_vendita: {total}")
    print(f"  Acquaviva: {acquaviva_count}")
    print(f"    - Con immagine: {con_img}")
    print(f"    - Con costo: {con_costo}")
    print(f"    - Con prezzo vendita: {con_prezzo}")
    print(f"  Interni: {interni_count}")

    # Sample acquaviva
    samples = await db.prodotti_vendita.find(
        {"fonte": "acquaviva"},
        {"_id": 0, "nome": 1, "costo_produzione": 1, "pezzi_cartone": 1, "immagine_url": 1, "categoria": 1}
    ).limit(5).to_list(5)
    print("\n  Sample Acquaviva:")
    for s in samples:
        img_short = (s.get("immagine_url") or "")[-40:]
        print(f"    {s['nome'][:40]:40s} cat={s.get('categoria','')[:25]:25s} costo={s.get('costo_produzione',0):.4f} img=...{img_short}")


if __name__ == "__main__":
    asyncio.run(main())
