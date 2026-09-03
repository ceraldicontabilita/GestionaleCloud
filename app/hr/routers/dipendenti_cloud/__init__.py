"""
Dipendenti in Cloud - Router Module
Sistema HR completo per gestione personale
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body
from fastapi.responses import Response
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
import re
import os
import io
import zipfile
import hashlib
import base64
import tempfile
import logging
from datetime import datetime, timezone, timedelta, date
from decimal import Decimal, InvalidOperation

from app.hr.database import Database

logger = logging.getLogger(__name__)

# Router principale
router = APIRouter(prefix="/dipendenti-cloud", tags=["Dipendenti Cloud"])

# ============ HELPERS ============

def get_db():
    """Get database instance"""
    return Database.get_db()

def generate_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def serialize_doc(doc):
    if doc and '_id' in doc:
        del doc['_id']
    return doc

# ============ MODELS ============

class DipendenteCloud(BaseModel):
    nome: str
    cognome: str
    matricola: Optional[str] = None
    codice_fiscale: Optional[str] = None
    data_nascita: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    indirizzo: Optional[str] = None
    ruolo: Optional[str] = None
    luogo_lavoro: Optional[str] = None
    contratto: str = "Indeterminato"
    data_assunzione: Optional[str] = None
    data_fine_contratto: Optional[str] = None
    iban: Optional[str] = None
    stato: str = "attivo"

class PresenzaCloud(BaseModel):
    dipendente_id: str
    data: str
    entrata: Optional[str] = None
    uscita: Optional[str] = None
    stato: str = "presente"
    giustificativo: Optional[str] = None
    ore_lavorate: float = 0
    note: Optional[str] = None

class FerieCloud(BaseModel):
    dipendente_id: str
    tipo: str  # Ferie, Permesso, Malattia, ROL
    data_inizio: str
    data_fine: str
    giorni: int = 1
    stato: str = "in_attesa"
    nota: Optional[str] = None

class TurnoCloud(BaseModel):
    nome: str
    orario_inizio: str
    orario_fine: str
    colore: str = "#3b82f6"

class BustaPagaCloud(BaseModel):
    dipendente_id: str
    mese: int
    anno: int
    lordo: float
    netto: float
    inps: float = 0
    irpef: float = 0
    trattenute: float = 0
    stato: str = "DA_PAGARE"
    data_pagamento: Optional[str] = None

class MissioneCloud(BaseModel):
    dipendente_id: str
    destinazione: str
    data_inizio: str
    data_fine: str
    scopo: str
    rimborso: float = 0
    stato: str = "in_attesa"

class DocumentoCloud(BaseModel):
    dipendente_id: str
    titolo: str
    tipo: str
    scadenza: Optional[str] = None
    file_url: Optional[str] = None

# ============ DIPENDENTI ============

@router.get("/dipendenti")
async def get_dipendenti():
    """Legge dalla collezione 'dipendenti' esistente nel database Gestionale"""
    dipendenti = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    # Normalizza i campi per compatibilità con il frontend.
    # ruolo/contratto: prima il valore inserito a mano, poi quello letto
    # dall'UNILAV (qualifica_unilav / tipo_contratto) — MAI un default fisso:
    # "Indeterminato" per chi non lo sappiamo e' un dato inventato, non ignoto.
    result = []
    for d in dipendenti:
        result.append({
            "id": d.get("id") or str(d.get("_id", "")),
            "nome": d.get("nome", ""),
            "cognome": d.get("cognome", ""),
            "codice_fiscale": d.get("codice_fiscale", ""),
            "stato": d.get("stato", "attivo"),
            "ruolo": d.get("ruolo") or d.get("qualifica_unilav") or d.get("mansione") or "",
            "iban": d.get("iban", ""),
            "email": d.get("email", ""),
            "telefono": d.get("telefono", ""),
            "contratto": d.get("contratto") or d.get("tipo_contratto") or "",
            "data_assunzione": d.get("data_assunzione", ""),
            "data_cessazione": d.get("data_cessazione", ""),
            "luogo_lavoro": d.get("luogo_lavoro", ""),
            "importo_stipendio": d.get("importo_stipendio", 0),
            "livello": d.get("livello", ""),
            "ore_settimanali": d.get("ore_settimanali"),
            "created_at": d.get("created_at", "")
        })
    return result

@router.get("/dipendenti/{dipendente_id}")
async def get_dipendente(dipendente_id: str):
    dip = await get_db().dipendenti.find_one({"id": dipendente_id}, {"_id": 0})
    if not dip:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return dip

@router.post("/dipendenti")
async def create_dipendente(dip: DipendenteCloud):
    dip_dict = dip.model_dump()
    dip_dict["id"] = generate_id()
    dip_dict["created_at"] = now_iso()
    await get_db().dipendenti.insert_one(dip_dict)
    return serialize_doc(dip_dict)

@router.put("/dipendenti/{dipendente_id}")
async def update_dipendente(dipendente_id: str, dip: DipendenteCloud):
    result = await get_db().dipendenti.update_one(
        {"id": dipendente_id},
        {"$set": dip.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return {"message": "Dipendente aggiornato"}

@router.delete("/dipendenti/{dipendente_id}")
async def delete_dipendente(dipendente_id: str):
    result = await get_db().dipendenti.delete_one({"id": dipendente_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return {"message": "Dipendente eliminato"}

@router.post("/dipendenti/{dipendente_id}/cessa")
async def cessa_dipendente(dipendente_id: str, data: dict = Body(default={})):
    """Cessa il rapporto: aggiorna lo stato e innesca l'iter completo di chiusura
    (termina contratti, rifiuta assenze future, annulla partite, risolve alert)
    tramite l'evento DIPENDENTE_CESSATO già agganciato all'handler dedicato."""
    db = get_db()
    dip = await db.dipendenti.find_one({"id": dipendente_id}, {"_id": 0})
    if not dip:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    data_cessazione = (data.get("data_cessazione") or now_iso()[:10])
    nome = dip.get("nome_completo") or f"{dip.get('cognome','')} {dip.get('nome','')}".strip()
    await db.dipendenti.update_one({"id": dipendente_id}, {"$set": {
        "stato": "cessato", "attivo": False,
        "data_dimissione": data_cessazione, "cessato_il": now_iso(),
        "motivo_cessazione": data.get("motivo") or "cessazione_manuale",
    }})
    try:
        from app.hr.services.event_bus import propagate_event, EventTypes
        risultati = await propagate_event(EventTypes.DIPENDENTE_CESSATO, {
            "dipendente_id": dipendente_id, "nome_completo": nome,
            "data_cessazione": data_cessazione,
        }, db, source_module="gestione", user="admin")
    except Exception as e:
        risultati = [{"error": str(e)}]
    return {"ok": True, "stato": "cessato", "data_cessazione": data_cessazione, "automazioni": risultati}

@router.get("/ordine-dipendenti")
async def get_ordine_dipendenti():
    doc = await get_db().dipendenti_ordine.find_one({"id": "ordine"}, {"_id": 0})
    return {"ordine": (doc or {}).get("lista", [])}

@router.post("/ordine-dipendenti")
async def set_ordine_dipendenti(data: dict):
    lista = data.get("ordine", [])
    await get_db().dipendenti_ordine.update_one(
        {"id": "ordine"}, {"$set": {"lista": lista}}, upsert=True)
    return {"ok": True}

# ============ PAGHE MENSILI (importo busta + bonifico + acconti) ============

@router.post("/paghe/sincronizza")
async def sincronizza_paghe_da_cedolini(anno: Optional[int] = None):
    """Popola il registro paghe dai cedolini e dai bonifici reali gia' in
    archivio, invece di lasciarlo alla compilazione manuale. Non tocca un mese
    che qualcuno ha gia' modificato a mano."""
    from app.hr.services.sincronizza_paghe_mensili import sincronizza
    return await sincronizza(get_db(), anno)


@router.post("/paghe/sincronizza-bonifici-storici")
async def sincronizza_bonifici_storici():
    """Ponte una tantum (ma ripetibile: idempotente) tra la collezione `bonifici`
    (dove un import passato dei PDF storici ha salvato ~800 bonifici reali con
    dipendente_id + competenza "YYYY-MM", ma NESSUN cedolino_id) e il motore unico
    di "Cedolini & Bonifici" / stato paga, che legge solo `pagamenti_esiti`.
    Senza questo ponte quei bonifici — già nell'archivio, già con il PDF allegato —
    restavano invisibili nella pagina di riconciliazione e le buste corrispondenti
    risultavano ancora "da pagare". Non tocca i pagamenti "una_tantum" (TFR/
    transazioni di fine rapporto: competenza non è un mese di stipendio)."""
    db = get_db()
    # Esclude i bonifici già scritti a mano dalla coda "Bonifici da associare"
    # (associa_bonifico li ha già messi in pagamenti_esiti con la chiave
    # "beneficiari-diversi:*"): reimportarli qui creerebbe una seconda riga in
    # pagamenti_esiti per lo stesso pagamento, raddoppiando il bonifico del mese.
    bonifici = await db.bonifici.find(
        {"categoria": "DIPENDENTE", "dipendente_id": {"$ne": None}, "competenza": {"$ne": None},
         "assegnato_da_bonifico_diversi": {"$exists": False}},
        {"_id": 0}).to_list(5000)

    # Prefetch in blocco (stesso motivo delle altre correzioni in questo file):
    # un find_one per bonifico dentro il ciclo, su una tabella non indicizzata,
    # è di nuovo un incrocio N×M — con ~800 bonifici storici è la stessa causa
    # dei 502 già visti in produzione. pdf_data escluso: qui serve solo la
    # tripletta dipendente/data/importo per il controllo duplicati.
    pagamenti_esistenti = set()
    async for p in db.pagamenti_esiti.find({}, {"_id": 0, "pdf_data": 0, "causale": 0, "beneficiario": 0}):
        pagamenti_esistenti.add((p.get("dipendente_id"), p.get("data"), p.get("importo")))

    affected = set()
    importati = duplicati = 0
    for b in bonifici:
        competenza = str(b.get("competenza") or "")
        m = re.match(r"^(\d{4})-(\d{1,2})$", competenza)
        if not m:
            continue
        anno, mese = int(m.group(1)), int(m.group(2))
        dip_id = b.get("dipendente_id")
        importo = b.get("importo")
        data_pag = b.get("data")
        if not dip_id or not importo:
            continue
        # Difesa aggiuntiva: stesso dipendente/data/importo già presente in
        # pagamenti_esiti (da CSV, da un'altra corsa di questo stesso ponte, o
        # da un altro percorso) -> non è un nuovo pagamento, salta.
        if (dip_id, data_pag, importo) in pagamenti_esistenti:
            duplicati += 1
            continue
        pagamenti_esistenti.add((dip_id, data_pag, importo))
        key = f"bonifici-coll:{b.get('id')}"
        doc = {"key": key, "cro": None, "dipendente_id": dip_id,
               "data": data_pag, "importo": importo,
               "causale": b.get("fonte") or "Archivio bonifici PDF",
               "beneficiario": b.get("dipendente_nome"),
               "mese": mese, "anno": anno, "origine": "bonifici-storico"}
        if b.get("pdf_data"):
            doc["pdf_data"] = b["pdf_data"]
            doc["ha_pdf"] = True
        await db.pagamenti_esiti.update_one({"key": key}, {"$set": doc}, upsert=True)
        await db.paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": anno, "mese": mese},
            {"$set": {"dipendente_id": dip_id, "anno": anno, "mese": mese,
                      "updated_at": now_iso()}}, upsert=True)
        affected.add((dip_id, mese, anno))
        importati += 1

    for dip_id, mese, anno in affected:
        tot = 0.0
        async for p in db.pagamenti_esiti.find({"dipendente_id": dip_id, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1}):
            tot += p.get("importo") or 0
        await db.paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": anno, "mese": mese},
            {"$set": {"bonifico_importo": round(tot, 2), "bonifico_ricevuto": tot > 0,
                      "bonifico_da_esiti": True, "updated_at": now_iso()}})
        await _ricalcola_stato_paga(db, dip_id, anno, mese)

    return {"bonifici_esaminati": len(bonifici), "importati_in_pagamenti_esiti": importati,
            "duplicati_saltati": duplicati, "mesi_aggiornati": len(affected)}


@router.get("/paghe")
async def get_paghe(anno: int, mese: int):
    return await get_db().paghe_mensili.find(
        {"anno": int(anno), "mese": int(mese)}, {"_id": 0}).to_list(500)

@router.post("/paghe")
async def upsert_pagha(data: dict):
    dip = data.get("dipendente_id"); anno = data.get("anno"); mese = data.get("mese")
    if not dip or not anno or not mese:
        raise HTTPException(status_code=400, detail="dipendente_id, anno, mese obbligatori")
    # Normalizza gli acconti: massimo 3, solo importo+data
    acconti = []
    for a in (data.get("acconti") or [])[:3]:
        acconti.append({"importo": a.get("importo"), "data": a.get("data")})
    doc = {
        "dipendente_id": dip, "anno": int(anno), "mese": int(mese),
        "importo_busta": data.get("importo_busta"),
        "bonifico_ricevuto": bool(data.get("bonifico_ricevuto", False)),
        "bonifico_importo": data.get("bonifico_importo"),
        "bonifico_data": data.get("bonifico_data"),
        "acconti": acconti,
        "updated_at": now_iso(),
    }
    await get_db().paghe_mensili.update_one(
        {"dipendente_id": dip, "anno": int(anno), "mese": int(mese)},
        {"$set": doc}, upsert=True)
    await _ricalcola_stato_paga(get_db(), dip, int(anno), int(mese))
    return {"ok": True, "pagha": doc}


async def _ricalcola_stato_paga(db, dip, anno, mese):
    """MOTORE UNICO buste↔bonifici. Aggancia i pagamenti bancari già arrivati
    (pagamenti_esiti) come bonifico del mese e ricalcola lo stato:
    in_attesa_pagamento (busta senza pagamento) / parziale / pagato / vuoto.
    Chiamato da OGNI ingresso (busta da LUL/email, prima nota, CSV, modifica manuale),
    così il popolamento di un dato aggiorna automaticamente gli altri."""
    anno, mese = int(anno), int(mese)
    p = await db.paghe_mensili.find_one({"dipendente_id": dip, "anno": anno, "mese": mese})
    if not p:
        return None
    tot_esiti, n_esiti = 0.0, 0
    async for e in db.pagamenti_esiti.find({"dipendente_id": dip, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1}):
        tot_esiti += e.get("importo") or 0
        n_esiti += 1
    bonifico = round(tot_esiti, 2) if n_esiti else float(p.get("bonifico_importo") or 0)
    busta = float(p.get("importo_busta") or 0)
    acc = sum(float(a.get("importo") or 0) for a in (p.get("acconti") or []))
    erogato = bonifico + acc
    if busta <= 0 and erogato <= 0:
        stato = "vuoto"
    elif erogato <= 0:
        stato = "in_attesa_pagamento"
    elif erogato + 0.5 >= busta:
        stato = "pagato"
    else:
        stato = "parziale"
    upd = {"stato_pagamento": stato, "saldo": round(busta - erogato, 2), "updated_at": now_iso()}
    if n_esiti:
        upd["bonifico_importo"] = bonifico
        upd["bonifico_ricevuto"] = bonifico > 0
    await db.paghe_mensili.update_one({"dipendente_id": dip, "anno": anno, "mese": mese}, {"$set": upd})
    return stato

# ============ BONIFICI DA ASSOCIARE ============
# Bonifici bancari "BENEFICIARI DIVERSI": la banca li emette come un unico
# addebito cumulativo su piu' persone, senza nominarne nessuna nel documento.
# Non c'e' modo di attribuirli automaticamente: qui restano in coda, con
# menu a tendina dipendente + periodo, finche' qualcuno non li assegna a mano.

@router.get("/bonifici-da-associare")
async def lista_bonifici_da_associare():
    """Elenco leggero (senza il PDF) di chi aspetta un'assegnazione manuale."""
    righe = await get_db().bonifici_da_associare.find(
        {"stato": "da_associare"}, {"_id": 0, "pdf_data": 0}).to_list(500)
    righe.sort(key=lambda r: r.get("data") or "", reverse=True)
    return righe


@router.get("/paghe/pagamento-esito/{key}/pdf")
async def pdf_pagamento_esito(key: str):
    """PDF sorgente di un bonifico già associato a una busta (pagamenti_esiti),
    per verificarlo prima di premere Conferma — solo se è stato importato con
    l'allegato (Drive o ponte bonifici storici; i CSV banca non hanno PDF)."""
    doc = await get_db().pagamenti_esiti.find_one(
        {"key": key}, {"_id": 0, "pdf_data": 1})
    if not doc or not doc.get("pdf_data"):
        raise HTTPException(404, "PDF non disponibile per questo pagamento")
    pdf_bytes = base64.b64decode(doc["pdf_data"])
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": 'inline; filename="bonifico.pdf"'})


@router.get("/bonifici-da-associare/{bonifico_id}/pdf")
async def pdf_bonifico_da_associare(bonifico_id: str):
    doc = await get_db().bonifici_da_associare.find_one(
        {"id": bonifico_id}, {"_id": 0, "pdf_data": 1, "pdf_filename": 1})
    if not doc or not doc.get("pdf_data"):
        raise HTTPException(404, "PDF non trovato")
    pdf_bytes = base64.b64decode(doc["pdf_data"])
    fname = doc.get("pdf_filename") or "bonifico.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{fname}"'})


@router.post("/bonifici-da-associare/{bonifico_id}/associa")
async def associa_bonifico(bonifico_id: str, data: Dict[str, Any] = Body(...)):
    """Assegna un bonifico in coda a un dipendente e a un periodo, scelti a
    mano. Lo trasforma in un bonifico vero (stessa collezione degli altri,
    stessa logica di riconciliazione) col PDF portato dietro come prova, e lo
    toglie dalla coda."""
    dipendente_id = data.get("dipendente_id")
    mese = data.get("mese")
    anno = data.get("anno")
    if not dipendente_id or not mese or not anno:
        raise HTTPException(400, "dipendente_id, mese, anno obbligatori")

    db = get_db()
    in_coda = await db.bonifici_da_associare.find_one({"id": bonifico_id}, {"_id": 0})
    if not in_coda:
        raise HTTPException(404, "Bonifico non trovato in coda")
    dip = await db.dipendenti.find_one({"id": dipendente_id}, {"_id": 0})
    if not dip:
        raise HTTPException(404, "Dipendente non trovato")

    nuovo = {
        "id": str(uuid.uuid4()), "dipendente_id": dipendente_id,
        "dipendente_nome": dip.get("nome_completo"),
        "data": in_coda.get("data"), "importo": in_coda.get("importo"),
        "competenza": "%s-%02d" % (int(anno), int(mese)),
        "categoria": "DIPENDENTE",
        "pdf_filename": in_coda.get("pdf_filename"), "pdf_data": in_coda.get("pdf_data"),
        "fonte": in_coda.get("fonte"),
        "assegnato_manualmente": True,
        "assegnato_da_bonifico_diversi": bonifico_id,
        "created_at": now_iso(),
    }
    await db.bonifici.insert_one(nuovo)
    await db.bonifici_da_associare.update_one(
        {"id": bonifico_id},
        {"$set": {"stato": "associato", "associato_a": dipendente_id,
                  "associato_competenza": nuovo["competenza"], "associato_il": now_iso()}})

    # Aggancia anche al MOTORE UNICO paghe (pagamenti_esiti + paghe_mensili):
    # senza questo passo l'associazione restava confinata alla collezione
    # `bonifici` e non si vedeva mai né in "Cedolini & Bonifici" né sulla busta
    # come pagata, perché quella vista/lo stato paga leggono solo pagamenti_esiti.
    anno_i, mese_i = int(anno), int(mese)
    key = f"beneficiari-diversi:{bonifico_id}"
    await db.pagamenti_esiti.update_one(
        {"key": key},
        {"$set": {"key": key, "cro": None, "dipendente_id": dipendente_id,
                  "data": in_coda.get("data"), "importo": in_coda.get("importo") or 0,
                  "causale": in_coda.get("causale") or "Bonifico beneficiari diversi",
                  "beneficiario": dip.get("nome_completo"),
                  "mese": mese_i, "anno": anno_i}}, upsert=True)
    await db.paghe_mensili.update_one(
        {"dipendente_id": dipendente_id, "anno": anno_i, "mese": mese_i},
        {"$set": {"dipendente_id": dipendente_id, "anno": anno_i, "mese": mese_i,
                  "updated_at": now_iso()}}, upsert=True)
    await _ricalcola_stato_paga(db, dipendente_id, anno_i, mese_i)
    return {"ok": True, "bonifico": nuovo}


@router.post("/bonifici-da-associare/{bonifico_id}/ignora")
async def ignora_bonifico_da_associare(bonifico_id: str):
    """Non e' un pagamento a un dipendente (es. lotto di fornitori): esce
    dalla coda senza creare nulla."""
    res = await get_db().bonifici_da_associare.update_one(
        {"id": bonifico_id}, {"$set": {"stato": "ignorato", "ignorato_il": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(404, "Bonifico non trovato in coda")
    return {"ok": True}


_MOVIMENTO_NON_STIPENDIO_RE = re.compile(
    r"MUTUO|FINANZIAMENTO|QUIETANZA DI PAGAMENTO|CONTABILE DI FILIALE|"
    r"VERSAMENTO|UFFICIO SERVIZI VARI|AGENZIA DELLE ENTRATE|REGIONE \w+|"
    r"COMUNE DI |SPESE LIBR|ADDEBITO DIRETTO|SDD\b|INC\.?POS|"
    r"COMM\.?\s*SU BONIFICI|\bS\.?R\.?L\.?\b|\bS\.?P\.?A\.?\b|\bS\.?N\.?C\.?\b|\bS\.?A\.?S\.?\b",
    re.IGNORECASE)


def _e_movimento_non_stipendio(text: str) -> bool:
    """La cartella Drive dei bonifici (estratto conto aziendale) contiene di tutto:
    mutui, fornitori, tasse, versamenti contanti, utenze — non solo stipendi. Questi
    movimenti non vanno importati né messi in coda di revisione: inquinerebbero
    "Bonifici da associare" con roba che non è mai stata uno stipendio."""
    return bool(_MOVIMENTO_NON_STIPENDIO_RE.search(text or ""))


def _parse_bonifico_pdf(text: str, filename: str) -> Dict[str, Any]:
    """Estrae importo/data/causale da un PDF di bonifico (distinta stipendi banca o
    contabile bonifico singolo — layout diversi, best-effort). Ritorna anche
    n_beneficiari: se >1 (distinta cumulativa su più persone) chi chiama deve
    mettere il bonifico in coda manuale, MAI indovinare a chi appartiene."""
    t = text or ""

    n_benef = 1
    m = re.search(r"n\.?\s*stipendi\s*:?\s*(\d+)", t, re.IGNORECASE)
    if m:
        n_benef = int(m.group(1))

    importo = None
    m = re.search(r"tot\.?\s*distinta\s*:?\s*([\d.,]+)\s*eur", t, re.IGNORECASE)
    if m:
        importo = m.group(1)
    if importo is None:
        m = re.search(r"eur\s*([\d.,]+)\s*\d{2}/\d{2}/\d{4}", t, re.IGNORECASE)
        if m:
            importo = m.group(1)
    if importo is None:
        m = re.search(r"importo\s*:?\s*([\d.,]+)\s*(?:eur|€)", t, re.IGNORECASE)
        if m:
            importo = m.group(1)
    importo_val = None
    if importo:
        try:
            importo_val = float(importo.replace(".", "").replace(",", "."))
        except ValueError:
            importo_val = None

    data_val = None
    m = re.search(r"data\s+esecuzione\s*:?\s*(\d{2}/\d{2}/\d{4})", t, re.IGNORECASE)
    if not m:
        m = re.search(r"\b(\d{2}/\d{2}/\d{4})\s+data\b", t, re.IGNORECASE)
    if not m:
        m = re.search(r"\b(\d{2}/\d{2}/20\d{2})\b", t)
    if m:
        try:
            data_val = datetime.strptime(m.group(1), "%d/%m/%Y")
        except ValueError:
            data_val = None

    causale = None
    m = re.search(r"causale\s*:?\s*\n?\s*([^\n]{4,80})", t, re.IGNORECASE)
    if m:
        causale = m.group(1).strip()

    # Mese di competenza: quasi sempre nel nome file/causale ("bonifico marzo…",
    # "stip giugno 2022"), NON nella data di esecuzione (il bonifico parte
    # qualche giorno dopo la fine del mese di competenza).
    haystack = re.sub(r"\s+", " ", f"{filename} {causale or ''}").lower()
    mese_comp, anno_comp = None, None
    for nome, num in _MESI_IT.items():
        if nome in haystack:
            mese_comp = num
            break
    anno_m = re.search(r"(20\d{2})", haystack)
    if anno_m:
        anno_comp = int(anno_m.group(1))
    if mese_comp and not anno_comp and data_val:
        anno_comp = data_val.year
        if mese_comp == 12 and data_val.month == 1:
            anno_comp -= 1
    if not mese_comp and data_val:
        mese_comp, anno_comp = data_val.month, data_val.year

    # Segnale esplicito "è uno stipendio": nella cartella (estratto conto
    # aziendale generico) c'è anche, per es., "CAUSALE vincenzo ceraldi" senza
    # la parola stipendio — un bonifico al nome di un dipendente ma non
    # necessariamente il suo stipendio (potrebbe essere un prelievo personale
    # del titolare, un rimborso...). Senza questa parola non si associa da
    # solo: va in coda, decide una persona.
    # ATTENZIONE: NON cercare "stipend" in tutto il testo — l'intestazione
    # "N. stipendi: 1" (marcatore già usato per n_beneficiari) è presente su
    # OGNI distinta di questo tipo, causale vera o no, e farebbe scattare il
    # segnale sempre. Si guarda solo il titolo del documento ("Distinta
    # Stipendi", vero per costruzione su quel template) o la causale estratta
    # per l'altro formato.
    causale_stipendio = (bool(re.search(r"distinta\s+stipend", t, re.IGNORECASE))
                         or bool(re.search(r"stipend|\bstip\b|\bstip\.", causale or "", re.IGNORECASE)))

    return {"importo": importo_val, "data": data_val, "causale": causale or filename,
            "mese": mese_comp, "anno": anno_comp, "n_beneficiari": n_benef,
            "causale_stipendio": causale_stipendio}


@router.post("/paghe/importa-bonifici-drive")
async def importa_bonifici_drive(body: Dict[str, Any] = Body(default={})):
    """Importa i PDF dei bonifici stipendi dalla cartella Google Drive (service account,
    vedi services/google_drive_sa.py — stessa cartella del link "Cartella Drive bonifici").
    ATTENZIONE: quella cartella è l'estratto conto aziendale generico (mutui, fornitori,
    tasse, versamenti, utenze, oltre agli stipendi), non solo bonifici stipendi — i
    movimenti chiaramente non salariali vengono scartati subito (_e_movimento_non_stipendio),
    mai messi in coda.
    Per ogni PDF residuo: estrae beneficiario (dal testo o dal nome file, stesso indice
    nome/cognome degli altri importer), importo e mese di competenza. Entra direttamente
    nei pagamenti reali (pagamenti_esiti, motore unico paghe) SOLO se il dipendente è
    univoco E la causale dice esplicitamente "stipendio" — un bonifico intestato a un
    dipendente ma senza quella parola (es. un prelievo personale del titolare) non si
    associa da solo. Tutto il resto (distinta cumulativa, dipendente non univoco, nessun
    segnale di stipendio, PDF illeggibile) finisce nella coda "Bonifici da associare"
    per l'assegnazione manuale — mai indovinato. Il PDF sorgente viene sempre allegato
    (pagamenti_esiti.pdf_data o bonifici_da_associare.pdf_data) così si può riaprire e
    verificare prima di confermare.

    A LOTTI: questa cartella può avere centinaia di PDF (mutui, fornitori, tasse...
    non solo stipendi) — scaricarli e leggerli tutti in una sola richiesta HTTP è
    già andato in timeout in produzione (573s, 502). Ogni chiamata processa al più
    `limit` file MAI VISTI PRIMA (default 120, tracciati in drive_bonifici_visti per
    id Drive, non per contenuto: anche un file scartato come non-stipendio non viene
    riletto al giro successivo). Se `restanti > 0` nella risposta, richiamare di
    nuovo lo stesso endpoint per continuare (il frontend lo fa in automatico)."""
    from app.hr.services.google_drive_sa import elenca_pdf_cartella, scarica_per_id

    folder = str(body.get("folder_id") or os.environ.get("DRIVE_BONIFICI_FOLDER_ID")
                or "1yl55742cu9i-AFLxu2s0QnMvXG6kVkJC")
    limite = int(body.get("limit") or 120)

    db = get_db()
    tutti = await elenca_pdf_cartella(folder)
    visti = set()
    async for v in db.drive_bonifici_visti.find({}, {"_id": 0, "id": 1}):
        visti.add(v.get("id"))
    da_fare = [f for f in tutti if f.get("id") not in visti]
    lotto = da_fare[:limite]
    scaricati, falliti_id = await scarica_per_id([f["id"] for f in lotto])
    id_per_nome = {f["id"]: f.get("name") for f in lotto}

    indici = await _indici_dipendenti(db)
    by_nome, by_cogn = indici["nome"], indici["cogn"]

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()

    def trova_dipendente(testo, filename):
        haystack = norm(f"{testo} {filename}")
        # Nomi completi citati nel testo: se ne compare più di uno (es. layout
        # con più beneficiari che il marcatore "n. stipendi" non ha intercettato),
        # è ambiguo — meglio la coda manuale che indovinare la persona sbagliata
        # su un documento che vale come prova di pagamento.
        per_nome = {}
        for nome_n, d in by_nome.items():
            if nome_n in haystack:
                per_nome[d["id"]] = d
        if len(per_nome) == 1:
            return next(iter(per_nome.values()))
        if len(per_nome) > 1:
            return None
        candidati = {}
        for cogn, lst in by_cogn.items():
            if cogn in haystack and len(lst) == 1:
                candidati[lst[0]["id"]] = lst[0]
        return next(iter(candidati.values())) if len(candidati) == 1 else None

    # Hash già visti, letti una volta sola (non un find_one per file nel ciclo):
    # stesso motivo delle altre correzioni in questo file, l'adattatore non ha
    # indici. Proiezioni ad esclusione: pdf_data può essere grande su entrambe
    # le collezioni ora che i bonifici stipendio lo portano con sé.
    hash_esistenti = set()
    async for e in db.pagamenti_esiti.find({}, {"_id": 0, "pdf_data": 0, "beneficiario": 0, "causale": 0}):
        if e.get("hash"):
            hash_esistenti.add(e["hash"])
    async for b in db.bonifici_da_associare.find({}, {"_id": 0, "pdf_data": 0}):
        if b.get("hash"):
            hash_esistenti.add(b["hash"])

    importati, in_coda, duplicati, esclusi_non_stipendio, falliti_lettura = 0, 0, 0, 0, []
    affected = set()
    for drive_id, nome_file, raw in scaricati:
        async def segna_visto():
            await db.drive_bonifici_visti.update_one(
                {"id": drive_id}, {"$set": {"id": drive_id, "nome": nome_file, "visto_il": now_iso()}},
                upsert=True)

        if not raw or raw[:4] != b"%PDF":
            falliti_lettura.append(nome_file)
            await segna_visto()
            continue
        h = hashlib.sha256(raw).hexdigest()
        try:
            import pdfplumber
            testo = ""
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for p in pdf.pages[:3]:
                    testo += (p.extract_text() or "") + "\n"
        except Exception:
            testo = ""
        if _e_movimento_non_stipendio(testo) or _e_movimento_non_stipendio(nome_file):
            esclusi_non_stipendio += 1
            await segna_visto()
            continue
        if h in hash_esistenti:
            duplicati += 1
            await segna_visto()
            continue
        hash_esistenti.add(h)
        dati = _parse_bonifico_pdf(testo, nome_file)
        dip = None
        if dati["n_beneficiari"] == 1:
            dip = trova_dipendente(testo, nome_file)
        pdf_b64 = base64.b64encode(raw).decode()

        if dip and dati["importo"] and dati["mese"] and dati["anno"] and dati["causale_stipendio"]:
            key = f"drive:{h[:24]}"
            await db.pagamenti_esiti.update_one(
                {"key": key},
                {"$set": {"key": key, "hash": h, "cro": None, "dipendente_id": dip["id"],
                          "data": dati["data"].strftime("%Y-%m-%d") if dati["data"] else None,
                          "importo": dati["importo"], "causale": dati["causale"],
                          "beneficiario": dip.get("nome_completo") or f"{dip.get('cognome','')} {dip.get('nome','')}".strip(),
                          "mese": dati["mese"], "anno": dati["anno"], "origine": "drive-pdf",
                          "pdf_data": pdf_b64, "ha_pdf": True}},
                upsert=True)
            await db.paghe_mensili.update_one(
                {"dipendente_id": dip["id"], "anno": dati["anno"], "mese": dati["mese"]},
                {"$set": {"dipendente_id": dip["id"], "anno": dati["anno"], "mese": dati["mese"],
                          "updated_at": now_iso()}}, upsert=True)
            affected.add((dip["id"], dati["mese"], dati["anno"]))
            importati += 1
        else:
            await db.bonifici_da_associare.insert_one({
                "id": str(uuid.uuid4()), "hash": h,
                "data": dati["data"].strftime("%Y-%m-%d") if dati["data"] else None,
                "importo": dati["importo"], "causale": dati["causale"] or nome_file,
                "pdf_filename": nome_file, "pdf_data": pdf_b64,
                "fonte": "drive-pdf", "stato": "da_associare", "created_at": now_iso(),
            })
            in_coda += 1
        await segna_visto()

    for dip_id, mese, anno in affected:
        tot = 0.0
        async for p in db.pagamenti_esiti.find({"dipendente_id": dip_id, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1, "pdf_data": 0}):
            tot += p.get("importo") or 0
        await db.paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": anno, "mese": mese},
            {"$set": {"bonifico_importo": round(tot, 2), "bonifico_ricevuto": tot > 0,
                      "bonifico_da_esiti": True, "updated_at": now_iso()}})
        await _ricalcola_stato_paga(db, dip_id, anno, mese)

    falliti_nomi = [id_per_nome.get(fid, fid) for fid in falliti_id]
    return {"trovati_totale": len(tutti), "lavorati_ora": len(lotto),
            "restanti": max(0, len(da_fare) - len(lotto)),
            "importati": importati, "in_coda_da_associare": in_coda, "duplicati": duplicati,
            "esclusi_non_stipendio": esclusi_non_stipendio,
            "falliti_download": falliti_nomi[:50], "falliti_lettura": falliti_lettura[:50]}


@router.delete("/paghe")
async def delete_pagha(dipendente_id: str, anno: int, mese: int):
    res = await get_db().paghe_mensili.delete_one(
        {"dipendente_id": dipendente_id, "anno": int(anno), "mese": int(mese)})
    return {"ok": True, "eliminati": res.deleted_count}


_MESI_IMPORT_SALARI = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
    "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}


def _mese_import_salari(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        numeric = Decimal(str(value))
        return int(numeric) if numeric == numeric.to_integral_value() and 1 <= numeric <= 12 else None
    text = str(value).strip().lower()
    if text in _MESI_IMPORT_SALARI:
        return _MESI_IMPORT_SALARI[text]
    try:
        numeric = Decimal(text)
    except InvalidOperation:
        return None
    return int(numeric) if numeric == numeric.to_integral_value() and 1 <= numeric <= 12 else None


def _importo_import_salari(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


@router.post("/paghe/importa-excel-salari")
async def importa_excel_salari(file: UploadFile = File(...)):
    """Importa l'Excel 'prima nota salari' (colonne: DIPENDENTE, MESE, ANNO,
    STIPENDIO NETTO, IMPORTO EROGATO). Il netto fissa il valore atteso della busta,
    l'erogato il valore atteso del bonifico. I flag di riconciliazione partono a False
    e diventano True quando arriva il PDF (busta o ricevuta bonifico) con importo che
    combacia. I dipendenti non presenti in anagrafica vengono solo segnalati."""
    import openpyxl
    nome_file = (file.filename or "").lower()
    if not nome_file.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Serve un file Excel (.xlsx)")

    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel non leggibile: {e}")
    ws = wb.active

    dips = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    anag = {}
    for d in dips:
        cg = (d.get("cognome") or "").upper().strip()
        nm = (d.get("nome") or "").upper().strip()
        if cg or nm:
            anag[f"{cg} {nm}".strip()] = d   # Cognome Nome
            anag[f"{nm} {cg}".strip()] = d   # Nome Cognome (ordine invertito)

    try:
        await get_db().paghe_mensili.create_index(
            [("dipendente_id", 1), ("anno", 1), ("mese", 1)], unique=True, name="uniq_dip_anno_mese")
    except Exception:
        pass

    importati, mesi_set, righe_lette = 0, set(), 0
    non_trovati, scartati = {}, []
    aggregati = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        righe_lette += 1
        nome = str(row[0]).strip()
        mese = _mese_import_salari(row[1]) if len(row) > 1 else None
        try:
            anno_value = Decimal(str(row[2])) if len(row) > 2 and row[2] not in (None, "") else None
            anno = int(anno_value) if anno_value is not None and anno_value == anno_value.to_integral_value() else None
        except Exception:
            anno = None
        netto = _importo_import_salari(row[3]) if len(row) > 3 else None
        erogato = _importo_import_salari(row[4]) if len(row) > 4 else None

        dip = anag.get(nome.upper())
        if not dip:
            non_trovati[nome] = non_trovati.get(nome, 0) + 1
            continue
        if not mese or not anno or anno not in ANNI_AMMESSI:
            scartati.append({"nome": nome, "motivo": f"periodo non valido ({row[1]} {row[2]})"})
            continue

        key = (dip["id"], anno, mese)
        aggregato = aggregati.setdefault(key, {
            "netto": None,
            "erogato": Decimal("0"),
            "erogato_presente": False,
        })
        if netto is not None:
            if aggregato["netto"] is None:
                aggregato["netto"] = netto
            elif aggregato["netto"] != netto:
                scartati.append({
                    "nome": nome,
                    "motivo": f"stipendio netto incoerente nello stesso periodo ({aggregato['netto']} / {netto})",
                })
        if erogato is not None:
            aggregato["erogato"] += erogato
            aggregato["erogato_presente"] = True

    for (dipendente_id, anno, mese), aggregato in aggregati.items():
        netto = aggregato["netto"]
        erogato = aggregato["erogato"]
        set_doc = {
            "dipendente_id": dipendente_id,
            "anno": anno,
            "mese": mese,
            "fonte_excel": True,
            "updated_at": now_iso(),
        }
        if netto is not None:
            set_doc["importo_busta"] = float(netto)
            set_doc["netto_atteso"] = float(netto)
        if aggregato["erogato_presente"]:
            set_doc["bonifico_importo"] = float(erogato)
            set_doc["erogato_atteso"] = float(erogato)
            set_doc["bonifico_da_prima_nota"] = erogato > 0
        await get_db().paghe_mensili.update_one(
            {"dipendente_id": dipendente_id, "anno": anno, "mese": mese},
            {"$set": set_doc,
             "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
            upsert=True)
        await _ricalcola_stato_paga(get_db(), dipendente_id, anno, mese)
        importati += 1
        mesi_set.add((anno, mese))

    mesi = sorted([{"anno": y, "mese": m} for (y, m) in mesi_set], key=lambda x: (x["anno"], x["mese"]))
    return {"importati": importati,
            "righe_lette": righe_lette,
            "righe_aggregate": len(aggregati),
            "mesi": mesi,
            "dipendenti_non_in_anagrafica": [{"nome": k, "righe": v} for k, v in sorted(non_trovati.items())],
            "scartati": scartati}


# --- Import automatico Libro Unico (PDF) → divide per dipendente e memorizza i netti ---

_CF_RE = re.compile(r'\b([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])\b')
_MESI = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
         "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}
# Regola atomica: si importano SOLO questi anni. Tutto il resto è bloccato.
ANNI_AMMESSI = {2023, 2024, 2025, 2026}

def _lul_netto(text):
    m = re.findall(r'([\d]{1,3}(?:\.\d{3})*,\d{2})\s*€', text)
    return m[-1] if m else None

def _lul_acconto(text):
    """Rileva acconti/anticipi erogati durante il mese e trattenuti nel cedolino
    (righe con 'acconto', 'anticipo', 'rec. acconto', escluso il TFR). Serve a sapere
    quanto è già stato dato, così il bonifico del solo saldo chiude comunque la busta.
    Ritorna l'importo totale o None."""
    tot = 0.0
    for line in text.split("\n"):
        low = line.lower()
        if ("acconto" in low or "anticipo" in low) and "tfr" not in low and "trattamento fine" not in low:
            nums = re.findall(r'([\d]{1,3}(?:\.\d{3})*,\d{2})', line)
            if nums:
                tot += _to_float(nums[-1]) or 0
    return round(tot, 2) if tot > 0 else None

def _acconto_cedolino_plausibile(acconto, netto_totale):
    """Filtro anti-falso-positivo: '_lul_acconto' può intercettare per errore una
    trattenuta minima non correlata (es. quota associativa, conguaglio di pochi
    euro) che contiene la parola 'acconto'/'anticipo' ma non è un vero anticipo
    già erogato al dipendente. Un acconto reale è una cifra significativa
    rispetto al netto, non poche decine di euro: sotto soglia lo scartiamo
    invece di mostrare un 'saldo da pagare' sbagliato per pochi euro di
    differenza (stessa soglia usata in services/libro_unico_parser.py)."""
    if not acconto or acconto <= 0 or not netto_totale:
        return False
    soglia = max(80.0, netto_totale * 0.10)
    return acconto >= soglia

def _lul_periodo(text):
    m = re.search(r'(Gennaio|Febbraio|Marzo|Aprile|Maggio|Giugno|Luglio|Agosto|Settembre|Ottobre|Novembre|Dicembre)\s+(\d{4})', text, re.I)
    if m:
        return _MESI[m.group(1).lower()], int(m.group(2))
    return None, None

_LUL_NUM = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2,6}|-?\d+,\d{2,6}')


def _lul_dati_busta(text: str) -> dict:
    """Estrae dal testo della busta i dati chiave (per codice voce o descrizione).
    Robusto sul prefisso (C/F/Z…). L'ultimo numero della riga voce = importo competenza."""
    voci = []
    voci_obj = []
    for line in text.split("\n"):
        m = re.match(r'^\s*([A-Z]\d{4,5})\b\s*(.*)$', line)
        if m:
            resto, valori = m.group(2), _LUL_NUM.findall(m.group(2))
            voci.append((m.group(1), resto, valori))
            voci_obj.append({"codice": m.group(1),
                             "descrizione": _LUL_NUM.split(resto)[0].strip(' .-'),
                             "valori": valori})

    def find(codici=None, testo=None):
        for codice, resto, valori in voci:
            if (codici and codice in codici) or (testo and testo.lower() in resto.lower()):
                return valori[-1] if valori else None
        return None

    dati = {
        "rateo_13ma": find(codici={"C50000", "Z50000"}, testo="13ma Mensilit"),
        "rateo_14ma": find(codici={"C50022", "Z50022"}, testo="14ma Mensilit"),
        "indennita_l207_24": find(codici={"F02703"}),
        "indennita_l207_24_cng_ann": find(codici={"F09088"}),
        "tratt_integrativo_l21": find(codici={"F09081"}),
        "tratt_integrativo_l21_rata": find(codici={"F09083"}),
        "tratt_integrativo_l21_cng": find(codici={"F09084"}),
        # tutte le voci del cedolino (codici+descrizione+importi) per il motore di ricerca
        "voci": voci_obj or None,
    }
    # Rimborso da 730 (residuo + importo del mese)
    for codice, resto, valori in voci:
        if "730" in resto:
            dati["rimborso_730"] = valori[-1] if valori else None
            if len(valori) >= 2:
                dati["rimborso_730_residuo"] = valori[0]
            break
    # Ore lavorate + giorni retribuiti (riquadro 'Lavorato', best effort)
    lav = re.search(r'(?:Lavorato|Ore\s*lavorat\w*)\D{0,15}?(\d{1,3},\d{2})\s+(\d{1,2})\b', text, re.IGNORECASE)
    if lav:
        dati["ore_lavorate"] = lav.group(1)
        dati["giorni_retribuiti"] = lav.group(2)
    # Giorni effettivamente lavorati: righe del foglio presenze con ore (LU 19 6,40 ...)
    gg = set()
    for line in text.split("\n"):
        pm = re.search(r'\b(LU|MA|ME|GI|VE|SA|DO)\s+(\d{1,2})\s+\d{1,2},\d{2}\b', line)
        if pm:
            gg.add(pm.group(2))
    if gg:
        dati["giorni_lavorati"] = len(gg)
    return {k: v for k, v in dati.items() if v is not None}


def _parse_lul(pdf_path):
    """Raggruppa le pagine per codice fiscale (gestisce 1, 2 o 3 pagine a dipendente).
    Tiene anche traccia degli indici di pagina di ciascun dipendente, così l'import
    può ritagliare il PDF reale del suo cedolino."""
    import pdfplumber
    ced = {}
    page_count = 0
    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        cur = None
        for idx, page in enumerate(pdf.pages):
            t = page.extract_text() or ""
            cfs = _CF_RE.findall(t)
            mese, anno = _lul_periodo(t)
            if cfs:
                cur = cfs[0]
                d = ced.setdefault(cur, {"nome": None, "netto": None, "mese": None, "anno": None, "pagine": []})
                if mese:
                    d["mese"], d["anno"] = mese, anno
                for line in t.split("\n"):
                    mm = re.search(r'\b0[0-9]{6}\b\s+([A-ZÀ-Ù\' ]{4,}?)\s+[A-Z]{6}\d{2}[A-Z]', line)
                    if mm:
                        d["nome"] = mm.group(1).strip()
                        break
            if cur:
                ced[cur].setdefault("pagine", []).append(idx)
                n = _lul_netto(t)
                if n:
                    ced[cur]["netto"] = n
                acc = _lul_acconto(t)
                if acc:
                    ced[cur]["acconto"] = round((ced[cur].get("acconto") or 0) + acc, 2)
                if not ced[cur].get("mese") and mese:
                    ced[cur]["mese"], ced[cur]["anno"] = mese, anno
                # Dati chiave della busta (rateo 13/14, indennità L.207/24, tratt. integ. L.21, giorni)
                for k, v in _lul_dati_busta(t).items():
                    ced[cur][k] = v
    if len(ced) <= 1 and (not ced or not next(iter(ced.values())).get("netto")):
        try:
            from app.hr.parsers.busta_paga_multi_template import parse_busta_paga_multi
            parsed = parse_busta_paga_multi(pdf_path)
            dipendente = parsed.get("dipendente") or {}
            periodo = parsed.get("periodo") or {}
            totali = parsed.get("totali") or {}
            tax_code = (dipendente.get("codice_fiscale") or "").upper()
            if ced:
                record = next(iter(ced.values()))
            elif tax_code:
                record = ced.setdefault(tax_code, {
                    "nome": None, "netto": None, "mese": None, "anno": None,
                    "pagine": list(range(page_count)),
                })
            else:
                record = None
            if record is not None and parsed.get("parse_success"):
                if totali.get("netto") is not None:
                    record["netto"] = f"{float(totali['netto']):.2f}".replace(".", ",")
                record["nome"] = record.get("nome") or dipendente.get("nome_completo")
                record["mese"] = record.get("mese") or periodo.get("mese")
                record["anno"] = record.get("anno") or periodo.get("anno")
        except Exception:
            pass
    return ced

def _to_float(s):
    if isinstance(s, (int, float)):
        return float(s)
    return float(s.replace(".", "").replace(",", ".")) if s else None


def _ritaglia_pdf(pdf_path, pagine):
    """Estrae le pagine indicate dal PDF originale e le restituisce come bytes:
    è il cedolino reale del singolo dipendente dentro il Libro Unico."""
    import fitz
    src = fitz.open(pdf_path)
    out = fitz.open()
    for i in sorted(set(pagine)):
        if 0 <= i < src.page_count:
            out.insert_pdf(src, from_page=i, to_page=i)
    data = out.tobytes()
    out.close(); src.close()
    return data

def _estrai_testo(pdf_path):
    import pdfplumber
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)

def _classifica_doc(text):
    """Distingue: bonifico (ricevuta bancaria), presenze (LUL ore/timbrature),
    cedolino (busta paga con netto). Default: cedolino (per il Libro Unico multi-dipendente)."""
    T = (text or "").upper()
    if "RICEVUTA PER ORDINANTE" in T or "A VOSTRO DEBITO A FAVORE DI" in T or ("BONIFICO" in T and "IBAN BENEFICIARIO" in T):
        return "bonifico"
    ha_netto = "NETTO DEL MESE" in T or "NETTOSDELSMESE" in T
    if ha_netto:
        return "cedolino"
    if ("PERIODO DI RIFERIMENTO" in T or "TIMBRATURE" in T or "ORE ORDINARIE" in T):
        return "presenze"
    return "cedolino"

def _competenza_da_causale(causale):
    """Estrae mese/anno SOLO se dichiarati esplicitamente nella causale."""
    c = (causale or "").lower()
    m = re.search(r'(gennaio|febbraio|marzo|aprile|maggio|giugno|luglio|agosto|settembre|ottobre|novembre|dicembre)\s*(\d{4})?', c)
    if m:
        return _MESI[m.group(1)], (int(m.group(2)) if m.group(2) else None), True
    m = re.search(r'\b(0?[1-9]|1[0-2])[-/](\d{4})\b', c)
    if m:
        return int(m.group(1)), int(m.group(2)), True
    return None, None, False

def _parse_bonifico(text):
    imp = None
    m = re.search(r'EUR\s+([\d.]+,\d{2})', text) or re.search(r'IMPORTO\s+([\d.]+,\d{2})', text)
    if m:
        imp = _to_float(m.group(1))
    data = None
    md = re.search(r'DATA\s+(\d{2})/(\d{2})/(\d{4})', text)
    if md:
        data = f"{md.group(3)}-{md.group(2)}-{md.group(1)}"
    caus = None
    mc = re.search(r'CAUSALE\s*\n\s*([^\n]+)', text)
    if mc:
        caus = mc.group(1).strip()
    cro = None
    mr = re.search(r'(MB0B\w+)', text)
    if mr:
        cro = mr.group(1).strip()
    mese_c, anno_c, esplicita = _competenza_da_causale(caus)
    is_tfr = bool(re.search(r'\btfr\b|trattamento fine rapporto|anticipo\s+t\.?f\.?r', (caus or "").lower()))
    return {"importo": imp, "data": data, "causale": caus, "cro": cro,
            "mese_causale": mese_c, "anno_causale": anno_c, "esplicita": esplicita, "is_tfr": is_tfr}


async def _importa_documenti(pdf_items, errori_iniziali=None, forza=False):
    """Pipeline condivisa: riceve una lista di (origine, pdf_bytes) già espansi (da upload
    file o da posta elettronica), li classifica e li importa in paghe_mensili / prestiti.
    L'anti-duplicazione per hash evita di re-importare gli stessi documenti."""
    dips = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    by_cf = {(d.get("codice_fiscale") or "").upper(): d for d in dips if d.get("codice_fiscale")}
    by_nome = {}
    for d in dips:
        cg = (d.get("cognome") or "").upper().strip()
        nm = (d.get("nome") or "").upper().strip()
        if cg or nm:
            by_nome[f"{cg} {nm}".strip()] = d
            by_nome[f"{nm} {cg}".strip()] = d

    # Vincolo: una sola busta per (dipendente, anno, mese) — i duplicati diventano impossibili
    try:
        await get_db().paghe_mensili.create_index(
            [("dipendente_id", 1), ("anno", 1), ("mese", 1)], unique=True, name="uniq_dip_anno_mese")
    except Exception:
        pass
    # Registro documenti importati (anti-duplicazione): impronta del file + chiave logica
    try:
        await get_db().documenti_importati.create_index([("hash", 1)], unique=True, name="uniq_hash")
        await get_db().documenti_importati.create_index([("chiave", 1)], name="idx_chiave")
    except Exception:
        pass

    async def _registra_doc(h, tipo, chiave, origine):
        try:
            await get_db().documenti_importati.update_one(
                {"hash": h},
                {"$set": {"hash": h, "tipo": tipo, "chiave": chiave, "file": origine,
                          "imported_at": now_iso()}}, upsert=True)
        except Exception:
            pass

    async def _imputa_competenza(dip_id, b):
        """Determina (mese, anno, fonte) di competenza del bonifico secondo le regole:
        1) mese esplicito in causale; 2) match per importo con la busta (acconto=busta o
        somma cumulativa=busta) nella finestra mese precedente→mese stesso; 3) ripiego sul
        mese precedente. Sfondamento d'anno (gen→dic anno prima) solo dal 2024 (2023 blindato)."""
        if b["esplicita"] and b["mese_causale"]:
            anno = b["anno_causale"] or (int(b["data"][:4]) if b.get("data") else None)
            return b["mese_causale"], anno, "causale"
        data = b.get("data")
        if not data:
            return None, None, "data assente"
        y, mo = int(data[:4]), int(data[5:7])
        pm, py = (mo - 1, y) if mo > 1 else (12, y - 1)
        finestra = []
        if not (mo == 1 and py < 2023):   # 2023 blindato: gennaio 2023 non sfonda a dic 2022
            finestra.append((py, pm))     # mese precedente (priorità)
        finestra.append((y, mo))          # mese stesso
        for (a, m) in finestra:
            rec = await get_db().paghe_mensili.find_one(
                {"dipendente_id": dip_id, "anno": a, "mese": m})
            if not rec:
                continue
            busta = rec.get("importo_busta") or rec.get("netto_atteso")
            if busta:
                if abs(busta - b["importo"]) <= 1:
                    return m, a, "importo (= busta)"
                # acconto già dato nel mese (rilevato nel cedolino o bonifico precedente):
                # acconto + questo bonifico = busta  ->  saldo che chiude la busta
                gia = (rec.get("bonifico_importo") or 0) + (rec.get("acconto_cedolino") or 0)
                if abs((gia + b["importo"]) - busta) <= 1:
                    return m, a, "importo (acconto+saldo = busta)"
                # il bonifico copre esattamente il saldo residuo dopo l'acconto
                residuo = rec.get("saldo_residuo")
                if residuo and abs(residuo - b["importo"]) <= 1:
                    return m, a, "importo (= saldo dopo acconto)"
        a, m = finestra[0]
        return m, a, "mese precedente (dedotta)"

    async def _processa_pdf(pdfbytes, origine):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdfbytes)
            path = tmp.name
        ass, dac, bon, pres, dup, tfr, prestiti = [], [], [], [], [], [], []
        # Anti-duplicazione 1: stesso file già importato (impronta del contenuto)
        h = hashlib.sha256(pdfbytes).hexdigest()
        try:
            if not forza and await get_db().documenti_importati.find_one({"hash": h}):
                dup.append({"file": origine, "motivo": "documento già importato (stesso file)"})
                try: os.unlink(path)
                except Exception: pass
                return ass, dac, bon, pres, dup, tfr, prestiti
        except Exception:
            pass
        try:
            text = _estrai_testo(path)
            tipo = _classifica_doc(text)

            # ---- BONIFICO (ricevuta bancaria) ----
            if tipo == "bonifico":
                b = _parse_bonifico(text)
                # match dipendente: "COGNOME NOME" presente nel testo; fallback cognome nella causale
                T = text.upper()
                dip = None
                for cand in dips:
                    cg = (cand.get("cognome") or "").upper().strip()
                    nm = (cand.get("nome") or "").upper().strip()
                    if cg and nm and f"{cg} {nm}" in T:
                        dip = cand; break
                if not dip:
                    cau = (b.get("causale") or "").upper()
                    for cand in dips:
                        cg = (cand.get("cognome") or "").upper().strip()
                        if cg and cg in cau:
                            dip = cand; break
                manca = []
                if not dip: manca.append("dipendente non riconosciuto")
                if not b.get("importo"): manca.append("importo")
                if manca:
                    dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                "motivo": "bonifico: " + ", ".join(manca)})
                else:
                    caus_low = (b.get("causale") or "").lower()
                    if "prestito" in caus_low:
                        # PRESTITO: non imputare a buste paga; mastrino prestiti con saldo progressivo
                        data = b.get("data")
                        if not data:
                            dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                        "motivo": "prestito: data assente"})
                        elif int(data[:4]) not in ANNI_AMMESSI:
                            dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                        "motivo": f"anno {data[:4]} non ammesso — bloccato (solo 2023-2026)"})
                        else:
                            pa, pm = int(data[:4]), int(data[5:7])
                            cro = b.get("cro")
                            gia = await get_db().documenti_importati.find_one({"chiave": f"cro:{cro}"}) if cro else None
                            if gia:
                                dup.append({"file": origine, "motivo": f"prestito già importato (CRO {cro})"})
                            else:
                                await get_db().prestiti_dipendenti.insert_one({
                                    "id": str(uuid.uuid4()), "dipendente_id": dip["id"],
                                    "importo": b["importo"], "data": data, "mese": pm, "anno": pa,
                                    "causale": b.get("causale"), "cro": cro, "pdf": origine,
                                    "created_at": now_iso()})
                                saldo = await _ricalcola_saldo_prestiti(dip["id"])
                                await _registra_doc(h, "prestito",
                                    f"cro:{cro}" if cro else f"pre:{dip['id']}:{pa}:{pm}:{b['importo']}", origine)
                                prestiti.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                                 "importo": b["importo"], "mese": pm, "anno": pa,
                                                 "data": data, "saldo": saldo})
                        return ass, dac, bon, pres, dup, tfr, prestiti
                    mese, anno, fonte = await _imputa_competenza(dip["id"], b)
                    if not mese or not anno:
                        dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                    "motivo": "bonifico: competenza non determinabile"})
                    elif anno not in ANNI_AMMESSI:
                        dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                    "motivo": f"anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    else:
                        cro = b.get("cro")
                        gia = await get_db().documenti_importati.find_one({"chiave": f"cro:{cro}"}) if cro else None
                        if gia:
                            dup.append({"file": origine, "motivo": f"bonifico già importato (CRO {cro})"})
                        elif b.get("is_tfr"):
                            # Anticipo TFR: fuori dal saldo stipendi
                            await get_db().paghe_mensili.update_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                                {"$set": {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                                          "tfr_anticipo_importo": b["importo"], "tfr_anticipo_data": b.get("data"),
                                          "tfr_anticipo_pdf": origine, "updated_at": now_iso()},
                                 "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
                                upsert=True)
                            await _registra_doc(h, "tfr", f"cro:{cro}" if cro else f"tfr:{dip['id']}:{anno}:{mese}", origine)
                            tfr.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                        "importo": b["importo"], "mese": mese, "anno": anno, "data": b.get("data")})
                        else:
                            esist = await get_db().paghe_mensili.find_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese}, {"erogato_atteso": 1})
                            atteso = (esist or {}).get("erogato_atteso")
                            discrep = atteso if (atteso is not None and abs(atteso - b["importo"]) > 1) else None
                            set_doc = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                                       "bonifico_importo": b["importo"], "bonifico_data": b.get("data"),
                                       "bonifico_ricevuto": True, "bonifico_causale": b.get("causale"),
                                       "bonifico_cro": cro, "bonifico_pdf": origine,
                                       "bonifico_riconciliato": True, "updated_at": now_iso()}
                            await get_db().paghe_mensili.update_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                                {"$set": set_doc, "$setOnInsert": {"busta_riconciliata": False}}, upsert=True)
                            await _registra_doc(h, "bonifico", f"cro:{cro}" if cro else f"bon:{dip['id']}:{anno}:{mese}", origine)
                            bon.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                        "importo": b["importo"], "mese": mese, "anno": anno,
                                        "causale": b.get("causale"), "data": b.get("data"),
                                        "riconciliato": True, "discrepanza": discrep, "fonte": fonte})
                return ass, dac, bon, pres, dup, tfr, prestiti

            # ---- FOGLIO PRESENZE (ore/timbrature, non è una busta) ----
            if tipo == "presenze":
                cf = (_CF_RE.findall(text) or [None])[0]
                mese, anno = _lul_periodo(text)
                if anno and anno not in ANNI_AMMESSI:
                    dac.append({"nome": cf or "?", "origine": origine,
                                "motivo": f"presenze anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    return ass, dac, bon, pres, dup, tfr, prestiti
                dip = by_cf.get((cf or "").upper())
                await _registra_doc(h, "presenze", f"pres:{cf}:{anno}:{mese}", origine)
                pres.append({"dipendente": (f"{dip.get('cognome')} {dip.get('nome')}".strip() if dip else (cf or "?")),
                             "mese": mese, "anno": anno, "origine": origine})
                return ass, dac, bon, pres, dup, tfr, prestiti

            # ---- CEDOLINO / LIBRO UNICO multi-dipendente (netti) ----
            ced = _parse_lul(path)
            for cf, info in ced.items():
                dip = by_cf.get(cf)
                metodo = "codice fiscale"
                if not dip:
                    dip = by_nome.get((info.get("nome") or "").upper())
                    metodo = "nome (CF non combacia)"
                netto = _to_float(info.get("netto"))
                mese, anno = info.get("mese"), info.get("anno")
                if not dip or not mese:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": "dipendente non trovato" if not dip else "periodo non rilevato"})
                    continue
                if not netto or netto <= 0:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": "netto non rilevato (non salvato)"})
                    continue
                if anno not in ANNI_AMMESSI:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": f"anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    continue
                esistente = await get_db().paghe_mensili.find_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese}, {"netto_atteso": 1})
                atteso = (esistente or {}).get("netto_atteso")
                discrep = atteso if (atteso is not None and abs(atteso - netto) > 1) else None
                acconto = info.get("acconto")
                acconto_valido = _acconto_cedolino_plausibile(acconto, netto)
                set_doc = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                           "importo_busta": netto, "busta_da_lul": True,
                           "busta_riconciliata": True, "updated_at": now_iso()}
                if acconto_valido:
                    set_doc["acconto_cedolino"] = acconto
                    set_doc["saldo_residuo"] = round(netto - acconto, 2)
                await get_db().paghe_mensili.update_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                    {"$set": set_doc}, upsert=True)
                # Motore unico: busta arrivata → aggancia il pagamento o la mette in attesa
                await _ricalcola_stato_paga(get_db(), dip["id"], anno, mese)
                # Cedolino (fonte del portale): salvo il PDF REALE ritagliato dal Libro
                # Unico + il netto, così il dipendente scarica la sua busta vera.
                ced_set = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                           "netto": netto,
                           "dipendente_nome": f"{dip.get('cognome','')} {dip.get('nome','')}".strip(),
                           "updated_at": now_iso()}
                if acconto_valido:
                    ced_set["acconto_cedolino"] = acconto
                    ced_set["saldo_residuo"] = round(netto - acconto, 2)
                # Dati chiave estratti dalla busta (salvati nel cedolino)
                for k in ("rateo_13ma", "rateo_14ma", "indennita_l207_24",
                          "indennita_l207_24_cng_ann", "tratt_integrativo_l21",
                          "tratt_integrativo_l21_rata", "tratt_integrativo_l21_cng",
                          "rimborso_730", "rimborso_730_residuo",
                          "ore_lavorate", "giorni_retribuiti", "giorni_lavorati", "voci"):
                    if info.get(k) is not None:
                        ced_set[k] = info[k]
                try:
                    if info.get("pagine"):
                        ced_set["pdf_data"] = base64.b64encode(_ritaglia_pdf(path, info["pagine"])).decode()
                        ced_set["filename"] = f"busta_{anno}_{str(mese).zfill(2)}.pdf"
                except Exception:
                    pass
                await get_db().cedolini.update_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                    {"$set": ced_set,
                     "$setOnInsert": {"id": generate_id(), "created_at": now_iso(), "stato": "importato"}},
                    upsert=True)
                ass.append({"dipendente_id": dip["id"],
                            "dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                            "netto": netto, "metodo": metodo, "mese": mese, "anno": anno,
                            "riconciliata": True, "discrepanza": discrep,
                            "acconto": acconto, "saldo_residuo": (round(netto - acconto, 2) if acconto else None)})
            if ass:
                await _registra_doc(h, "cedolino", f"file:{origine}", origine)
        finally:
            try:
                os.unlink(path)
            except Exception:
                pass
        return ass, dac, bon, pres, dup, tfr, prestiti

    associati, da_controllare, errori, bonifici, presenze, duplicati, tfr_list, prestiti_list = [], [], [], [], [], [], [], []
    errori = list(errori_iniziali or [])
    file_pdf = 0
    for (nome, data) in pdf_items:
        try:
            a, d, b, p, du, tf, pr = await _processa_pdf(data, nome)
            associati += a; da_controllare += d; bonifici += b; presenze += p; duplicati += du; tfr_list += tf; prestiti_list += pr; file_pdf += 1
        except Exception as e:
            errori.append(f"{nome}: {e}")

    if file_pdf == 0:
        raise HTTPException(status_code=400, detail="Nessun PDF elaborabile. " + ("; ".join(errori) if errori else ""))

    # Dedup: se lo stesso dipendente/mese è arrivato da più file, tieni una riga sola
    visti = {}
    for a in associati:
        visti[(a["dipendente_id"], a["anno"], a["mese"])] = a
    associati = list(visti.values())

    mesi_set = sorted({(a["anno"], a["mese"]) for a in associati})
    mesi = [{"anno": y, "mese": m, "n": sum(1 for a in associati if a["anno"] == y and a["mese"] == m)}
            for (y, m) in mesi_set]
    associati.sort(key=lambda x: (x["anno"], x["mese"], x["dipendente"]))
    bonifici.sort(key=lambda x: (x["anno"], x["mese"], x["dipendente"]))
    return {"associati": associati, "da_controllare": da_controllare,
            "totale_associati": len(associati), "file_pdf": file_pdf,
            "mesi": mesi, "errori": errori,
            "bonifici": bonifici, "presenze": presenze, "duplicati": duplicati, "tfr": tfr_list, "prestiti": prestiti_list}


async def _ricalcola_saldo_prestiti(dip_id):
    """Riporto continuo: azzera i campi prestito da tutti i mesi del dipendente, poi somma
    i movimenti in ordine cronologico riscrivendo erogato del mese e saldo cumulativo.
    Ritorna il saldo totale corrente."""
    await get_db().paghe_mensili.update_many(
        {"dipendente_id": dip_id},
        {"$unset": {"prestito_importo": "", "prestito_saldo": ""}})
    movs = await get_db().prestiti_dipendenti.find({"dipendente_id": dip_id}).to_list(2000)
    erog = {}
    for mv in movs:
        k = (mv["anno"], mv["mese"])
        erog[k] = erog.get(k, 0) + (mv.get("importo") or 0)
    saldo = 0
    for (a, m) in sorted(erog.keys()):
        saldo += erog[(a, m)]
        await get_db().paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": a, "mese": m},
            {"$set": {"dipendente_id": dip_id, "anno": a, "mese": m,
                      "prestito_importo": erog[(a, m)], "prestito_saldo": saldo,
                      "updated_at": now_iso()},
             "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
            upsert=True)
    return saldo


@router.get("/_unif_diag")
async def diagnostica_unificazione():
    """SOLA LETTURA. Fotografa cedolini vs paghe_mensili per pianificare l'unificazione:
    conteggi, sovrapposizioni per (dipendente_id, anno, mese), confronto netto vs importo_busta,
    record presenti solo in paghe_mensili, e dump completo di paghe_mensili per backup."""
    db = get_db()
    ced = await db.cedolini.find({}, {"_id": 0, "pdf_data": 0}).to_list(5000)
    pm = await db.paghe_mensili.find({}, {"_id": 0}).to_list(5000)
    ced_idx = {}
    for c in ced:
        ced_idx.setdefault((c.get("dipendente_id"), c.get("anno"), c.get("mese")), c)
    solo_in_pm, con_match, mismatch_netto = [], 0, []
    for p in pm:
        k = (p.get("dipendente_id"), p.get("anno"), p.get("mese"))
        c = ced_idx.get(k)
        if not c:
            solo_in_pm.append({"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")})
        else:
            con_match += 1
            nb = p.get("importo_busta") or p.get("netto_atteso")
            nc = c.get("netto")
            if nb is not None and nc is not None and abs(float(nb) - float(nc)) > 1:
                mismatch_netto.append({"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"),
                                       "mese": p.get("mese"), "paghe_mensili": nb, "cedolini": nc})
    # campi accessori presenti in paghe_mensili (riconciliazione)
    campi = set()
    for p in pm:
        campi.update(p.keys())
    pm_con_riconciliazione = [p for p in pm if any(p.get(k) for k in
        ("bonifico_importo", "acconti", "prestito_importo", "tfr_anticipo_importo",
         "busta_riconciliata", "bonifico_riconciliato"))]
    return {
        "cedolini_totali": len(ced),
        "paghe_mensili_totali": len(pm),
        "paghe_mensili_con_match_in_cedolini": con_match,
        "paghe_mensili_solo_loro": solo_in_pm,
        "mismatch_netto": mismatch_netto,
        "campi_presenti_in_paghe_mensili": sorted(campi),
        "paghe_mensili_con_dati_riconciliazione": len(pm_con_riconciliazione),
        "backup_paghe_mensili": pm,
    }


_RICON_FIELDS = ["bonifico_importo", "bonifico_data", "bonifico_ricevuto", "bonifico_causale",
                 "bonifico_cro", "bonifico_pdf", "bonifico_riconciliato", "busta_riconciliata",
                 "busta_da_lul", "acconti", "acconto_cedolino", "saldo_residuo", "netto_atteso",
                 "erogato_atteso", "fonte_excel", "tfr_anticipo_importo", "tfr_anticipo_data",
                 "tfr_anticipo_pdf", "prestito_importo", "prestito_saldo"]


@router.post("/_unif_esegui")
async def esegui_unificazione(dry_run: bool = True, limit: int = 25):
    """Unifica paghe_mensili dentro cedolini, A PICCOLI BATCH leggeri: processa solo i record
    non ancora migrati (flag _migrato sul documento paghe_mensili), così ogni chiamata è veloce.
    Chiamare ripetutamente finché completato=True. NON cancella paghe_mensili."""
    db = get_db()
    pendenti = await db.paghe_mensili.find({"_migrato": {"$ne": True}}).to_list(limit)
    if not pendenti:
        return {"dry_run": dry_run, "fatti_ora": 0, "restanti_da_fare": 0, "completato": True}
    arricchiti = creati = saltati = 0
    for p in pendenti:
        k = {"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")}
        ricon = {f: p[f] for f in _RICON_FIELDS if f in p and p[f] is not None}
        c = await db.cedolini.find_one(k, {"_id": 0, "id": 1})
        if dry_run:
            if c: arricchiti += 1
            elif (p.get("importo_busta") or p.get("netto_atteso") or 0) > 0: creati += 1
            else: saltati += 1
            continue
        if c:
            upd = dict(ricon); upd["unif_arricchito"] = True
            await db.cedolini.update_one({"id": c["id"]}, {"$set": upd})
            arricchiti += 1
        else:
            netto = p.get("importo_busta") or p.get("netto_atteso")
            if not netto or float(netto) <= 0:
                saltati += 1
            else:
                dip = await db.dipendenti.find_one({"id": p.get("dipendente_id")},
                                                   {"_id": 0, "nome": 1, "cognome": 1, "nome_completo": 1})
                nome = (dip or {}).get("nome_completo") or (f"{(dip or {}).get('cognome','')} {(dip or {}).get('nome','')}".strip() if dip else "")
                nuovo = {"id": str(uuid.uuid4()), "dipendente_id": p.get("dipendente_id"),
                         "dipendente_nome": nome, "anno": p.get("anno"), "mese": p.get("mese"),
                         "netto": float(netto), "stato": "importato",
                         "origine_unificazione": True, "unif_arricchito": True, "created_at": now_iso()}
                nuovo.update(ricon)
                await db.cedolini.insert_one(nuovo)
                creati += 1
        await db.paghe_mensili.update_one(
            {"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")},
            {"$set": {"_migrato": True}})
    restanti = await db.paghe_mensili.count_documents({"_migrato": {"$ne": True}})
    return {"dry_run": dry_run, "arricchiti_ora": arricchiti, "creati_ora": creati,
            "saltati_ora": saltati, "restanti_da_fare": restanti, "completato": restanti == 0}


@router.get("/prestiti")
async def lista_prestiti(dipendente_id: Optional[str] = None):
    """Mastrino prestiti: movimenti con saldo progressivo. Filtrabile per dipendente."""
    q = {"dipendente_id": dipendente_id} if dipendente_id else {}
    movs = await get_db().prestiti_dipendenti.find(q, {"_id": 0}).to_list(2000)
    movs.sort(key=lambda x: (x.get("anno", 0), x.get("mese", 0), x.get("data") or ""))
    # saldo progressivo per dipendente
    saldi = {}
    for mv in movs:
        d = mv["dipendente_id"]
        saldi[d] = saldi.get(d, 0) + (mv.get("importo") or 0)
        mv["saldo"] = saldi[d]
    return movs


@router.delete("/prestiti/{prestito_id}")
async def elimina_prestito(prestito_id: str):
    """Elimina un movimento di prestito e ricalcola il saldo progressivo del dipendente."""
    mv = await get_db().prestiti_dipendenti.find_one({"id": prestito_id})
    if not mv:
        raise HTTPException(status_code=404, detail="Prestito non trovato")
    await get_db().prestiti_dipendenti.delete_one({"id": prestito_id})
    # libera anche l'anti-dup così un eventuale re-import è possibile
    if mv.get("cro"):
        await get_db().documenti_importati.delete_many({"chiave": f"cro:{mv['cro']}"})
    saldo = await _ricalcola_saldo_prestiti(mv["dipendente_id"])
    return {"ok": True, "saldo_aggiornato": saldo}


def _espandi_in_pdf(nome, data):
    """Espande un allegato/file in lista di (origine, pdf_bytes): PDF diretto, oppure
    PDF contenuti in uno ZIP. Ritorna (items, errori)."""
    items, errori = [], []
    low = (nome or "").lower()
    if low.endswith(".pdf"):
        items.append((nome, data))
    elif low.endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                interni = [n for n in z.namelist() if n.lower().endswith(".pdf") and "__MACOSX" not in n]
                if not interni:
                    errori.append(f"{nome}: ZIP senza PDF")
                for zi in interni:
                    items.append((f"{nome} › {zi}", z.read(zi)))
        except zipfile.BadZipFile:
            errori.append(f"{nome}: ZIP non valido")
        except Exception as e:
            errori.append(f"{nome}: {e}")
    else:
        errori.append(f"{nome}: tipo non supportato (servono PDF o ZIP)")
    return items, errori


@router.post("/paghe/importa-lul")
async def importa_libro_unico(files: List[UploadFile] = File(...), forza: bool = False):
    """Importa uno o più PDF (anche dentro ZIP) caricati dall'utente: buste paga,
    fogli presenze, bonifici (acconti, saldi, TFR, prestiti). Vedi _importa_documenti."""
    pdf_items, errori = [], []
    for uf in files:
        nome = uf.filename or ""
        try:
            data = await uf.read()
        except Exception:
            errori.append(f"{nome}: lettura fallita")
            continue
        its, err = _espandi_in_pdf(nome, data)
        pdf_items += its
        errori += err
    if not pdf_items:
        raise HTTPException(status_code=400,
            detail="Nessun PDF valido trovato. " + ("; ".join(errori) if errori else ""))
    return await _importa_documenti(pdf_items, errori, forza=forza)


@router.post("/paghe/importa-email")
async def importa_da_email(cartella: Optional[str] = None, solo_non_letti: bool = False):
    """Scarica gli allegati PDF dalla casella di posta (INBOX + tutte le cartelle) e li
    importa con la stessa pipeline. Credenziali dalle variabili ambiente Render:
    IMAP_HOST, IMAP_PORT (default 993), IMAP_USER, IMAP_PASSWORD.
    L'anti-duplicazione per hash evita di re-importare email già lette in passato."""
    import imaplib, email
    host = os.getenv("IMAP_HOST") or os.getenv("IMAP_SERVER")
    user = os.getenv("IMAP_USER") or os.getenv("IMAP_EMAIL")
    pwd = os.getenv("IMAP_PASSWORD") or os.getenv("IMAP_PASS")
    port = int(os.getenv("IMAP_PORT") or 993)
    mancano = [n for n, v in [("IMAP_HOST", host), ("IMAP_USER", user), ("IMAP_PASSWORD", pwd)] if not v]
    if mancano:
        raise HTTPException(status_code=400,
            detail="Variabili ambiente IMAP mancanti su Render: " + ", ".join(mancano) +
                   ". Servono IMAP_HOST, IMAP_USER, IMAP_PASSWORD (IMAP_PORT opzionale, default 993).")
    try:
        M = imaplib.IMAP4_SSL(host, port)
        M.login(user, pwd)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Connessione/login IMAP fallito ({host}:{port}): {e}")

    pdf_items, errori, cartelle_lette = [], [], []
    try:
        # Elenco cartelle: una specifica se richiesta, altrimenti tutte
        if cartella:
            target = [cartella]
        else:
            target = []
            typ, data = M.list()
            if typ == "OK":
                for raw in data:
                    line = raw.decode(errors="ignore") if isinstance(raw, bytes) else str(raw)
                    # l'ultimo token tra virgolette è il nome cartella
                    nome_c = line.split(' "')[-1].strip().strip('"') if '"' in line else line.split()[-1]
                    if nome_c and "\\Noselect" not in line:
                        target.append(nome_c)
            if "INBOX" not in target:
                target.insert(0, "INBOX")
        for box in target:
            try:
                typ, _ = M.select(f'"{box}"', readonly=True)
                if typ != "OK":
                    continue
                crit = "(UNSEEN)" if solo_non_letti else "ALL"
                typ, msgnums = M.search(None, crit)
                if typ != "OK":
                    continue
                ids = msgnums[0].split()
                cartelle_lette.append({"cartella": box, "messaggi": len(ids)})
                for num in ids:
                    typ, msgdata = M.fetch(num, "(RFC822)")
                    if typ != "OK" or not msgdata or not msgdata[0]:
                        continue
                    msg = email.message_from_bytes(msgdata[0][1])
                    for part in msg.walk():
                        if part.get_content_maintype() == "multipart":
                            continue
                        fn = part.get_filename()
                        if not fn:
                            continue
                        try:
                            payload = part.get_payload(decode=True)
                        except Exception:
                            continue
                        if not payload:
                            continue
                        its, err = _espandi_in_pdf(fn, payload)
                        pdf_items += [(f"[{box}] {o}", d) for (o, d) in its]
                        errori += err
            except Exception as e:
                errori.append(f"cartella {box}: {e}")
    finally:
        try:
            M.logout()
        except Exception:
            pass

    if not pdf_items:
        return {"associati": [], "da_controllare": [], "totale_associati": 0, "file_pdf": 0,
                "mesi": [], "errori": errori, "bonifici": [], "presenze": [], "duplicati": [],
                "tfr": [], "prestiti": [], "cartelle_lette": cartelle_lette,
                "documenti": {"caricati": 0, "non_assegnati": 0, "duplicati": 0},
                "messaggio": "Nessun allegato PDF trovato nella casella."}
    res = await _importa_documenti(pdf_items, errori)
    res["cartelle_lette"] = cartelle_lette
    # Oltre a paghe/bonifici, archivia OGNI allegato nelle cartelle Documenti del dipendente
    # (UNILAV, Certificazione Unica, contratti, codice fiscale…): stesso motore dell'upload massivo.
    db_doc = get_db()
    indici = await _indici_dipendenti(db_doc)
    doc_caricati, doc_non_ass, doc_dup = 0, 0, 0
    for origine, raw in pdf_items:
        try:
            esito, _cat, _nome = await _archivia_documento_cloud(db_doc, origine, raw, indici=indici, origine="email")
            if esito == "caricato":
                doc_caricati += 1
            elif esito == "non_assegnato":
                doc_non_ass += 1
            elif esito == "duplicato":
                doc_dup += 1
        except Exception as e:
            errori.append(f"archivio doc {origine}: {e}")
    res["documenti"] = {"caricati": doc_caricati, "non_assegnati": doc_non_ass, "duplicati": doc_dup}
    return res


# ============ PRESENZE ============

@router.get("/presenze")
async def get_presenze(anno: Optional[int] = None, mese: Optional[int] = None, dipendente_id: Optional[str] = None):
    """
    Recupera presenze dalla collezione 'presenze' (dati storici dal Libro Unico).
    Le presenze sono raggruppate per mese con un array 'giorni'.
    """
    db = get_db()
    
    # Prima leggi da presenze_cloud (inserimenti manuali)
    query_cloud = {}
    if dipendente_id:
        query_cloud["dipendente_id"] = dipendente_id
    if anno and mese:
        query_cloud["data"] = {"$regex": f"^{anno}-{str(mese).zfill(2)}"}
    
    presenze_cloud = await db.presenze_cloud.find(query_cloud, {"_id": 0}).to_list(5000)
    
    # Poi leggi da presenze (dati storici dal LUL - struttura diversa)
    query_lul = {}
    if anno:
        query_lul["anno"] = anno
    if mese:
        query_lul["mese"] = mese
    
    presenze_lul = await db.presenze.find(query_lul, {"_id": 0}).to_list(500)
    
    # Converti presenze LUL in formato giornaliero
    result = list(presenze_cloud)
    cloud_keys = {(p.get("dipendente_id"), p.get("data")) for p in presenze_cloud}
    
    for p_lul in presenze_lul:
        cf = p_lul.get("codice_fiscale", "")
        anno_p = p_lul.get("anno", 2026)
        mese_p = p_lul.get("mese", 1)
        giorni = p_lul.get("giorni", [])
        
        # Trova l'ID dipendente dal codice fiscale
        dip = await db.dipendenti.find_one({"codice_fiscale": cf})
        dip_id = dip.get("id", cf) if dip else cf
        
        for g in giorni:
            giorno_num = g.get("giorno", 1)
            data_str = f"{anno_p}-{str(mese_p).zfill(2)}-{str(giorno_num).zfill(2)}"
            
            key = (dip_id, data_str)
            if key in cloud_keys:
                continue  # Già presente nei dati manuali
            
            # Determina lo stato dal giustificativo
            giust = g.get("giustificativo", "")
            ore = g.get("ore_ordinarie", 0)
            
            if giust:
                stato = giust  # AI, FE, MA, RL, etc.
            elif ore > 0:
                stato = "presente"
            else:
                stato = "assente"
            
            result.append({
                "id": f"{cf}_{data_str}",
                "dipendente_id": dip_id,
                "data": data_str,
                "entrata": None,
                "uscita": None,
                "stato": stato,
                "giustificativo": giust,
                "ore_lavorate": ore,
                "note": ""
            })
    
    return result

@router.post("/presenze")
async def create_presenza(presenza: PresenzaCloud):
    pres_dict = presenza.model_dump()
    pres_dict["id"] = generate_id()
    pres_dict["created_at"] = now_iso()
    
    # Calculate hours worked
    if pres_dict.get("entrata") and pres_dict.get("uscita"):
        try:
            ent = datetime.strptime(pres_dict["entrata"], "%H:%M")
            usc = datetime.strptime(pres_dict["uscita"], "%H:%M")
            pres_dict["ore_lavorate"] = round((usc - ent).seconds / 3600, 2)
        except:
            pass
    
    await get_db().presenze_cloud.insert_one(pres_dict)
    return serialize_doc(pres_dict)

@router.put("/presenze/{presenza_id}")
async def update_presenza(presenza_id: str, presenza: PresenzaCloud):
    pres_dict = presenza.model_dump()
    
    if pres_dict.get("entrata") and pres_dict.get("uscita"):
        try:
            ent = datetime.strptime(pres_dict["entrata"], "%H:%M")
            usc = datetime.strptime(pres_dict["uscita"], "%H:%M")
            pres_dict["ore_lavorate"] = round((usc - ent).seconds / 3600, 2)
        except:
            pass
    
    result = await get_db().presenze_cloud.update_one(
        {"id": presenza_id},
        {"$set": pres_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Presenza non trovata")
    return {"message": "Presenza aggiornata"}

@router.delete("/presenze/{presenza_id}")
async def delete_presenza(presenza_id: str):
    result = await get_db().presenze_cloud.delete_one({"id": presenza_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Presenza non trovata")
    return {"message": "Presenza eliminata"}

@router.post("/presenze/batch")
async def create_presenze_batch(presenze: List[PresenzaCloud]):
    created = []
    for p in presenze:
        pres_dict = p.model_dump()
        pres_dict["id"] = generate_id()
        pres_dict["created_at"] = now_iso()
        
        existing = await get_db().presenze_cloud.find_one({
            "dipendente_id": pres_dict["dipendente_id"],
            "data": pres_dict["data"]
        })
        
        if existing:
            await get_db().presenze_cloud.update_one(
                {"id": existing["id"]},
                {"$set": pres_dict}
            )
        else:
            await get_db().presenze_cloud.insert_one(pres_dict)
        created.append(pres_dict)

    return {"message": f"Inserite/aggiornate {len(created)} presenze"}


_MESI_PRES = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio",
              "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
# Colori codici presenza (RGB 0-1) coerenti col frontend (niente blu/viola)
_COL_PRES = {
    "P": (0.24, 0.51, 0.41), "AS": (0.83, 0.37, 0.30), "F": (0.36, 0.48, 0.42),
    "PE": (0.49, 0.33, 0.15), "M": (0.96, 0.62, 0.04), "R": (0.54, 0.60, 0.36),
    "RS": (0.61, 0.64, 0.58), "CH": (0.42, 0.45, 0.50), "FNL": (0.65, 0.45, 0.29),
    "X": (0.22, 0.25, 0.20),
}


def _csv_presenze(anno, mese, giorni, righe):
    sep = ";"
    intest = ["Dipendente"] + [str(i + 1) for i in range(giorni)]
    out = [f"Presenze {_MESI_PRES[mese - 1]} {anno} - Ceraldi Group S.r.l.", "", sep.join(intest)]
    for r in righe:
        out.append(sep.join([str(r.get("nome", ""))] + [str(c or "") for c in (r.get("celle") or [])]))
    out.append("")
    out.append("Legenda: P=Presente · AS=Assente · F=Ferie · PE=Permesso · M=Malattia · R=ROL · RS=Riposo · CH=Chiuso · FNL=Festivita non lav.")
    return "\n".join(out)


def _pdf_presenze(anno, mese, giorni, righe):
    """Foglio presenze del mese in PDF, UNA SOLA PAGINA orizzontale (A4 landscape)."""
    import fitz
    W, H = 842, 595  # A4 orizzontale in punti
    pdf = fitz.open()
    page = pdf.new_page(width=W, height=H)
    mL, mR, mT, mB = 24, 24, 60, 60
    page.insert_text((mL, 34), f"Presenze {_MESI_PRES[mese - 1]} {anno}", fontsize=15, fontname="hebo")
    page.insert_text((mL, 50), "Ceraldi Group S.r.l.", fontsize=9, fontname="helv", color=(0.4, 0.4, 0.4))

    n = max(1, len(righe))
    name_w = 120
    grid_w = W - mL - mR - name_w
    col_w = grid_w / max(1, giorni)
    grid_h = H - mT - mB
    row_h = min(22, grid_h / (n + 1))
    fs = max(4.5, min(8, row_h - 3))
    x0, y0 = mL, mT

    # Intestazione giorni
    page.insert_text((x0 + 4, y0 - 4), "Dipendente", fontsize=7, fontname="hebo")
    for g in range(giorni):
        cx = x0 + name_w + g * col_w
        page.insert_text((cx + col_w / 2 - 3, y0 - 4), str(g + 1), fontsize=6, fontname="helv", color=(0.4, 0.4, 0.4))

    for ri, r in enumerate(righe):
        ry = y0 + ri * row_h
        # nome
        nome = str(r.get("nome", ""))[:22]
        page.draw_rect(fitz.Rect(x0, ry, x0 + name_w, ry + row_h), color=(0.9, 0.88, 0.83), width=0.3)
        page.insert_text((x0 + 3, ry + row_h - 4), nome, fontsize=fs, fontname="helv")
        celle = r.get("celle") or []
        for g in range(giorni):
            code = str(celle[g]) if g < len(celle) and celle[g] else ""
            cx = x0 + name_w + g * col_w
            rect = fitz.Rect(cx, ry, cx + col_w, ry + row_h)
            col = _COL_PRES.get(code)
            if col:
                page.draw_rect(rect, color=col, fill=col, width=0)
                page.insert_text((cx + col_w / 2 - fs * 0.55, ry + row_h - 4), code, fontsize=fs, fontname="hebo", color=(1, 1, 1))
            else:
                page.draw_rect(rect, color=(0.9, 0.88, 0.83), width=0.3)

    # Legenda in fondo
    ly = H - mB + 16
    page.insert_text((mL, ly), "Legenda:", fontsize=7, fontname="hebo")
    lx = mL + 42
    for code, lab in [("P", "Presente"), ("AS", "Assente"), ("F", "Ferie"), ("PE", "Permesso"),
                      ("M", "Malattia"), ("R", "ROL"), ("RS", "Riposo"), ("CH", "Chiuso"), ("FNL", "Fest.")]:
        col = _COL_PRES.get(code, (0.6, 0.6, 0.6))
        page.draw_rect(fitz.Rect(lx, ly - 7, lx + 9, ly + 1), color=col, fill=col, width=0)
        page.insert_text((lx + 12, ly), f"{code} {lab}", fontsize=6, fontname="helv")
        lx += 62
    return pdf.tobytes()


# ---- Opzione C: documento combinato per il commercialista ----
# (riepilogo totali per dipendente + dettaglio periodi con le date) — vedi
# mockup discusso col titolare: più leggero della griglia giorno-per-giorno,
# che resta per l'uso interno.
_MAPPA_RIEPILOGO = {"P": "lav", "F": "ferie", "PE": "perm", "M": "malat", "R": "rol", "RS": "riposi"}
_CODICI_EVENTO = {"F": "Ferie", "PE": "Permesso", "M": "Malattia", "R": "ROL",
                  "AS": "Assenza", "CH": "Chiusura", "FNL": "Festività"}


def _riepilogo_da_celle(celle):
    cont = {"lav": 0, "ferie": 0, "perm": 0, "malat": 0, "rol": 0, "riposi": 0, "altro": 0}
    for c in celle:
        chiave = _MAPPA_RIEPILOGO.get(c)
        if chiave:
            cont[chiave] += 1
        else:
            cont["altro"] += 1  # AS, CH, FNL, X, cella vuota…
    cont["tot"] = len(celle)
    return cont


def _periodi_da_celle(celle, note=None):
    """Raggruppa le celle in periodi consecutivi per i codici che meritano
    annotazione (il riposo settimanale e la presenza normale sono routine,
    non compaiono)."""
    note = note or []
    eventi, i, n = [], 0, len(celle)
    while i < n:
        code = celle[i] or ""
        if code not in _CODICI_EVENTO:
            i += 1
            continue
        j = i
        while j + 1 < n and (celle[j + 1] or "") == code:
            j += 1
        nota = next((note[k] for k in range(i, j + 1) if k < len(note) and note[k]), "")
        eventi.append({"tipo": code, "label": _CODICI_EVENTO[code],
                       "dal": i + 1, "al": j + 1, "giorni": j - i + 1, "nota": nota})
        i = j + 1
    return eventi


def _pdf_riepilogo_periodi(anno, mese, giorni, righe):
    """Documento 'per il commercialista' (Opzione C): riepilogo totali per
    dipendente + dettaglio dei periodi di assenza con le date, su una o più
    pagine A4 verticali (si estende da sola se i periodi sono tanti)."""
    import fitz
    W, H = 595, 842  # A4 verticale
    mL, mR, mT, mB = 32, 32, 70, 40
    pdf = fitz.open()
    page = pdf.new_page(width=W, height=H)
    y = [mT]

    def intestazione(continua=False):
        page.insert_text((mL, 34), f"Presenze {_MESI_PRES[mese - 1]} {anno}"
                         + (" (continua)" if continua else ""), fontsize=15, fontname="hebo")
        page.insert_text((mL, 50), "Ceraldi Group S.r.l.", fontsize=9, fontname="helv", color=(0.4, 0.4, 0.4))

    def nuova_pagina(continua=True):
        nonlocal page
        page = pdf.new_page(width=W, height=H)
        intestazione(continua)
        y[0] = mT

    def spazio(h):
        if y[0] + h > H - mB:
            nuova_pagina()

    intestazione()

    # ---- Sezione 1: riepilogo totali ----
    page.insert_text((mL, y[0]), "1 · RIEPILOGO DEL MESE", fontsize=9.5, fontname="hebo", color=(0.36, 0.48, 0.42))
    y[0] += 16
    cols = [("Dipendente", 150), ("Lav.", 44), ("Ferie", 44), ("Perm.", 44),
            ("Malat.", 46), ("ROL", 40), ("Riposi", 44), ("Altro", 42), ("Tot.", 40)]
    x = mL
    for lab, w in cols:
        page.insert_text((x + (0 if lab == "Dipendente" else w - 4 - len(lab) * 3.2), y[0]),
                         lab, fontsize=7.5, fontname="hebo", color=(0.42, 0.45, 0.4))
        x += w
    y[0] += 4
    page.draw_line((mL, y[0]), (mL + sum(w for _, w in cols), y[0]), color=(0.85, 0.82, 0.76), width=0.6)
    y[0] += 12
    totali = {"lav": 0, "ferie": 0, "perm": 0, "malat": 0, "rol": 0, "riposi": 0, "altro": 0, "tot": 0}
    riepiloghi = {}
    for r in righe:
        celle = r.get("celle") or []
        rp = _riepilogo_da_celle(celle)
        riepiloghi[r.get("nome", "")] = rp
        for k in totali:
            totali[k] += rp[k]
        spazio(14)
        x = mL
        vals = [r.get("nome", ""), rp["lav"], rp["ferie"], rp["perm"], rp["malat"], rp["rol"], rp["riposi"], rp["altro"], rp["tot"]]
        for (lab, w), v in zip(cols, vals):
            testo = str(v)
            if lab == "Dipendente":
                page.insert_text((x, y[0]), testo[:26], fontsize=8, fontname="helv")
            else:
                page.insert_text((x + w - 4 - len(testo) * 4, y[0]), testo, fontsize=8, fontname="helv")
            x += w
        y[0] += 14
    # riga totale
    spazio(18)
    page.draw_line((mL, y[0] - 4), (mL + sum(w for _, w in cols), y[0] - 4), color=(0.85, 0.82, 0.76), width=0.6)
    x = mL
    vals = ["Totale azienda", totali["lav"], totali["ferie"], totali["perm"], totali["malat"],
            totali["rol"], totali["riposi"], totali["altro"], totali["tot"]]
    for (lab, w), v in zip(cols, vals):
        testo = str(v)
        if lab == "Dipendente":
            page.insert_text((x, y[0]), testo, fontsize=8, fontname="hebo")
        else:
            page.insert_text((x + w - 4 - len(testo) * 4, y[0]), testo, fontsize=8, fontname="hebo")
        x += w
    y[0] += 26

    # ---- Sezione 2: dettaglio periodi ----
    spazio(20)
    page.insert_text((mL, y[0]), "2 · DETTAGLIO DEI PERIODI", fontsize=9.5, fontname="hebo", color=(0.65, 0.45, 0.29))
    y[0] += 6
    page.insert_text((mL, y[0] + 10), "Il riposo settimanale non compare: è regolare e non richiede annotazione.",
                     fontsize=7, fontname="helv", color=(0.5, 0.5, 0.5))
    y[0] += 22
    qualcuno = False
    for r in righe:
        eventi = _periodi_da_celle(r.get("celle") or [], r.get("note") or [])
        if not eventi:
            continue
        qualcuno = True
        # Riservo lo spazio per il nome + almeno il primo evento, così il nome
        # non resta da solo in fondo pagina separato dai suoi eventi.
        spazio(16 + 13)
        page.insert_text((mL, y[0]), str(r.get("nome", "")), fontsize=8.5, fontname="hebo")
        y[0] += 13
        for e in eventi:
            spazio(13)
            col = _COL_PRES.get(e["tipo"], (0.6, 0.6, 0.6))
            page.draw_rect(fitz.Rect(mL + 4, y[0] - 6, mL + 14, y[0] + 1), color=col, fill=col, width=0)
            # Il font base PyMuPDF (helv) non ha il glifo "→": uso un trattino ASCII.
            periodo = f"{e['dal']:02d}" if e["dal"] == e["al"] else f"{e['dal']:02d}-{e['al']:02d}"
            testo = f"{e['label']}: {periodo}/{mese:02d} ({e['giorni']} gg)"
            if e["nota"]:
                testo += f" — {e['nota'][:60]}"
            page.insert_text((mL + 18, y[0]), testo, fontsize=7.5, fontname="helv")
            y[0] += 13
        y[0] += 6
    if not qualcuno:
        page.insert_text((mL, y[0]), "Nessuna assenza da segnalare questo mese.", fontsize=8, fontname="helv", color=(0.5, 0.5, 0.5))
    return pdf.tobytes()


@router.post("/presenze/riepilogo-dati")
async def presenze_riepilogo_dati(data: dict = Body(...)):
    """Stessi dati dell'Opzione C (riepilogo totali + periodi) ma in JSON, per
    l'anteprima diretta in pagina — senza dover scaricare il PDF."""
    righe = data.get("righe") or []
    out_righe, out_periodi = [], []
    totali = {"lav": 0, "ferie": 0, "perm": 0, "malat": 0, "rol": 0, "riposi": 0, "altro": 0, "tot": 0}
    for r in righe:
        celle = r.get("celle") or []
        rp = _riepilogo_da_celle(celle)
        out_righe.append({"nome": r.get("nome", ""), **rp})
        for k in totali:
            totali[k] += rp[k]
        eventi = _periodi_da_celle(celle, r.get("note") or [])
        if eventi:
            out_periodi.append({"nome": r.get("nome", ""), "eventi": eventi})
    return {"righe": out_righe, "totali": totali, "periodi": out_periodi}


@router.post("/presenze/pdf-riepilogo")
async def presenze_pdf_riepilogo(data: dict = Body(...)):
    """Opzione C: documento 'per il commercialista' — riepilogo totali +
    dettaglio periodi, al posto della griglia giorno-per-giorno."""
    from fastapi.responses import StreamingResponse
    import io as _io
    anno = int(data.get("anno") or datetime.now().year)
    mese = int(data.get("mese") or datetime.now().month)
    giorni = int(data.get("giorni") or 31)
    righe = data.get("righe") or []
    try:
        pdf_bytes = _pdf_riepilogo_periodi(anno, mese, giorni, righe)
    except Exception as e:
        raise HTTPException(500, f"Errore generazione documento: {e}")
    fname = f"presenze_riepilogo_{anno}_{str(mese).zfill(2)}.pdf"
    return StreamingResponse(_io.BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/presenze/pdf")
async def presenze_pdf(data: dict = Body(...)):
    """Genera il PDF del foglio presenze del mese (una pagina), dai dati passati dal frontend."""
    from fastapi.responses import StreamingResponse
    import io as _io
    anno = int(data.get("anno") or datetime.now().year)
    mese = int(data.get("mese") or datetime.now().month)
    giorni = int(data.get("giorni") or 31)
    righe = data.get("righe") or []
    try:
        pdf_bytes = _pdf_presenze(anno, mese, giorni, righe)
    except Exception as e:
        raise HTTPException(500, f"Errore generazione PDF: {e}")
    fname = f"presenze_{anno}_{str(mese).zfill(2)}.pdf"
    return StreamingResponse(_io.BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/presenze/email-commercialista")
async def get_email_commercialista():
    """Email del commercialista salvata in app: 'Invia' la usa in automatico,
    senza doverla ridigitare ogni volta."""
    doc = await get_db().impostazioni.find_one({"id": "email_commercialista"}, {"_id": 0}) or {}
    return {"email": doc.get("email") or os.getenv("COMMERCIALISTA_EMAIL") or None}


@router.post("/presenze/email-commercialista")
async def salva_email_commercialista(data: dict = Body(...)):
    """Body: {email: str|null}. null = torna a chiedere/usare l'env."""
    email = (data.get("email") or "").strip() or None
    await get_db().impostazioni.update_one(
        {"id": "email_commercialista"},
        {"$set": {"id": "email_commercialista", "email": email, "updated_at": now_iso()}},
        upsert=True)
    return {"ok": True, "email": email}


@router.get("/presenze/invii")
async def lista_invii_presenze(anno: Optional[int] = None, mese: Optional[int] = None):
    """Storico degli invii del foglio presenze: a chi e quando."""
    q = {}
    if anno:
        q["anno"] = int(anno)
    if mese:
        q["mese"] = int(mese)
    invii = await get_db().presenze_invii.find(q, {"_id": 0}).sort("data_invio", -1).to_list(500)
    return {"invii": invii, "totale": len(invii)}


@router.post("/presenze/invia-commercialista")
async def invia_presenze_commercialista(data: dict = Body(...)):
    """Invia via email al commercialista il foglio presenze del mese (allegati PDF
    riepilogo+periodi + CSV). Salva lo storico dell'invio (destinatario + data).
    Destinatario dal body o dalla env COMMERCIALISTA_EMAIL. Credenziali email da
    services/email_smtp.py (SMTP_*/PEC_*/GMAIL_APP_PASSWORD — punto unico)."""
    from app.hr.services.email_smtp import credenziali_smtp, invia_email
    anno = int(data.get("anno") or datetime.now().year)
    mese = int(data.get("mese") or datetime.now().month)
    giorni = int(data.get("giorni") or 31)
    righe = data.get("righe") or []
    csv = data.get("csv") or (_csv_presenze(anno, mese, giorni, righe) if righe else "")
    if not csv.strip() and not righe:
        raise HTTPException(400, "Nessun dato presenze da inviare")
    dest_salvato = (await get_db().impostazioni.find_one(
        {"id": "email_commercialista"}, {"_id": 0, "email": 1}) or {}).get("email")
    dest = (data.get("destinatario") or dest_salvato or os.getenv("COMMERCIALISTA_EMAIL") or "").strip()
    if not dest:
        raise HTTPException(400, "Manca l'email del commercialista (impostala o inseriscila).")
    if not credenziali_smtp():
        raise HTTPException(400, "Email non configurata su Render (manca SMTP_HOST/PEC_HOST oppure "
                                 "GMAIL_APP_PASSWORD + ADMIN_EMAIL).")

    periodo = f"{_MESI_PRES[mese - 1]} {anno}"
    base = f"presenze_{anno}_{str(mese).zfill(2)}"
    # PDF leggibile (riepilogo + periodi, Opzione C) per la lettura umana; il CSV
    # (griglia giorno-per-giorno) resta come allegato per l'import nel software paghe.
    pdf_bytes = None
    if righe:
        try:
            pdf_bytes = _pdf_riepilogo_periodi(anno, mese, giorni, righe)
        except Exception:
            pdf_bytes = None

    allegati = []
    if pdf_bytes:
        allegati.append((pdf_bytes, "application", "pdf", f"{base}.pdf"))
    if csv.strip():
        allegati.append((csv.encode("utf-8"), "text", "csv", f"{base}.csv"))

    try:
        import asyncio
        await asyncio.to_thread(
            invia_email, dest, f"Presenze {periodo} — Ceraldi Group S.r.l.",
            f"In allegato il riepilogo presenze di {periodo} (PDF + CSV).\n\n"
            f"Messaggio generato automaticamente dal gestionale Ceraldi Group.",
            allegati)
    except Exception as e:
        # Log con traceback completo: l'errore esatto (auth Gmail, porta SMTP
        # bloccata dall'hosting, timeout...) va nei log di Render, il messaggio
        # corto va all'utente.
        logger.exception("Invio presenze al commercialista fallito")
        raise HTTPException(502, f"Invio email fallito: {type(e).__name__}: {e}")

    # Salva lo storico dell'invio (a chi, quando)
    rec = {"id": generate_id(), "anno": anno, "mese": mese, "periodo": periodo,
           "destinatario": dest, "data_invio": now_iso(),
           "n_dipendenti": len(righe), "con_pdf": bool(pdf_bytes)}
    await get_db().presenze_invii.insert_one(rec.copy())
    rec.pop("_id", None)
    return {"ok": True, "destinatario": dest, "periodo": periodo, "invio": rec}


@router.post("/presenze/consolida-da-turni")
async def consolida_presenze_da_turni(data: dict = Body(default={})):
    """Crea le presenze REALI a partire dai turni assegnati, per il mese indicato e SOLO
    per i giorni fino a oggi (i futuri non si segnano presenti). Mappa: turno di lavoro→
    presente (P), Riposo→RS, Ferie→F, Malattia→M. NON sovrascrive le presenze già inserite
    a mano. Serve ad avere le presenze pronte per export/buste partendo dai turni."""
    import calendar
    db = get_db()
    anno = int(data.get("anno") or datetime.now().year)
    mese = int(data.get("mese") or datetime.now().month)
    turni = await db.turni_cloud.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(200)
    nome_turno = {t["id"]: (t.get("nome") or "") for t in turni}
    GIORNI = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]  # 0=lunedì
    ndays = calendar.monthrange(anno, mese)[1]
    oggi = datetime.now().date()
    settimane = set()
    for g in range(1, ndays + 1):
        dt = datetime(anno, mese, g).date()
        settimane.add((dt - timedelta(days=dt.weekday())).isoformat())
    ass = await db.assegnazioni_turni_cloud.find({"settimana": {"$in": list(settimane)}}, {"_id": 0}).to_list(8000)
    ass_by_day = {}
    for a in ass:
        ass_by_day.setdefault((a.get("settimana"), a.get("giorno")), []).append((a.get("dipendente_id"), a.get("turno_id")))
    creati, saltati = 0, 0
    for g in range(1, ndays + 1):
        dt = datetime(anno, mese, g).date()
        if dt > oggi:
            break
        lun = (dt - timedelta(days=dt.weekday())).isoformat()
        gname = GIORNI[dt.weekday()]
        dstr = dt.isoformat()
        for dip_id, turno_id in ass_by_day.get((lun, gname), []):
            n = nome_turno.get(turno_id, "")
            if not n or not dip_id:
                continue
            if await db.presenze_cloud.find_one({"dipendente_id": dip_id, "data": dstr}):
                saltati += 1
                continue
            if n == "Riposo":
                stato, giust = "giustificato", "RS"
            elif n == "Ferie":
                stato, giust = "giustificato", "F"
            elif n == "Malattia":
                stato, giust = "giustificato", "M"
            else:
                stato, giust = "presente", "P"
            await db.presenze_cloud.insert_one({
                "id": generate_id(), "dipendente_id": dip_id, "data": dstr,
                "stato": stato, "giustificativo": giust,
                "origine": "consolidamento_turni", "created_at": now_iso()})
            creati += 1
    return {"ok": True, "anno": anno, "mese": mese, "creati": creati, "saltati": saltati}

# ============ FERIE E PERMESSI ============

@router.get("/ferie")
async def get_ferie(dipendente_id: Optional[str] = None, stato: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if stato:
        query["stato"] = stato
    ferie = await get_db().ferie_cloud.find(query, {"_id": 0}).to_list(1000)
    return ferie

@router.post("/ferie")
async def create_ferie(ferie: FerieCloud):
    ferie_dict = ferie.model_dump()
    ferie_dict["id"] = generate_id()
    ferie_dict["created_at"] = now_iso()
    await get_db().ferie_cloud.insert_one(ferie_dict)
    return serialize_doc(ferie_dict)

@router.post("/ferie-giorno")
async def set_ferie_giorno(data: dict):
    """Assegna/aggiorna/rimuove un'assenza di un singolo giorno dal calendario.
    tipo=None rimuove. Usato dalla vista calendario di Ferie & Permessi."""
    dip = data.get("dipendente_id")
    giorno = data.get("data")
    tipo = data.get("tipo")
    if not dip or not giorno:
        raise HTTPException(status_code=400, detail="dipendente_id e data obbligatori")
    existing = await get_db().ferie_cloud.find_one({
        "dipendente_id": dip, "data_inizio": giorno, "data_fine": giorno
    })
    if tipo:
        if existing:
            await get_db().ferie_cloud.update_one({"id": existing["id"]}, {"$set": {"tipo": tipo}})
        else:
            await get_db().ferie_cloud.insert_one({
                "id": generate_id(), "dipendente_id": dip, "tipo": tipo,
                "data_inizio": giorno, "data_fine": giorno, "giorni": 1,
                "stato": "approvata", "created_at": now_iso()
            })
    elif existing:
        await get_db().ferie_cloud.delete_one({"id": existing["id"]})
    return {"ok": True}

@router.put("/ferie/{ferie_id}/approva")
async def approva_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.update_one(
        {"id": ferie_id},
        {"$set": {"stato": "approvata"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta approvata"}

@router.put("/ferie/{ferie_id}/rifiuta")
async def rifiuta_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.update_one(
        {"id": ferie_id},
        {"$set": {"stato": "rifiutata"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta rifiutata"}

@router.delete("/ferie/{ferie_id}")
async def delete_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.delete_one({"id": ferie_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta eliminata"}

# ============ TURNI ============

@router.get("/turni")
async def get_turni():
    turni = await get_db().turni_cloud.find({}, {"_id": 0}).to_list(100)
    return turni

@router.post("/turni")
async def create_turno(turno: TurnoCloud):
    turno_dict = turno.model_dump()
    turno_dict["id"] = generate_id()
    await get_db().turni_cloud.insert_one(turno_dict)
    return serialize_doc(turno_dict)

@router.put("/turni/{turno_id}")
async def update_turno(turno_id: str, turno: TurnoCloud):
    result = await get_db().turni_cloud.update_one(
        {"id": turno_id},
        {"$set": turno.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Turno non trovato")
    return {"message": "Turno aggiornato"}

@router.delete("/turni/{turno_id}")
async def delete_turno(turno_id: str):
    result = await get_db().turni_cloud.delete_one({"id": turno_id})
    await get_db().assegnazioni_turni_cloud.delete_many({"turno_id": turno_id})
    return {"message": "Turno eliminato"}

@router.get("/assegnazioni-turni")
async def get_assegnazioni(settimana: Optional[str] = None):
    query = {"settimana": settimana} if settimana else {}
    assegnazioni = await get_db().assegnazioni_turni_cloud.find(query, {"_id": 0}).to_list(2000)
    return assegnazioni

@router.post("/assegnazioni-turni/migra")
async def migra_settimana_assegnazioni(data: dict):
    """Una-tantum: assegna una settimana ai record che non ce l'hanno."""
    settimana = data.get("settimana")
    if not settimana:
        raise HTTPException(status_code=400, detail="settimana obbligatoria")
    res = await get_db().assegnazioni_turni_cloud.update_many(
        {"$or": [{"settimana": {"$exists": False}}, {"settimana": None}]},
        {"$set": {"settimana": settimana}}
    )
    return {"migrati": res.modified_count}

@router.post("/assegnazioni-turni")
async def create_or_update_assegnazione(data: dict):
    dipendente_id = data.get("dipendente_id")
    giorno = data.get("giorno")
    turno_id = data.get("turno_id")
    settimana = data.get("settimana")
    
    if not dipendente_id or not giorno:
        raise HTTPException(status_code=400, detail="dipendente_id e giorno sono obbligatori")
    
    motivo = data.get("motivo")  # es. "onomastico" → reso visibile nei turni
    match = {"dipendente_id": dipendente_id, "giorno": giorno}
    if settimana:
        match["settimana"] = settimana
    existing = await get_db().assegnazioni_turni_cloud.find_one(match)

    if turno_id:
        if existing:
            upd = {"$set": {"turno_id": turno_id}}
            if motivo:
                upd["$set"]["motivo"] = motivo
            else:
                upd["$unset"] = {"motivo": ""}
            await get_db().assegnazioni_turni_cloud.update_one({"id": existing["id"]}, upd)
        else:
            ass = {
                "id": generate_id(),
                "dipendente_id": dipendente_id,
                "giorno": giorno,
                "turno_id": turno_id,
                "settimana": settimana,
            }
            if motivo:
                ass["motivo"] = motivo
            await get_db().assegnazioni_turni_cloud.insert_one(ass)
    else:
        if existing:
            await get_db().assegnazioni_turni_cloud.delete_one({"id": existing["id"]})
    
    return {"message": "Assegnazione salvata"}

# ============ CONFIG TURNI PER DIPENDENTE ============
# Per ogni dipendente: turno abituale (turno_id) + giorno di riposo fisso
# settimanale (riposo_giorno, nome italiano). Usati da "Genera settimana".
def _nome_norm_cfg(s) -> str:
    return re.sub(r"[^a-z]", "", str(s or "").lower())


@router.get("/turni-config")
async def get_turni_config():
    """Config turni per dipendente. Ripara le relazioni ORFANE: se una config punta
    a un dipendente_id che non esiste più (dipendente reimportato/ricreato con un
    nuovo id), la riaggancia per nome al dipendente attivo corrispondente — era la
    causa dei dipendenti che 'non generano turni' senza nessun errore visibile."""
    db = get_db()
    configs = await db.turni_config.find({}, {"_id": 0}).to_list(1000)
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1,
                                     "stato": 1, "attivo": 1}).to_list(1000)
    ids_validi = {d["id"] for d in dips}
    con_config = {c["dipendente_id"] for c in configs}
    per_nome = {}
    for d in dips:
        if d.get("attivo") is False or (d.get("stato") or "attivo") in ("cessato", "dimesso", "archiviato"):
            continue
        for k in {_nome_norm_cfg(f"{d.get('cognome','')}{d.get('nome','')}"),
                  _nome_norm_cfg(f"{d.get('nome','')}{d.get('cognome','')}"),
                  _nome_norm_cfg(d.get("nome_completo"))}:
            if k:
                per_nome.setdefault(k, d)
    for c in configs:
        if c["dipendente_id"] in ids_validi:
            continue
        k = _nome_norm_cfg(c.get("nome_riferimento"))
        d = per_nome.get(k) if k else None
        if d and d["id"] not in con_config:
            await db.turni_config.update_one(
                {"dipendente_id": c["dipendente_id"]},
                {"$set": {"dipendente_id": d["id"], "updated_at": now_iso()}})
            con_config.add(d["id"])
            c["dipendente_id"] = d["id"]
    return configs


@router.post("/turni-config")
async def save_turni_config(data: dict = Body(...)):
    """Body: {voci: [{dipendente_id, turno_id, riposo_giorno}]}. Salva anche il nome
    del dipendente (nome_riferimento) per poter riparare la relazione se in futuro
    l'anagrafica venisse reimportata con id nuovi."""
    db = get_db()
    dips = await db.dipendenti.find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1,
                                         "nome_completo": 1}).to_list(1000)
    nomi = {d["id"]: (d.get("nome_completo") or f"{d.get('cognome', '')} {d.get('nome', '')}".strip())
            for d in dips}
    for v in (data.get("voci") or []):
        if not v.get("dipendente_id"):
            continue
        await db.turni_config.update_one(
            {"dipendente_id": v["dipendente_id"]},
            {"$set": {"dipendente_id": v["dipendente_id"],
                      "nome_riferimento": nomi.get(v["dipendente_id"]) or None,
                      "turno_id": v.get("turno_id") or None,
                      "riposo_giorno": v.get("riposo_giorno") or None,
                      "lunga_giorni": v.get("lunga_giorni") or [],
                      "rotazione": v.get("rotazione") or None,
                      # lunedì della settimana in cui la fase è stata impostata:
                      # "inizia mattina" = mattina in QUELLA settimana, poi si
                      # inverte ogni lunedì (ancora per-dipendente, niente base globale)
                      "rotazione_ancora": v.get("rotazione_ancora") or None,
                      "sala": bool(v.get("sala")),
                      # abilitato a coprire il bar nelle sostituzioni (es. Taiano, Russo)
                      "sostituto_bar": bool(v.get("sostituto_bar")),
                      "updated_at": now_iso()}}, upsert=True)
    return {"ok": True, "salvati": len(data.get("voci") or [])}

@router.get("/turni-disponibilita-bar")
async def get_turni_disponibilita_bar(settimana: Optional[str] = None):
    """Disponibilità a coprire il bar (dal portale) che toccano la settimana
    indicata: 'Genera settimana' le applica (sostituto al bar nella sua fascia,
    sala coperta con una Lunga)."""
    q = {}
    if settimana:
        try:
            lun = datetime.strptime(settimana, "%Y-%m-%d").date()
            q = {"dal": {"$lte": (lun + timedelta(days=6)).isoformat()},
                 "al": {"$gte": settimana}}
        except ValueError:
            pass
    return await get_db().turni_disponibilita_bar.find(
        q, {"_id": 0}).sort("dal", 1).to_list(200)


@router.get("/turni-preferenze")
async def get_turni_preferenze(settimana: Optional[str] = None):
    """Preferenze del giorno di riposo inviate dai dipendenti dal portale
    (collezione `turni_preferenze_riposo`): chi compone i turni le vede
    nella pagina Turni della settimana corrispondente."""
    q = {"settimana": settimana} if settimana else {}
    return await get_db().turni_preferenze_riposo.find(
        q, {"_id": 0}).sort("aggiornata_il", -1).to_list(500)


@router.get("/turni-chiusura-pomeridiana")
async def get_chiusura_pomeridiana():
    """Periodo in cui il bar resta chiuso di pomeriggio (impostato nel modale
    Configura turni): in quelle settimane tutti i baristi in rotazione fanno la
    mattina e riposano la domenica, come il resto della squadra."""
    doc = await get_db().impostazioni.find_one({"id": "chiusura_pomeridiana"}, {"_id": 0}) or {}
    return {"attiva": bool(doc.get("attiva")), "dal": doc.get("dal"), "al": doc.get("al")}


@router.post("/turni-chiusura-pomeridiana")
async def save_chiusura_pomeridiana(data: dict = Body(...)):
    """Body: {attiva: bool, dal: YYYY-MM-DD, al: YYYY-MM-DD}."""
    await get_db().impostazioni.update_one(
        {"id": "chiusura_pomeridiana"},
        {"$set": {"id": "chiusura_pomeridiana", "attiva": bool(data.get("attiva")),
                  "dal": data.get("dal") or None, "al": data.get("al") or None,
                  "updated_at": now_iso()}}, upsert=True)
    return {"ok": True}


# ============ ONOMASTICI (riposo per onomastico nei turni) ============
# Date standard italiane (mese, giorno) per nome proprio. Prefillate e
# MODIFICABILI in gestione. I nomi non presenti sono "stranieri" → esclusi.
ONOMASTICI_DEFAULT = {
    "angela": (1, 27), "angelo": (10, 2), "anna": (7, 26), "antonella": (6, 13),
    "antonietta": (6, 13), "antonio": (6, 13), "carmela": (7, 16), "carmine": (7, 16),
    "caterina": (11, 25), "ciro": (1, 31), "domenico": (8, 8), "elena": (8, 18),
    "emanuele": (3, 26), "fabio": (5, 11), "francesca": (3, 9), "francesco": (10, 4),
    "gaetano": (8, 7), "gennaro": (9, 19), "giorgio": (4, 23), "giovanna": (5, 30),
    "giovanni": (6, 24), "giulia": (5, 22), "giuliano": (1, 9), "giuseppa": (3, 19),
    "giuseppe": (3, 19), "ignazio": (7, 31), "liliana": (7, 27), "lucia": (12, 13),
    "luigi": (6, 21), "luigia": (6, 21), "marcella": (1, 31), "marco": (4, 25),
    "margherita": (2, 22), "maria": (9, 12), "mariano": (8, 19), "marina": (7, 17),
    "mario": (1, 19), "michele": (9, 29), "ottavio": (11, 20), "pasquale": (5, 17),
    "paolo": (6, 29), "pietro": (6, 29), "raffaele": (9, 29), "rosa": (8, 23),
    "salvatore": (8, 6), "simone": (10, 28), "stefano": (12, 26), "teresa": (10, 15),
    "valerio": (1, 29), "vincenzo": (1, 22), "vincenza": (1, 22),
}
NOMI_GIORNO_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
# Dipendenti che NON seguono i turni → niente riposo onomastico (decisione titolare).
NON_TURNI = [("vincenzo", "ceraldi"), ("valerio", "ceraldi"),
             ("antonietta", "ceraldi"), ("marina", "liuzza")]


def _non_turni(d: dict) -> bool:
    f = f"{d.get('nome','')} {d.get('cognome','')} {d.get('nome_completo','')}".lower()
    return any(a in f and b in f for a, b in NON_TURNI)


def _nome_proprio(dip: dict) -> str:
    n = (dip.get("nome") or "").strip()
    if not n and dip.get("nome_completo"):
        n = dip["nome_completo"].split()[0]
    return n.split()[0].lower() if n else ""


@router.get("/onomastici")
async def get_onomastici():
    """Onomastico per ogni dipendente attivo: data (prefillata dal nome o salvata),
    attivo e flag 'straniero' (nome senza onomastico italiano)."""
    db = get_db()
    dips = await db.dipendenti.find(
        {"merged_into": {"$exists": False}}, {"_id": 0}).to_list(1000)
    salvati = {o["dipendente_id"]: o async for o in db.onomastici.find({}, {"_id": 0})}
    out = []
    for d in dips:
        if d.get("attivo") is False or (d.get("stato") or "attivo") in ("cessato", "dimesso", "archiviato"):
            continue
        nome = _nome_proprio(d)
        default = ONOMASTICI_DEFAULT.get(nome)
        straniero = default is None
        s = salvati.get(d.get("id"))
        if s:
            mese, giorno, attivo = s.get("mese"), s.get("giorno"), s.get("attivo", True)
        else:
            mese, giorno = (default if default else (None, None))
            attivo = (not straniero) and (not _non_turni(d))
        out.append({
            "dipendente_id": d.get("id"),
            "nome": d.get("nome_completo") or f"{d.get('cognome','')} {d.get('nome','')}".strip(),
            "mese": mese, "giorno": giorno, "attivo": bool(attivo), "straniero": straniero,
            "non_turni": _non_turni(d),
        })
    out.sort(key=lambda x: (x["nome"] or "").lower())
    return out


@router.post("/onomastici")
async def save_onomastici(data: dict = Body(...)):
    """Salva le date/attivo onomastico. Body: {voci: [{dipendente_id, mese, giorno, attivo}]}."""
    db = get_db()
    for v in (data.get("voci") or []):
        if not v.get("dipendente_id"):
            continue
        await db.onomastici.update_one(
            {"dipendente_id": v["dipendente_id"]},
            {"$set": {"dipendente_id": v["dipendente_id"],
                      "mese": v.get("mese"), "giorno": v.get("giorno"),
                      "attivo": bool(v.get("attivo", True)),
                      "updated_at": now_iso()}}, upsert=True)
    return {"ok": True, "salvati": len(data.get("voci") or [])}


@router.get("/onomastici/settimana")
async def onomastici_settimana(settimana: str):
    """Onomastici (idonei al riposo) che cadono nella settimana indicata (lunedì
    ISO). Esclude stranieri, esclusi (attivo=False) e la domenica (bar chiuso)."""
    try:
        lun = datetime.strptime(settimana, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="settimana deve essere YYYY-MM-DD (lunedì)")
    voci = await get_onomastici()
    giorni_sett = [(lun + timedelta(days=i)) for i in range(7)]
    out = []
    for v in voci:
        if not v["attivo"] or v["straniero"] or not v["mese"] or not v["giorno"]:
            continue
        for i, gd in enumerate(giorni_sett):
            if gd.month == v["mese"] and gd.day == v["giorno"] and i < 6:  # esclude domenica
                out.append({
                    "dipendente_id": v["dipendente_id"], "nome": v["nome"],
                    "data": gd.strftime("%Y-%m-%d"), "giorno_nome": NOMI_GIORNO_IT[i],
                    "data_label": gd.strftime("%d/%m"),
                })
    return out

# ============ MOTORE DI INTERROGAZIONE CEDOLINI ============

@router.get("/cedolini/cerca-voce")
async def cerca_voce(codice: Optional[str] = None, testo: Optional[str] = None,
                     anno: Optional[int] = None, dipendente_id: Optional[str] = None):
    """Cerca una voce in TUTTI i cedolini salvati (campo voci). Per codice (es. F09081)
    o per testo della descrizione (es. '730', '13ma'). Filtrabile per anno/dipendente."""
    if not codice and not testo:
        raise HTTPException(status_code=400, detail="Indica 'codice' (es. F09081) o 'testo' (es. 730) da cercare")
    q: dict = {}
    if anno:
        q["anno"] = anno
    if dipendente_id:
        q["dipendente_id"] = dipendente_id
    cod = (codice or "").upper().strip()
    txt = (testo or "").lower().strip()
    out = []
    async for c in get_db().cedolini.find(q, {"_id": 0, "dipendente_id": 1, "dipendente_nome": 1, "anno": 1, "mese": 1, "voci": 1}):
        for v in (c.get("voci") or []):
            if (cod and v.get("codice") == cod) or (txt and txt in (v.get("descrizione") or "").lower()):
                out.append({"dipendente_id": c.get("dipendente_id"), "dipendente": c.get("dipendente_nome"),
                            "anno": c.get("anno"), "mese": c.get("mese"),
                            "codice": v.get("codice"), "descrizione": v.get("descrizione"),
                            "importo": (v.get("valori") or [None])[-1], "valori": v.get("valori")})
    out.sort(key=lambda x: (x.get("anno") or 0, x.get("mese") or 0))
    return {"risultati": out, "totale": len(out)}


@router.post("/paghe/correggi-acconti-cedolino")
async def correggi_acconti_cedolino():
    """Una tantum: toglie gli 'acconto dal cedolino' già salvati che sono
    implausibili (poche decine di euro — trattenute minime non correlate
    intercettate per errore dal parser, non veri anticipi). Ricalcola il saldo
    residuo di conseguenza. Non tocca gli acconti registrati a mano
    (acconti_dipendenti) né quelli plausibili."""
    db = get_db()
    corretti = []
    for coll_name in ("paghe_mensili", "cedolini"):
        async for doc in db[coll_name].find(
                {"acconto_cedolino": {"$gt": 0}},
                {"_id": 0, "id": 1, "dipendente_id": 1, "anno": 1, "mese": 1,
                 "importo_busta": 1, "netto": 1, "acconto_cedolino": 1}):
            netto = doc.get("importo_busta") if doc.get("importo_busta") is not None else doc.get("netto")
            if _acconto_cedolino_plausibile(doc.get("acconto_cedolino"), netto):
                continue
            filtro = {"id": doc["id"]} if doc.get("id") else \
                     {"dipendente_id": doc["dipendente_id"], "anno": doc["anno"], "mese": doc["mese"]}
            await db[coll_name].update_one(filtro, {"$unset": {"acconto_cedolino": "", "saldo_residuo": ""}})
            corretti.append({"collezione": coll_name, "dipendente_id": doc.get("dipendente_id"),
                             "anno": doc.get("anno"), "mese": doc.get("mese"),
                             "acconto_scartato": doc.get("acconto_cedolino")})
    return {"corretti": len(corretti), "dettaglio": corretti}


@router.post("/cedolini/riscansiona")
async def riscansiona_cedolini(anno: Optional[int] = None, dipendente_id: Optional[str] = None):
    """Ri-estrae tutte le voci dai cedolini storici (2023→oggi) che hanno il PDF salvato,
    così il motore di ricerca trova ogni codice anche sulle buste già importate."""
    import io
    import pdfplumber
    db = get_db()
    q: dict = {"pdf_data": {"$exists": True}}
    if anno:
        q["anno"] = anno
    if dipendente_id:
        q["dipendente_id"] = dipendente_id
    aggiornati, errori = 0, 0
    async for c in db.cedolini.find(q, {"_id": 0, "id": 1, "pdf_data": 1}):
        try:
            raw = base64.b64decode(c["pdf_data"])
            text = ""
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for p in pdf.pages:
                    text += (p.extract_text() or "") + "\n"
            dati = _lul_dati_busta(text)
            if dati:
                await db.cedolini.update_one({"id": c["id"]}, {"$set": dati})
                aggiornati += 1
        except Exception:
            errori += 1
    return {"aggiornati": aggiornati, "errori": errori,
            "nota": "I cedolini senza PDF salvato non possono essere riscansionati: vanno re-importati dal Libro Unico."}


# ============ IMPORT PRIMA NOTA SALARI (Excel) ============

@router.post("/paghe/importa-prima-nota")
async def importa_prima_nota(file: UploadFile = File(...)):
    """Importa la 'Prima Nota Salari' (Excel: Dipendente, Mese, Anno, Stipendio Netto,
    Importo Erogato) oppure il tracciato 'PAGAMENTI' (Banca, Data contabile, Mese cedolino,
    Anno, Uscita, Entrata, NOME DIPENDENTE, Numero operazione — mese numerico, 13=tredicesima
    e 14=quattordicesima). Per ogni dipendente/mese/anno SOMMA gli Importi Erogati (più
    bonifici nello stesso mese) e li scrive in paghe_mensili.bonifico_importo. Riempie
    l'importo busta se mancante. Confronta col dato già in app e segnala differenze e
    nomi non trovati."""
    import io
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel non valido: {e}")
    ws = wb["Salari"] if "Salari" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Foglio vuoto")
    header = [(str(c).strip().lower() if c is not None else "") for c in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None
    ci_dip = col("dipendente", "nome dipendente")
    ci_mese = col("mese", "mese cedolino")
    ci_anno = col("anno")
    ci_netto = col("stipendio netto", "netto", "importo busta")
    ci_erog = col("importo erogato", "erogato", "bonifico", "uscita")
    if None in (ci_dip, ci_mese, ci_anno, ci_erog):
        raise HTTPException(400, "Colonne attese: Dipendente, Mese, Anno, Importo Erogato "
                                 "(oppure tracciato PAGAMENTI: NOME DIPENDENTE, Mese cedolino, Anno, Uscita)")

    MESI = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
            "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
            "tredicesima": 13, "quattordicesima": 14}

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()

    def parse_mese(v):
        vn = norm(v)
        if vn in MESI:
            return MESI[vn]
        try:
            m = int(float(vn))
            return m if 1 <= m <= 14 else None  # 13 = tredicesima, 14 = quattordicesima
        except (TypeError, ValueError):
            return None

    def fnum(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    agg = {}
    for r in rows[1:]:
        if ci_dip >= len(r) or not r[ci_dip]:
            continue
        mese = parse_mese(r[ci_mese])
        try:
            anno = int(float(r[ci_anno]))
        except (TypeError, ValueError):
            anno = None
        if not mese or not anno:
            continue
        k = (norm(r[ci_dip]), mese, anno)
        a = agg.setdefault(k, {"nome": str(r[ci_dip]).strip(), "netto": 0.0, "erogato": 0.0})
        a["erogato"] += fnum(r[ci_erog])
        if ci_netto is not None and ci_netto < len(r):
            a["netto"] += fnum(r[ci_netto])

    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)
    by_nome = {}
    for d in dips:
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v:
                by_nome[v] = d

    aggiornati, non_trovati, discrepanze = [], [], []
    for (nome_n, mese, anno), a in agg.items():
        erog, netto = round(a["erogato"], 2), round(a["netto"], 2)
        if erog <= 0 and netto <= 0:
            continue
        d = by_nome.get(nome_n)
        if not d:
            non_trovati.append({"nome": a["nome"], "mese": mese, "anno": anno, "bonifico": erog})
            continue
        existing = await db.paghe_mensili.find_one(
            {"dipendente_id": d["id"], "anno": anno, "mese": mese}, {"importo_busta": 1}) or {}
        set_doc = {"dipendente_id": d["id"], "anno": anno, "mese": mese,
                   "bonifico_importo": erog, "bonifico_ricevuto": erog > 0,
                   "bonifico_da_prima_nota": True, "updated_at": now_iso()}
        busta_app = existing.get("importo_busta")
        if (busta_app in (None, 0, "")) and netto > 0:
            set_doc["importo_busta"] = netto
        elif busta_app and netto > 0 and abs(float(busta_app) - netto) > 1:
            discrepanze.append({"dipendente": a["nome"], "mese": mese, "anno": anno,
                                "busta_app": round(float(busta_app), 2), "busta_excel": netto})
        await db.paghe_mensili.update_one(
            {"dipendente_id": d["id"], "anno": anno, "mese": mese}, {"$set": set_doc}, upsert=True)
        await _ricalcola_stato_paga(db, d["id"], anno, mese)
        aggiornati.append({"dipendente": a["nome"], "mese": mese, "anno": anno, "bonifico": erog})

    nomi_non_trovati = sorted({x["nome"] for x in non_trovati})
    return {"aggiornati": len(aggiornati),
            "righe_aggregate": len(agg),
            "non_trovati": len(non_trovati),
            "nomi_non_trovati": nomi_non_trovati,
            "discrepanze": sorted(discrepanze, key=lambda x: (x["anno"], x["mese"]))}


@router.post("/paghe/importa-storico-pagamenti")
async def importa_storico_pagamenti(file: UploadFile = File(...)):
    """Importa l'archivio storico dei pagamenti (un foglio Excel per dipendente: data del
    pagamento in colonna A, nome in colonna B, importo di busta in colonna C, importo
    effettivamente pagato in colonna D — le intestazioni di questi fogli sono spesso
    disallineate rispetto ai dati, quindi il formato si riconosce dal TIPO di dato in
    colonna A, non dal testo dell'header).
    Le righe finiscono in 'pagamenti_storico', un registro di SOLA CONSULTAZIONE per il
    periodo precedente all'app: non tocca 'paghe_mensili' né lo stato di pagamento dei
    cedolini correnti, perché qui si conosce solo la data del bonifico e non il mese di
    competenza della busta (attribuirlo al mese del bonifico rischierebbe di sfalsare il
    saldo di un mese). Import idempotente: le righe già presenti (stesso dipendente, data,
    busta, pagato) non vengono duplicate."""
    import io
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel non valido: {e}")

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()

    def fnum(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    righe = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        start = 0 if (rows[0] and isinstance(rows[0][0], (datetime, date))) else 1
        for r in rows[start:]:
            if not r or not isinstance(r[0], (datetime, date)):
                continue
            nome = r[1] if len(r) > 1 else None
            busta = fnum(r[2]) if len(r) > 2 else 0.0
            pagato = fnum(r[3]) if len(r) > 3 else 0.0
            if not nome or (busta <= 0 and pagato <= 0):
                continue
            d = r[0]
            data_iso = d.date().isoformat() if isinstance(d, datetime) else d.isoformat()
            righe.append({"nome": str(nome).strip(), "data": data_iso,
                          "busta": round(busta, 2), "pagato": round(pagato, 2)})

    if not righe:
        raise HTTPException(400, "Nessuna riga riconosciuta: attesa una data in colonna A per ogni pagamento")

    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)
    by_nome = {}
    for dd in dips:
        n, c = norm(dd.get("nome")), norm(dd.get("cognome"))
        for v in {norm(dd.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v:
                by_nome[v] = dd

    try:
        await db.pagamenti_storico.create_index(
            [("dipendente_id", 1), ("data", 1), ("busta", 1), ("pagato", 1)],
            unique=True, name="uniq_storico_riga")
    except Exception:
        pass

    importati, gia_presenti, non_trovati = 0, 0, {}
    for r in righe:
        dip = by_nome.get(norm(r["nome"]))
        if not dip:
            non_trovati[r["nome"]] = non_trovati.get(r["nome"], 0) + 1
            continue
        doc = {"id": generate_id(), "dipendente_id": dip["id"], "data": r["data"],
               "busta": r["busta"], "pagato": r["pagato"], "fonte": "excel_storico", "created_at": now_iso()}
        res = await db.pagamenti_storico.update_one(
            {"dipendente_id": dip["id"], "data": r["data"], "busta": r["busta"], "pagato": r["pagato"]},
            {"$setOnInsert": doc}, upsert=True)
        if res.upserted_id:
            importati += 1
        else:
            gia_presenti += 1

    return {"righe_lette": len(righe), "importati": importati, "gia_presenti": gia_presenti,
            "dipendenti_non_in_anagrafica": [{"nome": k, "righe": v} for k, v in sorted(non_trovati.items())]}


@router.get("/paghe/storico-pagamenti")
async def storico_pagamenti(dipendente_id: str):
    """Registro storico dei pagamenti ante-app (da Excel), in sola lettura, per data."""
    db = get_db()
    righe = await db.pagamenti_storico.find(
        {"dipendente_id": dipendente_id}, {"_id": 0}).sort("data", 1).to_list(2000)
    return {"righe": righe,
            "totale_busta": round(sum(r.get("busta", 0) for r in righe), 2),
            "totale_pagato": round(sum(r.get("pagato", 0) for r in righe), 2)}


_MESI_IT = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
            "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}


@router.post("/dipendenti/importa-anagrafica")
async def importa_anagrafica(file: UploadFile = File(...)):
    """Importa/aggiorna l'anagrafica da Excel (Cognome, Nome, CF, Data di nascita,
    Mansione, Telefono, Email, Indirizzo). Match per codice fiscale: aggiorna se esiste,
    altrimenti crea."""
    import io
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [(str(c).strip().lower() if c is not None else "") for c in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None
    ci = {"cognome": col("cognome"), "nome": col("nome"), "cf": col("cf", "codice fiscale"),
          "nascita": col("nascita"), "mansione": col("mansione"), "tel": col("telefono", "cell"),
          "email": col("email", "mail"), "indirizzo": col("indirizzo")}
    db = get_db()
    creati, aggiornati = 0, 0

    def val(r, k):
        i = ci.get(k)
        if i is None or i >= len(r) or r[i] is None:
            return None
        return str(r[i]).strip()
    for r in rows[1:]:
        cf = (val(r, "cf") or "").upper().replace(" ", "")
        nome, cognome = val(r, "nome"), val(r, "cognome")
        if not (cf or (nome and cognome)):
            continue
        campi = {"nome": nome, "cognome": cognome, "codice_fiscale": cf or None,
                 "data_nascita": (val(r, "nascita") or "")[:10] or None,
                 "mansione": val(r, "mansione"), "telefono": val(r, "tel"),
                 "email": val(r, "email"), "indirizzo": val(r, "indirizzo")}
        campi = {k: v for k, v in campi.items() if v}
        campi["nome_completo"] = f"{cognome or ''} {nome or ''}".strip()
        esistente = await db.dipendenti.find_one({"codice_fiscale": cf}) if cf else None
        if esistente:
            await db.dipendenti.update_one({"id": esistente["id"]}, {"$set": campi})
            aggiornati += 1
        else:
            campi.update({"id": generate_id(), "attivo": True, "stato": "attivo", "created_at": now_iso()})
            await db.dipendenti.insert_one(campi)
            creati += 1
    return {"creati": creati, "aggiornati": aggiornati}


@router.post("/riduzioni-orario")
async def save_riduzioni_orario(data: dict = Body(...)):
    """Salva la riduzione oraria per dipendente: ore/giorno ridotte, paga oraria,
    data inizio e data fine (scadenza sorvegliata). Body: {voci:[{dipendente_id,...}]}."""
    db = get_db()
    n = 0
    for v in (data.get("voci") or []):
        did = v.get("dipendente_id")
        if not did:
            continue

        def num(x):
            try:
                return float(str(x).replace(",", ".")) if x not in (None, "") else None
            except (TypeError, ValueError):
                return None
        rid = {"attiva": bool(v.get("attiva")),
               "ore_giorno": num(v.get("ore_giorno")),
               "paga_oraria": num(v.get("paga_oraria")),
               "data_inizio": v.get("data_inizio") or None,
               "data_fine": v.get("data_fine") or None,
               "note": (v.get("note") or "").strip(),
               "updated_at": now_iso()}
        await db.dipendenti.update_one({"id": did}, {"$set": {"riduzione_orario": rid}})
        n += 1
    return {"salvati": n}


@router.get("/riduzioni-orario/scadenze")
async def riduzioni_in_scadenza(giorni: int = 30):
    """Riduzioni attive con scadenza entro N giorni (o già scadute) — vigilanza contratto."""
    db = get_db()
    oggi = datetime.now(timezone.utc).date()
    out = []
    async for d in db.dipendenti.find({"riduzione_orario.attiva": True}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "riduzione_orario": 1}):
        rid = d.get("riduzione_orario") or {}
        df = rid.get("data_fine")
        if not df:
            continue
        try:
            scad = datetime.strptime(df[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        gg = (scad - oggi).days
        if gg <= giorni:
            out.append({"dipendente_id": d.get("id"),
                        "nome": f"{d.get('cognome','')} {d.get('nome','')}".strip(),
                        "data_fine": df, "giorni_alla_scadenza": gg,
                        "scaduta": gg < 0, "ore_giorno": rid.get("ore_giorno")})
    out.sort(key=lambda x: x["giorni_alla_scadenza"])
    return out


@router.post("/paghe/importa-pagamenti")
async def importa_pagamenti(file: UploadFile = File(...)):
    """Importa i bonifici/pagamenti dal CSV banca. Riconosce due formati dall'intestazione:
    1) ESITI bonifici (Esecuzione;Ordinante;Beneficiario;Importo;Div;Causale;CRO);
    2) ANDAMENTO conto (Ragione Sociale;Data contabile;Data valuta;Banca;Rapporto;Importo;
       Divisa;Descrizione;Categoria;Hashtag): tiene solo le USCITE (importo negativo),
       scarta commissioni bancarie, estrae il nominativo dal 'FAVORE <Nome>' nella descrizione.
    In entrambi i casi aggancia solo chi è in anagrafica (fornitori esclusi automaticamente).
    Mese di competenza dalla causale (es. '9-2025', 'luglio') o, in mancanza, dalla data del
    movimento. Idempotente (dedup per CRO o hash riga). Aggiorna il bonifico del mese = somma
    dei pagamenti di quel mese e ricalcola lo stato paga (alimenta la prima nota)."""
    import io
    import csv as _csv
    raw = await file.read()
    text = raw.decode("utf-8", errors="ignore")
    reader = _csv.reader(io.StringIO(text), delimiter=";")
    righe = list(reader)
    if not righe:
        raise HTTPException(400, "CSV vuoto")
    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()
    by_nome, by_cogn = {}, {}
    for d in dips:
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v and len(v) > 5:
                by_nome[v] = d
        if len(c) >= 4:
            by_cogn.setdefault(c, []).append(d)

    def trova_dip(beneficiario):
        b = norm(beneficiario)
        for nome_n, d in by_nome.items():
            if nome_n in b or b in nome_n:
                return d
        for cogn, lst in by_cogn.items():
            if cogn in b and len(lst) == 1:
                return lst[0]
        return None

    def to_float(s):
        try:
            return float(str(s).replace(".", "").replace(",", "."))
        except (TypeError, ValueError):
            return None

    def mese_anno(causale, data_dt):
        c = norm(causale)
        m = re.search(r'\b(\d{1,2})[-/](20\d{2})\b', c)
        if m:
            return int(m.group(1)), int(m.group(2))
        for nome, n in _MESI_IT.items():
            if nome in c:
                y = re.search(r'(20\d{2})', c)
                return n, int(y.group(1)) if y else data_dt.year
        return data_dt.month, data_dt.year

    def favore(s):
        m = re.search(r'favore\s+(.+?)(?:\s+-|\s+notprovide|$)', norm(s))
        return (m.group(1) if m else norm(s))[:50]

    # Rileva il formato dall'intestazione: ESITI bonifici o ESTRATTO CONTO (entrate/uscite)
    hdr = [norm(c) for c in (righe[0] if righe else [])]

    def col(*names):
        return next((i for i, h in enumerate(hdr) if any(n in h for n in names)), None)
    i_ben = col("beneficiario")
    if i_ben is not None:
        formato = "esiti"
        i_data = col("esecuzione", "data") if col("esecuzione", "data") is not None else 0
        i_imp = col("importo") if col("importo") is not None else 3
        i_caus = col("causale", "descrizione")
        i_cro = col("cro")
        i_cat = None
    else:
        formato = "andamento"
        i_data = col("data contabile", "data valuta", "data")
        i_imp = col("importo")
        i_caus = col("descrizione")
        i_cat = col("categoria")
        i_cro = None
        i_ben = i_caus

    importati, non_trovati, affected = 0, [], set()
    for r in righe[1:]:
        if i_imp is None or i_imp >= len(r) or i_data is None or i_data >= len(r):
            continue
        importo = to_float(r[i_imp])
        if importo is None:
            continue
        try:
            data_dt = datetime.strptime(str(r[i_data]).strip()[:10], "%d/%m/%Y")
        except (ValueError, TypeError):
            continue
        causale = (r[i_caus] if i_caus is not None and i_caus < len(r) else "") or ""
        if formato == "andamento":
            if importo >= 0:  # solo uscite = pagamenti
                continue
            cat = (r[i_cat] if i_cat is not None and i_cat < len(r) else "") or ""
            if "commission" in norm(cat) or norm(causale).startswith("comm"):
                continue  # niente commissioni bancarie
            importo = -importo
            beneficiario = favore(causale)
        else:
            if importo <= 0:
                continue
            beneficiario = (r[i_ben] if i_ben is not None and i_ben < len(r) else "") or ""
        if importo < 5:
            continue
        d = trova_dip(beneficiario if formato == "esiti" else causale)
        if not d:
            non_trovati.append(beneficiario or favore(causale))
            continue
        mese, anno = mese_anno(causale, data_dt)
        cro = (r[i_cro].strip() if i_cro is not None and i_cro < len(r) and r[i_cro] else "")
        key = cro or hashlib.sha1(f"{d['id']}|{r[i_data]}|{importo}|{causale}".encode()).hexdigest()
        await db.pagamenti_esiti.update_one(
            {"key": key},
            {"$set": {"key": key, "cro": cro, "dipendente_id": d["id"], "data": data_dt.strftime("%Y-%m-%d"),
                      "importo": importo, "causale": causale, "beneficiario": beneficiario,
                      "mese": mese, "anno": anno}}, upsert=True)
        affected.add((d["id"], mese, anno))
        importati += 1
    # ricalcola il bonifico del mese = somma dei pagamenti di quel mese
    for dip_id, mese, anno in affected:
        tot = 0.0
        async for p in db.pagamenti_esiti.find({"dipendente_id": dip_id, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1}):
            tot += p.get("importo") or 0
        await db.paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": anno, "mese": mese},
            {"$set": {"dipendente_id": dip_id, "anno": anno, "mese": mese,
                      "bonifico_importo": round(tot, 2), "bonifico_ricevuto": tot > 0,
                      "bonifico_da_esiti": True, "updated_at": now_iso()}}, upsert=True)
        await _ricalcola_stato_paga(db, dip_id, anno, mese)
    return {"importati": importati, "mesi_aggiornati": len(affected),
            "non_trovati": sorted(set(non_trovati))}


@router.get("/paghe/in-attesa")
async def paghe_in_attesa():
    """Buste in attesa di pagamento (o parziali): elenco per il pannello/avvisi."""
    db = get_db()
    dip_map = {d["id"]: f"{d.get('cognome','')} {d.get('nome','')}".strip()
               async for d in db.dipendenti.find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1})}
    out = []
    async for p in db.paghe_mensili.find(
            {"stato_pagamento": {"$in": ["in_attesa_pagamento", "parziale"]}}, {"_id": 0}):
        saldo = p.get("saldo")
        if saldo is None:
            saldo = round(float(p.get("importo_busta") or 0) - float(p.get("bonifico_importo") or 0), 2)
        if not saldo or saldo <= 0.5:
            continue
        out.append({"dipendente_id": p.get("dipendente_id"),
                    "dipendente": dip_map.get(p.get("dipendente_id"), p.get("dipendente_id")),
                    "anno": p.get("anno"), "mese": p.get("mese"),
                    "stato": p.get("stato_pagamento"),
                    "busta": round(float(p.get("importo_busta") or 0), 2),
                    "saldo": round(saldo, 2)})
    out.sort(key=lambda x: (x["anno"] or 0, x["mese"] or 0))
    return {"righe": out, "totale": len(out), "importo": round(sum(x["saldo"] for x in out), 2)}


@router.get("/paghe/prima-nota")
async def prima_nota(dipendente_id: str):
    """Prima nota salari di un dipendente: tutti i mesi con busta, erogato (bonifici+acconti)
    e saldo progressivo (cumulato busta − cumulato erogato; >0 = ancora da pagare)."""
    db = get_db()
    paghe = await db.paghe_mensili.find({"dipendente_id": dipendente_id}, {"_id": 0}).to_list(2000)
    paghe.sort(key=lambda p: (p.get("anno") or 0, p.get("mese") or 0))
    out, saldo = [], 0.0
    for p in paghe:
        busta = float(p.get("importo_busta") or 0)
        acc = sum(float(a.get("importo") or 0) for a in (p.get("acconti") or []))
        bon = float(p.get("bonifico_importo") or 0)
        erogato = bon + acc
        if busta == 0 and erogato == 0:
            continue
        saldo += busta - erogato
        out.append({"anno": p.get("anno"), "mese": p.get("mese"), "busta": round(busta, 2),
                    "bonifico": round(bon, 2), "acconti": round(acc, 2),
                    "erogato": round(erogato, 2), "saldo_progressivo": round(saldo, 2)})
    return {"righe": out, "saldo_finale": round(saldo, 2)}


@router.get("/paghe/associazioni-bonifici")
async def associazioni_bonifici(anno: Optional[int] = None, mese: Optional[int] = None,
                                stato: Optional[str] = None):
    """Vista UNICA cedolino↔bonifico. Per ogni busta del periodo mostra l'importo busta,
    i bonifici REALMENTE pagati (collezione pagamenti_esiti: data, importo, causale, riferimento/CRO),
    gli acconti, il saldo e lo stato di associazione:
      - pagato            = erogato (bonifici+acconti) ≥ busta
      - parziale          = erogato > 0 ma < busta
      - da_pagare         = busta presente, nessun pagamento
      - bonifico_senza_busta = pagamento presente ma nessuna busta
    Inoltre indica la 'fonte' del bonifico (banca/prima_nota/manuale), la 'qualita' del match
    (esatto/per_importo/aggregato/da_verificare) e se esiste il PDF del cedolino.
    Sorgente dati = sistema vivo paghe_mensili + pagamenti_esiti (nessun sistema parallelo)."""
    return await _calcola_associazioni_bonifici(get_db(), anno, mese, stato)


async def _calcola_associazioni_bonifici(db, anno: Optional[int] = None, mese: Optional[int] = None,
                                         stato: Optional[str] = None):
    q = {}
    if anno:
        q["anno"] = int(anno)
    if mese:
        q["mese"] = int(mese)

    dip_map = {}
    async for d in db.dipendenti.find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "codice_fiscale": 1}):
        dip_map[d["id"]] = d

    # Prefetch in blocco (UNA lettura per tabella) invece che una query per ogni
    # riga di paghe_mensili: l'adattatore Supabase non ha indici, ogni find/
    # find_one legge l'intera tabella e filtra in Python — con centinaia/migliaia
    # di buste, farlo dentro il ciclo era un incrocio N×M che portava la risposta
    # a decine o centinaia di secondi (e in produzione ha saturato il pool di
    # connessioni, con 502 a cascata anche su endpoint scollegati). Qui si legge
    # ogni tabella una sola volta e si indicizza in memoria.
    # pdf_data escluso qui per lo stesso motivo dei cedolini più sotto: alcuni
    # pagamenti (import da Drive, ponte bonifici storici) ora portano il PDF
    # allegato, un'inclusione lo trasferirebbe comunque per intero per ogni riga.
    esiti_idx: Dict[tuple, List[Dict[str, Any]]] = {}
    async for e in db.pagamenti_esiti.find({}, {"_id": 0, "pdf_data": 0}):
        esiti_idx.setdefault((e.get("dipendente_id"), e.get("mese"), e.get("anno")), []).append(e)
    for lst in esiti_idx.values():
        lst.sort(key=lambda e: e.get("data") or "")

    # NON includere pdf_data qui: l'adattatore Supabase ottimizza solo le
    # proiezioni "ad esclusione" (campo: 0) — una proiezione a inclusione
    # come questa arriverebbe comunque per intero, PDF in base64 compresi
    # (anche centinaia di MB su ~1500 cedolini). Il PDF si scarica a parte,
    # da qui basta sapere che il cedolino esiste (has_pdf = bool(ced)).
    cedolini_lista: List[Dict[str, Any]] = []
    ced_by_periodo: Dict[tuple, Dict[str, Any]] = {}
    async for c in db.cedolini.find({}, {"_id": 0, "pdf_data": 0}):
        cedolini_lista.append(c)
        chiave = (c.get("dipendente_id"), c.get("mese"), c.get("anno"))
        ced_by_periodo.setdefault(chiave, c)

    def trova_cedolino(dip_id, cognome, mese_p, anno_p):
        c = ced_by_periodo.get((dip_id, mese_p, anno_p))
        if c:
            return c
        if cognome:
            cg = cognome.lower()
            for cand in cedolini_lista:
                if cand.get("mese") == mese_p and cand.get("anno") == anno_p and cg in (cand.get("nome_dipendente") or "").lower():
                    return cand
        return None

    righe = []
    tot = {"buste": 0.0, "bonifici": 0.0, "acconti": 0.0, "saldo": 0.0,
           "pagati": 0, "parziali": 0, "da_pagare": 0, "senza_busta": 0,
           "associati": 0, "da_verificare": 0}

    async for p in db.paghe_mensili.find(q, {"_id": 0}):
        busta = float(p.get("importo_busta") or 0)
        bon = float(p.get("bonifico_importo") or 0)
        acc_list = p.get("acconti") or []
        acc = sum(float(a.get("importo") or 0) for a in acc_list)
        if busta <= 0 and bon <= 0 and acc <= 0:
            continue

        dip_id = p.get("dipendente_id")
        dip = dip_map.get(dip_id) or {}
        nome = f"{dip.get('cognome', '')} {dip.get('nome', '')}".strip() or dip_id

        # Bonifici reali pagati (esiti banca) per questo dipendente/mese/anno
        esiti = [{
            "data": e.get("data"),
            "importo": round(float(e.get("importo") or 0), 2),
            "causale": e.get("causale") or "",
            "beneficiario": e.get("beneficiario") or "",
            "riferimento": e.get("cro") or e.get("key") or "",
            "pdf_key": e.get("key") if e.get("ha_pdf") else None,
        } for e in esiti_idx.get((dip_id, p.get("mese"), p.get("anno")), [])]

        erogato = bon + acc
        if busta <= 0 and erogato > 0:
            st = "bonifico_senza_busta"
            tot["senza_busta"] += 1
        elif erogato <= 0:
            st = "da_pagare"
            tot["da_pagare"] += 1
        elif erogato + 0.5 >= busta:
            st = "pagato"
            tot["pagati"] += 1
        else:
            st = "parziale"
            tot["parziali"] += 1

        # Fonte del bonifico
        if esiti:
            fonte = "banca"
        elif p.get("bonifico_da_prima_nota"):
            fonte = "prima_nota"
        elif bon > 0:
            fonte = "manuale"
        else:
            fonte = None

        # Qualità dell'associazione (quanto è affidabile il legame busta↔bonifico)
        if bon <= 0:
            qualita = None
        elif esiti:
            if len(esiti) == 1 and busta > 0 and abs(esiti[0]["importo"] - busta) <= 0.5:
                qualita = "esatto"          # un solo bonifico che combacia con la busta
            elif busta > 0 and abs(bon - busta) <= 0.5:
                qualita = "per_importo"     # somma bonifici = busta
            elif len(esiti) > 1:
                qualita = "aggregato"       # più bonifici nello stesso mese
            else:
                qualita = "per_importo"
        else:
            qualita = "da_verificare"       # importo inserito a mano / da prima nota, senza prova banca

        associato = bool(p.get("bonifico_riconciliato")) or qualita in ("esatto", "per_importo")
        if st in ("pagato", "parziale", "bonifico_senza_busta"):
            if associato:
                tot["associati"] += 1
            else:
                tot["da_verificare"] += 1

        # Esiste il cedolino per questo periodo? (il PDF non è stato letto qui,
        # vedi nota sul prefetch sopra — quasi ogni cedolino importato ne ha uno)
        ced = trova_cedolino(dip_id, dip.get("cognome"), p.get("mese"), p.get("anno"))
        has_pdf = bool(ced)
        cedolino_id = ced.get("id") if ced else None

        if stato and st != stato:
            continue

        tot["buste"] += busta
        tot["bonifici"] += bon
        tot["acconti"] += acc
        tot["saldo"] += (busta - erogato)

        righe.append({
            "dipendente_id": dip_id,
            "dipendente": nome,
            "anno": p.get("anno"),
            "mese": p.get("mese"),
            "busta": round(busta, 2),
            "bonifico": round(bon, 2),
            "acconti": round(acc, 2),
            "erogato": round(erogato, 2),
            "saldo": round(busta - erogato, 2),
            "stato": st,
            "fonte": fonte,
            "qualita": qualita,
            "associato": associato,
            "riconciliato": bool(p.get("bonifico_riconciliato")),
            "bonifico_data": p.get("bonifico_data"),
            "bonifici": esiti,
            "n_bonifici": len(esiti),
            "cedolino_pdf": has_pdf,
            "cedolino_id": cedolino_id,
        })

    righe.sort(key=lambda r: ((r["anno"] or 0), (r["mese"] or 0), r["dipendente"]), reverse=True)
    for k in ("buste", "bonifici", "acconti", "saldo"):
        tot[k] = round(tot[k], 2)
    return {"righe": righe, "totali": tot, "count": len(righe)}


@router.get("/paghe/associazioni-bonifici/export-excel")
async def associazioni_bonifici_export_excel(anno: Optional[int] = None, mese: Optional[int] = None,
                                              stato: Optional[str] = None):
    """Esporta in Excel la stessa vista di /paghe/associazioni-bonifici: una riga per
    dipendente/periodo con dipendente, periodo cedolino, importo cedolino, importo bonifico
    e stato dell'associazione. Stessa fonte dati (nessun sistema parallelo)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    dati = await _calcola_associazioni_bonifici(get_db(), anno, mese, stato)
    mesi = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio",
            "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cedolini e Bonifici"
    intestazioni = ["Dipendente", "Periodo", "Importo Cedolino", "Importo Bonifico",
                    "Acconti", "Erogato", "Saldo", "Stato", "Qualità match", "Fonte",
                    "N. Bonifici", "Data ultimo bonifico", "PDF Cedolino"]
    ws.append(intestazioni)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B7A6B", end_color="5B7A6B", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    stati_lbl = {"pagato": "Pagato", "parziale": "Parziale", "da_pagare": "Da pagare",
                 "bonifico_senza_busta": "Bonifico senza busta"}
    qualita_lbl = {"esatto": "Match esatto", "per_importo": "Match per importo",
                   "aggregato": "Più bonifici", "da_verificare": "Da verificare"}
    for r in dati["righe"]:
        periodo = f"{mesi[r['mese'] - 1]} {r['anno']}" if r.get("mese") and 1 <= r["mese"] <= 12 else f"{r.get('mese')}/{r.get('anno')}"
        # "Data ultimo bonifico": paghe_mensili.bonifico_data non viene mai
        # popolato dagli import via pagamenti_esiti (CSV, Drive, ponte storico) —
        # la data vera sta sui singoli pagamenti (r["bonifici"], già ordinati per
        # data), non sul campo aggregato della busta.
        date_bonifici = [b.get("data") for b in (r.get("bonifici") or []) if b.get("data")]
        data_ultimo = max(date_bonifici) if date_bonifici else (r.get("bonifico_data") or "")
        ws.append([
            r.get("dipendente"), periodo,
            r.get("busta") or 0, r.get("bonifico") or 0, r.get("acconti") or 0,
            r.get("erogato") or 0, r.get("saldo") or 0,
            stati_lbl.get(r.get("stato"), r.get("stato")),
            qualita_lbl.get(r.get("qualita"), r.get("qualita") or ""),
            r.get("fonte") or "", r.get("n_bonifici") or 0,
            data_ultimo, "Sì" if r.get("cedolino_pdf") else "No",
        ])
    for col in ws.columns:
        larghezza = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max(larghezza, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_file = "cedolini_bonifici"
    if anno:
        nome_file += f"_{anno}"
    if mese:
        nome_file += f"_{int(mese):02d}"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_file}.xlsx"'})


@router.post("/paghe/conferma-associazione")
async def conferma_associazione(data: dict):
    """Conferma/annulla manualmente l'associazione bonifico↔cedolino di una busta.
    Imposta bonifico_riconciliato e traccia data/nota. Non crea record nuovi:
    agisce sul record paghe_mensili esistente (sistema unico)."""
    dip = data.get("dipendente_id")
    anno = data.get("anno")
    mese = data.get("mese")
    if not dip or not anno or not mese:
        raise HTTPException(status_code=400, detail="dipendente_id, anno, mese obbligatori")
    val = bool(data.get("riconciliato", True))
    set_doc = {"bonifico_riconciliato": val, "updated_at": now_iso()}
    if val:
        set_doc["associazione_confermata_at"] = now_iso()
    if data.get("nota") is not None:
        set_doc["associazione_nota"] = str(data.get("nota"))
    res = await get_db().paghe_mensili.update_one(
        {"dipendente_id": dip, "anno": int(anno), "mese": int(mese)}, {"$set": set_doc})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Busta non trovata per quel dipendente/mese")
    return {"ok": True, "riconciliato": val}


# ============ BUSTE PAGA ============

@router.get("/buste-paga")
async def get_buste_paga(anno: Optional[int] = None, mese: Optional[int] = None, dipendente_id: Optional[str] = None):
    """
    Recupera cedolini dalla collezione 'cedolini' (dati storici dal 2014).
    Se dipendente_id è fornito, cerca anche per nome del dipendente in dipendenti_cloud.
    """
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    
    # Se abbiamo un dipendente_id, dobbiamo trovare il nome per cercare nei cedolini
    if dipendente_id:
        dip = await get_db().dipendenti_cloud.find_one({"id": dipendente_id})
        if dip:
            nome_completo = f"{dip.get('nome', '')} {dip.get('cognome', '')}".strip().upper()
            query["$or"] = [
                {"dipendente_id": dipendente_id},
                {"nome_dipendente": {"$regex": dip.get('cognome', ''), "$options": "i"}}
            ]
    
    # Leggi dalla collezione cedolini (dati storici)
    cedolini = await get_db().cedolini.find(query, {"_id": 0}).sort([("anno", -1), ("mese", -1)]).to_list(1000)
    
    # Normalizza i campi per compatibilità con il frontend
    result = []
    for c in cedolini:
        result.append({
            "id": c.get("id", str(c.get("_id", ""))),
            "dipendente_id": c.get("dipendente_id", ""),
            "dipendente_nome": c.get("nome_dipendente") or c.get("dipendente_nome") or "",
            "mese": c.get("mese"),
            "anno": c.get("anno"),
            "lordo": c.get("lordo", 0),
            "netto": c.get("netto", 0),
            "inps": c.get("inps_dipendente", 0),
            "irpef": c.get("irpef", 0),
            "trattenute": c.get("trattenute", 0),
            "stato": c.get("stato_pagamento") or c.get("stato") or "DA_PAGARE",
            "created_at": c.get("created_at", "")
        })
    
    return result

@router.post("/buste-paga")
async def create_busta_paga(busta: BustaPagaCloud):
    busta_dict = busta.model_dump()
    busta_dict["id"] = generate_id()
    busta_dict["created_at"] = now_iso()
    await get_db().buste_paga_cloud.insert_one(busta_dict)
    return serialize_doc(busta_dict)

@router.post("/buste-paga/genera")
async def genera_buste_paga(data: dict):
    """Genera buste paga per tutti i dipendenti attivi per un mese specifico"""
    mese = data.get("mese")
    anno = data.get("anno")
    lordo_default = data.get("lordo", 1500)
    
    if not mese or not anno:
        raise HTTPException(status_code=400, detail="mese e anno sono obbligatori")
    
    dipendenti = await get_db().dipendenti_cloud.find({"stato": "attivo"}, {"_id": 0}).to_list(1000)
    created = 0
    
    for dip in dipendenti:
        existing = await get_db().buste_paga_cloud.find_one({
            "dipendente_id": dip["id"],
            "mese": mese,
            "anno": anno
        })
        
        if not existing:
            inps = round(lordo_default * 0.0919, 2)
            irpef = round((lordo_default - inps) * 0.23, 2)
            netto = round(lordo_default - inps - irpef, 2)
            
            busta = {
                "id": generate_id(),
                "dipendente_id": dip["id"],
                "mese": mese,
                "anno": anno,
                "lordo": lordo_default,
                "inps": inps,
                "irpef": irpef,
                "trattenute": 0,
                "netto": netto,
                "stato": "DA_PAGARE",
                "created_at": now_iso()
            }
            await get_db().buste_paga_cloud.insert_one(busta)
            created += 1
    
    return {"message": f"Generate {created} buste paga"}

@router.put("/buste-paga/{busta_id}/paga")
async def paga_busta(busta_id: str):
    result = await get_db().buste_paga_cloud.update_one(
        {"id": busta_id},
        {"$set": {"stato": "PAGATO", "data_pagamento": now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Busta paga non trovata")
    return {"message": "Busta paga marcata come pagata"}

# ============ MISSIONI ============

@router.get("/missioni")
async def get_missioni(dipendente_id: Optional[str] = None, stato: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if stato:
        query["stato"] = stato
    missioni = await get_db().missioni_cloud.find(query, {"_id": 0}).to_list(1000)
    return missioni

@router.post("/missioni")
async def create_missione(missione: MissioneCloud):
    miss_dict = missione.model_dump()
    miss_dict["id"] = generate_id()
    miss_dict["created_at"] = now_iso()
    await get_db().missioni_cloud.insert_one(miss_dict)
    return serialize_doc(miss_dict)

@router.put("/missioni/{missione_id}/approva")
async def approva_missione(missione_id: str):
    db = get_db()
    miss = await db.missioni_cloud.find_one({"id": missione_id}, {"_id": 0})
    if not miss:
        raise HTTPException(status_code=404, detail="Missione non trovata")
    await db.missioni_cloud.update_one(
        {"id": missione_id}, {"$set": {"stato": "approvata", "approvata_il": now_iso()}})

    automazioni = []
    rimborso = float(miss.get("rimborso") or 0)
    dip_id = miss.get("dipendente_id")
    dip = await db.dipendenti.find_one({"id": dip_id}, {"_id": 0, "nome_completo": 1, "nome": 1, "cognome": 1}) if dip_id else None
    nome = (dip or {}).get("nome_completo") or (f"{(dip or {}).get('cognome','')} {(dip or {}).get('nome','')}".strip() if dip else "")

    # Rimborso missione → partita aperta (tracciamento finanziario)
    if rimborso > 0 and dip_id:
        try:
            from app.hr.services.partite_aperte_engine import crea_partita, TipoPartita
            await crea_partita(
                tipo=TipoPartita.ALTRO, documento_id=missione_id,
                documento_collection="missioni_cloud", controparte_id=dip_id,
                controparte_nome=nome, importo=rimborso, db=db, data_documento=now_iso()[:10],
                extra={"categoria": "rimborso_missione", "destinazione": miss.get("destinazione")})
            automazioni.append("partita_rimborso")
        except Exception:
            pass
    # Notifica al dipendente
    if dip_id:
        try:
            from app.hr.services.notifiche import crea_notifica
            await crea_notifica(db, dip_id, "missione", "Missione approvata",
                                f"La missione a {miss.get('destinazione','')} è stata approvata"
                                + (f" · rimborso € {rimborso:.2f}" if rimborso > 0 else "") + ".",
                                extra={"missione_id": missione_id})
            automazioni.append("notifica_dipendente")
        except Exception:
            pass
    return {"message": "Missione approvata", "automazioni": automazioni}

@router.delete("/missioni/{missione_id}")
async def delete_missione(missione_id: str):
    result = await get_db().missioni_cloud.delete_one({"id": missione_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Missione non trovata")
    return {"message": "Missione eliminata"}

# ============ DOCUMENTI ============

@router.get("/documenti")
async def get_documenti(dipendente_id: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    documenti = await get_db().documenti_cloud.find(query, {"_id": 0}).to_list(1000)
    return documenti

@router.post("/documenti")
async def create_documento(doc: DocumentoCloud):
    doc_dict = doc.model_dump()
    doc_dict["id"] = generate_id()
    doc_dict["data_caricamento"] = now_iso()
    await get_db().documenti_cloud.insert_one(doc_dict)
    return serialize_doc(doc_dict)

@router.delete("/documenti/{documento_id}")
async def delete_documento(documento_id: str):
    result = await get_db().documenti_cloud.delete_one({"id": documento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    return {"message": "Documento eliminato"}


_CF_DOC_RE = re.compile(r'\b([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])\b')
CATEGORIE_DOC = ["UNILAV", "CERTIFICAZIONE_UNICA", "CONTRATTO", "RIDUZIONE_ORARIO", "BONIFICO",
                 "CODICE_FISCALE", "CARTA_IDENTITA", "BUSTA_PAGA", "ALTRO"]


def classifica_documento(text: str, filename: str = "") -> str:
    """Riconosce il tipo di documento dal testo e, in fallback, dal nome del file
    (utile per le scansioni-immagine senza testo). Diciture standard italiane."""
    t = (text or "").lower()
    fn = (filename or "").lower()

    def H(s, *ks):
        return any(k in s for k in ks)
    # 1) Segnali forti dal TESTO
    if H(t, "unilav", "comunicazione obbligatoria", "modello unificato lav", "centro per l'impiego"):
        return "UNILAV"
    if H(t, "certificazione unica", "redditi di lavoro dipendente e assimilati"):
        return "CERTIFICAZIONE_UNICA"
    if H(t, "contratto individuale di lavoro", "contratto di lavoro", "patto di prova", "lettera di assunzione"):
        return "CONTRATTO"
    if H(t, "bonifico", "ordinante", "beneficiario", "disposizione di pagamento", "sepa credit"):
        return "BONIFICO"
    if H(t, "busta paga", "cedolino", "netto in busta", "retribuzione lorda"):
        return "BUSTA_PAGA"
    if H(t, "carta di identità", "carta d'identità", "documento di identità", "carta d identita"):
        return "CARTA_IDENTITA"
    if H(t, "riduzione orario", "riduzione dell'orario", "riduzione dell orario", "trasformazione part-time", "riduzione part time"):
        return "RIDUZIONE_ORARIO"
    # 2) Nome FILE (per scansioni senza testo)
    if H(fn, "riduzione"):
        return "RIDUZIONE_ORARIO"
    if H(fn, "unilav"):
        return "UNILAV"
    if H(fn, "certificazione_unica", "certificazione unica", "_cu_", "cud"):
        return "CERTIFICAZIONE_UNICA"
    if H(fn, "contratto"):
        return "CONTRATTO"
    if H(fn, "bonific"):
        return "BONIFICO"
    if H(fn, "carta_di_identit", "carta d'identit", "carta identit", "carta_identit"):
        return "CARTA_IDENTITA"
    if H(fn, "codice_fiscale", "codice fiscale", "tessera_sanitaria", "tessera sanitaria"):
        return "CODICE_FISCALE"
    if H(fn, "busta", "cedolino"):
        return "BUSTA_PAGA"
    # 3) Segnale debole dal testo
    if H(t, "tessera sanitaria", "servizio sanitario nazionale"):
        return "CODICE_FISCALE"
    return "ALTRO"


async def _indici_dipendenti(db):
    """Indici per riconoscere il dipendente da CF/nome nei documenti."""
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1, "codice_fiscale": 1}).to_list(1000)

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()
    by_cf, by_nome, by_cogn = {}, {}, {}
    for d in dips:
        cf = (d.get("codice_fiscale") or "").upper().strip()
        if cf:
            by_cf[cf] = d
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v and len(v) > 6:
                by_nome[v] = d
        if len(c) >= 4:
            by_cogn.setdefault(c, []).append(d)
    return {"cf": by_cf, "nome": by_nome, "cogn": by_cogn}


async def _archivia_documento_cloud(db, filename, raw, contesto="", indici=None, origine="upload_massivo"):
    """Classifica un PDF (UNILAV, C.U., contratto, CF, busta…), trova il dipendente dal
    codice fiscale o dal nome, e lo archivia nella sua cartella (documenti_cloud).
    Anti-duplicati per hash. Riusato da upload massivo E import da email/Gmail.
    Ritorna (esito, categoria, dipendente_nome) con esito in
    'caricato'|'non_assegnato'|'duplicato'|'vuoto'."""
    import io
    if not raw:
        return ("vuoto", None, None)
    h = hashlib.sha256(raw).hexdigest()
    if await db.documenti_cloud.find_one({"hash": h}):
        return ("duplicato", None, None)
    text = ""
    if raw[:4] == b"%PDF":
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for p in pdf.pages[:6]:
                    text += (p.extract_text() or "") + "\n"
        except Exception:
            text = ""
    if indici is None:
        indici = await _indici_dipendenti(db)
    by_cf, by_nome, by_cogn = indici["cf"], indici["nome"], indici["cogn"]

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()
    categoria = classifica_documento(text, f"{contesto} {filename}".strip())
    d = None
    for cf in _CF_DOC_RE.findall((text or "").upper()):
        if cf in by_cf:
            d = by_cf[cf]
            break
    if not d:
        tl = norm(text)
        for nome_n, dd in by_nome.items():
            if nome_n in tl:
                d = dd
                break
    if not d:
        fn_norm = norm(f"{contesto} {filename}".replace("_", " ").replace("-", " "))
        for nome_n, dd in by_nome.items():
            if nome_n in fn_norm:
                d = dd
                break
        if not d:
            for cogn, lst in by_cogn.items():
                if cogn in fn_norm and len(lst) == 1:
                    d = lst[0]
                    break
    doc = {"id": generate_id(),
           "dipendente_id": (d or {}).get("id"),
           "dipendente_nome": (f"{d.get('cognome','')} {d.get('nome','')}".strip() if d else None),
           "titolo": filename, "filename": filename,
           "tipo": categoria, "categoria": categoria, "hash": h,
           "file_data": base64.b64encode(raw).decode(),
           "assegnato": bool(d), "origine": origine, "data_caricamento": now_iso()}
    await db.documenti_cloud.insert_one(doc)
    return ("caricato" if d else "non_assegnato", categoria, doc["dipendente_nome"])


@router.post("/documenti/upload-massivo")
async def upload_documenti_massivo(files: List[UploadFile] = File(...)):
    """Carica più documenti insieme: per ognuno riconosce il tipo (UNILAV, C.U., contratto,
    bonifico, codice fiscale…), trova il dipendente dal codice fiscale (o dal nome) nel testo,
    e lo archivia nella sua cartella. Anti-duplicati per hash del file."""
    import io
    db = get_db()
    indici = await _indici_dipendenti(db)

    import zipfile
    caricati, duplicati, non_assegnati, per_categoria = [], [], [], {}

    async def processa(filename, raw, contesto=""):
        esito, categoria, nome = await _archivia_documento_cloud(db, filename, raw, contesto=contesto, indici=indici)
        if esito == "duplicato":
            duplicati.append(filename)
        elif esito in ("caricato", "non_assegnato"):
            per_categoria[categoria] = per_categoria.get(categoria, 0) + 1
            (caricati if esito == "caricato" else non_assegnati).append({"file": filename, "categoria": categoria, "dipendente": nome})

    for f in files:
        raw = await f.read()
        fn = f.filename or ""
        if fn.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                    for nm in zf.namelist():
                        if nm.endswith("/"):
                            continue
                        await processa(nm.split("/")[-1], zf.read(nm), contesto=fn)
            except zipfile.BadZipFile:
                non_assegnati.append({"file": fn, "categoria": "ALTRO"})
        else:
            await processa(fn, raw, contesto="")
    return {"caricati": len(caricati), "duplicati": duplicati,
            "non_assegnati": non_assegnati, "per_categoria": per_categoria,
            "dettaglio": caricati[:300]}


@router.get("/documenti/{documento_id}/file")
async def download_documento(documento_id: str):
    from fastapi.responses import Response
    doc = await get_db().documenti_cloud.find_one({"id": documento_id}, {"_id": 0})
    if not doc or not doc.get("file_data"):
        raise HTTPException(status_code=404, detail="File non disponibile")
    data = base64.b64decode(doc["file_data"])
    fn = doc.get("filename") or "documento.pdf"
    media = "application/pdf" if fn.lower().endswith(".pdf") else "application/octet-stream"
    return Response(content=data, media_type=media, headers={"Content-Disposition": f'inline; filename="{fn}"'})

# ============ DASHBOARD STATS ============

@router.get("/dashboard/stats")
async def get_dashboard_stats():
    dipendenti = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    attivi = [d for d in dipendenti if d.get("attivo", True) is not False and d.get("stato", "attivo") not in ("cessato", "disattivo", "inattivo")]
    
    ferie_pending = await get_db().ferie_cloud.count_documents({"stato": "in_attesa"})
    missioni_pending = await get_db().missioni_cloud.count_documents({"stato": "in_attesa"})
    
    # Presenze oggi
    today = datetime.now().strftime("%Y-%m-%d")
    presenze_oggi = await get_db().presenze_cloud.count_documents({"data": today, "stato": "presente"})

    # Solo alert HR: la collezione `alerts` è condivisa con l'ERP contabile, qui
    # mostriamo soltanto i moduli del personale (niente fatture/fornitori/banca…).
    alert_aperti = await get_db().alerts.count_documents(
        {"stato": "aperto", "modulo": {"$in": MODULI_HR}})

    # Buste in attesa di pagamento (motore unico): busta presente ma non ancora pagata
    buste_attesa = 0
    importo_attesa = 0.0
    async for p in get_db().paghe_mensili.find(
            {"stato_pagamento": {"$in": ["in_attesa_pagamento", "parziale"]}},
            {"_id": 0, "saldo": 1, "importo_busta": 1, "bonifico_importo": 1}):
        saldo = p.get("saldo")
        if saldo is None:
            saldo = float(p.get("importo_busta") or 0) - float(p.get("bonifico_importo") or 0)
        if saldo and saldo > 0.5:
            buste_attesa += 1
            importo_attesa += saldo

    return {
        "totale_dipendenti": len(dipendenti),
        "dipendenti_attivi": len(attivi),
        "ferie_in_attesa": ferie_pending,
        "missioni_in_attesa": missioni_pending,
        "presenze_oggi": presenze_oggi,
        "alert_aperti": alert_aperti,
        "buste_in_attesa": buste_attesa,
        "importo_in_attesa": round(importo_attesa, 2),
    }


# Moduli di competenza HR (gli altri appartengono all'ERP contabile OpenClaw).
MODULI_HR = ["dipendenti", "cedolini"]

# ID interni (UUID) dentro i messaggi degli alert: vanno tradotti in testo
# leggibile per il titolare (data, importo e descrizione del movimento banca).
_UUID_RE = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")


def _data_it(iso: str) -> str:
    try:
        return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return iso or "?"


async def _uuid_in_testo_leggibile(mov_id: str) -> Optional[str]:
    """Se l'UUID corrisponde a un movimento bancario, ritorna «descrizione»
    con data e importo; altrimenti None (l'ID resta com'è)."""
    for coll, campo_desc in (("estratto_conto_movimenti", "descrizione_originale"),
                             ("prima_nota_banca", "descrizione")):
        m = await get_db()[coll].find_one({"id": mov_id}, {"_id": 0})
        if m:
            desc = (m.get(campo_desc) or m.get("descrizione") or "").strip()
            importo = m.get("importo")
            pezzi = [f"bancario del {_data_it(m.get('data', ''))}"]
            if importo is not None:
                pezzi.append(f"di € {abs(float(importo)):.2f}")
            if desc:
                pezzi.append(f"«{desc[:120]}»")
            return " ".join(pezzi)
    return None


async def _messaggio_leggibile(messaggio: str) -> str:
    """Sostituisce gli UUID nel messaggio con i dati veri del movimento banca."""
    if not messaggio:
        return messaggio
    for mov_id in set(_UUID_RE.findall(messaggio)):
        leggibile = await _uuid_in_testo_leggibile(mov_id)
        if leggibile:
            messaggio = messaggio.replace(mov_id, leggibile)
    return messaggio


@router.get("/alerts")
async def lista_alert(modulo: str = "", severita: str = "", stato: str = "aperto"):
    """Elenco degli alert HR (scadenze contratti/prova, contestazioni…).
    La collezione `alerts` è condivisa con la contabilità: qui filtriamo ai soli
    moduli del personale. `stato`: 'aperto' (default), 'risolto' (archivio),
    'tutti' (entrambi). Gli alert risolti NON vengono cancellati: restano in archivio."""
    q = {}
    if stato and stato != "tutti":
        q["stato"] = stato
    if modulo:
        if modulo not in MODULI_HR:
            return {"totale": 0, "alerts": []}
        q["modulo"] = modulo
    else:
        q["modulo"] = {"$in": MODULI_HR}
    if severita:
        q["severita"] = severita
    sort_field = "resolved_at" if stato == "risolto" else "created_at"
    alerts = await get_db().alerts.find(q, {"_id": 0}).sort(sort_field, -1).to_list(500)
    # Traduzione a lettura: anche gli alert vecchi già salvati con l'ID interno
    # diventano leggibili, senza toccare il dato in archivio.
    for a in alerts:
        try:
            a["messaggio"] = await _messaggio_leggibile(a.get("messaggio", ""))
        except Exception:
            pass
    return {"totale": len(alerts), "alerts": alerts}


@router.post("/alerts/{alert_id}/risolvi")
async def risolvi_alert_id(alert_id: str):
    """Segna un alert come risolto (manuale)."""
    r = await get_db().alerts.update_one(
        {"id": alert_id, "stato": "aperto"},
        {"$set": {"stato": "risolto", "risolto": True,
                  "resolved_at": now_iso(), "resolved_by": "admin"}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alert non trovato o già risolto")
    return {"ok": True, "stato": "risolto"}

# ============ SEED DATA ============

@router.post("/seed-data")
async def seed_data():
    """Crea dati di esempio se non esistono"""
    existing = await get_db().dipendenti_cloud.count_documents({})
    if existing > 0:
        return {"message": "Dati già presenti"}
    
    # Crea dipendenti di esempio
    dipendenti_sample = [
        {"nome": "Mario", "cognome": "Rossi", "ruolo": "Responsabile", "stato": "attivo", "contratto": "Indeterminato"},
        {"nome": "Lucia", "cognome": "Bianchi", "ruolo": "Cameriere", "stato": "attivo", "contratto": "Determinato"},
        {"nome": "Giuseppe", "cognome": "Verdi", "ruolo": "Barista", "stato": "attivo", "contratto": "Indeterminato"},
    ]
    
    for d in dipendenti_sample:
        d["id"] = generate_id()
        d["created_at"] = now_iso()
        await get_db().dipendenti_cloud.insert_one(d)
    
    # Crea turni di esempio
    turni_sample = [
        {"nome": "Mattina", "orario_inizio": "06:00", "orario_fine": "14:00", "colore": "#3b82f6"},
        {"nome": "Pomeriggio", "orario_inizio": "14:00", "orario_fine": "22:00", "colore": "#10b981"},
        {"nome": "Notte", "orario_inizio": "22:00", "orario_fine": "06:00", "colore": "#8b5cf6"},
    ]
    
    for t in turni_sample:
        t["id"] = generate_id()
        await get_db().turni_cloud.insert_one(t)
    
    return {"message": "Dati di esempio creati"}
