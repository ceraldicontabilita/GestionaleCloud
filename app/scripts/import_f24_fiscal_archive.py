"""Validate an F24 archive and optionally import its evidence ledger.

Dry-run is the default and never opens the live Sheets registry. Writes require
the explicit ``--esegui`` switch after the complete archive has validated.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import io
import json
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from app.config import settings
from app.database import Database
from app.services.f24_fiscal_evidence import (
    PARSER_KIND_PRINTABLE,
    PARSER_KIND_QUIETANZA,
    ingest_f24_evidence,
    normalize_f24_evidence_rows,
    parse_f24_evidence,
)


MANIFEST_NAME = "INDICE_UNICO_DOCUMENTI_F24.csv"
XLSX_INDEX_SHEET = "INDICE_F24_PDF"


def _kind(row: dict[str, str]) -> str:
    label = (row.get("tipo_documento") or "").casefold()
    return PARSER_KIND_PRINTABLE if "stampabile" in label else PARSER_KIND_QUIETANZA


def _year(value: str) -> int | None:
    for token in reversed((value or "").replace("-", "/").split("/")):
        if len(token) == 4 and token.isdigit():
            return int(token)
    return None


def _xlsx_index_rows(content: bytes) -> tuple[list[dict[str, str]], int | None]:
    """Converte l'indice Excel finale dello ZIP nel formato manifest interno."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if XLSX_INDEX_SHEET not in workbook.sheetnames:
            return [], None
        sheet = workbook[XLSX_INDEX_SHEET]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        required = {"Data", "Tipo", "Protocollo", "PDF", "SHA256"}
        if not required.issubset(headers):
            missing = ", ".join(sorted(required - set(headers)))
            raise ValueError(f"Colonne mancanti in {XLSX_INDEX_SHEET}: {missing}")
        indexes = {name: headers.index(name) for name in required}
        result: list[dict[str, str]] = []
        for values in rows:
            relative = values[indexes["PDF"]]
            if not relative:
                continue
            date_value = values[indexes["Data"]]
            if hasattr(date_value, "strftime"):
                date_value = date_value.strftime("%d/%m/%Y")
            result.append({
                "file": str(relative).strip(),
                "sha256": str(values[indexes["SHA256"]] or "").strip(),
                "tipo_documento": str(values[indexes["Tipo"]] or "").strip(),
                "data_versamento": str(date_value or "").strip(),
                "protocollo_telematico": str(
                    values[indexes["Protocollo"]] or ""
                ).strip(),
            })
        declared_f24_rows = None
        if "F24_TUTTI" in workbook.sheetnames:
            f24_rows = workbook["F24_TUTTI"].iter_rows(values_only=True)
            f24_headers = [str(value or "").strip() for value in next(f24_rows, ())]
            if "SHA256" in f24_headers:
                sha_index = f24_headers.index("SHA256")
                declared_f24_rows = sum(
                    1 for values in f24_rows
                    if values[sha_index] not in (None, "")
                )
        return result, declared_f24_rows
    finally:
        workbook.close()


def _archive_reader(root: Path):
    """Restituisce manifest e loader sicuro per cartella estratta o ZIP."""
    if root.is_file() and root.suffix.casefold() == ".zip":
        archive = zipfile.ZipFile(root)
        bad_member = archive.testzip()
        if bad_member:
            archive.close()
            raise ValueError(f"ZIP corrotto: {bad_member}")
        manifests = [
            name for name in archive.namelist()
            if Path(name).name == MANIFEST_NAME and not name.endswith("/")
        ]
        if len(manifests) == 1:
            manifest_name = manifests[0]
            manifest_parent = Path(manifest_name).parent
            source_rows = list(csv.DictReader(
                io.StringIO(archive.read(manifest_name).decode("utf-8-sig")), delimiter=";",
            ))
            manifest_format = "csv"
        elif not manifests:
            indexed_workbooks: list[tuple[str, list[dict[str, str]], int | None]] = []
            for name in archive.namelist():
                if name.endswith("/") or Path(name).suffix.casefold() != ".xlsx":
                    continue
                rows, declared_f24_rows = _xlsx_index_rows(archive.read(name))
                if rows:
                    indexed_workbooks.append((name, rows, declared_f24_rows))
            if len(indexed_workbooks) != 1:
                archive.close()
                raise FileNotFoundError(
                    f"Atteso un solo {MANIFEST_NAME} o indice Excel "
                    f"{XLSX_INDEX_SHEET}, trovati {len(indexed_workbooks)}"
                )
            manifest_name, source_rows, declared_f24_rows = indexed_workbooks[0]
            workbook_parent = Path(manifest_name).parent
            manifest_parent = (
                workbook_parent.parent
                if workbook_parent.name.casefold() == "01_excel"
                else workbook_parent
            )
            manifest_format = "xlsx"
        else:
            archive.close()
            raise FileNotFoundError(
                f"Atteso un solo {MANIFEST_NAME} nello ZIP, trovati {len(manifests)}"
            )

        def load(relative: Path) -> bytes:
            normalized = Path(*relative.parts)
            if normalized.is_absolute() or ".." in normalized.parts:
                raise ValueError("percorso PDF non sicuro nel manifest")
            member = (manifest_parent / normalized).as_posix()
            return archive.read(member)

        if manifest_format == "csv":
            declared_f24_rows = None
        return (
            source_rows, load, archive.close, f"zip:{root}", manifest_format,
            declared_f24_rows,
        )

    manifests = list(root.rglob(MANIFEST_NAME)) if root.is_dir() else []
    if len(manifests) == 1:
        manifest = manifests[0]
        manifest_parent = manifest.parent.resolve()
        with manifest.open("r", encoding="utf-8-sig", newline="") as stream:
            source_rows = list(csv.DictReader(stream, delimiter=";"))
        manifest_format = "csv"
        declared_f24_rows = None
    elif not manifests and root.is_dir():
        indexed_workbooks: list[tuple[Path, list[dict[str, str]], int | None]] = []
        for workbook_path in root.rglob("*.xlsx"):
            rows, row_count = _xlsx_index_rows(workbook_path.read_bytes())
            if rows:
                indexed_workbooks.append((workbook_path, rows, row_count))
        if len(indexed_workbooks) != 1:
            raise FileNotFoundError(
                f"Atteso un solo {MANIFEST_NAME} o indice Excel "
                f"{XLSX_INDEX_SHEET} sotto {root}, trovati {len(indexed_workbooks)}"
            )
        manifest, source_rows, declared_f24_rows = indexed_workbooks[0]
        manifest_parent = (
            manifest.parent.parent
            if manifest.parent.name.casefold() == "01_excel"
            else manifest.parent
        ).resolve()
        manifest_format = "xlsx"
    else:
        raise FileNotFoundError(
            f"Atteso un solo {MANIFEST_NAME} sotto {root}, trovati {len(manifests)}"
        )

    def load(relative: Path) -> bytes:
        path = (manifest_parent / relative).resolve()
        path.relative_to(manifest_parent)
        if not path.is_file():
            raise FileNotFoundError("PDF non trovato")
        return path.read_bytes()

    return (
        source_rows, load, lambda: None, str(manifest_parent), manifest_format,
        declared_f24_rows,
    )


def validate_archive(root: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    (
        source_rows, load_content, close_archive, archive_source, manifest_format,
        declared_f24_rows,
    ) = _archive_reader(root)

    validated: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    by_year: Counter[int] = Counter()
    by_type: Counter[str] = Counter()
    totals = defaultdict(float)
    credit_rows = 0
    hashes: Counter[str] = Counter()
    missing_protocols = 0
    missing_protocols_by_type: Counter[str] = Counter()

    try:
        for source in source_rows:
            relative = Path(source.get("file") or "")
            try:
                content = load_content(relative)
                path = relative
                if not content.startswith(b"%PDF"):
                    raise ValueError("contenuto non PDF")
                digest = hashlib.sha256(content).hexdigest()
                hashes[digest] += 1
                expected = (source.get("sha256") or "").strip().lower()
                if expected and digest != expected:
                    raise ValueError("SHA-256 diverso dal manifest")
                kind = _kind(source)
                parsed = parse_f24_evidence(content, document_kind=kind)
                rows = normalize_f24_evidence_rows(parsed)
                if not rows:
                    raise ValueError("nessuna riga fiscale estratta")
                general = parsed.get("dati_generali") or {}
                protocol = str(
                    general.get("protocollo_telematico")
                    or source.get("protocollo_telematico")
                    or ""
                ).strip()
                if protocol in ("", "-"):
                    missing_protocols += 1
                    missing_protocols_by_type[kind] += 1
                year = _year(source.get("data_versamento") or "")
                by_year[year or 0] += len(rows)
                by_type[kind] += 1
                totals["debits"] += sum(item["debit_amount"] for item in rows)
                totals["credits"] += sum(item["credit_amount"] for item in rows)
                credit_rows += sum(1 for item in rows if item["credit_amount"] > 0)
                validated.append({
                    "path": path,
                    "relative_path": relative.as_posix(),
                    "content": content,
                    "sha256": digest,
                    "document_kind": kind,
                    "manifest": source,
                    "parsed": parsed,
                    "rows": rows,
                    "protocol": protocol or None,
                })
            except Exception as exc:
                errors.append({"file": relative.as_posix(), "error": str(exc)})
    finally:
        close_archive()

    normalized_rows = sum(by_year.values())
    row_delta = (
        normalized_rows - declared_f24_rows
        if declared_f24_rows is not None else None
    )
    summary = {
        "mode": "dry-run",
        "archive_source": archive_source,
        "manifest_format": manifest_format,
        "manifest_rows": len(source_rows),
        "valid_documents": len(validated),
        "invalid_documents": len(errors),
        "normalized_rows": normalized_rows,
        "archive_index_normalized_rows": declared_f24_rows,
        "normalized_row_delta": row_delta,
        "archive_index_reconciliation": (
            "NOT_AVAILABLE"
            if declared_f24_rows is None
            else "MATCH" if row_delta == 0 else "REVIEW_REQUIRED_PDF_PARSER_DIFFERS"
        ),
        "credit_rows": credit_rows,
        "total_debits": round(totals["debits"], 2),
        "total_credits": round(totals["credits"], 2),
        "rows_by_payment_year": {str(key): value for key, value in sorted(by_year.items())},
        "documents_by_type": dict(sorted(by_type.items())),
        "archive_duplicate_documents": sum(count - 1 for count in hashes.values() if count > 1),
        "duplicate_sha256": sorted(digest for digest, count in hashes.items() if count > 1),
        "missing_protocols": missing_protocols,
        "missing_protocols_by_type": dict(sorted(missing_protocols_by_type.items())),
        "unexpected_missing_quietanza_protocols": missing_protocols_by_type.get(
            PARSER_KIND_QUIETANZA, 0
        ),
        "unbalanced_documents": sum(
            1 for item in errors if "non quadrato" in item["error"].casefold()
        ),
        "database_duplicate_check": "NOT_RUN_NO_DATABASE_CONNECTION",
        "errors": errors,
    }
    return validated, summary


async def run(args: argparse.Namespace) -> dict[str, Any]:
    root = Path(args.archive).resolve()
    validated, summary = validate_archive(root)
    if summary["invalid_documents"]:
        summary["write_status"] = "BLOCKED_ARCHIVE_VALIDATION"
        return summary
    if summary["archive_index_reconciliation"] == "REVIEW_REQUIRED_PDF_PARSER_DIFFERS":
        summary["write_status"] = "BLOCKED_INDEX_RECONCILIATION"
        return summary
    if getattr(args, "check_db", False):
        from app.db_collections import COLL_FISCAL_DOCUMENT_VERSIONS

        try:
            await Database.connect_db()
            db = Database.get_db()
            hashes = [item["sha256"] for item in validated]
            existing = await db[COLL_FISCAL_DOCUMENT_VERSIONS].find(
                {"company_id": args.company_id, "sha256": {"$in": hashes}},
                {"_id": 0, "sha256": 1},
            ).to_list(len(hashes))
            existing_hashes = {item.get("sha256") for item in existing}
            summary.update({
                "database_duplicate_check": "COMPLETED_READ_ONLY",
                "database_duplicate_documents": len(existing_hashes),
                "new_documents": len(validated) - len(existing_hashes),
            })
        except Exception as exc:
            summary["database_duplicate_check"] = f"FAILED_READ_ONLY: {exc}"
            summary["write_status"] = "BLOCKED_DATABASE_PREVIEW"
            return summary
        finally:
            await Database.close_db()
    if not args.esegui:
        summary["write_status"] = "NOT_REQUESTED"
        return summary

    await Database.connect_db()
    imported = duplicates = failed = 0
    failures: list[dict[str, str]] = []
    try:
        db = Database.get_db()
        for item in validated:
            try:
                result = await ingest_f24_evidence(
                    db,
                    content=item["content"],
                    filename=item["path"].name,
                    document_kind=item["document_kind"],
                    company_id=args.company_id,
                    source="archive_f24_manifest",
                    source_metadata={
                        **item["manifest"],
                        "archive_relative_path": item["relative_path"],
                    },
                    expected_sha256=item["sha256"],
                )
                duplicates += int(result["duplicate_document"])
                imported += int(not result["duplicate_document"])
            except Exception as exc:
                failed += 1
                failures.append({"file": item["relative_path"], "error": str(exc)})
    finally:
        await Database.close_db()
    summary.update({
        "mode": "write",
        "write_status": "COMPLETED" if not failed else "COMPLETED_WITH_ERRORS",
        "imported_documents": imported,
        "duplicate_documents": duplicates,
        "failed_documents": failed,
        "write_errors": failures,
    })
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archive", help=f"Cartella che contiene {MANIFEST_NAME}")
    parser.add_argument("--company-id", default=settings.FISCAL_COMPANY_ID)
    parser.add_argument(
        "--check-db", action="store_true",
        help="Confronta gli SHA-256 con il database in sola lettura",
    )
    parser.add_argument("--esegui", action="store_true", help="Abilita esplicitamente le scritture nel database")
    args = parser.parse_args()
    print(json.dumps(asyncio.run(run(args)), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
