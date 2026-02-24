"""
Excel Exporter.
Export data to Excel with professional formatting.
"""
from typing import List, Dict, Any
from io import BytesIO
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export data to Excel with formatting."""
    
    # Color scheme
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    def export_invoices(
        self,
        invoices: List[Dict[str, Any]],
        filename: str = None
    ) -> BytesIO:
        """
        Export invoices to Excel.
        
        Args:
            invoices: List of invoice dicts
            filename: Optional filename (metadata only)
            
        Returns:
            BytesIO with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Fatture"
        
        # Headers
        headers = [
            "Numero", "Data", "Fornitore", "Importo Netto",
            "IVA", "Totale", "Stato Pagamento", "Data Scadenza"
        ]
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        # Write data
        for row_num, invoice in enumerate(invoices, 2):
            data = [
                invoice.get('invoice_number', ''),
                invoice.get('date', ''),
                invoice.get('supplier_name', ''),
                invoice.get('total_amount', 0) - invoice.get('vat_amount', 0),
                invoice.get('vat_amount', 0),
                invoice.get('total_amount', 0),
                self._translate_payment_status(invoice.get('payment_status', '')),
                invoice.get('due_date', '')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.BORDER
                
                # Format currency
                if col_num in [4, 5, 6]:
                    cell.number_format = '€#,##0.00'
                
                # Align
                if col_num in [1, 2, 7, 8]:
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add totals row
        total_row = len(invoices) + 2
        ws.cell(total_row, 3, "TOTALE").font = Font(bold=True)
        
        total_net = sum(inv.get('total_amount', 0) - inv.get('vat_amount', 0) for inv in invoices)
        total_vat = sum(inv.get('vat_amount', 0) for inv in invoices)
        total_amount = sum(inv.get('total_amount', 0) for inv in invoices)
        
        ws.cell(total_row, 4, total_net).number_format = '€#,##0.00'
        ws.cell(total_row, 5, total_vat).number_format = '€#,##0.00'
        ws.cell(total_row, 6, total_amount).number_format = '€#,##0.00'
        
        for col in [3, 4, 5, 6]:
            ws.cell(total_row, col).font = Font(bold=True)
            ws.cell(total_row, col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(invoices)} invoices to Excel")
        
        return output
    
    def export_warehouse_inventory(
        self,
        products: List[Dict[str, Any]]
    ) -> BytesIO:
        """Export warehouse inventory to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        headers = [
            "Codice", "Nome", "Categoria", "Stock", "Unità",
            "Fornitore", "Costo Unitario", "Valore Totale", "Scorta Minima"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Write data
        for row_num, product in enumerate(products, 2):
            stock = product.get('stock', 0)
            unit_cost = product.get('unit_cost', 0)
            total_value = stock * unit_cost
            
            data = [
                product.get('code', ''),
                product.get('name', ''),
                product.get('category', ''),
                stock,
                product.get('unit', ''),
                product.get('supplier_name', ''),
                unit_cost,
                total_value,
                product.get('min_stock', 0)
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.BORDER
                
                # Format numbers
                if col_num == 4:
                    cell.number_format = '#,##0.00'
                elif col_num in [7, 8]:
                    cell.number_format = '€#,##0.00'
                
                # Highlight low stock
                min_stock = product.get('min_stock', 0)
                if col_num == 4 and stock < min_stock:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Totals
        total_row = len(products) + 2
        ws.cell(total_row, 7, "TOTALE").font = Font(bold=True)
        total_value = sum(p.get('stock', 0) * p.get('unit_cost', 0) for p in products)
        ws.cell(total_row, 8, total_value).number_format = '€#,##0.00'
        ws.cell(total_row, 8).font = Font(bold=True)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(products)} products to Excel")
        
        return output
    
    def export_employees(
        self,
        employees: List[Dict[str, Any]]
    ) -> BytesIO:
        """Export employees to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dipendenti"
        
        headers = [
            "Nome", "Cognome", "Codice Fiscale", "Ruolo",
            "Email", "Telefono", "Data Assunzione", "Attivo"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        for row_num, emp in enumerate(employees, 2):
            data = [
                emp.get('first_name', ''),
                emp.get('last_name', ''),
                emp.get('codice_fiscale', ''),
                emp.get('role', ''),
                emp.get('email', ''),
                emp.get('phone', ''),
                emp.get('hire_date', ''),
                "Sì" if emp.get('is_active', True) else "No"
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.BORDER
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(employees)} employees to Excel")
        
        return output
    
    def export_accounting_report(
        self,
        data: Dict[str, Any],
        month: str
    ) -> BytesIO:
        """Export monthly accounting report."""
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Riepilogo"
        
        # Title
        ws_summary.merge_cells('A1:D1')
        title_cell = ws_summary['A1']
        title_cell.value = f"Report Contabile - {month}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Summary data
        row = 3
        summary_items = [
            ("Fatture Totali", data.get('total_invoices', 0)),
            ("Imponibile", data.get('total_net', 0)),
            ("IVA", data.get('total_vat', 0)),
            ("Totale Fatturato", data.get('total_amount', 0)),
            ("Pagato", data.get('total_paid', 0)),
            ("Da Pagare", data.get('total_unpaid', 0))
        ]
        
        for label, value in summary_items:
            ws_summary.cell(row, 1, label).font = Font(bold=True)
            cell = ws_summary.cell(row, 2, value)
            if isinstance(value, (int, float)) and value != data.get('total_invoices', 0):
                cell.number_format = '€#,##0.00'
            row += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported accounting report for {month}")
        
        return output
    
    def _translate_payment_status(self, status: str) -> str:
        """Translate payment status to Italian."""
        translations = {
            'paid': 'Pagata',
            'unpaid': 'Da Pagare',
            'partial': 'Parziale',
            'overdue': 'Scaduta'
        }
        return translations.get(status, status)


# Singleton instance
excel_exporter = ExcelExporter() if OPENPYXL_AVAILABLE else None
