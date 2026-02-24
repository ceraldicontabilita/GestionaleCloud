"""
Balance Sheet Service
Business logic per bilanci e report contabili
"""
from typing import Dict, Any
from datetime import datetime, date
import logging
from io import BytesIO

from app.repositories.balance_sheet_repository import (
    BalanceSheetRepository,
    YearEndRepository,
    AccountBalanceRepository
)
from app.repositories.chart_repository import ChartOfAccountsRepository
from app.exceptions import ValidationError

logger = logging.getLogger(__name__)


class BalanceSheetService:
    """Service per bilanci e report contabili."""
    
    def __init__(
        self,
        balance_repo: BalanceSheetRepository,
        year_end_repo: YearEndRepository,
        account_balance_repo: AccountBalanceRepository,
        chart_repo: ChartOfAccountsRepository
    ):
        self.balance_repo = balance_repo
        self.year_end_repo = year_end_repo
        self.account_balance_repo = account_balance_repo
        self.chart_repo = chart_repo
    
    async def generate_balance_sheet(
        self,
        year: int,
        save: bool = True
    ) -> Dict[str, Any]:
        """
        Genera bilancio annuale (stato patrimoniale).
        """
        # Data di riferimento: 31 dicembre
        as_of_date = date(year, 12, 31)
        
        # Calcola totali per tipo
        assets_total = await self.account_balance_repo.get_total_by_type(
            'assets', as_of_date
        )
        liabilities_total = await self.account_balance_repo.get_total_by_type(
            'liabilities', as_of_date
        )
        equity_total = await self.account_balance_repo.get_total_by_type(
            'equity', as_of_date
        )
        
        # Recupera dettaglio conti
        assets = await self.account_balance_repo.get_accounts_by_type(
            'assets', as_of_date
        )
        liabilities = await self.account_balance_repo.get_accounts_by_type(
            'liabilities', as_of_date
        )
        equity = await self.account_balance_repo.get_accounts_by_type(
            'equity', as_of_date
        )
        
        balance_sheet = {
            'year': year,
            'date': as_of_date,
            'assets': {
                'total': assets_total,
                'accounts': assets
            },
            'liabilities': {
                'total': liabilities_total,
                'accounts': liabilities
            },
            'equity': {
                'total': equity_total,
                'accounts': equity
            },
            'balanced': abs(assets_total - (liabilities_total + equity_total)) < 0.01
        }
        
        if save:
            await self.balance_repo.save_balance_sheet(balance_sheet)
        
        return balance_sheet
    
    async def generate_trial_balance(
        self,
        balance_date: date
    ) -> Dict[str, Any]:
        """
        Genera bilancio di verifica.
        """
        # Calcola saldi di tutti i conti
        balances = await self.account_balance_repo.calculate_account_balances(
            balance_date
        )
        
        total_debit = sum(acc['debit'] for acc in balances)
        total_credit = sum(acc['credit'] for acc in balances)
        
        return {
            'date': balance_date,
            'accounts': balances,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'balanced': abs(total_debit - total_credit) < 0.01,
            'generated_at': datetime.now(timezone.utc)
        }
    
    async def generate_profit_loss(
        self,
        year: int
    ) -> Dict[str, Any]:
        """
        Genera conto economico (P&L).
        """
        as_of_date = date(year, 12, 31)
        
        # Ricavi
        revenue_total = await self.account_balance_repo.get_total_by_type(
            'revenue', as_of_date
        )
        revenue_accounts = await self.account_balance_repo.get_accounts_by_type(
            'revenue', as_of_date
        )
        
        # Costi
        expenses_total = await self.account_balance_repo.get_total_by_type(
            'expenses', as_of_date
        )
        expenses_accounts = await self.account_balance_repo.get_accounts_by_type(
            'expenses', as_of_date
        )
        
        # Calcola utile/perdita
        net_profit = revenue_total - expenses_total
        
        return {
            'year': year,
            'revenue': {
                'total': revenue_total,
                'accounts': revenue_accounts
            },
            'expenses': {
                'total': expenses_total,
                'accounts': expenses_accounts
            },
            'gross_profit': revenue_total - expenses_total,
            'net_profit': net_profit,
            'generated_at': datetime.now(timezone.utc)
        }
    
    async def generate_cash_flow(
        self,
        year: int
    ) -> Dict[str, Any]:
        """
        Genera rendiconto finanziario.
        """
        # Semplificato: calcola variazione liquidità
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        # Recupera conti di cassa/banca
        # Assumiamo che i conti con 'cassa' o 'banca' nel nome siano liquidità
        all_accounts = await self.chart_repo.get_all()
        cash_accounts = [
            acc for acc in all_accounts
            if 'cassa' in acc.get('name', '').lower() or
               'banca' in acc.get('name', '').lower() or
               'cash' in acc.get('name', '').lower() or
               'bank' in acc.get('name', '').lower()
        ]
        
        if not cash_accounts:
            return {
                'year': year,
                'opening_cash': 0.0,
                'closing_cash': 0.0,
                'net_cash_flow': 0.0,
                'note': 'No cash/bank accounts found'
            }
        
        cash_account_ids = [str(acc['_id']) for acc in cash_accounts]
        
        # Saldo iniziale (fine anno precedente)
        previous_year_end = date(year - 1, 12, 31)
        opening_balances = await self.account_balance_repo.calculate_account_balances(
            previous_year_end,
            cash_account_ids
        )
        opening_cash = sum(acc['balance'] for acc in opening_balances)
        
        # Saldo finale
        closing_balances = await self.account_balance_repo.calculate_account_balances(
            end_date,
            cash_account_ids
        )
        closing_cash = sum(acc['balance'] for acc in closing_balances)
        
        net_cash_flow = closing_cash - opening_cash
        
        return {
            'year': year,
            'opening_cash': opening_cash,
            'closing_cash': closing_cash,
            'net_cash_flow': net_cash_flow,
            'cash_accounts': len(cash_accounts),
            'generated_at': datetime.now(timezone.utc)
        }
    
    async def get_assets(self, year: int) -> Dict[str, Any]:
        """Recupera stato patrimoniale attivo."""
        as_of_date = date(year, 12, 31)
        
        assets = await self.account_balance_repo.get_accounts_by_type(
            'assets', as_of_date
        )
        total = sum(acc['balance'] for acc in assets)
        
        return {
            'year': year,
            'date': as_of_date,
            'total': total,
            'accounts': assets
        }
    
    async def get_liabilities(self, year: int) -> Dict[str, Any]:
        """Recupera stato patrimoniale passivo."""
        as_of_date = date(year, 12, 31)
        
        liabilities = await self.account_balance_repo.get_accounts_by_type(
            'liabilities', as_of_date
        )
        total = sum(acc['balance'] for acc in liabilities)
        
        return {
            'year': year,
            'date': as_of_date,
            'total': total,
            'accounts': liabilities
        }
    
    async def get_equity(self, year: int) -> Dict[str, Any]:
        """Recupera patrimonio netto."""
        as_of_date = date(year, 12, 31)
        
        equity = await self.account_balance_repo.get_accounts_by_type(
            'equity', as_of_date
        )
        total = sum(acc['balance'] for acc in equity)
        
        return {
            'year': year,
            'date': as_of_date,
            'total': total,
            'accounts': equity
        }
    
    async def close_year(
        self,
        year: int,
        user_id: str
    ) -> Dict[str, Any]:
        """
        Chiusura anno contabile.
        """
        # Verifica che non sia già chiuso
        existing = await self.year_end_repo.get_closure_by_year(year)
        if existing:
            raise ValidationError(f"Year {year} already closed")
        
        # Genera tutti i report
        balance_sheet = await self.generate_balance_sheet(year, save=True)
        profit_loss = await self.generate_profit_loss(year)
        cash_flow = await self.generate_cash_flow(year)
        
        # Salva chiusura
        closure_data = {
            'year': year,
            'closed_by': user_id,
            'closed_at': datetime.now(timezone.utc),
            'balance_sheet': balance_sheet,
            'profit_loss': profit_loss,
            'cash_flow': cash_flow
        }
        
        closure = await self.year_end_repo.create_year_end_closure(closure_data)
        
        logger.info(f"Year {year} closed by user {user_id}")
        
        return closure
    
    async def export_balance_excel(
        self,
        year: int
    ) -> BytesIO:
        """Export bilanci in Excel."""
        balance = await self.generate_balance_sheet(year, save=False)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            
            # Sheet bilancio
            ws = wb.active
            ws.title = f"Bilancio {year}"
            
            ws.append(['STATO PATRIMONIALE', year])
            ws.append([])
            ws.append(['ATTIVO', '', 'PASSIVO', ''])
            
            # Formatta
            for cell in ws[1]:
                cell.font = Font(bold=True, size=14)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        
        except ImportError:
            raise ValidationError("openpyxl non installato")
