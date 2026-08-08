"""Export non distruttivo dei dati paghe verso AppDipendenti."""
from __future__ import annotations

import base64
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


CURRENT_HEADERS = ["Dipendente", "Mese", "Anno", "Stipendio Netto", "Importo Erogato"]
HISTORICAL_HEADERS = ["Data bonifico", "Nome dipendente", "Importo di busta", "Importo effettivamente pagato"]


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0.00")


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            continue
    return None


def _period_in_range(year: Any, month: Any, date_from: date, date_to: date) -> bool:
    try:
        period = date(int(year), int(month), 1)
    except (TypeError, ValueError):
        return False
    return date_from.replace(day=1) <= period <= date_to.replace(day=1)


def _employee_name(record: dict[str, Any], cedolino: dict[str, Any] | None = None) -> str:
    return str(
        record.get("dipendente_nome") or record.get("dipendente") or
        record.get("nome_dipendente") or (cedolino or {}).get("nome_dipendente") or
        "Dipendente non rilevato"
    ).strip()


def _employee_parts(name: str) -> tuple[str, str]:
    cleaned = " ".join(name.upper().split())
    if "," in cleaned:
        surname, given = (part.strip() for part in cleaned.split(",", 1))
        return surname or "DIPENDENTE", given
    parts = cleaned.split()
    return (parts[0] if parts else "DIPENDENTE", " ".join(parts[1:]))


def _safe_component(value: str) -> str:
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1f]+", " ", value)
    return "_".join(value.split()).strip("._") or "NON_RILEVATO"


def _safe_sheet(value: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]+", " ", value).strip(" '") or "Dipendente"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f" ({index})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def _style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="24506F")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 34)


def _current_workbook(rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prima Nota"
    sheet.append(CURRENT_HEADERS)
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'
    _style_sheet(sheet)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _historical_workbook(rows_by_employee: dict[str, list[list[Any]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()
    if not rows_by_employee:
        sheet = workbook.create_sheet("Nessun pagamento")
        sheet.append(HISTORICAL_HEADERS)
        _style_sheet(sheet)
    for surname, rows in sorted(rows_by_employee.items(), key=lambda item: item[0].casefold()):
        sheet = workbook.create_sheet(_safe_sheet(surname, used))
        sheet.append(HISTORICAL_HEADERS)
        for row in rows:
            sheet.append(row)
        for row in sheet.iter_rows(min_row=2):
            row[0].number_format = "dd/mm/yyyy"
            row[2].number_format = '#,##0.00'
            row[3].number_format = '#,##0.00'
        _style_sheet(sheet)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_filename(year: int, month: int, name: str, multi: bool = False) -> str:
    if multi:
        return f"{year:04d}-{month:02d}_LIBRO_UNICO.pdf"
    surname, given = _employee_parts(name)
    return f"{year:04d}-{month:02d}_{_safe_component(surname)}_{_safe_component(given)}.pdf"


async def build_export_from_db(db, date_from: date, date_to: date) -> tuple[bytes, dict[str, Any]]:
    if date_from > date_to:
        raise ValueError("La data iniziale non puo superare la data finale")

    records = await db["prima_nota_salari"].find({}, {"_id": 0}).sort(
        [("anno", 1), ("mese", 1), ("dipendente", 1)]
    ).to_list(100000)
    cedolini = await db["cedolini"].find(
        {"pdf_data": {"$exists": True, "$nin": [None, ""]}}, {"_id": 0}
    ).to_list(100000)
    cedolini_by_id = {str(item.get("id")): item for item in cedolini if item.get("id")}
    cedolini_by_period = {}
    for item in cedolini:
        key = (str(item.get("codice_fiscale") or "").upper(), item.get("anno"), item.get("mese"))
        if key[0]:
            cedolini_by_period[key] = item

    current_rows: list[list[Any]] = []
    historical_by_employee: dict[str, list[list[Any]]] = {}
    employees: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    used_pdf: set[str] = set()
    pdf_entries: list[tuple[str, bytes]] = []

    for record in records:
        year = record.get("anno")
        month = record.get("mese")
        if not _period_in_range(year, month, date_from, date_to):
            continue
        cedolino = cedolini_by_id.get(str(record.get("cedolino_id")))
        if not cedolino and record.get("codice_fiscale"):
            cedolino = cedolini_by_period.get((str(record["codice_fiscale"]).upper(), year, month))
        name = _employee_name(record, cedolino)
        net = _decimal(record.get("importo_busta"))
        paid_actual = _decimal(record.get("importo_bonifico"))
        paid_documented = _decimal(record.get("importo_bonifico_documentato"))
        paid = paid_actual if paid_actual > 0 else paid_documented
        if paid <= 0:
            continue
        current_rows.append([name, int(month), int(year), float(net), float(paid)])
        identity = str(record.get("codice_fiscale") or name).upper()
        employee = employees.setdefault(identity, {"name": name, "total": Decimal("0.00")})
        employee["total"] += paid
        payment_date = _parse_date(record.get("data_bonifico_documentata") or record.get("data_bonifico"))
        if payment_date is None:
            warnings.append(f"Pagamento di {name} {month:02d}/{year} senza data di bonifico: escluso dallo storico")
        else:
            surname, _ = _employee_parts(name)
            historical_by_employee.setdefault(surname, []).append([payment_date, name, float(net), float(paid)])
        if paid_actual > 0 and paid_documented > 0 and paid_actual != paid_documented:
            warnings.append(f"Importi documentato e bancario diversi per {name} {month:02d}/{year}: usato il bancario")

    for cedolino in cedolini:
        try:
            year = int(cedolino.get("anno"))
            month = int(cedolino.get("mese"))
        except (TypeError, ValueError):
            continue
        if not _period_in_range(year, month, date_from, date_to):
            continue
        identifier = str(cedolino.get("id") or cedolino.get("file_hash") or cedolino.get("filename") or "")
        if not identifier or identifier in used_pdf:
            continue
        try:
            content = base64.b64decode(cedolino.get("pdf_data") or "", validate=True)
        except Exception:
            warnings.append(f"PDF cedolino non decodificabile: {cedolino.get('filename') or identifier}")
            continue
        if not content.startswith(b"%PDF"):
            warnings.append(f"PDF cedolino non valido: {cedolino.get('filename') or identifier}")
            continue
        used_pdf.add(identifier)
        names = cedolino.get("employee_names") or cedolino.get("dipendenti") or []
        if isinstance(names, str):
            names = [names]
        name = _employee_name({}, cedolino)
        pdf_entries.append((_pdf_filename(year, month, name, multi=len(names) > 1), content))

    current_xlsx = _current_workbook(sorted(current_rows, key=lambda row: (row[2], row[1], str(row[0]).casefold())))
    historical_xlsx = _historical_workbook(historical_by_employee)
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("export_cedolini/", b"")
        used_names: dict[str, int] = {}
        for filename, content in pdf_entries:
            stem = filename[:-4]
            count = used_names.get(stem, 0) + 1
            used_names[stem] = count
            actual = filename if count == 1 else f"{stem}_{count}.pdf"
            archive.writestr(f"export_cedolini/{actual}", content)
        archive.writestr("prima_nota_salari.xlsx", current_xlsx)
        archive.writestr("storico_pagamenti.xlsx", historical_xlsx)
    summary = {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "pdf_count": len(pdf_entries),
        "prima_nota_rows": len(current_rows),
        "storico_rows": sum(len(rows) for rows in historical_by_employee.values()),
        "employees": [
            {"employee_name": value["name"], "total_paid": format(value["total"], ".2f")}
            for value in sorted(employees.values(), key=lambda item: str(item["name"]).casefold())
        ],
        "warnings": sorted(set(warnings)),
    }
    return output.getvalue(), summary
