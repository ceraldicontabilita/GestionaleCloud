"""
VAT and F24 Service
Business logic per liquidazioni IVA e modelli F24
"""
from typing import List, Dict, Any, Optional
from datetime import date
import logging
from io import BytesIO
from calendar import monthrange

from app.repositories.vat_f24_repository import VATRepository, VATRegistryRepository, F24Repository
from app.repositories.invoice_repository import InvoiceRepository
from app.exceptions import NotFoundError, ValidationError

logger = logging.getLogger(__name__)


class VATService:
    """Service per liquidazioni IVA."""
    
    def __init__(
        self,
        vat_repo: VATRepository,
        registry_repo: VATRegistryRepository,
        invoice_repo: InvoiceRepository
    ):
        self.vat_repo = vat_repo
        self.registry_repo = registry_repo
        self.invoice_repo = invoice_repo
    
    async def calculate_liquidation(
        self,
        quarter: int,
        year: int
    ) -> Dict[str, Any]:
        """
        Calcola liquidazione IVA per trimestre.
        Analizza fatture passive del trimestre.
        """
        if quarter not in [1, 2, 3, 4]:
            raise ValidationError("Quarter must be 1-4")
        
        # Calcola date trimestre
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        
        start_date = date(year, start_month, 1)
        end_date = date(year, end_month, monthrange(year, end_month)[1])
        
        # Recupera fatture del trimestre
        invoices = await self.invoice_repo.find_many({
            'invoice_date': {
                '$gte': start_date,
                '$lte': end_date
            }
        })
        
        # Calcola IVA detraibile e a debito
        vat_deductible = 0.0
        vat_payable = 0.0
        
        for invoice in invoices:
            vat_amount = invoice.get('total_vat', 0)
            
            # Determina se detraibile (acquisti) o a debito (vendite)
            # Per ora assumiamo tutte passive (detraibili)
            vat_deductible += vat_amount
        
        # Recupera eventuale credito precedente
        previous_quarter = quarter - 1 if quarter > 1 else 4
        previous_year = year if quarter > 1 else year - 1
        
        previous_credit = 0.0
        previous_liq = await self.vat_repo.get_liquidation_by_period(
            previous_quarter,
            previous_year
        )
        
        if previous_liq and previous_liq.get('vat_balance', 0) < 0:
            previous_credit = abs(previous_liq['vat_balance'])
        
        # Calcola saldo
        vat_balance = vat_payable - vat_deductible
        to_pay = max(0, vat_balance - previous_credit)
        
        return {
            'quarter': quarter,
            'year': year,
            'period': f"Q{quarter}/{year}",
            'start_date': start_date,
            'end_date': end_date,
            'vat_deductible': vat_deductible,
            'vat_payable': vat_payable,
            'vat_balance': vat_balance,
            'previous_credit': previous_credit,
            'to_pay': to_pay,
            'invoices_count': len(invoices)
        }
    
    async def create_liquidation(
        self,
        quarter: int,
        year: int,
        user_id: str
    ) -> Dict[str, Any]:
        """Crea e salva liquidazione IVA."""
        # Verifica se esiste già
        existing = await self.vat_repo.get_liquidation_by_period(quarter, year)
        if existing:
            raise ValidationError(f"Liquidation Q{quarter}/{year} already exists")
        
        # Calcola
        calculation = await self.calculate_liquidation(quarter, year)
        
        # Salva
        liquidation_data = {
            **calculation,
            'paid': False,
            'payment_date': None,
            'payment_reference': None,
            'created_by': user_id,
            'notes': None
        }
        
        return await self.vat_repo.create_liquidation(liquidation_data)
    
    async def get_liquidation(
        self,
        quarter: int,
        year: int
    ) -> Dict[str, Any]:
        """Recupera liquidazione esistente."""
        liquidation = await self.vat_repo.get_liquidation_by_period(quarter, year)
        if not liquidation:
            raise NotFoundError(f"Liquidation Q{quarter}/{year} not found")
        return liquidation
    
    async def register_payment(
        self,
        liquidation_id: str,
        payment_date: date,
        payment_reference: Optional[str] = None
    ) -> Dict[str, Any]:
        """Registra pagamento IVA."""
        return await self.vat_repo.mark_as_paid(
            liquidation_id,
            payment_date,
            payment_reference
        )
    
    async def get_vat_registry(
        self,
        start_date: date,
        end_date: date,
        vat_type: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Recupera registro IVA."""
        return await self.registry_repo.get_registry_entries(
            start_date,
            end_date,
            vat_type
        )
    
    async def get_annual_report(self, year: int) -> Dict[str, Any]:
        """Genera report IVA annuale."""
        liquidations = await self.vat_repo.get_liquidations_by_year(year)
        
        total_deductible = sum(liq.get('vat_deductible', 0) for liq in liquidations)
        total_payable = sum(liq.get('vat_payable', 0) for liq in liquidations)
        total_paid = sum(
            liq.get('to_pay', 0) for liq in liquidations if liq.get('paid')
        )
        total_unpaid = sum(
            liq.get('to_pay', 0) for liq in liquidations if not liq.get('paid')
        )
        
        return {
            'year': year,
            'liquidations': liquidations,
            'summary': {
                'total_deductible': total_deductible,
                'total_payable': total_payable,
                'balance': total_payable - total_deductible,
                'total_paid': total_paid,
                'total_unpaid': total_unpaid
            }
        }
    
    async def export_vat_excel(
        self,
        quarter: int,
        year: int
    ) -> BytesIO:
        """Export liquidazione IVA in Excel."""
        liquidation = await self.get_liquidation(quarter, year)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"IVA Q{quarter}-{year}"
            
            # Header
            ws.append(['Liquidazione IVA', f'Q{quarter}/{year}'])
            ws.append([])
            ws.append(['Descrizione', 'Importo €'])
            
            # Dati
            ws.append(['IVA Detraibile', liquidation['vat_deductible']])
            ws.append(['IVA a Debito', liquidation['vat_payable']])
            ws.append(['Saldo IVA', liquidation['vat_balance']])
            ws.append(['Credito Precedente', liquidation['previous_credit']])
            ws.append(['Da Versare', liquidation['to_pay']])
            
            # Stile
            for cell in ws[1]:
                cell.font = Font(bold=True, size=14)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        
        except ImportError:
            raise ValidationError("openpyxl non installato")


class F24Service:
    """Service per modelli F24."""
    
    def __init__(self, f24_repo: F24Repository):
        self.f24_repo = f24_repo
    
    async def create_f24(
        self,
        f24_data: Dict[str, Any],
        user_id: str
    ) -> Dict[str, Any]:
        """Crea modello F24."""
        f24_data['created_by'] = user_id
        return await self.f24_repo.create_f24(f24_data)
    
    async def get_f24(self, f24_id: str) -> Dict[str, Any]:
        """Recupera F24 per ID."""
        f24 = await self.f24_repo.get_by_id(f24_id)
        if not f24:
            raise NotFoundError(f"F24 {f24_id} not found")
        return f24
    
    async def list_f24s(
        self,
        month: Optional[int] = None,
        year: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """Lista F24."""
        if month and year:
            return await self.f24_repo.get_f24_by_period(month, year)
        elif year:
            return await self.f24_repo.get_f24_by_year(year)
        else:
            # Ultimo anno
            current_year = date.today().year
            return await self.f24_repo.get_f24_by_year(current_year)
    
    async def update_f24(
        self,
        f24_id: str,
        update_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Aggiorna F24."""
        return await self.f24_repo.update_f24(f24_id, update_data)
    
    async def mark_as_paid(
        self,
        f24_id: str,
        payment_reference: str
    ) -> Dict[str, Any]:
        """Marca F24 come pagato."""
        return await self.f24_repo.mark_f24_as_paid(f24_id, payment_reference)
    
    async def generate_pdf(self, f24_id: str) -> BytesIO:
        """Genera PDF F24."""
        f24 = await self.get_f24(f24_id)
        
        # TODO: Implementare generazione PDF reale con reportlab
        raise NotImplementedError("PDF generation not implemented yet")
