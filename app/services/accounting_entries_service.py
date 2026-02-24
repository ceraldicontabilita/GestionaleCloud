"""
Accounting Entries Service
Business logic per gestione Prima Nota
"""
from typing import List, Dict, Any, Optional
from datetime import datetime, date
import logging
from io import BytesIO

from app.repositories.accounting_entries_repository import AccountingEntriesRepository
from app.repositories.chart_repository import ChartOfAccountsRepository
from app.exceptions import NotFoundError, ValidationError

logger = logging.getLogger(__name__)


class AccountingEntriesService:
    """Service per registrazioni contabili."""
    
    def __init__(
        self,
        entries_repo: AccountingEntriesRepository,
        chart_repo: ChartOfAccountsRepository
    ):
        self.entries_repo = entries_repo
        self.chart_repo = chart_repo
    
    async def create_entry(
        self,
        entry_data: Dict[str, Any],
        user_id: str
    ) -> Dict[str, Any]:
        """Crea nuova registrazione in prima nota."""
        # Valida che tutti i conti esistano
        for line in entry_data.get('lines', []):
            account_id = line.get('account_id')
            account_code = line.get('account_code')
            
            account = None
            
            # Se c'è ID, usalo
            if account_id:
                account = await self.chart_repo.get_by_id(account_id)
            # Se no, prova con codice
            elif account_code:
                account = await self.chart_repo.find_by_code(account_code, user_id)
                if account:
                    line['account_id'] = str(account['_id'])
            
            if not account:
                # Se non trovato, cerca conto generico o crea errore
                # Per ora errore
                raise NotFoundError(f"Account {account_id or account_code} not found")
            
            # Arricchisci con info conto se mancanti
            if not line.get('account_code'):
                line['account_code'] = account.get('code')
            if not line.get('account_name'):
                line['account_name'] = account.get('name')
        
        # Aggiungi user_id
        entry_data['created_by'] = user_id
        
        return await self.entries_repo.create_entry(entry_data)
    
    async def get_entry(self, entry_id: str) -> Dict[str, Any]:
        """Recupera registrazione per ID."""
        entry = await self.entries_repo.get_by_id(entry_id)
        if not entry:
            raise NotFoundError(f"Accounting entry {entry_id} not found")
        return entry
    
    async def list_entries(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        entry_type: Optional[str] = None,
        skip: int = 0,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """Lista registrazioni con filtri."""
        if not start_date:
            # Default: ultimo anno
            end_date = date.today()
            start_date = date(end_date.year - 1, 1, 1)
        elif not end_date:
            end_date = date.today()
        
        return await self.entries_repo.get_entries_by_date_range(
            start_date=start_date,
            end_date=end_date,
            entry_type=entry_type,
            skip=skip,
            limit=limit
        )
    
    async def update_entry(
        self,
        entry_id: str,
        update_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Aggiorna registrazione."""
        # Valida che la registrazione esista
        existing = await self.entries_repo.get_by_id(entry_id)
        if not existing:
            raise NotFoundError(f"Accounting entry {entry_id} not found")
        
        # Se ci sono nuove righe, valida conti
        if 'lines' in update_data:
            for line in update_data['lines']:
                account_id = line.get('account_id')
                account = await self.chart_repo.get_by_id(account_id)
                if not account:
                    raise NotFoundError(f"Account {account_id} not found")
                
                # Arricchisci con info conto
                line['account_code'] = account.get('code')
                line['account_name'] = account.get('name')
        
        return await self.entries_repo.update_entry(entry_id, update_data)
    
    async def delete_entry(self, entry_id: str) -> bool:
        """Elimina registrazione."""
        return await self.entries_repo.delete_entry(entry_id)
    
    async def get_entries_by_account(
        self,
        account_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> List[Dict[str, Any]]:
        """Recupera tutte le registrazioni di un conto."""
        # Valida che il conto esista
        account = await self.chart_repo.get_by_id(account_id)
        if not account:
            raise NotFoundError(f"Account {account_id} not found")
        
        return await self.entries_repo.get_entries_by_account(
            account_id=account_id,
            start_date=start_date,
            end_date=end_date
        )
    
    async def bulk_import_entries(
        self,
        entries: List[Dict[str, Any]],
        user_id: str
    ) -> Dict[str, Any]:
        """Import bulk di registrazioni."""
        # Valida tutti i conti
        validated_entries = []
        errors = []
        
        for i, entry in enumerate(entries):
            try:
                # Valida conti
                for line in entry.get('lines', []):
                    account_id = line.get('account_id')
                    account = await self.chart_repo.get_by_id(account_id)
                    if not account:
                        raise NotFoundError(f"Account {account_id} not found")
                    
                    line['account_code'] = account.get('code')
                    line['account_name'] = account.get('name')
                
                entry['created_by'] = user_id
                validated_entries.append(entry)
            
            except Exception as e:
                errors.append({
                    'row': i + 1,
                    'error': str(e),
                    'entry': entry.get('description', 'Unknown')
                })
        
        # Importa solo le registrazioni valide
        imported = []
        if validated_entries:
            imported = await self.entries_repo.bulk_create_entries(validated_entries)
        
        return {
            'total': len(entries),
            'imported': len(imported),
            'errors': len(errors),
            'entries': imported,
            'error_details': errors
        }
    
    async def export_entries_excel(
        self,
        start_date: date,
        end_date: date,
        entry_type: Optional[str] = None
    ) -> BytesIO:
        """Export registrazioni in Excel."""
        entries = await self.entries_repo.get_entries_by_date_range(
            start_date=start_date,
            end_date=end_date,
            entry_type=entry_type,
            limit=10000
        )
        
        # Prepara dati per Excel
        rows = []
        for entry in entries:
            for line in entry.get('lines', []):
                rows.append({
                    'Data': entry.get('date'),
                    'Tipo': entry.get('entry_type'),
                    'Descrizione': entry.get('description'),
                    'N. Documento': entry.get('document_number', ''),
                    'Conto': line.get('account_code'),
                    'Nome Conto': line.get('account_name'),
                    'Dare': line.get('debit', 0),
                    'Avere': line.get('credit', 0),
                    'Note': entry.get('notes', '')
                })
        
        # Crea Excel (import openpyxl qui per non averlo come dipendenza globale)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Prima Nota"
            
            # Header
            headers = list(rows[0].keys()) if rows else []
            ws.append(headers)
            
            # Stile header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Dati
            for row in rows:
                ws.append(list(row.values()))
            
            # Salva in BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        
        except ImportError:
            raise ValidationError("openpyxl non installato. Impossibile creare Excel.")
    
    async def export_entries_pdf(
        self,
        start_date: date,
        end_date: date
    ) -> BytesIO:
        """Export registrazioni in PDF."""
        entries = await self.entries_repo.get_entries_by_date_range(
            start_date=start_date,
            end_date=end_date,
            limit=10000
        )
        
        # TODO: Implementare generazione PDF con reportlab
        # Per ora ritorna None, da implementare quando serve
        raise NotImplementedError("PDF export non ancora implementato")
    
    async def get_account_balance(
        self,
        account_id: str,
        up_to_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """Calcola saldo di un conto."""
        # Valida conto
        account = await self.chart_repo.get_by_id(account_id)
        if not account:
            raise NotFoundError(f"Account {account_id} not found")
        
        balance = await self.entries_repo.get_account_balance(
            account_id=account_id,
            up_to_date=up_to_date or date.today()
        )
        
        # Arricchisci con info conto
        balance['account_code'] = account.get('code')
        balance['account_name'] = account.get('name')
        
        return balance
    
    async def generate_trial_balance(
        self,
        balance_date: date
    ) -> Dict[str, Any]:
        """Genera bilancio di verifica."""
        accounts = await self.entries_repo.get_trial_balance(balance_date)
        
        # Calcola totali
        total_debit = sum(acc['debit'] for acc in accounts)
        total_credit = sum(acc['credit'] for acc in accounts)
        
        return {
            'date': balance_date,
            'accounts': accounts,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'balanced': abs(total_debit - total_credit) < 0.01,
            'generated_at': datetime.now(timezone.utc)
        }
