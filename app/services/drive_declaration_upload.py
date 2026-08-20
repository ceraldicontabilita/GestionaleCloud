"""Acquisizione dichiarazioni con Google Drive come unico archivio canonico."""
from __future__ import annotations
import hashlib, io, re
from pathlib import PurePath
from typing import Any
from openpyxl import load_workbook
from app.services.fiscal_document_ingestion import CATEGORY_DOCUMENT_TYPES, extract_pdf_pages
from app.services.fiscal_domain import classify_document
from app.services import drive_document_index as index

FOLDER_MIME = "application/vnd.google-apps.folder"
TYPE_FOLDERS = {"MODELLO_770":"770","DICHIARAZIONE_IVA":"IVA","LIPE":"LIPE","REDDITI_SC":"Redditi_SC","DICHIARAZIONE_IRAP":"IRAP","ELENCO_PERCIPIENTI":"Percipienti"}

def _safe_filename(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9À-ÿ._() -]+", "_", PurePath(value or "dichiarazione.pdf").name).strip(" .")
    if not name.lower().endswith(".pdf"): raise ValueError("Sono ammessi solo PDF")
    return name or "dichiarazione.pdf"

def _folder(service, parent_id: str, name: str) -> str:
    matches=[x for x in index._drive._list_children(service,parent_id) if index._norm(x.get("name"))==index._norm(name) and x.get("mimeType")==FOLDER_MIME]
    if len(matches)>1: raise ValueError(f"Cartella Drive ambigua: {name}")
    if matches: return matches[0]["id"]
    return service.files().create(body={"name":name,"mimeType":FOLDER_MIME,"parents":[parent_id]},fields="id",supportsAllDrives=True).execute()["id"]

def _classification(content: bytes, filename: str, category: str) -> str:
    if category != "automatica":
        result=CATEGORY_DOCUMENT_TYPES.get(category)
        if not result: raise ValueError("Categoria dichiarazione non valida")
        return result
    pages=extract_pdf_pages(content)
    result=classify_document(filename,"\n".join(p["text"] for p in pages)).get("document_type")
    if result not in TYPE_FOLDERS: raise ValueError("Classificazione automatica incerta: scegliere manualmente il tipo")
    return str(result)

def upload_declaration(*,content:bytes,filename:str,category:str,filing_year:int,note:str|None=None,service=None)->dict[str,Any]:
    if not content.startswith(b"%PDF"): raise ValueError("PDF non valido")
    if not 2000<=int(filing_year)<=2100: raise ValueError("Anno dichiarazione non valido")
    filename=_safe_filename(filename); digest=hashlib.sha256(content).hexdigest(); service=service or index.build_drive_service()
    source,catalog=index.load_full_catalog(service)
    duplicates=[r for r in catalog["documents"] if index._norm(r.get("SHA-256"))==digest]
    if duplicates:
        row=duplicates[0]; return {"success":True,"duplicate":True,"document_id":row.get("ID documento"),"sha256":digest,"drive_path":row.get("Percorso Drive")}
    document_type=_classification(content,filename,category)
    folders=["01_DICHIARAZIONI_FISCALI",TYPE_FOLDERS[document_type],str(filing_year)]; parent=source["root_id"]
    for name in folders: parent=_folder(service,parent,name)
    from googleapiclient.http import MediaIoBaseUpload
    created=service.files().create(body={"name":filename,"parents":[parent],"description":note or ""},media_body=MediaIoBaseUpload(io.BytesIO(content),mimetype="application/pdf",resumable=False),fields="id,name,size,webViewLink",supportsAllDrives=True).execute()
    wb=load_workbook(io.BytesIO(index._download_index_sync(service,source["index"]["id"])))
    document_id=f"DOC-{digest[:24].upper()}"; drive_path="/".join([*folders,filename]); ws=wb[index.INDEX_SHEET_NAME]; headers=[str(c.value or "") for c in ws[1]]
    values={"ID documento":document_id,"Dominio":"FISCALE","Categoria":document_type,"Anno":filing_year,"Nome file":filename,"Estensione":".pdf","Dimensione byte":len(content),"SHA-256":digest,"Percorso Drive":drive_path,"Cartella Drive":"/".join(folders),"ZIP origine":"UPLOAD_GESTIONALE","Percorso nel pacchetto":filename,"Stato":"ATTIVO","Numero documento":""}
    ws.append([values.get(h,"") for h in headers]); wb["DICHIARAZIONI"].append([filing_year,document_type,"",drive_path])
    out=io.BytesIO(); wb.save(out); out.seek(0)
    service.files().update(fileId=source["index"]["id"],media_body=MediaIoBaseUpload(out,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",resumable=False),fields="id,modifiedTime",supportsAllDrives=True).execute()
    index._CACHE_KEY=None; index._CACHE_CATALOG=None
    verified=service.files().get(fileId=created["id"],fields="id,name,size,webViewLink,trashed",supportsAllDrives=True).execute()
    return {"success":True,"duplicate":False,"document_id":document_id,"document_type":document_type,"filing_year":filing_year,"sha256":digest,"drive_path":drive_path,"drive_file_id":verified["id"],"drive_url":verified.get("webViewLink")}
