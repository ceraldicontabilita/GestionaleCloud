"""Catalogo documentale Drive in sola lettura.

L'Excel e' il ponte tra Drive e il gestionale. Questo servizio non accede a
Drive/Sheets e non scarica i documenti indicizzati: scarica soltanto l'indice e,
quando richiesto, risolve il percorso fino al link Drive del file originale.
"""

from __future__ import annotations

import io
import hashlib
import posixpath
import re
import threading
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook

from app.config import settings
from app.services import drive_cedolini_ingest as _drive

FOLDER_MIME = "application/vnd.google-apps.folder"
INDEX_FOLDER_NAME = "INDICI GESTIONALE"
INDEX_FILE_NAME = "INDICE_DOCUMENTALE_DRIVE.xlsx"
INDEX_SHEET_NAME = "DOCUMENTI"
REQUIRED_HEADERS = (
    "ID documento", "Dominio", "Categoria", "Anno", "Nome file",
    "Estensione", "Dimensione byte", "SHA-256", "Percorso Drive",
    "Cartella Drive", "ZIP origine", "Percorso nel pacchetto", "Stato",
    "Numero documento",
)
F24_HEADERS = (
    "ID documento", "Anno pagamento", "Data pagamento", "Sezione",
    "Tipo riga", "Codice tributo", "Descrizione", "Periodo tributo",
    "Ente", "Debito", "Credito", "Protocollo", "Tipo documento",
    "SHA-256", "Percorso Drive", "Pagina", "Testo sorgente", "Fonte",
)
DECLARATION_HEADERS = ("Anno", "Tipo", "Protocollo", "Percorso archivio")
DUPLICATE_HEADERS = (
    "ZIP origine", "Percorso nel pacchetto", "Nome", "Estensione",
    "Dimensione byte", "SHA-256", "Esito", "Percorso Drive collegato",
)
_CACHE_LOCK = threading.Lock()
_CACHE_KEY: tuple[str, str | None] | None = None
_CACHE_CATALOG: dict[str, list[dict[str, Any]]] | None = None


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def build_drive_service():
    creds, error = _drive._load_credentials_cedolini()
    if creds is None:
        raise RuntimeError(f"Credenziali Google Drive non disponibili: {error}")
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _unique_named(items: Iterable[dict[str, Any]], name: str, *, folder: bool) -> dict[str, Any]:
    matches = [
        item for item in items
        if _norm(item.get("name")) == _norm(name)
        and (item.get("mimeType") == FOLDER_MIME) is folder
    ]
    if len(matches) != 1:
        raise ValueError(f"Elemento Drive assente o ambiguo: {name} ({len(matches)} corrispondenze)")
    return matches[0]


def _discover_index_file_sync(service, root_id: str) -> tuple[dict[str, Any], dict[str, Any]]:
    root = service.files().get(
        fileId=root_id,
        fields="id,name,mimeType,trashed",
        supportsAllDrives=True,
    ).execute()
    if root.get("trashed") or root.get("mimeType") != FOLDER_MIME:
        raise ValueError("La radice dell'archivio documentale non e' una cartella Drive attiva")
    folder = _unique_named(_drive._list_children(service, root_id), INDEX_FOLDER_NAME, folder=True)
    index_file = _unique_named(_drive._list_children(service, folder["id"]), INDEX_FILE_NAME, folder=False)
    metadata = service.files().get(
        fileId=index_file["id"],
        fields="id,name,mimeType,parents,trashed,modifiedTime,md5Checksum,size,webViewLink",
        supportsAllDrives=True,
    ).execute()
    if metadata.get("trashed"):
        raise ValueError("L'indice documentale risulta nel cestino")
    return folder, metadata


def _download_index_sync(service, file_id: str) -> bytes:
    from googleapiclient.http import MediaIoBaseDownload
    output = io.BytesIO()
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(output, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return output.getvalue()


def _parse_sheet_records(workbook, sheet_name: str, required_headers: tuple[str, ...]) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Foglio obbligatorio assente: {sheet_name}")
    rows = workbook[sheet_name].iter_rows(values_only=True)
    try:
        headers = tuple(str(value or "").strip() for value in next(rows))
    except StopIteration as exc:
        raise ValueError(f"Foglio vuoto: {sheet_name}") from exc
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"Colonne obbligatorie assenti in {sheet_name}: {', '.join(missing)}")
    records = []
    for values in rows:
        record = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if any(value is not None for value in record.values()):
            records.append(record)
    return records


def _parse_index_workbook(content: bytes) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    documents = _parse_sheet_records(workbook, INDEX_SHEET_NAME, REQUIRED_HEADERS)
    return {
        "documents": [record for record in documents if record.get("ID documento")],
        "f24_rows": _parse_sheet_records(workbook, "F24_RIGHE", F24_HEADERS),
        "declarations": _parse_sheet_records(workbook, "DICHIARAZIONI", DECLARATION_HEADERS),
        "duplicates": _parse_sheet_records(workbook, "DUPLICATI_SCARTI", DUPLICATE_HEADERS),
    }


def _parse_index_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Compatibilita' con i chiamanti che richiedono il solo catalogo."""
    return _parse_index_workbook(content)["documents"]


def _public_record(record: dict[str, Any]) -> dict[str, Any]:
    display = _record_display_fields(record)
    return {
        "document_id": record.get("ID documento"),
        "domain": record.get("Dominio"),
        "category": record.get("Categoria"),
        "year": record.get("Anno"),
        "filename": record.get("Nome file"),
        "extension": record.get("Estensione"),
        "size_bytes": record.get("Dimensione byte"),
        "sha256": record.get("SHA-256"),
        "drive_path": record.get("Percorso Drive"),
        "source_zip": record.get("ZIP origine"),
        "source_path": record.get("Percorso nel pacchetto"),
        "status": record.get("Stato"),
        "document_number": record.get("Numero documento"),
        **display,
    }


def _record_display_fields(record: dict[str, Any]) -> dict[str, Any]:
    """Etichette operative sintetiche; non fingono dati non letti dal PDF."""
    filename = str(record.get("Nome file") or "")
    path = str(record.get("Percorso Drive") or "")
    extension = _norm(record.get("Estensione")).lstrip(".")
    searchable = f"{filename} {path} {record.get('Categoria') or ''}"
    normalized = _norm(searchable)
    words = re.sub(r"[_-]+", " ", normalized)
    from app.services.personal_family_registry import match_family_person
    person = match_family_person(searchable)
    subject = person["display_name"] if person else (
        "Ceraldi Group Srl" if "ceraldi group" in normalized or "04523831214" in normalized else "Da identificare"
    )
    if extension == "zip":
        document_type = "Pacchetto sorgente"
        if "archivio fiscale pulito" in words:
            title = "Archivio fiscale verificato 2019-2026"
        elif "partenopay" in normalized:
            title = "Pacchetto PartenoPay completo"
        elif "5 mittenti" in words:
            title = "Raccolta PEC dei 5 mittenti"
        else:
            title = "Archivio ZIP originale"
        summary = "Contenitore di originali: consultare i singoli atti estratti nell'indice operativo."
    elif "r da 2023" in words or "definizione agevolata" in words:
        document_type = "Definizione agevolata AdeR"
        title = "Domanda Rottamazione-quater"
        summary = "Richiesta presentata ad AdeR: attendere esito, piano/importo e successive prove di pagamento."
    elif "verbale" in normalized or "polizia locale" in normalized:
        document_type = "Verbale"
        title = "Verbale Codice della strada"
        summary = "Estrarre numero, targa, date, importi, soggetti e scadenze; il pagamento resta da provare."
    elif "tari" in normalized or "tares" in normalized or "tarsu" in normalized:
        document_type = "Tributo locale"
        title = "TARI / tributo locale"
        summary = "Estrarre contribuente, posizione, immobile, anno, importi e scadenze."
    elif "dimission" in normalized or "unilav" in normalized:
        document_type = "Rapporto di lavoro"
        title = "Dimissioni / comunicazione lavoro"
        summary = "Documento aziendale da collegare al dipendente; non prova pagamenti."
    elif "cartelle esattoriali" in normalized or "agenzia riscossione" in normalized:
        document_type = "Atto AdeR"
        title = "Atto Agenzia Entrate-Riscossione"
        summary = "Leggere numeri di cartella/avviso, contribuente, importi, stato e relazioni documentali."
    else:
        document_type = str(record.get("Categoria") or "Documento")
        title = document_type
        summary = "Metadati catalogati; aprire l'originale o la sezione associata per la lavorazione."
    return {
        "display_title": title,
        "subject": subject,
        "document_type_label": document_type,
        "summary": summary,
        "is_source_package": extension == "zip",
    }


def _basename(value: Any) -> str:
    normalized = str(value or "").replace("\\", "/")
    return PurePosixPath(posixpath.normpath(normalized)).name.casefold()


def _amount(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _date_sort_key(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], pattern).strftime("%Y%m%d")
        except ValueError:
            continue
    return text


def _is_documentary_payment(row: dict[str, Any]) -> bool:
    """Una quietanza prova documentalmente il pagamento, non il riscontro bancario."""
    document_type = _norm(row.get("Tipo documento"))
    source_path = _norm(row.get("Percorso Drive"))
    return (
        "quietanza" in document_type
        or "formato stampabile" in document_type
        or "formato_stampabile" in source_path
    )


def _declaration_type(value: Any, filename: Any = None) -> str:
    """Converte le etichette del foglio nel vocabolario fiscale canonico."""
    normalized = re.sub(r"[^A-Z0-9]+", "_", str(value or "").upper()).strip("_")
    compact = normalized.replace("_", "")
    aliases = {
        "770": "MODELLO_770",
        "MODELLO770": "MODELLO_770",
        "IVA": "DICHIARAZIONE_IVA",
        "DICHIARAZIONEIVA": "DICHIARAZIONE_IVA",
        "LIPE": "LIPE",
        "REDDITI": "REDDITI_SC",
        "REDDITISC": "REDDITI_SC",
        "MODELLO760": "REDDITI_SC",
        "760": "REDDITI_SC",
        "IRAP": "DICHIARAZIONE_IRAP",
        "DICHIARAZIONEIRAP": "DICHIARAZIONE_IRAP",
        "ELENCOPERCIPIENTI": "ELENCO_PERCIPIENTI",
        "PERCIPIENTI": "ELENCO_PERCIPIENTI",
    }
    if compact in aliases:
        return aliases[compact]
    searchable = str(filename or "").upper().replace("\\", "/")
    filename_patterns = (
        (r"(?:^|[/_\s-])LIPE(?:[/_\s.-]|$)", "LIPE"),
        (r"(?:^|[/_\s-])770(?:[/_\s.-]|$)", "MODELLO_770"),
        (r"(?:^|[/_\s-])(?:760|REDDITI(?:_SC)?)(?:[/_\s.-]|$)", "REDDITI_SC"),
        (r"(?:^|[/_\s-])IRAP(?:[/_\s.-]|$)", "DICHIARAZIONE_IRAP"),
        (r"(?:^|[/_\s-])IVA(?:[/_\s.-]|$)", "DICHIARAZIONE_IVA"),
        (r"(?:^|[/_\s-])PERCIPIENTI?(?:[/_\s.-]|$)", "ELENCO_PERCIPIENTI"),
    )
    for pattern, canonical in filename_patterns:
        if re.search(pattern, searchable):
            return canonical
    return normalized


def _stable_f24_row_id(row: dict[str, Any], ordinal: int) -> str:
    identity = "|".join(str(row.get(field) or "").strip() for field in (
        "ID documento", "Anno pagamento", "Data pagamento", "Sezione",
        "Tipo riga", "Codice tributo", "Periodo tributo", "Ente",
        "Debito", "Credito", "Protocollo", "Pagina", "Testo sorgente",
    ))
    digest = hashlib.sha256(f"{identity}|{ordinal}".encode("utf-8")).hexdigest()[:24]
    return f"drive-f24-row:{digest}"


def validate_relations(catalog: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    documents = catalog["documents"]
    f24_rows = catalog["f24_rows"]
    declarations = catalog["declarations"]
    duplicates = catalog["duplicates"]
    by_id = {str(record.get("ID documento")): record for record in documents}
    names: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in documents:
        names[_basename(record.get("Nome file"))].append(record)

    checks = {
        "document_ids_unique": len(by_id) == len(documents),
        "document_hashes_unique": len({_norm(item.get("SHA-256")) for item in documents}) == len(documents),
        "drive_paths_unique": len({_norm(item.get("Percorso Drive")) for item in documents}) == len(documents),
        "all_f24_documents_exist": all(str(row.get("ID documento")) in by_id for row in f24_rows),
        "all_f24_sha_match_document": all(
            str(row.get("ID documento")) in by_id
            and _norm(row.get("SHA-256")) == _norm(by_id[str(row.get("ID documento"))].get("SHA-256"))
            for row in f24_rows
        ),
        "all_f24_paths_match_document": all(
            str(row.get("ID documento")) in by_id
            and _norm(row.get("Percorso Drive")) == _norm(by_id[str(row.get("ID documento"))].get("Percorso Drive"))
            for row in f24_rows
        ),
        "f24_amounts_nonnegative": all(
            _amount(row.get("Debito")) >= 0 and _amount(row.get("Credito")) >= 0
            for row in f24_rows
        ),
        "all_declarations_link_exactly_one_document": all(
            len(names.get(_basename(row.get("Percorso archivio")), [])) == 1
            for row in declarations
        ),
    }
    return {
        "all_true": all(checks.values()),
        "checks": checks,
        "counts": {
            "documents": len(documents),
            "f24_rows": len(f24_rows),
            "f24_documents": len({str(row.get("ID documento")) for row in f24_rows}),
            "documentary_payment_rows": sum(1 for row in f24_rows if _is_documentary_payment(row)),
            "documentary_payment_documents": len({
                str(row.get("ID documento")) for row in f24_rows if _is_documentary_payment(row)
            }),
            "tax_debit_rows": sum(1 for row in f24_rows if _amount(row.get("Debito")) > 0),
            "declarations": len(declarations),
            "duplicates_and_discards": len(duplicates),
        },
    }


def search_records(
    records: list[dict[str, Any]], *, q: str | None = None,
    domain: str | None = None, year: str | None = None,
    extension: str | None = None, status: str | None = None,
    limit: int = 100,
) -> list[dict[str, Any]]:
    query = _norm(q)
    matches = []
    for record in records:
        if domain and _norm(record.get("Dominio")) != _norm(domain):
            continue
        if year and _norm(record.get("Anno")) != _norm(year):
            continue
        if extension and _norm(record.get("Estensione")).lstrip(".") != _norm(extension).lstrip("."):
            continue
        if status and _norm(record.get("Stato")) != _norm(status):
            continue
        display = _record_display_fields(record)
        if query and query not in _norm(" ".join([
            *(str(value or "") for value in record.values()),
            *(str(value or "") for value in display.values()),
        ])):
            continue
        matches.append(_public_record(record))
    matches.sort(key=lambda item: (
        item.get("is_source_package", False),
        _norm(item.get("subject")),
        str(item.get("year") or "9999"),
        _norm(item.get("document_type_label")),
        _norm(item.get("filename")),
    ))
    return matches[:limit]


def _resolve_path_sync(service, root_id: str, drive_path: str) -> dict[str, Any]:
    parts = [part for part in re.split(r"[\\/]+", drive_path or "") if part]
    if not parts:
        raise ValueError("Percorso Drive mancante")
    parent_id = root_id
    item: dict[str, Any] | None = None
    for position, part in enumerate(parts):
        expected_folder = position < len(parts) - 1
        item = _unique_named(_drive._list_children(service, parent_id), part, folder=expected_folder)
        parent_id = item["id"]
    metadata = service.files().get(
        fileId=item["id"],
        fields="id,name,mimeType,parents,trashed,modifiedTime,md5Checksum,size,webViewLink",
        supportsAllDrives=True,
    ).execute()
    if metadata.get("trashed"):
        raise ValueError("Il documento Drive risulta nel cestino")
    return metadata


def load_catalog(service=None) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    source, catalog = load_full_catalog(service)
    return source, catalog["documents"]


def load_full_catalog(service=None) -> tuple[dict[str, Any], dict[str, list[dict[str, Any]]]]:
    global _CACHE_KEY, _CACHE_CATALOG
    service = service or build_drive_service()
    root_id = settings.DRIVE_DOCUMENT_INDEX_ROOT_FOLDER_ID.strip()
    folder, metadata = _discover_index_file_sync(service, root_id)
    cache_key = (str(metadata["id"]), metadata.get("modifiedTime"))
    with _CACHE_LOCK:
        if _CACHE_KEY == cache_key and _CACHE_CATALOG is not None:
            catalog = _CACHE_CATALOG
        else:
            catalog = _parse_index_workbook(_download_index_sync(service, metadata["id"]))
            _CACHE_KEY = cache_key
            _CACHE_CATALOG = catalog
    return {"folder": folder, "index": metadata, "root_id": root_id}, catalog


def get_status(service=None) -> dict[str, Any]:
    source, catalog = load_full_catalog(service)
    metadata = source["index"]
    validation = validate_relations(catalog)
    return {
        "status": "ok",
        "mode": "drive_index_read_only",
        "documents": len(catalog["documents"]),
        "validation": validation,
        "index_name": metadata.get("name"),
        "index_modified_time": metadata.get("modifiedTime"),
        "index_size_bytes": metadata.get("size"),
        "index_url": metadata.get("webViewLink"),
        "stores_document_binaries_in_database": False,
    }


def get_overview(service=None) -> dict[str, Any]:
    _, catalog = load_full_catalog(service)
    documents = catalog["documents"]
    validation = validate_relations(catalog)
    return {
        "validation": validation,
        "domains": [
            {"name": name, "documents": count}
            for name, count in sorted(Counter(str(item.get("Dominio") or "Senza dominio") for item in documents).items())
        ],
        "years": [
            {"year": year, "documents": count}
            for year, count in sorted(
                Counter(str(item.get("Anno") or "Non indicato") for item in documents).items(),
                reverse=True,
            )
        ],
        "semantics": {
            "originals": "Google Drive",
            "database_binary_storage": False,
            "f24_model_is_not_bank_payment": True,
            "quietanza_is_documentary_evidence": True,
            "ambiguous_relations_are_confirmed": False,
        },
    }


def search_catalog(service=None, **filters: Any) -> dict[str, Any]:
    _, records = load_catalog(service)
    results = search_records(records, **filters)
    return {"total_indexed": len(records), "returned": len(results), "results": results}


_ADMINISTRATIVE_AREA_TERMS = {
    "verbali": ("verbali auto", "notifiche polizia locale", "notifica polizia locale"),
    "tributi_locali": ("tributi locali", "tari", "tares", "tarsu"),
    "riscossione": ("cartelle esattoriali", "agenzia riscossione", "ader"),
    "personale": ("dimission", "unilav"),
    "famiglia": (),
}


# Eccezioni documentali verificate sul PDF originale. La chiave SHA-256 evita
# che omonimie, importi o la cartella PEC trasformino un documento personale in
# un costo aziendale. Questi metadati sono solo descrittivi: non generano
# movimenti contabili, pagamenti o riconciliazioni.
_PERSONAL_FAMILY_DOCUMENTS = {
    "d3edc9fd5c999343a4370d441bf2e67fe672d41eb6c370e4a43b589d01bcd45a": {
        "contribuente": "Ceraldi Antonietta",
        "codice_contribuente": "1804135",
        "anno_tributo": "2024",
        "oggetto": "Avviso di pagamento TARI - Acconto 2024",
        "immobile": "Via Cavallerizza 46, Napoli",
    },
}


def _administrative_area(record: dict[str, Any]) -> str | None:
    if _norm(record.get("SHA-256")) in _PERSONAL_FAMILY_DOCUMENTS:
        return "famiglia"
    # La collocazione operativa Drive decide l'area. Il percorso storico nel
    # pacchetto puo' descrivere solo l'email a cui un allegato apparteneva (per
    # esempio un documento d'identita' allegato a una pratica TARI) e non deve
    # quindi promuovere quell'allegato ad atto amministrativo.
    searchable_raw = " ".join(str(record.get(field) or "") for field in (
        "Dominio", "Categoria", "Nome file", "Percorso Drive",
    ))
    searchable = _norm(searchable_raw)
    from app.services.personal_family_registry import (
        is_company_context, is_employment_context, match_family_person,
    )
    # Un documento di rapporto di lavoro resta aziendale anche quando il
    # lavoratore e' contemporaneamente un familiare.
    if is_employment_context(searchable_raw):
        if any(term in searchable for term in _ADMINISTRATIVE_AREA_TERMS["personale"]):
            return "personale"
        return None
    person = match_family_person(searchable_raw)
    # Pane Giuseppina compare anche come legale rappresentante. Nell'indice
    # Drive (che non contiene il testo del PDF) un atto AdeR col solo suo CF e'
    # ambiguo e resta aziendale/documentale; il parser del contenuto potra'
    # spostarlo in Famiglia solo quando l'intestatario persona fisica e' certo.
    if person and person["person_id"] == "pane-giuseppina" and any(
        term in searchable for term in _ADMINISTRATIVE_AREA_TERMS["riscossione"]
    ):
        return "riscossione"
    if person and not is_company_context(searchable_raw):
        return "famiglia"
    for area, terms in _ADMINISTRATIVE_AREA_TERMS.items():
        if any(term in searchable for term in terms):
            return area
    return None


def list_administrative_documents(
    service=None, *, area: str | None = None, year: str | None = None,
    q: str | None = None, review_only: bool = False, limit: int = 500,
) -> dict[str, Any]:
    """Atti amministrativi dall'indice Drive, senza copiare i PDF nel DB."""
    _, records = load_catalog(service)
    overview_counts = {key: 0 for key in _ADMINISTRATIVE_AREA_TERMS}
    overview_review = 0
    matches: list[dict[str, Any]] = []
    query = _norm(q)
    from app.services.personal_family_registry import family_search_terms, match_family_person
    query_terms = family_search_terms(q) if query else set()

    for record in records:
        if _norm(record.get("Estensione")).lstrip(".") != "pdf":
            continue
        record_area = _administrative_area(record)
        if not record_area:
            continue
        status = _norm(record.get("Stato"))
        category = _norm(record.get("Categoria"))
        requires_review = (
            status in {"da verificare", "da_verificare", "errore"}
            or not str(record.get("Anno") or "").strip()
            or category in {"esistente su drive", "documento"}
        )
        overview_counts[record_area] += 1
        overview_review += int(requires_review)

        if area and record_area != area:
            continue
        if year and _norm(record.get("Anno")) != _norm(year):
            continue
        if review_only and not requires_review:
            continue
        record_search = _norm(" ".join(str(value or "") for value in record.values()))
        if query and not any(term.casefold() in record_search for term in query_terms):
            continue

        public = _public_record(record)
        personal_metadata = _PERSONAL_FAMILY_DOCUMENTS.get(_norm(public["sha256"]), {})
        family_person = match_family_person(record_search)
        if family_person:
            personal_metadata = {
                **personal_metadata,
                "person_id": family_person["person_id"],
                "persona": family_person["display_name"],
                "identity_matched_by": family_person["matched_by"],
                "lavoratore_cf": next(iter(family_person.get("identifiers", {}).get("codice_fiscale", ())), None),
            }
        accounting_excluded = record_area == "famiglia"
        matches.append({
            "id": public["document_id"],
            "filename": public["filename"],
            "category": public["category"],
            "category_label": public["category"] or public["domain"],
            "administrative_area": record_area,
            "document_date_display": public["year"],
            "status": public["status"] or "indicizzato_drive",
            "sha256": public["sha256"],
            "source_kind": "drive_index",
            "source_label": "Google Drive",
            "accounting_scope": "personal_family" if accounting_excluded else "business_documentary",
            "accounting_excluded": accounting_excluded,
            "accounting_exclusion_reason": (
                "Documento personale/familiare: escluso da bilanci, costi, Prima Nota e riconciliazioni aziendali."
                if accounting_excluded else None
            ),
            "source_context": {
                "archive_path": public["drive_path"],
                "source_zip": public["source_zip"],
                "source_path": public["source_path"],
            },
            "parsed_metadata": {
                "requires_review": requires_review,
                "numero_documento": public["document_number"],
                **personal_metadata,
            },
        })

    matches.sort(key=lambda item: (
        str(item.get("document_date_display") or ""), str(item.get("filename") or "").casefold(),
    ), reverse=True)
    return {
        "items": matches[:limit],
        "total": len(matches),
        "counts": dict(Counter(item["administrative_area"] for item in matches)),
        "requires_review": sum(bool(item["parsed_metadata"]["requires_review"]) for item in matches),
        "overview": {
            "counts": overview_counts,
            "total": sum(overview_counts.values()),
            "requires_review": overview_review,
        },
        "source": "drive_excel_index",
    }


def list_f24_documents(
    service=None, *, q: str | None = None, year: str | None = None,
    tax_code: str | None = None, limit: int = 200,
) -> dict[str, Any]:
    _, catalog = load_full_catalog(service)
    by_id = {str(item.get("ID documento")): item for item in catalog["documents"]}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in catalog["f24_rows"]:
        if year and _norm(row.get("Anno pagamento")) != _norm(year):
            continue
        if tax_code and _norm(row.get("Codice tributo")) != _norm(tax_code):
            continue
        if q and _norm(q) not in _norm(" ".join(str(value or "") for value in row.values())):
            document = by_id.get(str(row.get("ID documento")), {})
            if _norm(q) not in _norm(" ".join(str(value or "") for value in document.values())):
                continue
        grouped[str(row.get("ID documento"))].append(row)

    results = []
    for document_id, rows in grouped.items():
        document = by_id.get(document_id)
        if not document:
            continue
        codes = sorted({str(row.get("Codice tributo") or "") for row in rows if row.get("Codice tributo")})
        results.append({
            "document": _public_record(document),
            "payment_year": rows[0].get("Anno pagamento"),
            "payment_date": rows[0].get("Data pagamento"),
            "protocol": rows[0].get("Protocollo"),
            "document_type": rows[0].get("Tipo documento"),
            "tax_codes": codes,
            "tax_rows": len(rows),
            "total_debit": round(sum(_amount(row.get("Debito")) for row in rows), 2),
            "total_credit": round(sum(_amount(row.get("Credito")) for row in rows), 2),
            "evidence_state": (
                "QUIETANZA_DOCUMENTALE_NON_PROVA_BANCARIA"
                if _is_documentary_payment(rows[0])
                else "MODELLO_F24_NON_PROVA_BANCARIA"
            ),
        })
    results.sort(key=lambda item: (_date_sort_key(item.get("payment_date")), item["document"]["filename"]), reverse=True)
    return {"returned": min(len(results), limit), "total_matching": len(results), "results": results[:limit]}


def list_f24_rows(
    service=None, *, year: str | None = None, tax_code: str | None = None,
    document_id: str | None = None, credits_only: bool = False,
    offset: int = 0, limit: int = 1000,
) -> dict[str, Any]:
    """Righe F24 del foglio Drive, senza inferire quietanza o pagamento bancario."""
    _, catalog = load_full_catalog(service)
    by_id = {str(item.get("ID documento")): item for item in catalog["documents"]}
    matches: list[dict[str, Any]] = []
    ordinals: dict[str, int] = defaultdict(int)
    for row in catalog["f24_rows"]:
        current_document_id = str(row.get("ID documento") or "")
        ordinals[current_document_id] += 1
        ordinal = ordinals[current_document_id]
        if year and _norm(row.get("Anno pagamento")) != _norm(year):
            continue
        if tax_code and _norm(row.get("Codice tributo")) != _norm(tax_code):
            continue
        if document_id and _norm(current_document_id) != _norm(document_id):
            continue
        credit = _amount(row.get("Credito"))
        if credits_only and credit <= 0:
            continue
        document = by_id.get(current_document_id, {})
        matches.append({
            "id": _stable_f24_row_id(row, ordinal),
            "document_id": current_document_id,
            "ordinal": ordinal,
            "source_kind": "DRIVE_EXCEL_INDEX_F24_ROW",
            "payment_year": row.get("Anno pagamento"),
            "payment_date": row.get("Data pagamento"),
            "section": row.get("Sezione"),
            "row_type": row.get("Tipo riga"),
            "tax_code": row.get("Codice tributo"),
            "description": row.get("Descrizione"),
            "reference_period": row.get("Periodo tributo"),
            "entity": row.get("Ente"),
            "debit_amount": _amount(row.get("Debito")),
            "credit_amount": credit,
            "protocol": row.get("Protocollo"),
            "document_type": row.get("Tipo documento"),
            "sha256": row.get("SHA-256") or document.get("SHA-256"),
            "drive_path": row.get("Percorso Drive") or document.get("Percorso Drive"),
            "page": row.get("Pagina"),
            "source_text": row.get("Testo sorgente"),
            "source": row.get("Fonte"),
            "filename": document.get("Nome file"),
            "evidence_state": (
                "QUIETANZA_DOCUMENTALE_NON_PROVA_BANCARIA"
                if _is_documentary_payment(row)
                else "MODELLO_F24_NON_PROVA_BANCARIA"
            ),
        })
    matches.sort(key=lambda item: (
        _date_sort_key(item.get("payment_date")), str(item.get("filename") or ""),
        str(item.get("tax_code") or ""),
    ), reverse=True)
    page = matches[offset:offset + limit]
    return {
        "items": page,
        "total": len(matches),
        "offset": offset,
        "limit": limit,
        "source": "drive_excel_index",
    }


def list_documented_tax_payments(
    service=None, *, offset: int = 0, limit: int = 5000,
) -> dict[str, Any]:
    """Tutte le righe delle deleghe documentate da quietanza Drive.

    Una quietanza F24 puo' contenere contemporaneamente righe a debito e righe
    a credito.  Escludere queste ultime altera il documento e il saldo netto,
    percio' l'intera delega viene restituita mantenendo ogni riga del PDF.
    """
    _, catalog = load_full_catalog(service)
    documented_ids = {
        str(row.get("ID documento") or "")
        for row in catalog["f24_rows"]
        if _is_documentary_payment(row)
    }
    rows = list_f24_rows(service=service, offset=0, limit=5000)["items"]
    items = []
    for row in rows:
        if str(row.get("document_id") or "") not in documented_ids:
            continue
        items.append({
            **row,
            "payment_status": "DOCUMENTATO_DA_QUIETANZA",
            "documentary_payment_status": "QUIETANZA_PRESENTE",
            "bank_status": "DA_VERIFICARE",
        })
    return {
        "items": items[offset:offset + limit],
        "total": len(items),
        "offset": offset,
        "limit": limit,
        "source": "drive_excel_index",
    }


def list_tax_obligations(
    service=None, *, offset: int = 0, limit: int = 5000,
) -> dict[str, Any]:
    """Deleghe F24 complete, con debiti e crediti e stato prova esplicito.

    La vista raggruppa le righe per documento: eliminare qui le righe a credito
    renderebbe falsi i totali e il saldo della quietanza visualizzata.
    """
    rows = list_f24_rows(service=service, offset=0, limit=5000)["items"]
    items = []
    for row in rows:
        documentary = row.get("evidence_state") == "QUIETANZA_DOCUMENTALE_NON_PROVA_BANCARIA"
        items.append({
            **row,
            "payment_status": "DOCUMENTATO_DA_QUIETANZA" if documentary else "MODELLO_F24_PRESENTE",
            "documentary_payment_status": "QUIETANZA_PRESENTE" if documentary else "DA_VERIFICARE",
            "bank_status": "DA_VERIFICARE",
        })
    return {
        "items": items[offset:offset + limit],
        "total": len(items),
        "offset": offset,
        "limit": limit,
        "source": "drive_excel_index",
    }


def list_declarations(
    service=None, *, year: str | None = None, declaration_type: str | None = None,
    q: str | None = None, limit: int = 200,
) -> dict[str, Any]:
    _, catalog = load_full_catalog(service)
    names: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for document in catalog["documents"]:
        names[_basename(document.get("Nome file"))].append(document)
    results = []
    for row in catalog["declarations"]:
        if year and _norm(row.get("Anno")) != _norm(year):
            continue
        canonical_type = _declaration_type(row.get("Tipo"), row.get("Percorso archivio"))
        if declaration_type and canonical_type != _declaration_type(declaration_type):
            continue
        if q and _norm(q) not in _norm(" ".join(str(value or "") for value in row.values())):
            continue
        matches = names.get(_basename(row.get("Percorso archivio")), [])
        filing_year = int(row.get("Anno")) if str(row.get("Anno") or "").isdigit() else row.get("Anno")
        document = _public_record(matches[0]) if len(matches) == 1 else None
        results.append({
            "id": document.get("document_id") if document else None,
            "document_id": document.get("document_id") if document else None,
            "source_kind": "DRIVE_EXCEL_INDEX_DECLARATION",
            "filing_year": filing_year,
            "tax_year": filing_year if canonical_type == "LIPE" else (filing_year - 1 if isinstance(filing_year, int) else None),
            "document_type": canonical_type,
            "filename": document.get("filename") if document else _basename(row.get("Percorso archivio")),
            "sha256": document.get("sha256") if document else None,
            "drive_path": document.get("drive_path") if document else row.get("Percorso archivio"),
            "year": row.get("Anno"),
            "type": row.get("Tipo"),
            "protocol": row.get("Protocollo"),
            "archive_path": row.get("Percorso archivio"),
            "relation_state": "CONFERMATA_NOME_UNIVOCO_E_INDICE_VERIFICATO" if len(matches) == 1 else "AMBIGUA",
            "document": document,
            "f24_links": [],
            "f24_confirmed_count": 0,
            "f24_candidate_count": 0,
        })
    results.sort(key=lambda item: (str(item.get("year") or ""), str(item.get("archive_path") or "")), reverse=True)
    return {"returned": min(len(results), limit), "total_matching": len(results), "results": results[:limit]}


def load_declaration_pdf(document_id: str, service=None) -> dict[str, Any]:
    """Scarica un originale dichiarativo solo dopo il legame univoco dell'indice."""
    service = service or build_drive_service()
    source, catalog = load_full_catalog(service)
    documents = [
        row for row in catalog["documents"]
        if _norm(row.get("ID documento")) == _norm(document_id)
    ]
    if len(documents) != 1:
        raise ValueError(f"Documento dichiarativo assente o ambiguo: {document_id}")
    document = documents[0]
    declaration_rows = [
        row for row in catalog["declarations"]
        if _basename(row.get("Percorso archivio")) == _basename(document.get("Nome file"))
    ]
    if len(declaration_rows) != 1:
        raise ValueError(
            f"Relazione dichiarazione-documento assente o ambigua: {document_id} ({len(declaration_rows)})"
        )
    public = _public_record(document)
    if _norm(public.get("extension")).lstrip(".") != "pdf":
        raise ValueError(f"Originale dichiarativo non PDF: {document_id}")
    metadata = _resolve_path_sync(service, source["root_id"], str(document.get("Percorso Drive") or ""))
    from app.services.fiscal_document_ingestion import download_drive_file

    content = download_drive_file(service, metadata["id"])
    digest = hashlib.sha256(content).hexdigest()
    expected = str(document.get("SHA-256") or "").strip().casefold()
    if expected and digest.casefold() != expected:
        raise ValueError(f"Hash originale Drive non coincide con l'indice: {document_id}")
    declaration = declaration_rows[0]
    canonical_type = _declaration_type(declaration.get("Tipo"), declaration.get("Percorso archivio"))
    return {
        "content": content,
        "document": public,
        "declaration": {
            "document_id": document_id,
            "document_type": canonical_type,
            "filing_year": declaration.get("Anno"),
            "protocol": declaration.get("Protocollo"),
            "archive_path": declaration.get("Percorso archivio"),
        },
        "drive_file_id": metadata["id"],
        "drive_url": metadata.get("webViewLink") or f"https://drive.google.com/open?id={metadata['id']}",
        "sha256": digest,
    }


def get_document(document_id: str, service=None) -> dict[str, Any]:
    service = service or build_drive_service()
    source, catalog = load_full_catalog(service)
    records = catalog["documents"]
    matches = [record for record in records if _norm(record.get("ID documento")) == _norm(document_id)]
    if len(matches) != 1:
        raise ValueError(f"Documento assente o ambiguo: {document_id}")
    record = _public_record(matches[0])
    metadata = _resolve_path_sync(service, source["root_id"], str(matches[0].get("Percorso Drive") or ""))
    record.update({
        "drive_file_id": metadata.get("id"),
        "drive_url": metadata.get("webViewLink") or f"https://drive.google.com/open?id={metadata.get('id')}",
        "drive_modified_time": metadata.get("modifiedTime"),
        "drive_md5": metadata.get("md5Checksum"),
        "drive_size_bytes": metadata.get("size"),
    })
    f24_rows = [row for row in catalog["f24_rows"] if _norm(row.get("ID documento")) == _norm(document_id)]
    declaration_rows = [
        row for row in catalog["declarations"]
        if _basename(row.get("Percorso archivio")) == _basename(matches[0].get("Nome file"))
    ]
    record["relations"] = {
        "f24_rows": [{
            "payment_year": row.get("Anno pagamento"),
            "payment_date": row.get("Data pagamento"),
            "section": row.get("Sezione"),
            "row_type": row.get("Tipo riga"),
            "tax_code": row.get("Codice tributo"),
            "description": row.get("Descrizione"),
            "tax_period": row.get("Periodo tributo"),
            "entity": row.get("Ente"),
            "debit": _amount(row.get("Debito")),
            "credit": _amount(row.get("Credito")),
            "protocol": row.get("Protocollo"),
            "page": row.get("Pagina"),
            "source": row.get("Fonte"),
        } for row in f24_rows],
        "declarations": [{
            "year": row.get("Anno"), "type": row.get("Tipo"),
            "protocol": row.get("Protocollo"), "archive_path": row.get("Percorso archivio"),
        } for row in declaration_rows],
        "bank_payment_confirmed": False,
        "relation_note": "Il modello o la quietanza non confermano da soli il pagamento bancario.",
    }
    return record
