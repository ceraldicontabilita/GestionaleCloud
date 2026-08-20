"""
Gestione Estratto Conto
Salva e visualizza tutti i movimenti bancari importati con campi strutturati.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Depends
from app.utils.dependencies import get_current_admin_user
from typing import Dict, Any, List, Optional
from datetime import datetime, date, timezone
import logging
import io
import re
import csv
from collections import Counter, defaultdict

from app.database import Database
from app.utils.error_handler import handle_errors
from app.routers.prima_nota_module.common import (
    aggrega_saldo_prima_nota,
    entra_in_prima_nota,
)
from app.routers.prima_nota_module.sync import costruisci_campi_movimento_fattura
from app.services.scritture_contabili import scrivi_movimento
from app.services.bank_evidence import EVIDENZA_UFFICIALE, campi_evidenza

logger = logging.getLogger(__name__)
router = APIRouter()


def _occorrenza_gia_importata(
    existing_counts: Counter, incoming_occurrences: Counter, key: Any,
) -> bool:
    """Deduplica un reimport preservando due righe bancarie identiche reali."""
    incoming_occurrences[key] += 1
    return existing_counts[key] >= incoming_occurrences[key]


def estrai_numero_fattura(descrizione: str) -> Optional[str]:
    """Estrae il numero/i fattura dalla descrizione dopo NOTPROVIDE."""
    if not descrizione:
        return None
    
    # Pattern: dopo "NOTPROVIDE - " c'è il riferimento fatture
    match = re.search(r'NOTPROVIDE\s*-?\s*(.+)$', descrizione, re.IGNORECASE)
    if match:
        riferimento = match.group(1).strip()
        # Pulisci e restituisci
        # Rimuovi prefissi comuni
        riferimento = re.sub(r'^(saldo|pagamento)\s+(fattur[ae]|ft)\s*', '', riferimento, flags=re.IGNORECASE)
        return riferimento.strip()[:200] if riferimento.strip() else None
    
    # Pattern alternativo: "fattura/e" seguito da numeri
    match = re.search(r'fattur[ae]\s+(.+?)(?:\s*$|\s+-)', descrizione, re.IGNORECASE)
    if match:
        return match.group(1).strip()[:200]
    
    return None


def is_versamento_contanti(descrizione: str) -> bool:
    """Riconosce la causale di un versamento di contanti in banca.

    Bug segnalato dall'utente 15/07/2026: l'export reale della banca (Banco
    BPM) usa l'abbreviazione "VERS." (es. "VERS. CONTANTI - VVVVV"), MAI la
    parola intera "VERSAMENTO" — verificato su un export reale di 4287
    movimenti: 96 causali "VERS. CONTANTI", zero contenenti "VERSAMENTO".
    Riconosce entrambe le forme.
    """
    desc_upper = (descrizione or "").upper()
    # Uno storno non e' un nuovo versamento: e' una rettifica bancaria che
    # deve restare visibile e da verificare, senza creare movimenti di cassa.
    if "STORNO" in desc_upper:
        return False
    if "CONTANT" not in desc_upper:
        return False
    return "VERSAMENTO" in desc_upper or bool(re.search(r"VERS\.?\s*CONTANT", desc_upper))


def is_storno_versamento(descrizione: str) -> bool:
    """Riconosce la rettifica/storno di un versamento contanti.

    Lo storno non prova che il contante sia rientrato in cassa e non deve
    essere trattato come un secondo deposito. Resta nell'estratto conto in
    attesa di verifica/associazione esplicita.
    """
    desc_upper = (descrizione or "").upper()
    if "STORNO" not in desc_upper or "CONTANT" not in desc_upper:
        return False
    return "VERSAMENTO" in desc_upper or bool(re.search(r"VERS\.?\s*CONTANT", desc_upper))


# REGOLA (utente 17/07/2026): in Prima Nota Banca la categoria dice
# L'OPERAZIONE, sintetica al massimo — mai la classificazione della banca.
# Il dettaglio (chi, cosa, riferimento) sta nella DESCRIZIONE. Il
# riconoscimento parte dalla causale (PayPal, versamento, prelievo,
# utenza), poi ricade sulla tassonomia del CSV bancario. NON tocca i dati
# salvati: l'estratto conto resta immutabile, il nome operativo viene
# calcolato al volo nella risposta (campo `categoria_canonica`).
_CATEGORIE_EC_ESATTE = {
    # il deposito di contanti È il versamento (doppia scrittura cassa/banca)
    "Ricavi - Deposito contanti": "Versamento Banca",
    # il leasing viaggia con fatture periodiche del concedente
    "Altre passività - Leasing": "Fatture",
    "Ricavi - Rimborsi diversi": "Rimborso",
}
_CATEGORIE_EC_PREFISSI = {
    "Fornitori": "Fatture",
    "Utenze": "Utenze",
    "Servizi": "Fatture",
    "Assicurazione": "Fatture",
    "Operazioni Finanziarie": "Commissioni bancarie",
    "Tasse": "F24",
    "Risorse Umane": "Stipendi",
    "Ricavi": "Rimborso",
    "Altre passività": "Altro",
    "Altre spese": "Altro",
    "Intercompany in entrata": "Altro",
    "Intercompany in uscita": "Altro",
    "Intercompany": "Altro",
}


# La STESSA operazione bancaria arriva con causali diverse a seconda del
# file: l'export sintetico scrive "SDD CORE: M-...", quello analitico
# "ADDEBITO DIRETTO SDD - SDD CORE: M-...". Con la descrizione grezza nella
# chiave di deduplica le due forme risultavano operazioni diverse, e lo
# stesso addebito WORLDPAY entrava due volte (segnalazione utente 07/08/2026,
# tre coppie da 1,79 nello stesso giorno).
#
# Si normalizzano SOLO i prefissi di canale, mai il contenuto: il mandato,
# il beneficiario e i riferimenti restano parte della chiave. Se la banca ha
# davvero addebitato 3 volte 1,79 nello stesso giorno, ogni file porta 3
# righe e il conteggio per occorrenza le conserva tutte e tre.
_PREFISSI_CANALE_EC = re.compile(
    r"^(ADDEBITO DIRETTO SDD\s*-\s*|ADDEBITO SDD\s*-\s*|PAGAMENTO\s*-\s*)",
    re.IGNORECASE,
)


def normalizza_descrizione_ec(desc: str) -> str:
    """Descrizione canonica di una riga EC ai fini della deduplica."""
    compatta = re.sub(r"\s+", " ", (desc or "").strip())
    return _PREFISSI_CANALE_EC.sub("", compatta)[:80]


def mappa_categoria_ec(categoria: Optional[str], descrizione: Optional[str] = None) -> Optional[str]:
    """Categoria OPERATIVA di Prima Nota per una riga di estratto conto.

    Prima la causale (che batte la classificazione della banca: PayPal
    finisce classificato dalla banca in 3 modi diversi), poi la tassonomia
    del CSV. Ritorna None se non riconosciuta (resta l'originale)."""
    desc = (descrizione or "").upper()
    if desc:
        if "PAYPAL" in desc:
            return "Pagamento PayPal"
        if is_storno_versamento(desc):
            return "Storno versamento"
        if is_versamento_contanti(desc):
            return "Versamento Banca"
        if is_prelievo_contanti(desc):
            return "Prelevamento Banca"
        if "UTENZ" in desc:
            return "Utenze"
    if not categoria:
        return None
    if categoria in _CATEGORIE_EC_ESATTE:
        return _CATEGORIE_EC_ESATTE[categoria]
    prefisso = categoria.split(" - ")[0].strip()
    return _CATEGORIE_EC_PREFISSI.get(prefisso)


def is_prelievo_contanti(descrizione: str) -> bool:
    """Riconosce la causale di un prelievo di contanti dal conto (bancomat,
    sportello, ATM): il movimento opposto del versamento — il denaro esce
    dalla banca ed entra in cassa."""
    desc_upper = (descrizione or "").upper()
    if "PRELIEVO" not in desc_upper and "PRELEV" not in desc_upper:
        return False
    return any(k in desc_upper for k in ("BANCOMAT", "CONTANT", "SPORTELLO", " ATM"))


def estrai_fornitore_pulito(descrizione: str) -> Optional[str]:
    """Estrae il nome fornitore dalla descrizione, pulendolo."""
    if not descrizione:
        return None
    
    desc_upper = descrizione.upper()
    
    if "FAVORE" in desc_upper:
        idx = desc_upper.find("FAVORE")
        after = descrizione[idx + 7:].strip()
        
        # Prendi fino a "NOTPROVIDE" o " - " o fine
        for sep in ["NOTPROVIDE", " - ADD.", " - "]:
            if sep.upper() in after.upper():
                idx_sep = after.upper().find(sep.upper())
                after = after[:idx_sep].strip()
                break
        
        # Rimuovi forme societarie alla fine per pulizia display
        # Ma mantienile se sono parte del nome
        nome = after.strip()

        return nome if nome else None

    # Addebiti diretti SDD (CORE o B2B): niente "FAVORE", il beneficiario è
    # il testo dopo il codice mandato/riferimento (senza spazi), es. "SDD
    # CORE: <codice mandato> Nome Beneficiario" o "SDD B2B : <codice>
    # Nome Fornitore Srl". Bug segnalato dall'utente 15/07/2026: su un
    # export reale di 4287 movimenti, 240 righe SDD (utenze, leasing,
    # assicurazioni, fornitori ricorrenti) non estraevano MAI il fornitore,
    # perdendo il bonus di punteggio "nome in causale" nel matching con le
    # fatture.
    match_sdd = re.search(r"SDD\s+(?:CORE|B2B)\s*:\s*\S+\s+(.+)$", descrizione, re.IGNORECASE)
    if match_sdd:
        nome = match_sdd.group(1).strip()
        return nome if nome else None

    return None


@router.post("/import")
@handle_errors
async def import_estratto_conto(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa estratto conto bancario e salva tutti i movimenti con campi strutturati.
    
    Formato CSV atteso (delimitatore ';'):
    - Ragione Sociale: nome azienda
    - Data contabile: data in formato DD/MM/YYYY
    - Data valuta: data valuta in formato DD/MM/YYYY
    - Banca: nome banca e codice
    - Rapporto: numero rapporto/conto
    - Importo: importo con virgola decimale (es: 254,5)
    - Divisa: valuta (EUR)
    - Descrizione: descrizione del movimento
    - Categoria/sottocategoria: categoria del movimento
    - Hashtag: tag opzionale
    
    Evita duplicati controllando data + importo + descrizione.
    """
    db = Database.get_db()
    # Il job Drive puo' incontrare archivi storici gia' importati. In quel
    # canale non ripete le riparazioni globali per ogni file interamente
    # duplicato; upload manuali conservano invece il comportamento storico.
    skip_duplicate_repairs = bool(getattr(file, "skip_duplicate_repairs", False))
    drive_file_id = getattr(file, "drive_file_id", None)
    drive_source_path = getattr(file, "source_path", None)
    
    filename_originale = file.filename or "estratto-conto"
    filename = filename_originale.lower()
    evidenza = campi_evidenza(filename_originale)
    fonte_ufficiale = evidenza["livello_evidenza"] == EVIDENZA_UFFICIALE
    contents = await file.read()
    
    movimenti = []
    
    if filename.endswith('.pdf'):
        from app.parsers.estratto_conto_bpm_parser import parse_estratto_conto_bpm
        from app.parsers.estratto_conto_bnl_parser import parse_estratto_conto_bnl
        from app.parsers.estratto_conto_nexi_parser import EstrattoContoNexiParser

        # Un solo ingresso per tutti i PDF: prova i parser deterministici in
        # ordine e accetta soltanto un risultato con movimenti reali. Prima
        # questo endpoint provava esclusivamente BPM, quindi i PDF BNL e carta
        # finivano in Errori anche se i parser esistevano gia' nel progetto.
        attempts = []
        parser_errors = []
        for parser_name, parser in (
            ("Banco BPM", parse_estratto_conto_bpm),
            ("BNL", parse_estratto_conto_bnl),
            ("Nexi", EstrattoContoNexiParser().parse_pdf),
        ):
            try:
                attempts.append(parser(contents))
            except Exception as exc:
                # Un formato non riconosciuto da un parser non deve impedire
                # agli altri parser di esaminare lo stesso PDF.
                parser_errors.append(f"{parser_name}: {exc}")
        parsed = next(
            (candidate for candidate in attempts
             if candidate.get("success") and candidate.get("transazioni")),
            None,
        )
        if parsed is None:
            errors = [candidate.get("error") for candidate in attempts if candidate.get("error")]
            errors.extend(parser_errors)
            raise HTTPException(
                status_code=400,
                detail="; ".join(errors) or "PDF bancario/carta non riconosciuto",
            )

        def _date_obj(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            raw = str(value or "").strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(raw[:10], fmt).date()
                except ValueError:
                    continue
            return None

        banca_parser = parsed.get("banca") or (
            "BNL" if "bnl" in (parsed.get("tipo_documento") or "").lower()
            else "Nexi/Banco BPM" if "nexi" in (parsed.get("tipo_documento") or "").lower()
            else "Banco BPM"
        )
        for transazione in parsed.get("transazioni") or []:
            data_contabile = _date_obj(
                transazione.get("data") or transazione.get("data_contabile")
            )
            if data_contabile is None:
                continue
            data_valuta = _date_obj(transazione.get("data_valuta"))
            descrizione = transazione.get("descrizione") or f"Movimento {banca_parser}"
            importo = float(transazione.get("importo") or 0)
            tipo_parser = str(transazione.get("tipo") or "").lower()
            if tipo_parser in {"uscita", "addebito", "carta_credito"}:
                importo = -abs(importo)
            elif tipo_parser in {"entrata", "accredito"}:
                importo = abs(importo)
            movimenti.append({
                "data": data_contabile,
                "ragione_sociale": None,
                "fornitore": estrai_fornitore_pulito(descrizione),
                "importo": importo,
                "numero_fattura": estrai_numero_fattura(descrizione),
                "data_pagamento": data_valuta,
                "categoria": transazione.get("categoria") or "",
                "descrizione_originale": descrizione,
                "banca": banca_parser,
                "rapporto": (parsed.get("metadata") or {}).get("numero_conto"),
                "divisa": transazione.get("divisa") or "EUR",
                "hashtag": None,
                "tipo": "uscita" if importo < 0 else "entrata",
            })

    elif filename.endswith('.csv'):
        # Prova diversi encoding
        text = None
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                text = contents.decode(encoding)
                break
            except (UnicodeDecodeError, Exception):
                continue
        
        if not text:
            raise HTTPException(status_code=400, detail="Impossibile decodificare il file CSV")
        
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        
        for row in reader:
            # Estrai dati con supporto per varianti di nomi colonna
            # Supporta sia formato con virgolette che senza
            
            # Ragione sociale
            ragione_sociale = row.get('Ragione Sociale', '') or row.get('"Ragione Sociale"', '')
            
            # Data contabile
            data_contabile = (
                row.get('Data contabile', '') or 
                row.get('"Data contabile"', '') or
                row.get('Data', '')
            ).strip().strip('"')
            
            # Data valuta
            data_valuta = (
                row.get('Data valuta', '') or 
                row.get('"Data valuta"', '') or
                row.get('Data valut.', '')
            ).strip().strip('"')
            
            # Banca
            banca = (row.get('Banca', '') or row.get('"Banca"', '')).strip().strip('"')
            
            # Rapporto (numero conto)
            rapporto = (row.get('Rapporto', '') or row.get('"Rapporto"', '')).strip().strip('"')
            
            # Importo - può essere con virgola come separatore decimale
            importo_str = (
                row.get('Importo', '0') or 
                row.get('"Importo"', '0')
            ).strip().strip('"')
            
            # Parse importo: rimuovi punti migliaia, sostituisci virgola con punto
            importo_str = importo_str.replace('.', '').replace(',', '.')
            try:
                importo = float(importo_str)
            except (ValueError, TypeError):
                continue
            
            # Divisa
            divisa = (row.get('Divisa', 'EUR') or row.get('"Divisa"', 'EUR')).strip().strip('"')
            
            # Descrizione
            descrizione = (
                row.get('Descrizione', '') or 
                row.get('"Descrizione"', '') or
                row.get('Descrizion', '')
            ).strip().strip('"')
            
            # Categoria
            categoria = (
                row.get('Categoria/sottocategoria', '') or 
                row.get('"Categoria/sottocategoria"', '') or
                row.get('Categoria', '') or
                row.get('"Categoria"', '')
            ).strip().strip('"')
            
            # Hashtag
            hashtag = (row.get('Hashtag', '') or row.get('"Hashtag"', '')).strip().strip('"')
            
            # Parse data contabile (DD/MM/YYYY)
            try:
                if '/' in data_contabile:
                    parts = data_contabile.split('/')
                    data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    continue
            except (ValueError, TypeError, IndexError):
                continue
            
            # Parse data valuta
            data_pagamento = None
            try:
                if '/' in data_valuta:
                    parts = data_valuta.split('/')
                    data_pagamento = date(int(parts[2]), int(parts[1]), int(parts[0]))
            except (ValueError, TypeError, IndexError):
                pass
            
            # Estrai fornitore/beneficiario dalla descrizione
            fornitore = estrai_fornitore_pulito(descrizione)
            
            # Estrai numero fattura dalla descrizione
            numero_fattura = estrai_numero_fattura(descrizione)
            
            movimenti.append({
                "data": data_obj,
                "ragione_sociale": ragione_sociale.strip().strip('"') if ragione_sociale else None,
                "fornitore": fornitore,
                "importo": importo,
                "numero_fattura": numero_fattura,
                "data_pagamento": data_pagamento,
                "categoria": categoria,
                "descrizione_originale": descrizione,
                "banca": banca,
                "rapporto": rapporto,
                "divisa": divisa,
                "hashtag": hashtag,
                "tipo": "uscita" if importo < 0 else "entrata"
            })
    
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            
            # Mappa header originali a chiavi normalizzate
            headers_raw = [str(cell.value or '') for cell in sheet[1]]
            headers = [h.lower().strip() for h in headers_raw]
            
            for row_num in range(2, sheet.max_row + 1):
                row_data = {headers[i]: sheet.cell(row=row_num, column=i+1).value 
                           for i in range(len(headers))}
                
                # Trova colonne con supporto per varianti
                data_contabile = None
                importo = None
                descrizione = ""
                categoria = ""
                data_valuta = None
                ragione_sociale = ""
                banca = ""
                rapporto = ""
                divisa = "EUR"
                hashtag = ""
                
                for h, v in row_data.items():
                    if not v:
                        continue
                    h_lower = h.lower()
                    
                    # Ragione Sociale
                    if 'ragione sociale' in h_lower:
                        ragione_sociale = str(v).strip()
                    
                    # Data contabile
                    elif 'data contabile' in h_lower or (h_lower == 'data' and not data_contabile):
                        if isinstance(v, (datetime, date)):
                            data_contabile = v if isinstance(v, date) else v.date()
                        elif '/' in str(v):
                            parts = str(v).split('/')
                            try:
                                data_contabile = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            except (ValueError, TypeError, IndexError):
                                pass
                    
                    # Data valuta
                    elif 'data valuta' in h_lower or 'data valut' in h_lower:
                        if isinstance(v, (datetime, date)):
                            data_valuta = v if isinstance(v, date) else v.date()
                        elif '/' in str(v):
                            parts = str(v).split('/')
                            try:
                                data_valuta = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            except (ValueError, TypeError, IndexError):
                                pass
                    
                    # Banca
                    elif h_lower == 'banca':
                        banca = str(v).strip()
                    
                    # Rapporto
                    elif h_lower == 'rapporto':
                        rapporto = str(v).strip()
                    
                    # Importo
                    elif 'importo' in h_lower:
                        if isinstance(v, (int, float)):
                            importo = float(v)
                        else:
                            try:
                                importo = float(str(v).replace('.', '').replace(',', '.'))
                            except (ValueError, TypeError):
                                pass
                    
                    # Divisa
                    elif h_lower == 'divisa':
                        divisa = str(v).strip()
                    
                    # Descrizione
                    elif 'descri' in h_lower:
                        descrizione = str(v).strip()
                    
                    # Categoria
                    elif 'categoria' in h_lower:
                        categoria = str(v).strip()
                    
                    # Hashtag
                    elif h_lower == 'hashtag':
                        hashtag = str(v).strip()
                
                if data_contabile and importo is not None:
                    movimenti.append({
                        "data": data_contabile,
                        "ragione_sociale": ragione_sociale if ragione_sociale else None,
                        "fornitore": estrai_fornitore_pulito(descrizione),
                        "importo": importo,
                        "numero_fattura": estrai_numero_fattura(descrizione),
                        "data_pagamento": data_valuta,
                        "categoria": categoria,
                        "descrizione_originale": descrizione,
                        "banca": banca,
                        "rapporto": rapporto,
                        "divisa": divisa,
                        "hashtag": hashtag,
                        "tipo": "uscita" if importo < 0 else "entrata"
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa PDF, CSV o Excel.")
    
    # Salva nel database, evitando duplicati con un singolo query bulk
    import hashlib as _hashlib
    import uuid as _uuid
    inserted = 0
    duplicates = 0
    
    if not movimenti:
        return {
            "success": True,
            "stats": {"nuovi": 0, "duplicati": 0, "totale_letti": 0},
            "message": "Nessun movimento trovato nel file."
        }
    
    # Ordina per data contabile ASCENDENTE prima di inserire
    movimenti.sort(key=lambda x: x["data"].isoformat()[:10] if hasattr(x["data"], "isoformat") else str(x["data"]))
    
    # Recupera tutte le chiavi dedup già in DB nel range di date del CSV (1 sola query)
    date_nel_csv = sorted(set(
        m["data"].isoformat()[:10] if hasattr(m["data"], "isoformat") else str(m["data"])[:10]
        for m in movimenti
    ))
    data_min, data_max = date_nel_csv[0], date_nel_csv[-1]
    
    # Carica le chiavi dedup esistenti per quel range
    existing_cursor = db["estratto_conto_movimenti"].find(
        {"data": {"$gte": data_min, "$lte": data_max}},
        {"data": 1, "importo": 1, "descrizione_originale": 1, "descrizione": 1,
         "id": 1, "livello_evidenza": 1, "evidenza_bancaria_ufficiale": 1, "_id": 0}
    )
    # chiave whitespace-insensibile: gli export bancari variano gli spazi
    # interni ("NUMIA-INTER  DEL" vs "NUMIA-INTER DEL") e senza normalizzare
    # un re-import duplicherebbe centinaia di movimenti identici
    def _norm_desc(desc: str) -> str:
        return normalizza_descrizione_ec(desc)

    existing_by_key: Dict[Any, List[Dict[str, Any]]] = defaultdict(list)
    async for rec in existing_cursor:
        dstr = rec.get("data", "")[:10]
        imp  = abs(float(rec.get("importo", 0)))
        desc = _norm_desc(rec.get("descrizione_originale") or rec.get("descrizione") or "")
        existing_by_key[(dstr, round(imp, 2), desc)].append(rec)
    
    records_to_insert = []
    records_promossi = []
    incoming_occurrences: Counter = Counter()
    for mov in movimenti:
        data_str    = mov["data"].isoformat()[:10] if hasattr(mov["data"], "isoformat") else str(mov["data"])[:10]
        importo_abs = abs(mov["importo"])
        desc_raw    = _norm_desc(mov.get("descrizione_originale") or mov.get("descrizione") or "")
        
        # Determina tipo da segno importo
        tipo_mov = "entrata" if mov["importo"] >= 0 else "uscita"
        if any(kw in desc_raw.upper() for kw in ["DISPOSIZIONE", "VS.DISP", "ADD.TOT"]):
            tipo_mov = "uscita"
        if any(kw in desc_raw.upper() for kw in ["I24 AGENZIA ENTRATE", "BOLL.CBILL", "PAG. UTENZE"]):
            tipo_mov = "uscita"
        
        # Dedup uniforme: vale anche per commissioni e piccoli importi.
        dedup_key = (data_str, round(importo_abs, 2), desc_raw)
        
        incoming_occurrences[dedup_key] += 1
        occurrence = incoming_occurrences[dedup_key]
        existing_rows = existing_by_key.get(dedup_key, [])
        if len(existing_rows) >= occurrence:
            existing = existing_rows[occurrence - 1]
            if fonte_ufficiale and not (
                existing.get("evidenza_bancaria_ufficiale") is True
                or existing.get("livello_evidenza") == EVIDENZA_UFFICIALE
            ):
                records_promossi.append(existing)
            duplicates += 1
            continue
        
        row_uuid    = str(_uuid.uuid4())
        fingerprint = _hashlib.md5(f"{data_str}|{importo_abs:.2f}|{desc_raw}|{row_uuid}".encode()).hexdigest()
        mov_id      = f"EC-{data_str}-{importo_abs:.2f}-{fingerprint[:8]}"
        
        records_to_insert.append({
            "id": mov_id,
            "data": data_str,
            "ragione_sociale": mov.get("ragione_sociale"),
            "fornitore": mov.get("fornitore"),
            "importo": importo_abs,
            "numero_fattura": mov.get("numero_fattura"),
            "data_pagamento": mov["data_pagamento"].isoformat() if mov.get("data_pagamento") else None,
            "categoria": mov.get("categoria", ""),
            "descrizione_originale": mov["descrizione_originale"],
            "descrizione": mov.get("descrizione") or mov["descrizione_originale"],
            "banca": mov.get("banca"),
            "rapporto": mov.get("rapporto"),
            "divisa": mov.get("divisa", "EUR"),
            "hashtag": mov.get("hashtag"),
            "tipo": tipo_mov,
            "descrizione_hash": desc_raw[:50],
            "fingerprint": fingerprint,
            "riconciliato": False,
            "source_filename": filename_originale,
            "drive_file_id": drive_file_id,
            "drive_source_path": drive_source_path,
            **evidenza,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    promoted_ids = [record.get("id") for record in records_promossi if record.get("id")]
    if promoted_ids:
        await db["estratto_conto_movimenti"].update_many(
            {"id": {"$in": promoted_ids}},
            {"$set": {
                **evidenza,
                "source_filename_ufficiale": filename_originale,
                "drive_file_id": drive_file_id,
                "drive_source_path": drive_source_path,
                "riconciliato": False,
                "promosso_a_ufficiale_at": datetime.now(timezone.utc).isoformat(),
            }, "$unset": {
                "riconciliato_provvisoriamente": "",
            }},
        )
    
    # Inserimento bulk unico
    if records_to_insert:
        await db["estratto_conto_movimenti"].insert_many(records_to_insert, ordered=False)
        inserted = len(records_to_insert)

    has_material_changes = bool(records_to_insert or promoted_ids)

    riconciliazione_operativa = None
    if not fonte_ufficiale and records_to_insert:
        from app.services.riconciliazione_operativa_banca import annota_movimenti_operativi
        riconciliazione_operativa = await annota_movimenti_operativi(
            db, [record["id"] for record in records_to_insert]
        )

    
    # ===== RICONCILIAZIONE AUTOMATICA =====
    # Dopo l'import, avvia la riconciliazione automatica
    riconciliazione_results = None
    try:
        from app.services.riconciliazione_bancaria import riconcilia_movimenti_banca
        ids_da_riconciliare = list(dict.fromkeys(
            [record["id"] for record in records_to_insert if record.get("id")]
            + promoted_ids
        ))
        riconciliazione_results = (
            await riconcilia_movimenti_banca(
                movimento_ids=ids_da_riconciliare,
            )
            if fonte_ufficiale and has_material_changes
            else {
                "success": True,
                "skipped": True,
                "message": "Nessun nuovo movimento bancario da riconciliare",
            }
            if fonte_ufficiale
            else {
                "success": True,
                "provvisoria": True,
                "message": "Abbinamenti in attesa dell'estratto bancario ufficiale",
            }
        )
    except Exception as e:
        logger.error(f"Errore riconciliazione automatica: {e}")
        riconciliazione_results = {"error": str(e)}
    
    # ===== RICONCILIAZIONE AUTOMATICA PAGHE (Stipendi + F24) =====
    riconciliazione_paghe = None
    try:
        from app.services.paghe_riconciliazione import esegui_riconciliazione_paghe_completa
        riconciliazione_paghe = (
            await esegui_riconciliazione_paghe_completa(db)
            if fonte_ufficiale and has_material_changes
            else {"skipped": True, "riconciliati": 0}
            if fonte_ufficiale
            else {"provvisoria": True, "riconciliati": 0}
        )
    except Exception as e:
        logger.error(f"Errore riconciliazione paghe: {e}")
    
    # ===== RICONCILIAZIONE FATTURE PROVVISORIE =====
    # Cerca pagamenti nell'EC per fatture non ancora pagate.
    # Il match e la registrazione passano dall'helper condiviso (stessa logica
    # dell'import fatture): segno importo tollerato, finestra date, movimento
    # EC consumato (riconciliato=True), flag fattura coerenti.
    provvisori_riconciliati = 0
    try:
        import uuid as _uuid
        from app.routers.invoices.fatture_upload import find_ec_match_for_invoice
        provvisori = await db["invoices"].find({
            "total_amount": {"$gt": 0},
            # "sospesa" = bloccata manualmente in Prima Nota Provvisoria:
            # esclusa dal matching automatico, come in riconciliazione_bancaria.
            "stato_pagamento": {"$nin": ["pagata", "paid", "sospesa"]},
            "pagato": {"$ne": True},
            "$or": [{"prima_nota_id": None}, {"prima_nota_id": {"$exists": False}}, {"prima_nota_id": ""}]
        }, {"_id": 0, "id": 1, "supplier_name": 1, "supplier_vat": 1, "total_amount": 1,
            "invoice_date": 1, "invoice_number": 1, "tipo_documento": 1}).to_list(500)

        for f in provvisori if fonte_ufficiale and has_material_changes else []:
            importo = float(f.get("total_amount", 0))
            match = await find_ec_match_for_invoice(
                db, importo, f.get("supplier_name", ""), f.get("invoice_date", ""),
                f.get("invoice_number", ""),
            )
            if match:
                # Dedup: se esiste già un movimento prima nota per questa fattura, non duplicare
                rif = f"FATT-{f['id']}"
                existing_pn = await db["prima_nota_banca"].find_one({
                    "$or": [{"riferimento": rif}, {"fattura_id": f["id"]}],
                    "status": {"$nin": ["deleted", "archived"]},
                })
                if existing_pn:
                    pn_id = existing_pn.get("id")
                else:
                    pn_id = str(_uuid.uuid4())
                    # Nota di credito (TD04/TD08) NON è un pagamento a
                    # fornitore in uscita: stessa regola già applicata in
                    # prima_nota_module/sync.py (bug segnalato dall'utente
                    # 14/07/2026, qui era hardcoded "uscita"/"Fatture").
                    await scrivi_movimento(db, "banca", {
                        "id": pn_id, "data": match.get("data") or match.get("data_contabile") or f.get("invoice_date", ""),
                        **costruisci_campi_movimento_fattura(f, importo),
                        "fattura_id": f["id"],
                        "riferimento": rif,
                        "fornitore_piva": f.get("supplier_vat", ""),
                        "estratto_conto_id": match.get("id"),
                        "source": "riconciliazione_ec_auto",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                await db["invoices"].update_one({"id": f["id"]}, {"$set": {
                    "prima_nota_id": pn_id, "prima_nota_tipo": "banca", "prima_nota_banca_id": pn_id,
                    "stato_pagamento": "pagata", "pagato": True, "paid": True,
                    "data_pagamento": match.get("data") or match.get("data_contabile") or f.get("invoice_date"),
                    "riconciliato_con_ec": match.get("id"),
                }})
                await db["estratto_conto_movimenti"].update_one(
                    {"id": match.get("id")},
                    {"$set": {
                        "riconciliato": True,
                        "tipo_riconciliazione": "fattura_provvisoria",
                        "dettagli_riconciliazione": {"fattura_id": f["id"], "prima_nota_id": pn_id},
                    }}
                )
                provvisori_riconciliati += 1

                # --- EVENT BUS: propaga FATTURA_PAGATA (riconciliazione provvisori EC) ---
                try:
                    from app.services.event_bus import propagate_event, EventTypes
                    await propagate_event(EventTypes.FATTURA_PAGATA, {
                        "fattura_id": f["id"],
                        "metodo_pagamento": "banca",
                        "data_pagamento": match.get("data") or match.get("data_contabile") or f.get("invoice_date"),
                        "movimento_id": match.get("id"),
                        "importo": importo,
                    }, db, source_module="ec_riconciliazione_provvisori")
                except Exception:
                    logger.exception("Errore propagazione fattura.pagata (riconcilia provvisori EC)")

        if provvisori_riconciliati > 0:
            logger.info(f"[EC Import] Riconciliate {provvisori_riconciliati} fatture provvisorie con EC")
    except Exception as e:
        logger.error(f"Errore riconciliazione provvisori: {e}")
        riconciliazione_paghe = {"error": str(e)}

    # ===== SYNC ASSEGNI DA ESTRATTO CONTO =====
    # Prima era un bottone manuale nella pagina Assegni ("Sync da E/C"):
    # ora scatta automaticamente subito dopo ogni import dell'estratto conto.
    assegni_sync = None
    # Va eseguito anche quando il CSV e' interamente duplicato: in passato un
    # movimento gia importato ma non trasformato in assegno restava bloccato
    # per sempre, perche inserted=0 impediva la riparazione.
    try:
        from app.routers.bank.assegni import sync_assegni_da_estratto_conto
        assegni_sync = (
            await sync_assegni_da_estratto_conto(
                movimento_ids=ids_da_riconciliare,
            )
            if fonte_ufficiale and (has_material_changes or not skip_duplicate_repairs)
            else {
                "skipped": True,
                "message": "File Drive interamente duplicato: riparazione assegni non ripetuta",
            }
            if fonte_ufficiale
            else {"provvisoria": True, "assegni_riconciliati": 0}
        )
    except Exception as e:
        logger.error(f"Errore sync assegni da estratto conto: {e}")

    # ===== SYNC GENERICO IN PRIMA NOTA BANCA (+ CASSA per prelievi/versamenti) =====
    # Le fasi sopra (fatture/F24/stipendi/assegni) registrano in Prima Nota
    # solo i movimenti abbinati con certezza a un documento noto: tutti gli
    # altri restavano visibili solo nell'Estratto Conto e non comparivano mai
    # in Prima Nota Banca. Ora ogni movimento rimasto senza match viene
    # comunque registrato in Prima Nota Banca (categoria generica,
    # riclassificabile a mano in seguito), con provenienza tracciata
    # (estratto_conto_id + source) così un reimport non lo duplica e
    # un'eventuale associazione manuale successiva può aggiornare la riga
    # invece di crearne una seconda.
    # In più: prelievi bancomat e versamenti di contanti generano anche il
    # movimento speculare in Prima Nota Cassa (un prelievo fa entrare
    # contanti in cassa, un versamento li fa uscire dalla cassa verso banca).
    sync_generico = {"inseriti_banca": 0, "inseriti_cassa": 0}
    pos_storico = None
    if records_to_insert:
        try:
            ec_ids = [m["id"] for m in records_to_insert]
            stato_aggiornato: Dict[str, Any] = {}
            async for m in db["estratto_conto_movimenti"].find(
                {"id": {"$in": ec_ids}},
                {"_id": 0, "id": 1, "riconciliato": 1, "riconciliato_paghe": 1}
            ):
                stato_aggiornato[m["id"]] = m

            KEYWORDS_PRELIEVO = ["BANCOMAT", "CONTANT", "SPORTELLO", " ATM"]
            # P0-1 (verifica Contabilità): un accredito POS (NUMIA) è la stessa
            # moneta della quota "Corrispettivi POS" già registrata in
            # prima_nota_banca all'import del corrispettivo. Se lo inserissimo
            # anche qui come nuova entrata, il POS verrebbe contato DUE VOLTE nel
            # Bilancio. Perciò gli accrediti POS NON generano una nuova entrata:
            # marcano l'EC riconciliato e chiudono (best-effort) le entrate
            # sintetiche "Corrispettivi POS" ancora aperte.
            KEYWORDS_ACCREDITO_POS = ["NUMIA", "ACCREDITO POS", "INCASSO POS",
                                      "POS ACQUIRING", "ACCR. POS", "ACCRED POS"]

            banca_batch = []
            cassa_batch = []
            ec_da_marcare = []
            ec_pos_accrediti = []  # id EC accrediti POS (riconciliati senza duplicare)
            ec_in_attesa = []      # id EC che aspettano il documento a cui agganciarsi
            for mov in records_to_insert:
                mid = mov["id"]
                stato = stato_aggiornato.get(mid, {})
                if stato.get("riconciliato") or stato.get("riconciliato_paghe"):
                    continue  # già registrato/gestito dalle fasi precedenti

                desc_upper = (mov.get("descrizione_originale") or mov.get("descrizione") or "").upper()
                now_iso = datetime.now(timezone.utc).isoformat()

                # Il versamento nasce dalla registrazione manuale in Cassa,
                # non dalla sola lettura dell'estratto conto. Se la gamba
                # manuale manca, la riga resta da verificare e non inventiamo
                # due movimenti. Gli storni non sono mai nuovi versamenti.
                if is_versamento_contanti(desc_upper) or is_storno_versamento(desc_upper):
                    continue

                # REGOLA CANONICA POS (utente 18/07/2026): l'accredito POS
                # dell'estratto conto NON crea un'entrata — RICONCILIA il
                # trasferimento cassa→banca del suo giorno di vendita,
                # sommando i circuiti della stessa giornata (BNCMT, AMEX,
                # INTER, per ogni punto vendita).
                #
                # Vale anche per l'export CSV (07/08/2026, richiesta utente):
                # prima scattava solo col PDF ufficiale, e nel frattempo le
                # righe NUMIA del CSV finivano in Prima Nota come "Rimborso" —
                # entrate mai avvenute, visto che il denaro era gia' contato
                # nel trasferimento. Se poi arriva il PDF, la promozione a
                # ufficiale azzera `riconciliato` e il gruppo viene riverificato
                # con l'evidenza piena.
                if mov.get("tipo") == "entrata" and any(k in desc_upper for k in KEYWORDS_ACCREDITO_POS):
                    try:
                        from app.services.scritture_contabili import riconcilia_accredito_pos_ec
                        if await riconcilia_accredito_pos_ec(db, mov):
                            ec_pos_accrediti.append(mid)
                    except Exception as e:
                        logger.warning(f"Accredito POS non riconciliato ({mid}): {e}")
                    continue

                categoria_pn = (mappa_categoria_ec(mov.get("categoria"), desc_upper)
                                or mov.get("categoria") or "Altro")

                # Un pagamento che dovrebbe avere un documento e non l'ha
                # trovato NON entra in Prima Nota: resta nella coda da
                # riconciliare, dove lo si aggancia alla sua fattura o al suo
                # cedolino. Prima finiva qui come riga grezza "da verificare",
                # ed e' il motivo per cui la Prima Nota Banca sembrava una
                # fotocopia dell'estratto conto.
                if not entra_in_prima_nota(categoria_pn):
                    ec_in_attesa.append(mid)
                    continue

                banca_batch.append({
                    "id": str(_uuid.uuid4()),
                    "data": mov["data"],
                    "tipo": mov["tipo"],
                    "importo": mov["importo"],
                    "descrizione": mov.get("descrizione") or mov.get("descrizione_originale") or "",
                    # in Prima Nota entra il nome operativo; l'originale
                    # bancario resta sulla riga di estratto conto
                    "categoria": categoria_pn,
                    "estratto_conto_id": mid,
                    "source": "estratto_conto_auto" if fonte_ufficiale else "export_bancario_operativo",
                    "provvisorio": not fonte_ufficiale,
                    "riconciliato": False,
                    "in_attesa_estratto_ufficiale": not fonte_ufficiale,
                    "stato": "da_verificare" if fonte_ufficiale else "in_attesa_estratto_bancario_ufficiale",
                    "created_at": now_iso,
                })

                is_prelievo = "PRELIEVO" in desc_upper and any(k in desc_upper for k in KEYWORDS_PRELIEVO)
                is_versamento = is_versamento_contanti(desc_upper)
                if fonte_ufficiale and is_prelievo:
                    cassa_batch.append({
                        "id": str(_uuid.uuid4()), "data": mov["data"], "tipo": "entrata",
                        "importo": mov["importo"],
                        "descrizione": f"Prelevamento da banca - {(mov.get('descrizione') or '')[:100]}",
                        "categoria": "Prelevamento Banca",
                        "estratto_conto_id": mid,
                        "source": "estratto_conto_auto_prelievo",
                        "created_at": now_iso,
                    })
                elif fonte_ufficiale and is_versamento:
                    cassa_batch.append({
                        "id": str(_uuid.uuid4()), "data": mov["data"], "tipo": "uscita",
                        "importo": mov["importo"],
                        "descrizione": f"Versamento in banca - {(mov.get('descrizione') or '')[:100]}",
                        "categoria": "Versamento Banca",
                        "estratto_conto_id": mid,
                        "source": "estratto_conto_auto_versamento",
                        "created_at": now_iso,
                    })

                ec_da_marcare.append(mid)

            for _mov_batch in banca_batch:
                await scrivi_movimento(db, "banca", _mov_batch)
            for _mov_batch in cassa_batch:
                await scrivi_movimento(db, "cassa", _mov_batch)
            if ec_da_marcare:
                await db["estratto_conto_movimenti"].update_many(
                    {"id": {"$in": ec_da_marcare}},
                    {"$set": {
                        "importato_prima_nota": True,
                        "riconciliato": False,
                        "stato_riconciliazione": (
                            "da_verificare" if fonte_ufficiale
                            else "in_attesa_estratto_bancario_ufficiale"
                        ),
                    }, "$unset": {"tipo_riconciliazione": ""}}
                )
            if ec_in_attesa:
                # Marcati, non nascosti: sono il motivo per cui la Prima Nota
                # e il saldo della banca non coincidono ancora, e devono
                # potersi contare.
                await db["estratto_conto_movimenti"].update_many(
                    {"id": {"$in": ec_in_attesa}},
                    {"$set": {
                        "importato_prima_nota": False,
                        "riconciliato": False,
                        "stato_riconciliazione": "in_attesa_documento",
                    }}
                )
            # (le entrate banca degli accrediti POS sono già state create
            # dal motore unico riga per riga, con marcatura EC inclusa)
            sync_generico = {
                "inseriti_banca": len(banca_batch),
                "inseriti_cassa": len(cassa_batch),
                "accrediti_pos_riconciliati": len(ec_pos_accrediti),
                "in_attesa_documento": len(ec_in_attesa),
            }
        except Exception as e:
            logger.error(f"Errore sync generico estratto conto -> prima nota: {e}")

    # Recupero storico richiesto dall'utente: dopo avere salvato tutte le
    # righe, somma i circuiti NUMIA dello stesso giorno vendita (causale DEL)
    # e popola il POS provvisorio solo se non esiste gia' il totale manuale.
    # L'XML RT non partecipa mai a questa scrittura.
    if fonte_ufficiale and records_to_insert:
        try:
            from app.services.scritture_contabili import recupera_pos_storico_da_estratto
            anni_pos = sorted({
                int(str(m.get("data") or "")[:4])
                for m in records_to_insert
                if str(m.get("data") or "")[:4].isdigit()
            })
            esiti_pos = [
                await recupera_pos_storico_da_estratto(db, anno_pos)
                for anno_pos in anni_pos
            ]
            pos_storico = {"anni": esiti_pos}
        except Exception as e:
            logger.error(f"Errore recupero storico POS da estratto conto: {e}")

    # ── EVENTO: pubblica sul bus unico per matching automatico ──
    try:
        from app.services.event_bus import propagate_event, EventTypes
        if fonte_ufficiale:
            await propagate_event(EventTypes.ESTRATTO_CONTO_IMPORTATO, {
                "movimenti": [
                    {
                        "id":          m.get("id"),
                        "data":        str(m.get("data", ""))[:10],
                        "importo":     abs(float(m.get("importo", 0))),
                        "tipo":        m.get("tipo", ""),
                        "descrizione": m.get("descrizione", ""),
                    }
                    for m in records_to_insert
                ],
                "banca": records_to_insert[0].get("banca", "") if records_to_insert else "",
                "inseriti": inserted,
            }, db, source_module="estratto_conto_import")
    except Exception as _ev:
        logger.debug(f"[EstrattoContoRouter] Event Bus: {_ev}")

    nexi_verifica = None
    try:
        from app.services.nexi_carta import verifica_addebiti_nexi
        nexi_verifica = await verifica_addebiti_nexi(db)
    except Exception as _nexi_err:
        logger.debug(f"[EstrattoContoRouter] Verifica Nexi: {_nexi_err}")

    return {
        "success": True,
        "message": "Importazione estratto conto completata",
        "nexi_verifica": nexi_verifica,
        "movimenti_trovati": len(movimenti),
        "movimenti_importati": inserted,
        "inseriti": inserted,
        "duplicati_saltati": duplicates,
        "movimenti_promossi_a_ufficiali": len(promoted_ids),
        "livello_evidenza": evidenza["livello_evidenza"],
        "in_attesa_estratto_ufficiale": not fonte_ufficiale,
        "stats": {
            "nuovi": inserted,
            "duplicati": duplicates,
            "totale_letti": len(movimenti),
        },
        "riconciliazione_automatica": riconciliazione_results,
        "riconciliazione_operativa": riconciliazione_operativa,
        "riconciliazione_summary": (riconciliazione_results or {}).get("summary"),
        "riconciliazione_paghe": riconciliazione_paghe,
        "provvisori_riconciliati": provvisori_riconciliati,
        "assegni_sync": assegni_sync,
        "sync_prima_nota": sync_generico,
        "recupero_pos_storico": pos_storico,
    }


@router.post("/pulizia-non-in-csv")
@handle_errors
async def pulizia_movimenti_non_in_csv(
    file: UploadFile = File(...),
    dry_run: bool = Query(True, description="Sola lettura; false è bloccato"),
    _admin: Dict[str, Any] = Depends(get_current_admin_user),
) -> Dict[str, Any]:
    """Confronta l'export ufficiale con la fonte importata, in sola lettura.

    I movimenti bancari sono evidenze immutabili: l'assenza da un CSV non
    autorizza cancellazioni, scollegamenti o riaperture automatiche. Il
    risultato è un report di anomalie da esaminare puntualmente.
    """
    from collections import Counter

    if not dry_run:
        raise HTTPException(
            status_code=409,
            detail=(
                "Operazione bloccata: il confronto con il CSV è solo lettura; "
                "i movimenti bancari e i collegamenti contabili restano immutati"
            ),
        )

    db = Database.get_db()
    contents = await file.read()
    text = None
    for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            text = contents.decode(encoding)
            break
        except (UnicodeDecodeError, Exception):
            continue
    if not text:
        raise HTTPException(status_code=400, detail="Impossibile decodificare il file CSV")

    def _chiave_desc(descr: str) -> str:
        # gli export della banca variano gli spazi interni ("NUMIA-INTER  DEL"
        # vs "NUMIA-INTER DEL"): la chiave di confronto è whitespace-insensibile
        return re.sub(r"\s+", " ", (descr or "").strip())[:80]

    attesi: Counter = Counter()
    date_csv = []
    has_entrate = has_uscite = False
    for row in csv.DictReader(io.StringIO(text), delimiter=';'):
        data_it = (row.get('Data contabile') or row.get('Data') or '').strip().strip('"')
        importo_str = (row.get('Importo') or '').strip().strip('"').replace('.', '').replace(',', '.')
        descr = (row.get('Descrizione') or '').strip().strip('"')
        try:
            parts = data_it.split('/')
            data_iso = f"{int(parts[2]):04d}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            importo = float(importo_str)
        except (ValueError, TypeError, IndexError):
            continue
        if importo >= 0:
            has_entrate = True
        else:
            has_uscite = True
        date_csv.append(data_iso)
        attesi[(data_iso, round(abs(importo), 2), _chiave_desc(descr))] += 1

    if not date_csv:
        raise HTTPException(status_code=400, detail="Nessun movimento leggibile nel CSV")
    data_min, data_max = min(date_csv), max(date_csv)

    movimenti_db = await db["estratto_conto_movimenti"].find(
        {"data": {"$gte": data_min, "$lte": data_max}},
        {"_id": 0, "id": 1, "data": 1, "importo": 1, "tipo": 1, "banca": 1,
         "descrizione_originale": 1, "descrizione": 1, "riconciliato": 1},
    ).sort("created_at", 1).to_list(50000)

    rimasti = Counter(attesi)
    da_eliminare = []
    esempi = []
    esclusi_altra_banca = 0
    for m in movimenti_db:
        # nella stessa collezione vivono anche gli estratti PayPal e di altre
        # banche (banca="PayPal (Europe)…"): l'export Banco BPM non può mai
        # decidere la loro sorte
        banca_mov = (m.get("banca") or "").upper()
        if banca_mov and "BPM" not in banca_mov and "BANCO" not in banca_mov:
            esclusi_altra_banca += 1
            continue
        e_uscita = m.get("tipo") == "uscita" or float(m.get("importo") or 0) < 0
        if (e_uscita and not has_uscite) or ((not e_uscita) and not has_entrate):
            continue  # segno non coperto dall'export: mai toccato
        chiave = ((m.get("data") or "")[:10],
                  round(abs(float(m.get("importo") or 0)), 2),
                  _chiave_desc(m.get("descrizione_originale") or m.get("descrizione") or ""))
        if rimasti.get(chiave, 0) > 0:
            rimasti[chiave] -= 1  # presente nel CSV: resta (link intatti)
        else:
            da_eliminare.append(m)
            if len(esempi) < 20:
                esempi.append({"data": chiave[0], "importo": chiave[1],
                               "descrizione": chiave[2][:60],
                               "riconciliato": bool(m.get("riconciliato"))})

    mancanti_nel_db = sum(v for v in rimasti.values() if v > 0)
    return {
        "dry_run": True,
        "sola_lettura": True,
        "intervallo": [data_min, data_max],
        "segni_nel_csv": {"entrate": has_entrate, "uscite": has_uscite},
        "movimenti_csv": len(date_csv),
        "movimenti_db_in_scope": len(movimenti_db),
        "anomalie_non_presenti_nel_csv": len(da_eliminare),
        "esclusi_altra_banca_o_paypal": esclusi_altra_banca,
        "movimenti_modificati": 0,
        "collegamenti_modificati": 0,
        "mancanti_nel_db_da_importare": mancanti_nel_db,
        "esempi": esempi,
    }


@router.post("/force-reimport")
@router.post("/reimport")  # alias onesto: NON cancella nulla (vedi P0.6)
@handle_errors
async def force_reimport_estratto_conto(file: UploadFile = File(...), _admin: Dict[str, Any] = Depends(get_current_admin_user)) -> Dict[str, Any]:
    """
    Re-import ADDITIVO dell'estratto conto da CSV (NON distruttivo).
    - NON cancella alcun record: i movimenti esistenti e le riconciliazioni
      restano intatti (`record_cancellati` è sempre 0).
    - Inserisce solo i movimenti NUOVI, saltando i duplicati per fingerprint.
    - Sincronizza gli assegni dai nuovi movimenti.

    NB (P0.6): il nome storico "force-reimport" è fuorviante — l'endpoint non
    forza né sovrascrive; è un import additivo con deduplica. Il contratto della
    risposta è quello reale (nessuna cancellazione).
    """
    import hashlib as _hashlib
    import uuid as _uuid

    db = Database.get_db()
    
    filename = file.filename.lower()
    contents = await file.read()
    
    if not filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa CSV.")
    
    # Decode
    text = None
    for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            text = contents.decode(encoding)
            break
        except (UnicodeDecodeError, Exception):
            continue
    
    if not text:
        raise HTTPException(status_code=400, detail="Impossibile decodificare il file CSV")
    
    reader = csv.DictReader(io.StringIO(text), delimiter=';')
    movimenti = []
    
    for row in reader:
        ragione_sociale = row.get('Ragione Sociale', '') or row.get('"Ragione Sociale"', '')
        
        data_contabile = (
            row.get('Data contabile', '') or 
            row.get('"Data contabile"', '') or
            row.get('Data', '')
        ).strip().strip('"')
        
        data_valuta = (
            row.get('Data valuta', '') or 
            row.get('"Data valuta"', '') or
            row.get('Data valut.', '')
        ).strip().strip('"')
        
        banca = (row.get('Banca', '') or row.get('"Banca"', '')).strip().strip('"')
        rapporto = (row.get('Rapporto', '') or row.get('"Rapporto"', '')).strip().strip('"')
        
        importo_str = (row.get('Importo', '0') or row.get('"Importo"', '0')).strip().strip('"')
        importo_str = importo_str.replace('.', '').replace(',', '.')
        try:
            importo = float(importo_str)
        except (ValueError, TypeError):
            continue
        
        divisa = (row.get('Divisa', 'EUR') or row.get('"Divisa"', 'EUR')).strip().strip('"')
        
        descrizione = (
            row.get('Descrizione', '') or 
            row.get('"Descrizione"', '') or
            row.get('Descrizion', '')
        ).strip().strip('"')
        
        categoria = (
            row.get('Categoria/sottocategoria', '') or 
            row.get('"Categoria/sottocategoria"', '') or
            row.get('Categoria', '') or
            row.get('"Categoria"', '')
        ).strip().strip('"')
        
        hashtag = (row.get('Hashtag', '') or row.get('"Hashtag"', '')).strip().strip('"')
        
        try:
            if '/' in data_contabile:
                parts = data_contabile.split('/')
                data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                continue
        except (ValueError, TypeError, IndexError):
            continue
        
        data_pagamento = None
        try:
            if '/' in data_valuta:
                parts = data_valuta.split('/')
                data_pagamento = date(int(parts[2]), int(parts[1]), int(parts[0]))
        except (ValueError, TypeError, IndexError):
            pass
        
        movimenti.append({
            "data": data_obj,
            "ragione_sociale": ragione_sociale.strip().strip('"') if ragione_sociale else None,
            "fornitore": estrai_fornitore_pulito(descrizione),
            "importo": importo,
            "numero_fattura": estrai_numero_fattura(descrizione),
            "data_pagamento": data_pagamento,
            "categoria": categoria,
            "descrizione_originale": descrizione,
            "banca": banca,
            "rapporto": rapporto,
            "divisa": divisa,
            "hashtag": hashtag,
            "tipo": "uscita" if importo < 0 else "entrata"
        })
    
    if not movimenti:
        raise HTTPException(status_code=400, detail="Nessun movimento valido trovato nel CSV")
    
    # SICUREZZA: NON cancellare MAI i record esistenti.
    # Se il CSV sovrappone un periodo già importato, inserisci SOLO i nuovi.
    # I record già riconciliati NON vengono toccati.
    
    # Calcola range date del CSV
    date_nel_csv = sorted(set(
        mov["data"].isoformat()[:10] for mov in movimenti
    ))
    data_min_csv, data_max_csv = date_nel_csv[0], date_nel_csv[-1]
    
    # Carica chiavi dedup esistenti per il range del CSV
    existing_counts: Counter = Counter()
    existing_cursor = db["estratto_conto_movimenti"].find(
        {"data": {"$gte": data_min_csv, "$lte": data_max_csv}},
        {"data": 1, "importo": 1, "descrizione_originale": 1, "descrizione": 1, "_id": 0}
    )
    async for rec in existing_cursor:
        dstr = rec.get("data", "")[:10]
        imp = abs(float(rec.get("importo", 0)))
        desc = (rec.get("descrizione_originale") or rec.get("descrizione") or "")[:80]
        existing_counts[(dstr, round(imp, 2), desc)] += 1
    
    cancellati = 0  # Non cancelliamo nulla
    
    # Ordina per data contabile ascendente
    movimenti.sort(key=lambda x: x["data"].isoformat()[:10])
    
    # Inserisce SOLO i nuovi (dedup con chiavi esistenti)
    records = []
    duplicates_skipped = 0
    incoming_occurrences: Counter = Counter()
    for mov in movimenti:
        data_str = mov["data"].isoformat()[:10]
        importo_abs = abs(mov["importo"])
        desc_raw = (mov.get("descrizione_originale") or "")[:80]
        
        # Dedup: se esiste già, salta
        dedup_key = (data_str, round(importo_abs, 2), desc_raw)
        if _occorrenza_gia_importata(existing_counts, incoming_occurrences, dedup_key):
            duplicates_skipped += 1
            continue
        
        tipo_mov = "entrata" if mov["importo"] >= 0 else "uscita"
        if any(kw in desc_raw.upper() for kw in ["DISPOSIZIONE", "VS.DISP", "ADD.TOT", "I24 AGENZIA ENTRATE", "BOLL.CBILL", "PAG. UTENZE"]):
            tipo_mov = "uscita"
        
        row_uuid = str(_uuid.uuid4())
        fingerprint = _hashlib.md5(f"{data_str}|{importo_abs:.2f}|{desc_raw}|{row_uuid}".encode()).hexdigest()
        mov_id = f"EC-{data_str}-{importo_abs:.2f}-{fingerprint[:8]}"
        
        record = {
            "id": mov_id,
            "data": data_str,
            "ragione_sociale": mov.get("ragione_sociale"),
            "fornitore": mov.get("fornitore"),
            "importo": importo_abs,
            "numero_fattura": mov.get("numero_fattura"),
            "data_pagamento": mov["data_pagamento"].isoformat() if mov.get("data_pagamento") else None,
            "categoria": mov.get("categoria", ""),
            "descrizione_originale": mov["descrizione_originale"],
            "descrizione": mov.get("descrizione_originale"),
            "banca": mov.get("banca"),
            "rapporto": mov.get("rapporto"),
            "divisa": mov.get("divisa", "EUR"),
            "hashtag": mov.get("hashtag"),
            "tipo": tipo_mov,
            "descrizione_hash": desc_raw[:50],
            "fingerprint": fingerprint,
            "riconciliato": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        records.append(record)
    
    if records:
        await db["estratto_conto_movimenti"].insert_many(records)

    # Sync assegni: prima era un bottone manuale, ora scatta ad ogni import.
    assegni_sync = None
    try:
        from app.routers.bank.assegni import sync_assegni_da_estratto_conto
        assegni_sync = await sync_assegni_da_estratto_conto()
    except Exception as e:
        logger.error(f"Errore sync assegni da estratto conto: {e}")

    # Calcola statistiche per la risposta
    entrate = sum(r["importo"] for r in records if r["tipo"] == "entrata")
    uscite = sum(r["importo"] for r in records if r["tipo"] == "uscita")

    return {
        "message": f"Import completato: solo nuovi movimenti aggiunti",
        "periodo_csv": f"{data_min_csv} → {data_max_csv}",
        "record_nel_csv": len(movimenti),
        "duplicati_saltati": duplicates_skipped,
        "movimenti_nuovi_importati": len(records),
        "record_cancellati": 0,
        "totale_entrate": round(entrate, 2),
        "totale_uscite": round(uscite, 2),
        "assegni_sync": assegni_sync,
        "saldo": round(entrate - uscite, 2),
        "nota": "I record esistenti e le riconciliazioni NON sono stati toccati"
    }


@router.get("/movimenti")
@handle_errors
async def get_movimenti(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),  # "entrata" | "uscita"
    limit: int = Query(500, le=10000),
    offset: int = Query(0)
) -> Dict[str, Any]:
    """
    Recupera i movimenti dell'estratto conto con filtri.
    Ordinati per data decrescente.

    NOTA: il filtro usa il campo stringa `data` (YYYY-MM-DD), che è quello
    scritto da TUTTI gli importer (parser universale, email monitor, API).
    Il vecchio filtro su `data_contabile_obj` (datetime) non trovava nulla
    perché nessun importer valorizza quel campo.
    """
    import calendar as _calendar
    db = Database.get_db()

    query = {}

    if anno:
        # `data` è una stringa YYYY-MM-DD: il range lessicografico è corretto
        if mese:
            last_day = _calendar.monthrange(anno, mese)[1]
            query["data"] = {
                "$gte": f"{anno}-{mese:02d}-01",
                "$lte": f"{anno}-{mese:02d}-{last_day:02d}"
            }
        else:
            query["data"] = {
                "$gte": f"{anno}-01-01",
                "$lte": f"{anno}-12-31"
            }

    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}

    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}

    if tipo:
        query["tipo"] = tipo

    # Count totale
    total = await db["estratto_conto_movimenti"].count_documents(query)

    # Recupera movimenti — più recenti prima
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).skip(offset).limit(limit).to_list(limit)

    # Serializza i campi datetime per la risposta JSON
    for m in movimenti:
        if isinstance(m.get("data_contabile_obj"), datetime):
            m["data_contabile_obj"] = m["data_contabile_obj"].strftime("%Y-%m-%d")
        if isinstance(m.get("imported_at"), datetime):
            m["imported_at"] = m["imported_at"].isoformat()
        # Converti importo in float (arriva come stringa in alcuni doc)
        try:
            m["importo"] = float(m.get("importo", 0))
        except (ValueError, TypeError):
            m["importo"] = 0.0
        # Categoria operativa per la vista Banca (calcolata, non salvata)
        canonica = mappa_categoria_ec(
            m.get("categoria"),
            m.get("descrizione_originale") or m.get("descrizione"),
        )
        if canonica:
            m["categoria_canonica"] = canonica

    # §6.4: totali/riporto/saldo dalla funzione UNICA di saldo (stessa formula
    # di cassa/banca). query_base_precedente={} riproduce il comportamento
    # storico dell'estratto conto: il riporto considera TUTTI i movimenti
    # prima dell'anno (qui non esistono soft-delete né categorie escluse).
    saldi = await aggrega_saldo_prima_nota(
        db, "estratto_conto_movimenti", query, anno, query_base_precedente={}
    )

    return {
        "movimenti": movimenti,
        "totale": total,
        "offset": offset,
        "limit": limit,
        "totale_entrate": saldi["totale_entrate"],
        "totale_uscite": saldi["totale_uscite"],
        "saldo_anno": saldi["saldo_anno"],
        "saldo_precedente": saldi["saldo_precedente"],
        "saldo": saldi["saldo"],
        "anno": anno
    }


@router.get("/categorie")
@handle_errors
async def get_categorie() -> List[str]:
    """Restituisce lista categorie uniche."""
    db = Database.get_db()
    categorie = await db["estratto_conto_movimenti"].distinct("categoria")
    return sorted([c for c in categorie if c])


@router.get("/fornitori")
@handle_errors
async def get_fornitori_unici() -> List[str]:
    """Restituisce lista fornitori unici."""
    db = Database.get_db()
    fornitori = await db["estratto_conto_movimenti"].distinct("fornitore")
    return sorted([f for f in fornitori if f])


@router.get("/riepilogo")
@handle_errors
async def get_riepilogo(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo estratto conto con filtri."""
    db = Database.get_db()
    
    query = {}
    if anno:
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
        else:
            query["data"] = {"$regex": f"^{anno}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    total = await db["estratto_conto_movimenti"].count_documents(query)
    
    # Totali per tipo
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$tipo",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }}
    ]
    by_tipo = await db["estratto_conto_movimenti"].aggregate(pipeline).to_list(10)
    
    entrate = next((t for t in by_tipo if t["_id"] == "entrata"), {"totale": 0, "count": 0})
    uscite = next((t for t in by_tipo if t["_id"] == "uscita"), {"totale": 0, "count": 0})
    
    # Movimenti per categoria (top 10)
    pipeline_cat = [
        {"$match": query},
        {"$group": {
            "_id": "$categoria",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"totale": -1}},
        {"$limit": 10}
    ]
    by_categoria = await db["estratto_conto_movimenti"].aggregate(pipeline_cat).to_list(10)
    
    return {
        "totale_movimenti": total,
        "entrate": {"count": entrate["count"], "totale": round(entrate["totale"], 2)},
        "uscite": {"count": uscite["count"], "totale": round(uscite["totale"], 2)},
        "saldo": round(entrate["totale"] - uscite["totale"], 2),
        "per_categoria": [{"categoria": c["_id"] or "N/D", "totale": round(c["totale"], 2), "count": c["count"]} for c in by_categoria]
    }


@router.delete("/clear")
@handle_errors
async def clear_estratto_conto(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """La fonte bancaria e' immutabile: lo svuotamento non e' consentito."""
    raise HTTPException(
        status_code=409,
        detail="Operazione bloccata: i movimenti bancari non si cancellano; usare esclusione o archivio duplicato",
    )


@router.delete("/{movimento_id}")
@handle_errors
async def elimina_singolo_movimento(movimento_id: str) -> Dict[str, Any]:
    """La fonte bancaria e' immutabile: nessuna cancellazione fisica."""
    raise HTTPException(
        status_code=409,
        detail="Operazione bloccata: escludere il movimento dalla coda conservando la fonte",
    )


@router.get("/export-excel")
@handle_errors
async def export_estratto_conto_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None)
):
    """
    Esporta i movimenti dell'estratto conto in formato Excel.
    Applica gli stessi filtri della visualizzazione.
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    db = Database.get_db()
    
    # Costruisci query con filtri
    query = {}
    
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    # Recupera tutti i movimenti (senza paginazione per export)
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).to_list(10000)
    
    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estratto Conto"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Data", "Fornitore", "Importo (€)", "Tipo", "N. Fattura", "Data Pag.", "Categoria"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Larghezze colonne
    col_widths = [12, 35, 15, 10, 30, 12, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Dati
    entrata_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    uscita_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    totale_entrate = 0
    totale_uscite = 0
    
    for row_num, mov in enumerate(movimenti, 2):
        # Formatta data
        data_str = mov.get("data", "")
        if data_str:
            try:
                parts = data_str.split("-")
                data_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except (ValueError, TypeError, IndexError):
                data_formatted = data_str
        else:
            data_formatted = ""
        
        data_pag = mov.get("data_pagamento", "")
        if data_pag:
            try:
                parts = data_pag.split("-")
                data_pag_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except (ValueError, TypeError, IndexError):
                data_pag_formatted = data_pag
        else:
            data_pag_formatted = ""
        
        importo = mov.get("importo", 0)
        tipo_mov = "Entrata" if importo >= 0 else "Uscita"
        
        if importo >= 0:
            totale_entrate += importo
        else:
            totale_uscite += abs(importo)
        
        row_data = [
            data_formatted,
            mov.get("fornitore") or "",
            abs(importo),
            tipo_mov,
            mov.get("numero_fattura") or "",
            data_pag_formatted,
            mov.get("categoria") or ""
        ]
        
        row_fill = entrata_fill if importo >= 0 else uscita_fill
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = row_fill
            
            # Formato numerico per importo
            if col == 3:
                cell.number_format = '#,##0.00'
    
    # Riga totali
    last_row = len(movimenti) + 2
    totals_row = last_row + 1
    
    ws.cell(row=totals_row, column=1, value="TOTALI")
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=2, value=f"Entrate: € {totale_entrate:,.2f}")
    ws.cell(row=totals_row, column=2).font = Font(bold=True, color="16A34A")
    
    ws.cell(row=totals_row, column=3, value=f"Uscite: € {totale_uscite:,.2f}")
    ws.cell(row=totals_row, column=3).font = Font(bold=True, color="DC2626")
    
    saldo = totale_entrate - totale_uscite
    ws.cell(row=totals_row, column=4, value=f"Saldo: € {saldo:,.2f}")
    ws.cell(row=totals_row, column=4).font = Font(bold=True, color="16A34A" if saldo >= 0 else "DC2626")
    
    # Salva in memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome file
    filename_parts = ["estratto_conto"]
    if anno:
        filename_parts.append(str(anno))
    if mese:
        mesi_nomi = ['gen', 'feb', 'mar', 'apr', 'mag', 'giu', 'lug', 'ago', 'set', 'ott', 'nov', 'dic']
        filename_parts.append(mesi_nomi[mese - 1])
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ==================== RICONCILIAZIONE STIPENDI ====================

@router.post("/riconcilia-stipendi")
@handle_errors
async def riconcilia_stipendi_automatico(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """
    Riconcilia automaticamente i bonifici stipendio con la prima nota salari.
    Cerca i movimenti "VOSTRA DISPOSIZIONE" con nomi di dipendenti e li collega.
    """
    db = Database.get_db()
    # Unico motore autorizzato: nome completo nella causale, importo esatto,
    # periodo/data compatibili e candidato univoco. Il vecchio percorso per
    # singola parola/cognome non deve mai essere raggiunto.
    from app.services.stipendi_bonifici import associa_bonifici_stipendi
    return await associa_bonifici_stipendi(db, anno=anno)
    
    # Carica nomi dipendenti dalla prima_nota_salari (fonte più affidabile)
    nomi_dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    dipendenti_map = {}
    for nome in nomi_dipendenti:
        if nome:
            nome_upper = nome.strip().upper()
            dipendenti_map[nome_upper] = {"nome": nome}
            # Aggiungi anche versione invertita (Nome Cognome)
            parts = nome_upper.split()
            if len(parts) >= 2:
                nome_inv = " ".join(reversed(parts))
                dipendenti_map[nome_inv] = {"nome": nome}
            # Aggiungi anche singoli cognomi/nomi (per match parziale)
            for p in parts:
                if len(p) > 3:  # Solo parole significative
                    dipendenti_map[p] = {"nome": nome}
    
    # Se vuota, prova con collezione employees (nome corretto)
    if not dipendenti_map:
        dipendenti = await db["dipendenti"].find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)
        for d in dipendenti:
            nome_completo = d.get("nome_completo") or f"{d.get('cognome', '')} {d.get('nome', '')}".strip().upper()
            nome_inv = f"{d.get('nome', '')} {d.get('cognome', '')}".strip().upper()
            if nome_completo:
                dipendenti_map[nome_completo.upper()] = d
            if nome_inv:
                dipendenti_map[nome_inv] = d
    
    logger.info(f"Caricati {len(dipendenti_map)} nomi dipendenti")
    
    # Cerca i bonifici "VOSTRA DISPOSIZIONE" con nome dipendente non ancora riconciliati
    query = {
        "descrizione": {"$regex": "VOSTRA DISPOSIZIONE.*FAVORE|FAVORE.*", "$options": "i"},
        "$or": [
            {"riconciliato_salario": {"$exists": False}},
            {"riconciliato_salario": False}
        ]
    }
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    bonifici = await db["estratto_conto_movimenti"].find(query, {"_id": 0}).to_list(5000)
    
    riconciliati = 0
    non_trovati = []
    
    for bonifico in bonifici:
        desc = bonifico.get("descrizione", "").upper()
        importo = abs(bonifico.get("importo", 0))
        data = bonifico.get("data", "")
        
        # Estrai nome dopo "FAVORE"
        nome_trovato = None
        dip_trovato = None
        
        if "FAVORE" in desc:
            idx = desc.find("FAVORE")
            after = desc[idx + 7:].strip()
            # Pulisci fino a separatori comuni
            for sep in [" - ", " ADD", " NOTPROVIDE", "/", "COMM."]:
                if sep in after:
                    after = after[:after.find(sep)].strip()
                    break
            
            after_clean = after.strip()
            
            # Cerca il dipendente con vari metodi
            # 1. Match esatto
            if after_clean in dipendenti_map:
                dip_trovato = dipendenti_map[after_clean]
                nome_trovato = after_clean
            else:
                # 2. Match per singola parola del nome estratto (cerca cognome)
                after_parts = [p.strip() for p in after_clean.split() if len(p.strip()) > 2]
                for part in after_parts:
                    if part in dipendenti_map:
                        dip_trovato = dipendenti_map[part]
                        nome_trovato = part
                        break
                
                # 3. Match parziale se non trovato
                if not dip_trovato:
                    for nome, dip in dipendenti_map.items():
                        if nome in after_clean or after_clean in nome:
                            dip_trovato = dip
                            nome_trovato = nome
                            break
                        # Match per cognome
                        parts = nome.split()
                        for p in parts:
                            if len(p) > 3 and p in after_clean:
                                dip_trovato = dip
                                nome_trovato = nome
                                break
                        if dip_trovato:
                            break
        
        if dip_trovato:
            # Aggiorna l'estratto conto
            await db["estratto_conto_movimenti"].update_one(
                {"id": bonifico.get("id")},
                {"$set": {
                    "riconciliato_salario": True,
                    "dipendente_nome": dip_trovato.get("nome", nome_trovato),
                    "categoria": "Stipendi",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Cerca/aggiorna prima nota salari corrispondente
            mese_bonifico = int(data[5:7]) if len(data) >= 7 else None
            anno_bonifico = int(data[:4]) if len(data) >= 4 else None
            
            if mese_bonifico and anno_bonifico:
                nome_orig = dip_trovato.get("nome", "")
                cognome_search = nome_orig.split()[0] if nome_orig else ""
                
                if cognome_search:
                    pns = await db["prima_nota_salari"].find_one({
                        "dipendente": {"$regex": cognome_search, "$options": "i"},
                        "anno": anno_bonifico,
                        "mese": mese_bonifico,
                        "tipo": "bonifico"
                    })
                    
                    if pns:
                        await db["prima_nota_salari"].update_one(
                            {"id": pns.get("id")},
                            {"$set": {
                                "riconciliato": True,
                                "estratto_conto_id": bonifico.get("id"),
                                "importo_bonifico": importo,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
            
            riconciliati += 1
        else:
            non_trovati.append({
                "data": data,
                "importo": importo,
                "descrizione": bonifico.get("descrizione", "")[:80]
            })
    
    return {
        "success": True,
        "riconciliati": riconciliati,
        "non_trovati_count": len(non_trovati),
        "non_trovati_sample": non_trovati[:10],
        "message": f"Riconciliati {riconciliati} bonifici stipendio"
    }


@router.get("/movimenti-stipendi")
@handle_errors
async def get_movimenti_stipendi(
    anno: Optional[int] = Query(None),
    solo_non_riconciliati: bool = Query(False)
) -> Dict[str, Any]:
    """
    Restituisce i movimenti dell'estratto conto che sembrano essere stipendi.
    """
    db = Database.get_db()
    
    query = {
        "descrizione": {"$regex": "VOSTRA DISPOSIZIONE|VS\\.DISP", "$options": "i"},
        "tipo": "uscita"
    }
    
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    if solo_non_riconciliati:
        query["$or"] = [
            {"riconciliato_salario": {"$exists": False}},
            {"riconciliato_salario": False}
        ]
    
    movimenti = await db["estratto_conto_movimenti"].find(query, {"_id": 0}).sort("data", -1).to_list(1000)
    
    # Raggruppa per dipendente se riconciliato
    per_dipendente = {}
    non_riconciliati = []
    
    for m in movimenti:
        if m.get("riconciliato_salario") and m.get("dipendente_nome"):
            nome = m.get("dipendente_nome")
            if nome not in per_dipendente:
                per_dipendente[nome] = {"count": 0, "totale": 0, "movimenti": []}
            per_dipendente[nome]["count"] += 1
            per_dipendente[nome]["totale"] += abs(m.get("importo", 0))
            if len(per_dipendente[nome]["movimenti"]) < 5:
                per_dipendente[nome]["movimenti"].append(m)
        else:
            non_riconciliati.append(m)
    
    return {
        "totale": len(movimenti),
        "riconciliati": len(movimenti) - len(non_riconciliati),
        "non_riconciliati": len(non_riconciliati),
        "per_dipendente": [
            {"nome": k, **v} for k, v in sorted(per_dipendente.items(), key=lambda x: x[1]["totale"], reverse=True)
        ],
        "non_riconciliati_sample": non_riconciliati[:20]
    }



@router.post("/ricategorizza-batch")
@handle_errors
async def ricategorizza_batch_movimenti() -> Dict[str, Any]:
    """Ricategorizza automaticamente i movimenti bancari in base a pattern noti."""
    db = Database.get_db()

    PATTERN_CATEGORIE = {
        "stipend": "stipendi",
        "salary": "stipendi",
        "inps": "contributi",
        "f24": "tributi",
        "erario": "tributi",
        "affitto": "affitto",
        "canone": "affitto",
        "enel": "utenze",
        "telecom": "utenze",
        "vodafone": "utenze",
        "assicuraz": "assicurazioni",
    }

    movimenti = await db["bank_movements"].find(
        {"$or": [{"categoria": None}, {"categoria": ""}, {"categoria": {"$exists": False}}]},
        {"_id": 0, "id": 1, "descrizione": 1, "causale": 1}
    ).to_list(5000)

    aggiornati = 0
    for m in movimenti:
        testo = ((m.get("descrizione") or "") + " " + (m.get("causale") or "")).lower()
        for pattern, cat in PATTERN_CATEGORIE.items():
            if pattern in testo:
                await db["bank_movements"].update_one(
                    {"id": m["id"]},
                    {"$set": {"categoria": cat, "auto_categorizzato": True}}
                )
                aggiornati += 1
                break

    return {"totale_analizzati": len(movimenti), "aggiornati": aggiornati}


@router.post("/ripara-versamenti-cassa")
@handle_errors
async def ripara_versamenti_cassa(anno: int = Query(None, description="Anno (opzionale). Se omesso ripara tutti gli anni")) -> Dict[str, Any]:
    """Materializza e riconcilia le due gambe dei trasferimenti di contante.

    La causale esplicita dell'estratto conto prova il trasferimento: un
    versamento genera uscita Cassa + entrata Banca, un prelievo il contrario.
    Gli ID del movimento EC e dell'operazione rendono l'operazione idempotente.
    """
    import uuid as _uuid
    db = Database.get_db()

    query: Dict[str, Any] = {}
    if anno:
        query["data"] = {"$regex": f"^{anno}"}

    movimenti = await db["estratto_conto_movimenti"].find(query, {"_id": 0}).to_list(100000)

    # Tutti i nomi storicamente usati per le gambe di versamenti/prelievi
    # (prima dell'unificazione del 17/07/2026 in "Versamento Banca" e
    # "Prelevamento Banca"): servono nel dedup per non duplicare mai le
    # registrazioni gia' esistenti, comprese quelle manuali dell'utente.
    CATEGORIE_VERSAMENTO_CASSA = ["Versamento", "Versamento Banca"]
    CATEGORIE_PRELIEVO = ["Prelievo", "Prelevamento Banca", "trasferimento_interno"]

    creati_cassa = 0
    creati_banca = 0
    gia_presenti_cassa = 0
    gia_presenti_banca = 0
    analizzati = 0
    prelievi_trovati = 0
    versamenti_senza_registrazione_cassa = 0
    ec_da_marcare: List[str] = []

    for mov in movimenti:
        desc = mov.get("descrizione_originale") or mov.get("descrizione") or ""
        versamento = is_versamento_contanti(desc)
        prelievo = not versamento and is_prelievo_contanti(desc)
        if not versamento and not prelievo:
            continue
        analizzati += 1
        if prelievo:
            prelievi_trovati += 1
        mid = mov.get("id")
        data = mov.get("data")
        importo = abs(mov.get("importo") or 0)
        if not data or not importo:
            continue
        now_iso = datetime.now(timezone.utc).isoformat()
        operation_id = f"trasferimento-contanti:{mid}"

        if versamento:
            # versamento: USCITA cassa (contante via dal cassetto),
            # ENTRATA banca (lo stesso denaro arriva sul conto)
            tipo_cassa, tipo_banca = "uscita", "entrata"
            cat = "Versamento Banca"
            desc_cassa = f"Versamento in banca - {desc[:100]}"
            desc_banca = f"Versamento contanti da cassa - {desc[:100]}"
            dedup_cat_cassa = CATEGORIE_VERSAMENTO_CASSA
            dedup_cat_banca = CATEGORIE_VERSAMENTO_CASSA + ["Ricavi - Deposito contanti"]
            tipo_riconciliazione = "versamento_contanti"
        else:
            # prelievo (regola utente 17/07/2026): ENTRATA in cassa
            # ("prelevamento banca") e USCITA in banca ("prelevamento
            # verso cassa")
            tipo_cassa, tipo_banca = "entrata", "uscita"
            cat = "Prelevamento Banca"
            desc_cassa = f"Prelevamento da banca - {desc[:100]}"
            desc_banca = f"Prelevamento verso cassa - {desc[:100]}"
            dedup_cat_cassa = CATEGORIE_PRELIEVO
            dedup_cat_banca = CATEGORIE_PRELIEVO
            tipo_riconciliazione = "prelievo_contanti"

        # Dedup SOLO contro righe che contano nei saldi: le copie legacy
        # con source escluso (es. estratto_conto_sync) sono invisibili in
        # Prima Nota — contarle come "già presente" lasciava il registro
        # SENZA la gamba del versamento (bug segnalato dall'utente
        # 17/07/2026: "in banca trovo i versamenti ma in cassa non trovo
        # l'uscita" — e viceversa: 11 entrate banca mai create).
        from app.routers.prima_nota_module.common import SOURCES_ESCLUSE

        # ── Gamba 1: CASSA ──
        esistente_cassa = await db["prima_nota_cassa"].find_one({
            "tipo": tipo_cassa,
            "source": {"$nin": SOURCES_ESCLUSE},
            "status": {"$nin": ["deleted", "archived"]},
            "$or": [
                {"estratto_conto_id": mid},
                {"data": data, "importo": importo,
                 "categoria": {"$in": dedup_cat_cassa}},
            ],
        })
        if esistente_cassa:
            gia_presenti_cassa += 1
            await db["prima_nota_cassa"].update_one(
                {"id": esistente_cassa["id"]},
                {"$set": {
                    "estratto_conto_id": mid,
                    "trasferimento_operation_id": operation_id,
                    "riconciliato": True,
                    "in_banca": True,
                    "updated_at": now_iso,
                }},
            )
        elif False and versamento:  # Compatibilita': ora la causale EC materializza entrambe le gambe.
            # L'estratto conto dimostra l'accredito in banca, non l'uscita
            # fisica del contante dalla cassa. La gamba cassa deve essere
            # registrata prima dall'utente e non può nascere da questo file.
            versamenti_senza_registrazione_cassa += 1
            continue
        else:
            await scrivi_movimento(db, "cassa", {
                "id": str(_uuid.uuid4()),
                "data": data,
                "tipo": tipo_cassa,
                "importo": importo,
                "descrizione": desc_cassa,
                "categoria": cat,
                "estratto_conto_id": mid,
                "trasferimento_operation_id": operation_id,
                "riconciliato": True,
                "in_banca": True,
                "source": "estratto_conto_auto_versamento_riparazione",
                "created_at": now_iso,
            })
            creati_cassa += 1

        # ── Gamba 2: BANCA ──
        esistente_banca = await db["prima_nota_banca"].find_one({
            "tipo": tipo_banca,
            "source": {"$nin": SOURCES_ESCLUSE},
            "status": {"$nin": ["deleted", "archived"]},
            "$or": [
                {"estratto_conto_id": mid},
                # gamba già creata dal sync generico all'import o a mano
                {"data": data, "importo": importo,
                 "categoria": {"$in": dedup_cat_banca}},
                {"data": data, "importo": importo, "source": "estratto_conto_auto"},
            ],
        })
        if esistente_banca:
            gia_presenti_banca += 1
            await db["prima_nota_banca"].update_one(
                {"id": esistente_banca["id"]},
                {"$set": {
                    "estratto_conto_id": mid,
                    "trasferimento_operation_id": operation_id,
                    "riconciliato": True,
                    "updated_at": now_iso,
                }},
            )
        else:
            await scrivi_movimento(db, "banca", {
                "id": str(_uuid.uuid4()),
                "data": data,
                "tipo": tipo_banca,
                "importo": importo,
                "descrizione": desc_banca,
                "categoria": cat,
                "estratto_conto_id": mid,
                "trasferimento_operation_id": operation_id,
                "riconciliato": True,
                "source": "estratto_conto_auto_versamento_riparazione",
                "created_at": now_iso,
            })
            creati_banca += 1

        ec_da_marcare.append((mid, tipo_riconciliazione))

    # Marca le righe EC come riconciliate: la doppia scrittura esiste,
    # nessun processo successivo deve ricrearla.
    for tipo_ric in ("versamento_contanti", "prelievo_contanti"):
        ids = [mid for mid, t in ec_da_marcare if t == tipo_ric]
        if ids:
            await db["estratto_conto_movimenti"].update_many(
                {"id": {"$in": ids}},
                {"$set": {"riconciliato": True, "tipo_riconciliazione": tipo_ric}},
            )

    # Conclude anche la riconciliazione delle fatture, ma soltanto quando
    # il movimento contiene riferimenti espliciti, i candidati sono univoci
    # e la somma quadra al centesimo. L'uguaglianza del solo importo non basta.
    from app.services.bank_payment_allocations import (
        reconcile_deterministic_invoice_allocations,
    )
    riconciliazione_fatture = await reconcile_deterministic_invoice_allocations(
        db, anno=anno,
    )

    return {
        "success": True,
        "anno": anno,
        "movimenti_versamento_trovati": analizzati - prelievi_trovati,
        "movimenti_prelievo_trovati": prelievi_trovati,
        "creati_cassa": creati_cassa,
        "creati_banca": creati_banca,
        "gia_presenti_cassa": gia_presenti_cassa,
        "gia_presenti_banca": gia_presenti_banca,
        "versamenti_senza_registrazione_cassa": versamenti_senza_registrazione_cassa,
        # compatibilità con la vecchia risposta
        "riparati": creati_cassa,
        "gia_presenti": gia_presenti_cassa,
        "riconciliazione_fatture": riconciliazione_fatture,
    }
