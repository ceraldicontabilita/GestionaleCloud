"""Export delle fatture SELEZIONATE in Excel o PDF (richiesta utente
16/07/2026: "permettimi di selezionare le fatture e di poterle scaricare
in pdf ed excel"). L'utente spunta le fatture in Archivio Fatture e scarica
un file con solo quelle: Excel per rielaborare, PDF come prospetto.
"""
from io import BytesIO
from typing import Any, Dict, List

from fastapi import Body, HTTPException
from fastapi.responses import StreamingResponse

from app.database import Database
from .crud import _normalizza_da_invoices

COLONNE = [
    ("Data", "data_documento"),
    ("Numero", "numero_documento"),
    ("Fornitore", "fornitore_ragione_sociale"),
    ("P.IVA", "fornitore_partita_iva"),
    ("Imponibile", "imponibile"),
    ("IVA", "iva"),
    ("Totale", "importo_totale"),
    ("Stato", "stato"),
    ("Metodo", "metodo_pagamento"),
]
_CAMPI_NUMERICI = {"imponibile", "iva", "importo_totale"}


async def _carica_fatture(ids: List[str]) -> List[Dict[str, Any]]:
    db = Database.get_db()
    docs = await db["invoices"].find(
        {"id": {"$in": ids}, "entity_status": {"$ne": "deleted"}, "status": {"$ne": "deleted"}},
        {"_id": 0},
    ).to_list(len(ids) + 1)
    fatture = [_normalizza_da_invoices(d) for d in docs]
    fatture.sort(key=lambda f: (f.get("data_documento") or "", f.get("numero_documento") or ""))
    return fatture


def _excel(fatture: List[Dict[str, Any]]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Fatture selezionate"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F2744")
    for col_idx, (label, _campo) in enumerate(COLONNE, start=1):
        cella = ws.cell(row=1, column=col_idx, value=label)
        cella.font = header_font
        cella.fill = header_fill

    for row_idx, f in enumerate(fatture, start=2):
        for col_idx, (_label, campo) in enumerate(COLONNE, start=1):
            valore = f.get(campo)
            if campo in _CAMPI_NUMERICI:
                valore = round(float(valore or 0), 2)
                cella = ws.cell(row=row_idx, column=col_idx, value=valore)
                cella.number_format = "#,##0.00"
            else:
                ws.cell(row=row_idx, column=col_idx, value=valore or "")

    riga_tot = len(fatture) + 2
    ws.cell(row=riga_tot, column=1, value="TOTALE").font = Font(bold=True)
    for col_idx, (_label, campo) in enumerate(COLONNE, start=1):
        if campo in _CAMPI_NUMERICI:
            cella = ws.cell(
                row=riga_tot, column=col_idx,
                value=round(sum(float(f.get(campo) or 0) for f in fatture), 2),
            )
            cella.font = Font(bold=True)
            cella.number_format = "#,##0.00"

    larghezze = [12, 14, 34, 14, 12, 10, 12, 12, 12]
    for i, w in enumerate(larghezze, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _eur(v) -> str:
    return f"{float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _pdf(fatture: List[Dict[str, Any]]) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm,
    )
    stili = getSampleStyleSheet()
    corpo = [
        Paragraph(f"Fatture selezionate ({len(fatture)})", stili["Title"]),
        Spacer(1, 6),
    ]

    intestazione = [label for label, _ in COLONNE]
    righe = [intestazione]
    for f in fatture:
        righe.append([
            f.get("data_documento") or "",
            str(f.get("numero_documento") or ""),
            (f.get("fornitore_ragione_sociale") or "")[:45],
            f.get("fornitore_partita_iva") or "",
            _eur(f.get("imponibile")),
            _eur(f.get("iva")),
            _eur(f.get("importo_totale")),
            f.get("stato") or "",
            f.get("metodo_pagamento") or "",
        ])
    righe.append([
        "", "", "TOTALE", "",
        _eur(sum(float(f.get("imponibile") or 0) for f in fatture)),
        _eur(sum(float(f.get("iva") or 0) for f in fatture)),
        _eur(sum(float(f.get("importo_totale") or 0) for f in fatture)),
        "", "",
    ])

    tab = Table(righe, repeatRows=1)
    tab.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f2744")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (4, 1), (6, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    corpo.append(tab)
    doc.build(corpo)
    buf.seek(0)
    return buf


async def export_fatture_selezionate(data: Dict[str, Any] = Body(...)) -> StreamingResponse:
    """POST /api/fatture-ricevute/export-selezione — body {ids: [...], formato: excel|pdf}."""
    ids = data.get("ids") or []
    formato = (data.get("formato") or "excel").lower()
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="Nessuna fattura selezionata")
    if len(ids) > 2000:
        raise HTTPException(status_code=400, detail="Troppe fatture selezionate (max 2000)")
    if formato not in ("excel", "pdf"):
        raise HTTPException(status_code=400, detail="Formato non valido: excel o pdf")

    fatture = await _carica_fatture([str(i) for i in ids])
    if not fatture:
        raise HTTPException(status_code=404, detail="Nessuna fattura trovata per gli id indicati")

    if formato == "excel":
        buf = _excel(fatture)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="fatture_selezionate.xlsx"'},
        )
    buf = _pdf(fatture)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="fatture_selezionate.pdf"'},
    )
