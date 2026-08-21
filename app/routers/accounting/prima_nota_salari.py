"""
Prima Nota Salari - Nuovo sistema di gestione stipendi.

Struttura dati:
- Dipendente
- Mese di competenza
- Anno
- Importo Busta (da file paghe.xlsx)
- Importo Bonifico (da file bonifici.xlsx)
- Saldo (differenza)
- Progressivo (riporto da mesi precedenti)
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body, Depends
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List, Optional
from datetime import date, datetime, timedelta, timezone
import uuid
import logging
import io
import hashlib
from decimal import Decimal, InvalidOperation

from app.database import Collections, Database
from app.services.salari_periodo import (
    filtro_periodo_prima_nota,
    periodo_ammesso_in_prima_nota,
)
from app.utils.dependencies import get_current_admin_user, get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()

# Mapping mesi italiano -> numero (include anche numeri come stringhe)
MESI_MAP = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
    "gen": 1, "feb": 2, "mar": 3, "apr": 4, "mag": 5, "giu": 6,
    "lug": 7, "ago": 8, "set": 9, "ott": 10, "nov": 11, "dic": 12,
    "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6,
    "7": 7, "8": 8, "9": 9, "10": 10, "11": 11, "12": 12,
    "01": 1, "02": 2, "03": 3, "04": 4, "05": 5, "06": 6,
    "07": 7, "08": 8, "09": 9,
    # Mensilità aggiuntive
    "tredicesima": 13, "quattordicesima": 14, "13": 13, "14": 14
}

MESI_NOMI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
             "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
             "Tredicesima", "Quattordicesima"]


def normalize_name(name: str) -> str:
    """Normalizza nome dipendente per matching."""
    if not name:
        return ""
    return " ".join(name.strip().upper().split())


def get_mese_numero(mese_str: str) -> int:
    """Converte nome mese o numero in numero."""
    if not mese_str:
        return 0
    # Prima prova a convertire direttamente in intero
    try:
        mese = int(float(mese_str))
        if 1 <= mese <= 12:
            return mese
    except (ValueError, TypeError):
        pass
    # Altrimenti cerca nel mapping
    return MESI_MAP.get(str(mese_str).lower().strip(), 0)


def _prima_colonna(
    colonne: List[str],
    nomi_esatti: List[str],
    parole: Optional[List[str]] = None,
    escluse: Optional[List[str]] = None,
) -> Optional[str]:
    """Trova una colonna senza dipendere dall'ordine del foglio Excel."""
    normalizzate = {str(colonna).strip().lower(): colonna for colonna in colonne}
    for nome in nomi_esatti:
        if nome in normalizzate:
            return normalizzate[nome]
    for colonna in colonne:
        testo = str(colonna).strip().lower()
        if escluse and any(parola in testo for parola in escluse):
            continue
        if parole and any(parola in testo for parola in parole):
            return colonna
    return None


def _importo_positivo(valore: Any) -> Optional[float]:
    """Converte importi Excel italiani; vuoti e zeri non sono movimenti."""
    if valore is None:
        return None
    try:
        import pandas as pd
        if pd.isna(valore):
            return None
    except (TypeError, ValueError):
        pass
    testo = str(valore).strip().replace("€", "").replace(" ", "")
    if not testo:
        return None
    if "," in testo:
        testo = testo.replace(".", "").replace(",", ".")
    try:
        importo = abs(Decimal(testo)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None
    return float(importo) if importo > 0 else None


def _data_excel_iso(valore: Any) -> Optional[str]:
    if valore is None:
        return None
    try:
        import pandas as pd
        if pd.isna(valore):
            return None
        data = pd.to_datetime(valore, dayfirst=True, errors="coerce")
        if pd.isna(data):
            return None
        return data.date().isoformat()
    except (TypeError, ValueError, OverflowError):
        return None


def _chiave_bonifico_excel(
    dipendente: str,
    anno: int,
    mese: int,
    data_bonifico: Optional[str],
    importo: float,
) -> str:
    base = "|".join([
        normalize_name(dipendente), str(anno), str(mese),
        data_bonifico or "", f"{importo:.2f}",
    ])
    return "bonifico_excel:" + hashlib.sha256(base.encode("utf-8")).hexdigest()


def _chiave_busta_excel(
    dipendente: str,
    anno: int,
    mese: int,
    importo_busta: float,
) -> str:
    """Chiave stabile per una riga che documenta solo il netto in busta."""
    base = "|".join([
        normalize_name(dipendente), str(anno), str(mese), f"{importo_busta:.2f}",
    ])
    return "busta_excel:" + hashlib.sha256(base.encode("utf-8")).hexdigest()


# ============== ENDPOINTS ==============

@router.get("/salari")
async def get_prima_nota_salari(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """
    Recupera la prima nota salari con tutti i campi:
    - dipendente, mese, anno
    - importo_busta, importo_bonifico, saldo, progressivo
    - riconciliato
    """
    db = Database.get_db()
    # Ripara in modo deterministico le righe storiche prive di nome usando
    # esclusivamente cedolino_id oppure CF+periodo.
    from app.services.bonifici_pdf_ingest import arricchisci_nomi_salari_da_cedolini
    await arricchisci_nomi_salari_da_cedolini(db)
    
    query = filtro_periodo_prima_nota()
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)

    # Espone soltanto un booleano leggero, mai il PDF base64 nella lista.
    query_pdf: Dict[str, Any] = {
        "pdf_data": {"$exists": True, "$nin": [None, ""]},
        **filtro_periodo_prima_nota(),
    }
    if anno:
        query_pdf["anno"] = anno
    if mese:
        query_pdf["mese"] = mese
    docs = await db["cedolini"].find(
        query_pdf,
        {
            "_id": 0, "id": 1, "codice_fiscale": 1, "mese": 1,
            "anno": 1, "tipo_cedolino": 1, "nome_dipendente": 1,
            "dipendente_id": 1,
        },
    ).to_list(5000)
    disponibili = {d.get("id") for d in docs if d.get("id")}
    disponibili_per_periodo = {
        (
            d.get("codice_fiscale"), d.get("mese"), d.get("anno"),
            d.get("tipo_cedolino") or "mensile",
        )
        for d in docs if d.get("codice_fiscale")
    }
    cedolini_per_id = {d.get("id"): d for d in docs if d.get("id")}
    cedolini_per_periodo = {
        (
            d.get("codice_fiscale"), d.get("mese"), d.get("anno"),
            d.get("tipo_cedolino") or "mensile",
        ): d
        for d in docs if d.get("codice_fiscale")
    }
    from app.services.stipendi_bonifici import riconciliazione_salario_verificata
    for salario in salari:
        chiave = (
            salario.get("codice_fiscale"), salario.get("mese"), salario.get("anno"),
            salario.get("tipo_cedolino") or "mensile",
        )
        cedolino = cedolini_per_id.get(salario.get("cedolino_id")) or cedolini_per_periodo.get(chiave)
        if cedolino and not (salario.get("dipendente_nome") or salario.get("dipendente")):
            salario["dipendente_nome"] = cedolino.get("nome_dipendente")
            salario["dipendente"] = cedolino.get("nome_dipendente")
            salario["dipendente_id"] = cedolino.get("dipendente_id")
        salario["cedolino_disponibile"] = (
            salario.get("cedolino_id") in disponibili or chiave in disponibili_per_periodo
        )
        stato_archiviato = salario.get("riconciliato") is True
        stato_verificato = await riconciliazione_salario_verificata(db, salario)
        salario["riconciliato"] = stato_verificato
        salario["riconciliazione_precedente_da_rivedere"] = (
            stato_archiviato and not stato_verificato
        )
        bonifico_ids = [
            str(value) for value in (salario.get("bonifico_documenti_ids") or [])
            if value
        ]
        salario["bonifico_documenti_count"] = len(set(bonifico_ids))
        salario["bonifico_documento_disponibile"] = bool(bonifico_ids)
    
    return salari


@router.get("/salari/{record_id}/cedolino-pdf")
async def get_cedolino_pdf(
    record_id: str,
    _current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Visualizza il PDF originale collegato alla riga del dipendente."""
    import base64

    db = Database.get_db()
    salario = await db["prima_nota_salari"].find_one(
        {"id": record_id},
        {
            "_id": 0, "cedolino_id": 1, "codice_fiscale": 1,
            "mese": 1, "anno": 1, "tipo_cedolino": 1,
        },
    )
    if not salario:
        raise HTTPException(status_code=404, detail="Riga salario non trovata")

    cedolino = None
    if salario.get("cedolino_id"):
        cedolino = await db["cedolini"].find_one(
            {"id": salario["cedolino_id"]},
            {"_id": 0, "pdf_data": 1},
        )
    if not cedolino or not cedolino.get("pdf_data"):
        from app.services.salari_unificati_v2 import _cedolino_identity_filter
        fallback_query = _cedolino_identity_filter(
            salario.get("codice_fiscale"),
            salario.get("mese"),
            salario.get("anno"),
            salario.get("tipo_cedolino") or "mensile",
        )
        fallback_query["pdf_data"] = {"$exists": True, "$nin": [None, ""]}
        cedolino = await db["cedolini"].find_one(
            fallback_query,
            {"_id": 0, "pdf_data": 1},
        )
    if not cedolino or not cedolino.get("pdf_data"):
        raise HTTPException(status_code=404, detail="PDF del cedolino non disponibile")

    try:
        pdf_bytes = base64.b64decode(cedolino["pdf_data"], validate=True)
    except Exception as exc:
        logger.warning("PDF cedolino non decodificabile per record %s: %s", record_id, exc)
        raise HTTPException(status_code=422, detail="PDF del cedolino non leggibile")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": "inline; filename=cedolino.pdf",
            "X-Content-Type-Options": "nosniff",
        },
    )


async def _leggi_pdf_allegato(file: UploadFile) -> bytes:
    """Legge un PDF con limiti espliciti, senza fidarsi del solo MIME type."""
    content = await file.read()
    if not content.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Il file selezionato non e' un PDF valido")
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Il PDF supera il limite di 20 MB")
    return content


def _nome_riga_salario(salario: Dict[str, Any]) -> str:
    return normalize_name(
        salario.get("dipendente_nome")
        or salario.get("dipendente")
        or salario.get("nome_dipendente")
        or ""
    )


@router.post("/salari/{record_id}/cedolino-pdf")
async def allega_cedolino_pdf(
    record_id: str,
    file: UploadFile = File(...),
    _current_user: Dict[str, Any] = Depends(get_current_admin_user),
) -> Dict[str, Any]:
    """Allega manualmente il cedolino alla riga scelta, senza alterare gli importi."""
    import base64

    db = Database.get_db()
    salario = await db["prima_nota_salari"].find_one({"id": record_id}, {"_id": 0})
    if not salario:
        raise HTTPException(status_code=404, detail="Riga salario non trovata")
    content = await _leggi_pdf_allegato(file)
    nome = _nome_riga_salario(salario)
    if not nome:
        raise HTTPException(
            status_code=422,
            detail="Indicare prima il dipendente della riga, poi allegare il cedolino",
        )

    cedolino = None
    if salario.get("cedolino_id"):
        cedolino = await db["cedolini"].find_one(
            {"id": salario["cedolino_id"]}, {"_id": 0, "id": 1}
        )
    if not cedolino and salario.get("codice_fiscale"):
        cedolino = await db["cedolini"].find_one(
            {
                "codice_fiscale": salario.get("codice_fiscale"),
                "mese": salario.get("mese"),
                "anno": salario.get("anno"),
            },
            {"_id": 0, "id": 1},
        )

    cedolino_id = (cedolino or {}).get("id") or salario.get("cedolino_id") or str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    await db["cedolini"].update_one(
        {"id": cedolino_id},
        {"$set": {
            "id": cedolino_id,
            "dipendente_id": salario.get("dipendente_id"),
            "nome_dipendente": nome,
            "codice_fiscale": salario.get("codice_fiscale"),
            "mese": salario.get("mese"),
            "anno": salario.get("anno"),
            "periodo": f"{int(salario.get('mese') or 0):02d}/{salario.get('anno')}",
            "filename": file.filename or "cedolino.pdf",
            "pdf_data": base64.b64encode(content).decode("ascii"),
            "pdf_disponibile": True,
            "source": "upload_manuale_pagina_salari",
            "updated_at": now,
        }},
        upsert=True,
    )
    await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": {"cedolino_id": cedolino_id, "updated_at": now}},
    )
    return {
        "success": True,
        "record_id": record_id,
        "cedolino_id": cedolino_id,
        "message": "Cedolino allegato; gli importi della riga non sono stati modificati",
    }


@router.post("/salari/{record_id}/bonifico-pdf")
async def allega_bonifico_pdf(
    record_id: str,
    file: UploadFile = File(...),
    _current_user: Dict[str, Any] = Depends(get_current_admin_user),
) -> Dict[str, Any]:
    """Allega il PDF scelto alla mensilita', ma attende ancora la prova bancaria."""
    db = Database.get_db()
    salario = await db["prima_nota_salari"].find_one({"id": record_id}, {"_id": 0})
    if not salario:
        raise HTTPException(status_code=404, detail="Riga salario non trovata")
    if not _nome_riga_salario(salario):
        raise HTTPException(
            status_code=422,
            detail="Indicare prima il dipendente della riga, poi allegare il bonifico",
        )

    content = await _leggi_pdf_allegato(file)
    from app.services.bonifici_pdf_ingest import importa_pdf_bonifico
    esito = await importa_pdf_bonifico(
        db,
        content,
        file.filename or "bonifico.pdf",
        source="upload_manuale_pagina_salari",
        auto_associa=False,
    )
    if esito.get("status") == "error" or not esito.get("transfer_id"):
        raise HTTPException(
            status_code=422,
            detail=esito.get("message") or "Il PDF bonifico non e' leggibile",
        )

    transfer_id = esito["transfer_id"]
    transfer = await db["bonifici_transfers"].find_one(
        {"id": transfer_id}, {"_id": 0}
    )
    if not transfer:
        raise HTTPException(status_code=422, detail="Bonifico archiviato ma non rileggibile")
    collegato_altrove = (
        transfer.get("salario_associato") is True
        and transfer.get("operazione_salario_id")
        and transfer.get("operazione_salario_id") != record_id
    )
    if collegato_altrove:
        raise HTTPException(
            status_code=409,
            detail="Questo PDF risulta gia' collegato a un'altra mensilita'",
        )

    ids = [str(value) for value in (salario.get("bonifico_documenti_ids") or []) if value]
    nuovo_collegamento = transfer_id not in ids
    if nuovo_collegamento:
        ids.append(transfer_id)

    importi_pdf = []
    for documento_id in dict.fromkeys(ids):
        documento = await db["bonifici_transfers"].find_one(
            {"id": documento_id}, {"_id": 0, "importo": 1}
        )
        if documento:
            importi_pdf.append(abs(float(documento.get("importo") or 0)))
    totale_pdf = round(sum(importi_pdf), 2)
    totale_precedente = round(float(salario.get("importo_bonifico_documentato") or 0), 2)
    # Il prospetto Excel e il PDF possono essere due prove dello stesso
    # acconto: non sommarli due volte. Conserviamo il totale piu' completo.
    totale_documentato = max(totale_precedente, totale_pdf)
    now = datetime.now(timezone.utc).isoformat()
    await db["bonifici_transfers"].update_one(
        {"id": transfer_id},
        {"$set": {
            "salario_associato": True,
            "operazione_salario_id": record_id,
            "dipendente_id": salario.get("dipendente_id"),
            "dipendente_nome": _nome_riga_salario(salario),
            "associazione_evidenze": ["selezione_manuale_utente"],
            "stato_riconciliazione": "documento_associato_attesa_banca",
            "updated_at": now,
        }},
    )
    await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": {
            "bonifico_documenti_ids": list(dict.fromkeys(ids)),
            "bonifico_documento_associato": True,
            "bonifico_documento_associato_manualmente": True,
            "importo_bonifico_documentato": totale_documentato,
            "stato_bonifico": "documentato_attesa_estratto_conto",
            "updated_at": now,
        }},
    )
    return {
        "success": True,
        "record_id": record_id,
        "transfer_id": transfer_id,
        "documento_duplicato": not nuovo_collegamento,
        "importo_bonifico_documentato": totale_documentato,
        "riconciliato": False,
        "message": "Bonifico allegato; riconciliazione bancaria ancora da verificare",
    }


@router.get("/salari/{record_id}/bonifico-pdf")
async def get_bonifico_pdf_salario(
    record_id: str,
    _current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Visualizza il primo PDF bonifico collegato alla riga selezionata."""
    import base64

    db = Database.get_db()
    salario = await db["prima_nota_salari"].find_one(
        {"id": record_id}, {"_id": 0, "bonifico_documenti_ids": 1}
    )
    if not salario:
        raise HTTPException(status_code=404, detail="Riga salario non trovata")
    documento = None
    for transfer_id in salario.get("bonifico_documenti_ids") or []:
        documento = await db["bonifici_transfers"].find_one(
            {"id": transfer_id, "pdf_data": {"$exists": True, "$nin": [None, ""]}},
            {"_id": 0, "pdf_data": 1, "source_file": 1},
        )
        if documento:
            break
    if not documento:
        raise HTTPException(status_code=404, detail="PDF del bonifico non disponibile")
    try:
        pdf_bytes = base64.b64decode(documento["pdf_data"], validate=True)
    except Exception as exc:
        logger.warning("PDF bonifico non decodificabile per record %s: %s", record_id, exc)
        raise HTTPException(status_code=422, detail="PDF del bonifico non leggibile")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": "inline; filename=bonifico.pdf",
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.get("/salari/riepilogo")
async def get_riepilogo_salari(
    anno: int = Query(...),
    mese: Optional[int] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo statistiche salari."""
    db = Database.get_db()
    
    query = {**filtro_periodo_prima_nota(), "anno": anno}
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(query, {"_id": 0}).to_list(5000)
    
    totale_buste = sum(s.get("importo_busta", 0) or 0 for s in salari)
    totale_bonifici = sum(s.get("importo_bonifico", 0) or 0 for s in salari)
    totale_saldo = sum(s.get("saldo", 0) or 0 for s in salari)
    
    # Conta dipendenti unici
    dipendenti_unici = len(set(s.get("dipendente", "") for s in salari))
    
    # Conta soltanto i riscontri sostenuti da un vero movimento bancario.
    from app.services.stipendi_bonifici import riconciliazione_salario_verificata
    stati = [await riconciliazione_salario_verificata(db, s) for s in salari]
    riconciliati = sum(1 for stato in stati if stato)
    
    return {
        "anno": anno,
        "mese": mese,
        "totale_records": len(salari),
        "dipendenti_unici": dipendenti_unici,
        "totale_buste": round(totale_buste, 2),
        "totale_bonifici": round(totale_bonifici, 2),
        "totale_saldo": round(totale_saldo, 2),
        "riconciliati": riconciliati,
        "da_riconciliare": len(salari) - riconciliati
    }


@router.post("/import-paghe")
async def import_paghe(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa file PAGHE (stipendi netti).
    Formato atteso: Dipendente | Mese | Anno | Stipendio Netto
    
    REGOLA: MAI AGGREGARE - Crea un record per ogni riga del file.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    
    for c in df.columns:
        if 'dipendente' in c or 'nome' in c or 'cognome' in c:
            col_dipendente = c
        elif 'mese' in c:
            col_mese = c
        elif 'anno' in c:
            col_anno = c
        elif any(x in c for x in ['stipendio', 'netto', 'busta', 'paga', 'retribuzione']):
            col_importo = c
    
    # Fallback
    if not col_importo:
        for c in df.columns:
            if 'importo' in c and 'bonifico' not in c and 'erogato' not in c:
                col_importo = c
                break
    
    if not all([col_dipendente, col_mese, col_anno, col_importo]):
        raise HTTPException(
            status_code=400, 
            detail=f"Colonne richieste non trovate. Trovate: {list(df.columns)}"
        )
    
    logger.info(f"IMPORT PAGHE - Righe nel file: {len(df)}")
    
    created = 0
    errors = []
    
    # CREA UN RECORD PER OGNI RIGA - MAI AGGREGARE
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            # Gestisci mese (può essere numero o data)
            mese_val = row[col_mese]
            if hasattr(mese_val, 'month'):
                mese = mese_val.month
            else:
                mese_str = str(mese_val).strip()
                mese = get_mese_numero(mese_str)
            if mese == 0:
                continue
            
            # Gestisci anno (può essere numero o data)
            anno_val = row[col_anno]
            if hasattr(anno_val, 'year'):
                anno = anno_val.year
            elif isinstance(anno_val, datetime):
                anno = anno_val.year
            else:
                anno = int(anno_val)
            if not periodo_ammesso_in_prima_nota(anno, mese):
                continue
            
            importo = float(row[col_importo]) if pd.notna(row[col_importo]) else 0
            
            # Crea nuovo record per OGNI riga
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= len(MESI_NOMI) else f"Mese {mese}",
                "importo_busta": round(importo, 2),
                "importo_bonifico": 0,
                "saldo": round(-importo, 2),
                "progressivo": 0,
                "riconciliato": False,
                "tipo": "busta",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record.copy())
            created += 1
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    logger.info(f"IMPORT PAGHE - Record creati: {created}")
    
    # Ricalcola progressivi
    await ricalcola_progressivi_tutti(db)
    
    return {
        "success": True,
        "message": "Import PAGHE completato",
        "created": created,
        "updated": 0,
        "righe_file": len(df),
        "errors": errors[:10] if errors else []
    }


@router.post("/import-bonifici")
async def import_bonifici(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa il prospetto storico BUSTE + BONIFICI come evidenza documentale.
    Formati supportati:
    - Dipendente | Mese | Anno | Importo Erogato
    - Dipendente | Mese (data completa) | Importo Erogato
    
    REGOLA: MAI AGGREGARE. IMPORTO BUSTA e IMPORTO ACCONTO hanno significati
    distinti e vengono conservati entrambi. L'import non dichiara una
    riconciliazione bancaria: l'acconto resta documentato ma non verificato
    finche' non viene trovato nell'estratto conto.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne (rimuovi spazi extra)
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    logger.info(f"IMPORT BONIFICI - Colonne trovate: {list(df.columns)}")
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    col_importo_busta = None
    col_data_bonifico = None

    colonne = list(df.columns)
    col_dipendente = _prima_colonna(
        colonne,
        ["dipendente", "nome dipendente", "beneficiario"],
        ["dipendente", "beneficiario", "nome", "cognome"],
    )
    col_mese = _prima_colonna(
        colonne, ["mese", "mese competenza", "competenza"], ["mese", "competenza"]
    )
    col_anno = _prima_colonna(colonne, ["anno", "anno competenza"], ["anno"])
    # Nel foglio storico IMPORTO ACCONTO e' il pagamento; IMPORTO BUSTA e'
    # il netto del cedolino. Sono entrambi importanti, ma non intercambiabili.
    col_importo = _prima_colonna(
        colonne,
        [
            "importo bonifico", "importo acconto", "importo erogato",
            "importo pagato", "importo versato", "accredito",
        ],
        ["bonifico", "acconto", "erogato", "pagato", "versato", "accredito"],
        ["busta", "stipendio netto", "netto cedolino"],
    )
    col_importo_busta = _prima_colonna(
        colonne,
        ["importo busta", "stipendio netto", "netto cedolino", "netto in busta"],
        ["importo busta", "stipendio netto", "netto cedolino", "netto in busta"],
        ["bonifico", "acconto", "erogato", "pagato", "versato", "accredito"],
    )
    col_data_bonifico = _prima_colonna(
        colonne,
        ["data bonifico", "data acconto", "data pagamento", "data erogazione"],
        ["data bonifico", "data acconto", "data pagamento", "data erogazione"],
    )
    
    if not col_dipendente or not col_mese or not (col_importo or col_importo_busta):
        raise HTTPException(
            status_code=400, 
            detail=(
                f"Colonne richieste non trovate. Trovate: {list(df.columns)}. "
                "Servono dipendente, mese/competenza e almeno una colonna tra "
                "IMPORTO BUSTA e IMPORTO ACCONTO/BONIFICO."
            )
        )
    
    logger.info(f"IMPORT BONIFICI - Righe nel file: {len(df)}")
    logger.info(
        "IMPORT BONIFICI - Mapping: dip=%s, mese=%s, anno=%s, busta=%s, bonifico=%s",
        col_dipendente, col_mese, col_anno, col_importo_busta, col_importo,
    )
    
    created = 0
    updated = 0
    duplicates = 0
    skipped = 0
    totale_documentato = 0.0
    totale_buste_documentato = 0.0
    righe_con_busta = 0
    errors = []
    chiavi_nel_file = set()
    
    # CREA UN RECORD PER OGNI RIGA - MAI AGGREGARE
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            # Gestisci mese e anno di competenza (non la data del bonifico).
            mese_val = row[col_mese]
            
            # Se è una data completa, estrai mese e anno
            if hasattr(mese_val, 'month') and hasattr(mese_val, 'year'):
                mese = mese_val.month
                anno = mese_val.year
            elif isinstance(mese_val, str) and '-' in mese_val:
                # Formato "2021-06-21" o simile
                try:
                    from dateutil import parser
                    dt = parser.parse(mese_val)
                    mese = dt.month
                    anno = dt.year
                except Exception:
                    mese = get_mese_numero(str(mese_val))
                    anno = int(row[col_anno]) if col_anno else None
            else:
                mese_str = str(mese_val).strip()
                mese = get_mese_numero(mese_str)
                # Anno da colonna separata o default
                if col_anno:
                    anno_val = row[col_anno]
                    if hasattr(anno_val, 'year'):
                        anno = anno_val.year
                    else:
                        anno = int(anno_val)
                else:
                    anno = None

            data_bonifico = _data_excel_iso(row[col_data_bonifico]) if col_data_bonifico else None
            if not anno and data_bonifico:
                anno = int(data_bonifico[:4])
            
            if mese == 0 or mese > 14 or not anno:
                errors.append(f"Riga {idx + 2}: mese non valido ({mese_val})")
                continue
            if not periodo_ammesso_in_prima_nota(anno, mese):
                skipped += 1
                continue

            importo = _importo_positivo(row[col_importo]) if col_importo else None
            importo_busta = (
                _importo_positivo(row[col_importo_busta]) if col_importo_busta else None
            )
            if importo is None and importo_busta is None:
                skipped += 1
                continue

            if importo is not None:
                # Mantiene la chiave della prima versione dell'importatore: una
                # seconda importazione aggiorna la busta sulle righe gia' create.
                import_key = _chiave_bonifico_excel(
                    dipendente, int(anno), mese, data_bonifico, importo
                )
            else:
                import_key = _chiave_busta_excel(
                    dipendente, int(anno), mese, importo_busta
                )
            if import_key in chiavi_nel_file:
                duplicates += 1
                continue
            chiavi_nel_file.add(import_key)

            # Totali del contenuto unico del file, indipendenti dal fatto che
            # la riga venga creata, aggiornata o risulti gia' presente.
            totale_documentato += importo or 0
            totale_buste_documentato += importo_busta or 0
            if importo_busta is not None:
                righe_con_busta += 1

            esistente = await db["prima_nota_salari"].find_one({"import_key": import_key})
            if esistente:
                busta_esistente = _importo_positivo(esistente.get("importo_busta"))
                if importo_busta is not None and busta_esistente != importo_busta:
                    await db["prima_nota_salari"].update_one(
                        {"import_key": import_key},
                        {"$set": {
                            "importo_busta": round(importo_busta, 2),
                            "importo_busta_documentato": round(importo_busta, 2),
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        }},
                    )
                    updated += 1
                else:
                    duplicates += 1
                continue
            
            # Bonifico documentato, non ancora verificato contro l'estratto conto.
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
                "importo_busta": round(importo_busta or 0, 2),
                "importo_busta_documentato": round(importo_busta or 0, 2),
                "importo_bonifico": 0,
                "importo_bonifico_documentato": round(importo or 0, 2),
                "data_bonifico_documentata": data_bonifico,
                "saldo": 0,
                "progressivo": 0,
                "riconciliato": False,
                "tipo": "prospetto_salario_documentato",
                "source": "excel_bonifici_storici",
                "source_filename": file.filename or "bonifici.xlsx",
                "source_row": idx + 2,
                "import_key": import_key,
                "descrizione": "Busta e bonifico storici da confrontare con cedolino ed estratto conto",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record.copy())
            created += 1
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    logger.info(f"IMPORT BONIFICI - Record creati: {created}")
    
    return {
        "success": True,
        "message": "Import BUSTE e BONIFICI documentati completato; confronti non ancora eseguiti",
        "created": created,
        "updated": updated,
        "duplicates": duplicates,
        "skipped": skipped,
        "totale_documentato": round(totale_documentato, 2),
        "totale_buste_documentato": round(totale_buste_documentato, 2),
        "righe_con_busta": righe_con_busta,
        "riconciliati": 0,
        "righe_file": len(df),
        "errors": errors[:10] if errors else []
    }


@router.post("/import-salari-verificati")
async def import_salari_verificati(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Importa netti cedolino verificati senza caricare pandas/openpyxl.

    Il payload e' idempotente per dipendente, periodo, importo e sorgente. Ogni
    riga resta una busta dovuta e non viene mai trasformata in pagamento.
    """
    db = Database.get_db()
    righe = data.get("righe") or []
    if not isinstance(righe, list) or not righe:
        raise HTTPException(status_code=400, detail="righe deve essere una lista non vuota")
    if len(righe) > 5000:
        raise HTTPException(status_code=413, detail="Massimo 5000 righe per richiesta")

    created = 0
    duplicates = 0
    skipped = 0
    errors: List[str] = []
    totale = 0.0
    nuovi_record: List[Dict[str, Any]] = []

    for idx, row in enumerate(righe, 1):
        try:
            if not isinstance(row, dict):
                raise ValueError("riga non valida")
            dipendente = normalize_name(str(row.get("employee") or row.get("dipendente") or ""))
            anno = int(row.get("year") or row.get("anno"))
            mese = int(row.get("month") or row.get("mese"))
            importo = _importo_positivo(row.get("net_amount", row.get("importo_busta")))
            status = str(row.get("status") or "NETTO_VERIFICATO_DA_CEDOLINO").strip().upper()
            source_url = str(row.get("source") or "").strip()

            if not dipendente or dipendente == "NAN":
                raise ValueError("dipendente mancante")
            if status != "NETTO_VERIFICATO_DA_CEDOLINO":
                skipped += 1
                continue
            if importo is None or importo <= 0:
                raise ValueError("netto mancante o non positivo")
            if not periodo_ammesso_in_prima_nota(anno, mese):
                skipped += 1
                continue

            import_key = _chiave_busta_excel(dipendente, anno, mese, importo)
            existing = await db["prima_nota_salari"].find_one(
                {"import_key": import_key}, {"_id": 0}
            )
            if existing:
                if (
                    existing.get("source") != "indice_cedolini_drive"
                    or (source_url and existing.get("source_url") != source_url)
                ):
                    await db["prima_nota_salari"].update_one(
                        {"id": existing["id"]},
                        {"$set": {
                            "source": "indice_cedolini_drive",
                            "source_url": source_url or existing.get("source_url"),
                            "document_status": status,
                            "payment_status": "DA_ASSEGNARE",
                            "importo_busta_documentato": round(importo, 2),
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        }},
                    )
                duplicates += 1
                continue

            now = datetime.now(timezone.utc).isoformat()
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1],
                "importo_busta": round(importo, 2),
                "importo_busta_documentato": round(importo, 2),
                "importo_bonifico": 0,
                "importo_bonifico_documentato": 0,
                "saldo": round(-importo, 2),
                "progressivo": 0,
                "riconciliato": False,
                "tipo": "busta",
                "source": "indice_cedolini_drive",
                "source_url": source_url or None,
                "document_status": status,
                "payment_status": "DA_ASSEGNARE",
                "import_key": import_key,
                "descrizione": "Netto verificato dal cedolino; pagamento da assegnare",
                "created_at": now,
                "updated_at": now,
            }
            nuovi_record.append(new_record)
        except Exception as exc:
            errors.append(f"Riga {idx}: {exc}")

    if nuovi_record:
        await db["prima_nota_salari"].insert_many(
            [record.copy() for record in nuovi_record]
        )
        created = len(nuovi_record)
        totale = sum(float(record["importo_busta"]) for record in nuovi_record)

    return {
        "success": not errors,
        "created": created,
        "duplicates": duplicates,
        "skipped": skipped,
        "righe_file": len(righe),
        "totale_importato": round(totale, 2),
        "progressivi_ricalcolati": False,
        "errors": errors[:20],
    }


async def ricalcola_progressivi_tutti(db, anno_inizio: int = None, dipendente_filtro: str = None, anni_esclusi: List[int] = None):
    """
    Ricalcola saldi e progressivi per tutti i dipendenti o uno specifico.
    
    Per ogni record:
    - Saldo = importo_bonifico - importo_busta (della singola riga)
    - Progressivo = Somma cumulativa di tutti i saldi dei record NON esclusi
    
    I record sono ordinati per anno, mese e data di creazione.
    Gli anni in anni_esclusi vengono completamente ignorati nel calcolo del progressivo.
    I record con vincolo=True non vengono modificati.
    """
    if anni_esclusi is None:
        anni_esclusi = []
    
    # Ottieni dipendenti da processare
    if dipendente_filtro:
        dipendenti = [dipendente_filtro]
    else:
        dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    
    for dipendente in dipendenti:
        # Ordina per anno, mese e data creazione (dal più vecchio al più recente)
        records = await db["prima_nota_salari"].find(
            {"dipendente": dipendente}
        ).sort([("anno", 1), ("mese", 1), ("created_at", 1)]).to_list(5000)
        
        # Progressivo parte da 0
        progressivo = 0
        
        for record in records:
            # Skip record con vincolo
            if record.get("vincolo"):
                continue
            
            # Saldo della singola riga: Bonifico - Busta
            importo_busta = record.get("importo_busta", 0) or 0
            importo_bonifico = record.get("importo_bonifico", 0) or 0
            saldo = importo_bonifico - importo_busta
            
            anno_record = record.get("anno", 0)
            
            # Se l'anno è escluso, imposta progressivo a 0 e NON lo aggiunge al cumulo
            if anno_record in anni_esclusi:
                prog_value = 0
                # NON aggiungiamo saldo al progressivo cumulativo
            elif anno_inizio is None or anno_record >= anno_inizio:
                # Aggiungi il saldo al progressivo cumulativo
                progressivo += saldo
                prog_value = round(progressivo, 2)
            else:
                prog_value = 0
            
            # Aggiorna il record con saldo e progressivo
            await db["prima_nota_salari"].update_one(
                {"_id": record["_id"]},
                {"$set": {
                    "saldo": round(saldo, 2),
                    "progressivo": prog_value
                }}
            )
    
    return True  # Conferma che il ricalcolo è avvenuto


@router.post("/ricalcola-progressivi")
async def ricalcola_progressivi(
    anno_inizio: Optional[int] = Query(None, description="Anno da cui iniziare il calcolo del progressivo (es. 2023)"),
    dipendente: Optional[str] = Query(None, description="Nome dipendente specifico (opzionale)"),
    force_reset: bool = Query(False, description="Forza il reset dei progressivi prima del ricalcolo"),
    anni_esclusi: Optional[str] = Query(None, description="Anni da escludere (separati da virgola, es. '2018,2019,2020')")
) -> Dict[str, Any]:
    """
    Ricalcola i progressivi per uno o tutti i dipendenti.
    Se anno_inizio è specificato, il progressivo parte da 0 a gennaio di quell'anno.
    Se dipendente è specificato, ricalcola solo per quel dipendente.
    Se force_reset è True, prima azzera tutti i progressivi e poi li ricalcola.
    Se anni_esclusi è specificato, quei record vengono ignorati nel calcolo.
    """
    db = Database.get_db()
    
    # Parse anni esclusi
    anni_esclusi_list = []
    if anni_esclusi:
        try:
            anni_esclusi_list = [int(a.strip()) for a in anni_esclusi.split(',') if a.strip()]
        except Exception:
            pass
    
    # Se force_reset, prima azzera i progressivi
    if force_reset:
        reset_query = {}
        if dipendente:
            reset_query["dipendente"] = dipendente
        await db["prima_nota_salari"].update_many(
            reset_query,
            {"$set": {"progressivo": 0, "saldo": 0}}
        )
    
    # Ricalcola
    await ricalcola_progressivi_tutti(db, anno_inizio, dipendente, anni_esclusi_list)
    
    return {
        "message": f"Progressivi ricalcolati{f' dal {anno_inizio}' if anno_inizio else ''}{f' per {dipendente}' if dipendente else ''}{f' (esclusi: {anni_esclusi})' if anni_esclusi else ''}",
        "anno_inizio": anno_inizio,
        "dipendente": dipendente,
        "force_reset": force_reset,
        "anni_esclusi": anni_esclusi_list
    }


@router.post("/salari/aggiustamento")
async def aggiungi_aggiustamento(
    data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Aggiunge una riga di aggiustamento per allineare il saldo con il commercialista.
    L'importo positivo aumenta il saldo (va nel bonifico), negativo lo diminuisce (va nella busta).
    """
    db = Database.get_db()
    
    dipendente = data.get("dipendente", "").strip().upper()
    anno = int(data.get("anno", datetime.now(timezone.utc).year))
    mese = int(data.get("mese", datetime.now(timezone.utc).month))
    importo_busta = float(data.get("importo_busta", 0))
    importo_bonifico = float(data.get("importo_bonifico", 0))
    descrizione = data.get("descrizione", "Aggiustamento saldo")
    
    if not dipendente:
        raise HTTPException(status_code=400, detail="Dipendente obbligatorio")
    if not periodo_ammesso_in_prima_nota(anno, mese):
        raise HTTPException(
            status_code=422,
            detail="La prima nota salari ammette soltanto dicembre 2025 e i mesi dal 2026 fino a oggi",
        )
    
    # Crea il record di aggiustamento
    new_record = {
        "id": str(uuid.uuid4()),
        "dipendente": dipendente,
        "anno": anno,
        "mese": mese,
        "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
        "importo_busta": round(importo_busta, 2),
        "importo_bonifico": round(importo_bonifico, 2),
        "saldo": round(importo_bonifico - importo_busta, 2),
        "progressivo": 0,  # Verrà ricalcolato
        "riconciliato": False,
        "tipo": "aggiustamento",
        "descrizione": descrizione,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["prima_nota_salari"].insert_one(new_record.copy())
    
    # Ricalcola i progressivi per questo dipendente
    await ricalcola_progressivi_tutti(db, None, dipendente)
    
    return {
        "success": True,
        "message": f"Aggiustamento inserito per {dipendente}",
        "record_id": new_record["id"]
    }


@router.get("/dipendenti-lista")
async def get_dipendenti_lista() -> List[str]:
    """Nomi noti dall'anagrafica dipendenti e dallo storico salari."""
    db = Database.get_db()
    storico = await db["prima_nota_salari"].distinct(
        "dipendente", filtro_periodo_prima_nota()
    )
    anagrafica = await db[Collections.EMPLOYEES].find(
        {
            "attivo": {"$ne": False},
            "merged_into": {"$exists": False},
        },
        {
            "_id": 0,
            "nome": 1,
            "cognome": 1,
            "nome_completo": 1,
        },
    ).to_list(5000)

    def chiave_identita(nome: str) -> str:
        # Cognome Nome e Nome Cognome devono indicare la stessa anagrafica.
        return "|".join(sorted(normalize_name(nome).split()))

    nomi_per_chiave: Dict[str, str] = {}
    for dipendente in anagrafica:
        nome = (
            dipendente.get("nome_completo")
            or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}".strip()
        )
        if nome and normalize_name(nome):
            nomi_per_chiave[chiave_identita(nome)] = str(nome).strip()
    for nome in storico:
        if nome and normalize_name(nome):
            nomi_per_chiave.setdefault(chiave_identita(nome), str(nome).strip())
    return sorted(nomi_per_chiave.values(), key=normalize_name)


@router.delete("/salari/reset")
async def reset_prima_nota_salari(
    tipo: Optional[str] = Query(None, description="Tipo di record da eliminare: 'busta', 'bonifico', 'aggiustamento' o None per tutti"),
    admin_user: Dict[str, Any] = Depends(get_current_admin_user),
) -> Dict[str, Any]:
    """
    Elimina i record della prima nota salari.
    - tipo=None: elimina TUTTI i record
    - tipo='busta': elimina solo i record delle paghe
    - tipo='bonifico': elimina solo i record dei bonifici
    - tipo='aggiustamento': elimina solo i record di aggiustamento
    """
    db = Database.get_db()
    
    query = {}
    if tipo:
        query["tipo"] = tipo
    
    result = await db["prima_nota_salari"].delete_many(query)
    
    tipo_desc = tipo if tipo else "tutti"
    return {
        "message": f"Eliminati {result.deleted_count} record ({tipo_desc})",
        "deleted_count": result.deleted_count,
        "tipo_eliminato": tipo_desc
    }


@router.delete("/pulisci-righe-vuote")
async def pulisci_righe_vuote() -> Dict[str, Any]:
    """
    Elimina i record "vuoti" dalla prima_nota_salari, cioè quelli dove SIA
    `importo_busta` SIA `importo_bonifico` sono 0, null o mancanti.

    Usato dal pulsante "Pulisci righe vuote" della UI PrimaNotaSalariTab per
    rimuovere residui di import falliti o placeholder orfani. Non tocca i
    record che hanno almeno uno dei due importi valorizzati.

    Returns:
        {"righe_eliminate": int, "message": str}
    """
    db = Database.get_db()

    # Una riga è "vuota" se importo_busta è None/0 E importo_bonifico è None/0
    query = {
        "$and": [
            {"$or": [
                {"importo_busta": {"$in": [None, 0, 0.0]}},
                {"importo_busta": {"$exists": False}},
            ]},
            {"$or": [
                {"importo_bonifico": {"$in": [None, 0, 0.0]}},
                {"importo_bonifico": {"$exists": False}},
            ]},
        ]
    }

    result = await db["prima_nota_salari"].delete_many(query)
    return {
        "righe_eliminate": result.deleted_count,
        "message": f"Eliminate {result.deleted_count} righe vuote da prima_nota_salari",
    }


@router.post("/consolida-record")
async def consolida_record() -> Dict[str, Any]:
    """
    Consolida i record esistenti: unisce buste e bonifici dello stesso mese/dipendente in un'unica riga.
    Utile dopo import separati o per pulire dati duplicati.
    """
    db = Database.get_db()
    
    # Ottieni tutti i record
    filtro_contabile = filtro_periodo_prima_nota()
    all_records = await db["prima_nota_salari"].find(
        filtro_contabile, {"_id": 0}
    ).to_list(10000)
    
    # Raggruppa per dipendente/anno/mese
    grouped = {}
    for r in all_records:
        key = (r.get("dipendente"), r.get("anno"), r.get("mese"))
        if key not in grouped:
            grouped[key] = {"busta": 0, "bonifico": 0, "records": []}
        grouped[key]["busta"] += r.get("importo_busta") or 0
        grouped[key]["bonifico"] += r.get("importo_bonifico") or 0
        grouped[key]["records"].append(r)
    
    # Conta quanti gruppi hanno più di un record
    duplicates = sum(1 for v in grouped.values() if len(v["records"]) > 1)
    
    if duplicates == 0:
        return {"message": "Nessun record da consolidare", "duplicates": 0}
    
    # Elimina tutti e ricrea consolidati
    await db["prima_nota_salari"].delete_many(filtro_contabile)
    
    created = 0
    for (dipendente, anno, mese), data in grouped.items():
        if not dipendente or not anno or not mese:
            continue
        if not periodo_ammesso_in_prima_nota(anno, mese):
            continue
        
        importo_busta = round(data["busta"], 2)
        importo_bonifico = round(data["bonifico"], 2)
        saldo = round(importo_bonifico - importo_busta, 2)
        
        new_record = {
            "id": str(uuid.uuid4()),
            "dipendente": dipendente,
            "anno": anno,
            "mese": mese,
            "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
            "importo_busta": importo_busta,
            "importo_bonifico": importo_bonifico,
            "saldo": saldo,
            "progressivo": 0,
            "riconciliato": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db["prima_nota_salari"].insert_one(new_record.copy())
        created += 1
    
    # Ricalcola progressivi
    await ricalcola_progressivi_tutti(db)
    
    return {
        "message": "Consolidamento completato",
        "record_originali": len(all_records),
        "record_consolidati": created,
        "duplicati_risolti": duplicates
    }


@router.delete("/salari/{record_id}")
async def delete_salario(record_id: str) -> Dict[str, str]:
    """Elimina un singolo record."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Record eliminato"}


@router.put("/salari/{record_id}")
async def update_salario(
    record_id: str,
    data: Dict[str, Any] = Body(...)
) -> Dict[str, str]:
    """Aggiorna un record della prima nota salari."""
    db = Database.get_db()
    
    # Calcola saldo: Bonifico - Busta
    importo_busta = float(data.get("importo_busta", 0))
    importo_bonifico = float(data.get("importo_bonifico", 0))
    saldo = importo_bonifico - importo_busta
    
    # Aggiorna il mese_nome se cambia il mese
    mese = int(data.get("mese", 1))
    mese_nome = MESI_NOMI[mese - 1] if 1 <= mese <= 12 else ""
    
    update_data = {
        "dipendente": data.get("dipendente", ""),
        "anno": int(data.get("anno", 2025)),
        "mese": mese,
        "mese_nome": mese_nome,
        "importo_busta": importo_busta,
        "importo_bonifico": importo_bonifico,
        "saldo": round(saldo, 2),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Aggiungi vincolo se presente
    if "vincolo" in data:
        update_data["vincolo"] = bool(data.get("vincolo"))
    
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Record aggiornato"}


@router.put("/salari/{record_id}/riconcilia")
async def riconcilia_salario(
    record_id: str,
    riconciliato: bool = Query(True)
) -> Dict[str, str]:
    """Riconcilia soltanto se il movimento bancario certo e' disponibile."""
    db = Database.get_db()
    if riconciliato:
        from app.services.stipendi_bonifici import associa_bonifici_stipendi
        # Una vecchia etichetta non verificata non deve impedire il nuovo
        # controllo deterministico.
        await db["prima_nota_salari"].update_one(
            {"id": record_id}, {"$set": {"riconciliato": False}}
        )
        esito = await associa_bonifici_stipendi(db, stipendio_id=record_id)
        if not esito.get("bonifici_associati"):
            raise HTTPException(
                status_code=409,
                detail="Nessun movimento bancario univoco con nome completo e importo esatto",
            )
        return {"message": "Riconciliazione bancaria verificata"}
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": {
            "riconciliato": False,
            "data_riconciliazione": None,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Stato riconciliazione aggiornato"}


@router.get("/export-excel")
async def export_prima_nota_salari_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None)
):
    """Esporta prima nota salari in Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    from app.services.stipendi_bonifici import riconciliazione_salario_verificata
    for salario in salari:
        salario["riconciliato"] = await riconciliazione_salario_verificata(db, salario)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prima Nota Salari"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    
    # Headers
    headers = ["Dipendente", "Anno", "Mese", "Importo Busta", "Importo Bonifico", "Saldo", "Progressivo", "Riconciliato"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, s in enumerate(salari, 2):
        ws.cell(row=row_num, column=1, value=s.get("dipendente", ""))
        ws.cell(row=row_num, column=2, value=s.get("anno", ""))
        ws.cell(row=row_num, column=3, value=s.get("mese_nome", ""))
        ws.cell(row=row_num, column=4, value=s.get("importo_busta", 0))
        ws.cell(row=row_num, column=5, value=s.get("importo_bonifico", 0))
        ws.cell(row=row_num, column=6, value=s.get("saldo", 0))
        ws.cell(row=row_num, column=7, value=s.get("progressivo", 0))
        ws.cell(row=row_num, column=8, value="Sì" if s.get("riconciliato") else "No")
    
    # Larghezze
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "prima_nota_salari"
    if anno:
        filename += f"_{anno}"
    if mese:
        filename += f"_{mese:02d}"
    filename += ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _appdipendenti_range(
    anno: Optional[int],
    mese: Optional[int],
    data_da: Optional[date],
    data_a: Optional[date],
) -> tuple[date, date]:
    if data_da and data_a:
        return data_da, data_a
    if anno and mese:
        first = date(anno, mese, 1)
        last = date(anno + 1, 1, 1) if mese == 12 else date(anno, mese + 1, 1)
        return first, last.replace(day=1) - timedelta(days=1)
    if anno:
        return date(anno, 1, 1), date(anno, 12, 31)
    return date(2000, 1, 1), date(2100, 12, 31)


@router.get("/export-appdipendenti/preview")
async def preview_export_appdipendenti(
    anno: Optional[int] = Query(None, ge=2000, le=2100),
    mese: Optional[int] = Query(None, ge=1, le=12),
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
):
    """Riepilogo dell'export senza modificare il database."""
    from app.services.appdipendenti_export import build_export_from_db

    data_inizio, data_fine = _appdipendenti_range(anno, mese, data_da, data_a)
    _, summary = await build_export_from_db(Database.get_db(), data_inizio, data_fine)
    return summary


@router.get("/export-appdipendenti/download")
async def download_export_appdipendenti(
    anno: Optional[int] = Query(None, ge=2000, le=2100),
    mese: Optional[int] = Query(None, ge=1, le=12),
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
):
    """Scarica ZIP con PDF originali e i due workbook AppDipendenti."""
    from app.services.appdipendenti_export import build_export_from_db

    data_inizio, data_fine = _appdipendenti_range(anno, mese, data_da, data_a)
    content, _ = await build_export_from_db(Database.get_db(), data_inizio, data_fine)
    filename = f"export_appdipendenti_{data_inizio.isoformat()}_{data_fine.isoformat()}.zip"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )
